import os
import argparse
from pathlib import Path
import subprocess
import sys
import shutil
import shlex
import geopandas as gpd
import numpy as np
import pandas as pd
from osgeo import gdal, ogr, osr, gdalconst
from sklearn.metrics import confusion_matrix, precision_recall_fscore_support
import openpyxl
from openpyxl.styles import Font
from concurrent.futures import ThreadPoolExecutor
import threading
import math

# Try importing SAM
try:
    import torch
    from segment_anything import sam_model_registry, SamAutomaticMaskGenerator
    HAS_SAM = True
except ImportError:
    HAS_SAM = False
    print("WARNING: segment-anything or torch not found. SAM segmentation will fail if called.")

# Base Paths
base_dir = Path("D:/AIML_CropMapper_Cloud/workingDir")
aux_dir = Path("D:/AIML_CropMapper_Cloud/auxiliary_files")
otb_dir = Path("D:/AIML_CropMapper_Cloud/bin/OTB-6.2.0-Win64")

TOTAL_STAGES = 11

# Crop aggregation mapping for the Netherlands (NL) to reduce semantic confusion
CROP_AGGREGATION_NL = {
    2: 5,   # Clover -> Grassland
    7: 5,   # Lucerne -> Grassland
}


class ProcessingPipeline:
    def __init__(self, track, mask_variant='3class'):
        self.track = track
        self.mask_variant = mask_variant

        if '/' in track or '\\' in track:
            normalized_track = track.replace('\\', '/')
            self.country = normalized_track.split('/')[0].upper()
        elif len(track) == 2:
            self.country = track.upper()
        else:
            print(f"Error: Track '{track}' does not contain country code and is not a 2-letter country code.")
            sys.exit(1)

        self.total_stages = TOTAL_STAGES
        print(f"Initializing pipeline for Track: {self.track}, Country: {self.country}")

        self.sanitized_track = self.track.replace('/', '_').replace('\\', '_')
        if self.sanitized_track.upper().startswith(self.country.upper() + "_"):
            self.file_prefix = self.sanitized_track
        else:
            self.file_prefix = f"{self.country}_{self.sanitized_track}"

        # Define paths
        self.base_dir = base_dir
        self.aux_dir = aux_dir
        self.proc_dir = self.base_dir / self.track / 'processed_raster'
        self.out_dir = self.base_dir / self.track / 'classification_results'
        self.samples_dir = self.out_dir / 'samples'
        self.model_dir = self.out_dir / 'train_model'
        self.seg_dir = self.out_dir / 'segmentation'
        self.class_dir = self.out_dir / 'classification'

        self._ensure_directories()

        # Resolve input raster
        search_patterns = [
            f"{self.sanitized_track}_*_VH_VV*.tif",
            f"*_{self.sanitized_track}_*_VH_VV*.tif",
            f"*{self.sanitized_track}*.tif",
            f"{self.track}_*_VH_VV*.tif",
            f"*_{self.track}_*_VH_VV*.tif",
            f"*{self.track}*.tif",
            f"{self.sanitized_track}_*_VH_VV*.hdr",
            f"*{self.sanitized_track}*.hdr",
        ]

        self.hdr = None
        if self.proc_dir.exists():
            for pattern in search_patterns:
                self.hdr = next(self.proc_dir.glob(pattern), None)
                if self.hdr:
                    break

            if not self.hdr:
                raise FileNotFoundError(f"No raster file (TIF/HDR) found for track {self.track} in {self.proc_dir}")

            self.ras = self._resolve_raster(self.hdr)
            print(f"Input raster found: {self.ras}")
        else:
            raise FileNotFoundError(f"Processing directory does not exist: {self.proc_dir}")

        # Distinct Output Files for OTB-SAM Hybrid
        self.seg_tif = self.seg_dir / f"{self.file_prefix}_segmentation.tif"
        self.seg_shp = self.seg_dir / f"{self.file_prefix}_segmentation_sam.shp"

        # Search for samples.shp
        samples_base = self.aux_dir / 'shapefiles_samples'
        candidate_paths = [
            samples_base / self.file_prefix / "samples.shp",
            samples_base / f"{self.country}_{self.sanitized_track}" / "samples.shp",
            samples_base / f"{self.country}_{self.track}" / "samples.shp",
            samples_base / self.sanitized_track / "samples.shp",
            samples_base / self.track / "samples.shp",
            samples_base / self.country / "samples.shp",
            samples_base / "samples.shp"
        ]

        self.sample_shp = None
        for p in candidate_paths:
            if p.exists():
                self.sample_shp = p
                print(f"Training samples found at: {self.sample_shp}")
                break

        if not self.sample_shp:
            print(f"\nCRITICAL WARNING: Could not find 'samples.shp' inside {samples_base}")
            self.sample_shp = samples_base / self.file_prefix / "samples.shp"

        # Stages output paths
        self.learn_shp = self.samples_dir / 'learn.shp'
        self.control_shp = self.samples_dir / 'control.shp'
        self.sel_shp = self.samples_dir / f"{self.file_prefix}_learn_selected_otb_sam.shp"
        self.class_shp = self.class_dir / f"{self.file_prefix}_classified_otb_sam.shp"
        self.class_tif = self.class_dir / f"{self.file_prefix}_classified_otb_sam.tif"
        self.conf_tif = self.class_dir / f"{self.file_prefix}_confidence_otb_sam.tif"
        self.footprint_mask = self.seg_dir / f"{self.file_prefix}_data_footprint.tif"
        self.masked_class = self.class_dir / f"{self.file_prefix}_classified_otb_sam_masked.tif"
        self.masked_conf = self.class_dir / f"{self.file_prefix}_confidence_otb_sam_masked.tif"
        self.metrics_fp = self.class_dir / f"{self.file_prefix}_metrics_otb_sam.xlsx"

        self.agri_mask = self._resolve_agri_mask()

        # Configurable Parameters
        self.stage1_params = {
            'method': 'python_sam',
            'tile_size': 2048,
            'buffer': 128,
            'sam_checkpoint': str(self.aux_dir / 'SAM_models' / 'sam_vit_h_4b8939.pth'),
            'sam_model_type': 'vit_h',
            'sam_device': 'cuda' if (HAS_SAM and torch.cuda.is_available()) else 'cpu'
        }
        self.stage3_params = {
            'learn_frac': 0.7, 'random_state': 42
        }
        self.stage5_params = {
            'classifier': 'rf',
            'rf_max': 110, 'rf_min': 2, 'rf_var': 16, 'rf_cat': 16, 'rf_acc': 0.01,
            'svm_c': 1.0, 'svm_k': 'linear',
            'ann_sizes': '100 50', 'ann_t': 'rprop', 'ann_a': 'sigmoid'
        }

        self.feat_str = ""

    def _ensure_directories(self):
        for d in [self.samples_dir, self.model_dir, self.seg_dir, self.class_dir]:
            d.mkdir(parents=True, exist_ok=True)

    def _resolve_raster(self, hdr):
        if hdr.suffix.lower() in ['.tif', '.tiff']: return hdr
        for ext in ['.img', '.tif', '.TIF']:
            p = hdr.with_suffix(ext)
            if p.exists(): return p
        p_no_ext = hdr.with_suffix('')
        if p_no_ext.exists() and p_no_ext.is_file(): return p_no_ext
        raise FileNotFoundError(f"No raster image (.img/.tif) found matching header {hdr.stem}")

    def _resolve_agri_mask(self) -> Path:
        raster_dir = self.aux_dir / 'raster_files'
        country_dir = raster_dir / 'AgriMasks' / self.country

        mask_3class = country_dir / f"{self.country}_agri_mask_3class_epsg3857.tif"
        mask_allcrops = country_dir / f"{self.country}_agri_mask_allcrops_epsg3857.tif"
        mask_eu = raster_dir / 'EU_arable_areas_mask_3857.tif'

        if self.mask_variant == 'allcrops':
            candidates = [mask_allcrops, mask_3class, mask_eu]
        else:
            candidates = [mask_3class, mask_allcrops, mask_eu]

        for p in candidates:
            if p.exists():
                print(f"Agricultural mask selected: {p}")
                return p

        print(f"[WARNING] No agricultural mask found for country '{self.country}'. Fallback to EU mask.")
        return mask_eu

    def _run_cmd(self, cmd, stage, desc, ram=None):
        print(f"[Stage {stage}/{self.total_stages}] {desc}")
        env = os.environ.copy()
        otb_bin = str(otb_dir / "bin")
        otb_lib = str(otb_dir / "lib")
        otb_apps = str(otb_dir / "lib" / "otb" / "applications")
        env["PATH"] = f"{otb_bin};{otb_lib};{env['PATH']}"
        env["OTB_APPLICATION_PATH"] = otb_apps
        env["GEOTIFF_CSV"] = str(otb_dir / "share" / "epsg_csv")
        env["GDAL_DATA"] = str(otb_dir / "share" / "gdal")

        if ram:
            env["OTB_MAX_RAM_HINT"] = str(ram)

        if isinstance(cmd, str):
            cmd = shlex.split(cmd, posix=os.name != 'nt')

        executable = shutil.which(cmd[0], path=env.get("PATH"))
        if executable:
            cmd[0] = executable

        proc = subprocess.Popen(cmd, shell=False, stdout=sys.stdout, stderr=sys.stderr, env=env)
        proc.communicate()
        if proc.returncode != 0:
            raise RuntimeError(f"Stage {stage} failed with return code {proc.returncode}: {cmd}")
        print(f"Completed stage {stage}/{self.total_stages}\n")

    def _create_summed_composite(self):
        print("    [INFO] Creating a log-domain (dB) summed composite of all SAR bands to smooth speckle...")
        composite_tif = self.seg_dir / f"{self.file_prefix}_summed_composite.tif"

        if composite_tif.exists():
            print(f"    [INFO] Summed composite already exists.")
            return composite_tif

        ds = gdal.Open(str(self.ras))
        cols, rows, nbands = ds.RasterXSize, ds.RasterYSize, ds.RasterCount
        gt, proj = ds.GetGeoTransform(), ds.GetProjection()

        driver = gdal.GetDriverByName('GTiff')
        out_ds = driver.Create(str(composite_tif), cols, rows, 1, gdal.GDT_Float32,
                               options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        out_ds.SetGeoTransform(gt)
        out_ds.SetProjection(proj)
        out_band = out_ds.GetRasterBand(1)
        out_band.SetNoDataValue(0)

        tile_size = 4096
        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                sum_arr = np.zeros((ysize, xsize), dtype=np.float32)
                valid_mask = np.zeros((ysize, xsize), dtype=bool)

                for b in range(1, nbands + 1):
                    band = ds.GetRasterBand(b)
                    arr = band.ReadAsArray(x, y, xsize, ysize)
                    arr = np.nan_to_num(arr)
                    nodata = band.GetNoDataValue()
                    
                    if nodata is not None:
                        mask = (arr != nodata) & (arr != 0)
                    else:
                        mask = (arr != 0)
                        
                    sum_arr[mask] += arr[mask]
                    valid_mask |= mask

                sum_arr[~valid_mask] = 0
                out_band.WriteArray(sum_arr, x, y)

        out_ds.FlushCache()
        out_ds = None
        ds = None
        return composite_tif

    # --- Stage 0: Footprint ---
    def stage_0_generate_footprint(self, force_recompute=False):
        self._ensure_directories()
        if self.footprint_mask.exists() and not force_recompute:
            print("[Stage 0] Data footprint mask already exists, skipping.")
            return

        print(f"[Stage 0/{self.total_stages}] Generating data footprint mask...")
        ds = gdal.Open(str(self.ras))
        cols, rows = ds.RasterXSize, ds.RasterYSize
        gt, proj = ds.GetGeoTransform(), ds.GetProjection()

        driver = gdal.GetDriverByName('GTiff')
        out_ds = driver.Create(str(self.footprint_mask), cols, rows, 1, gdal.GDT_Byte,
                               options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        out_ds.SetGeoTransform(gt)
        out_ds.SetProjection(proj)
        out_band = out_ds.GetRasterBand(1)

        tile_size = 4096
        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                combined_mask = None
                for b in range(1, min(ds.RasterCount + 1, 3)):
                    data = ds.GetRasterBand(b).ReadAsArray(x, y, xsize, ysize)
                    m = (np.abs(data) > 1e-7) & (~np.isnan(data))
                    if combined_mask is None:
                        combined_mask = m
                    else:
                        combined_mask |= m

                out_band.WriteArray(combined_mask.astype(np.uint8), x, y)

        out_ds.FlushCache()
        gdal.SieveFilter(out_band, None, out_band, threshold=100, connectedness=4)
        out_ds = None
        ds = None
        print(f"    Footprint mask saved to {self.footprint_mask}\n")

    # --- Stage 1: SAM Segmentation ---
    def stage_1_segmentation(self, **kwargs):
        self._ensure_directories()
        params = self.stage1_params.copy()
        params.update(kwargs)
        stage = 1

        if self.seg_tif.exists():
            print(f"[Stage {stage}/{self.total_stages}] SAM Segmentation Raster already exists, skipping.\n")
            return

        print(f"[Stage {stage}/{self.total_stages}] Running Tiled Python SAM Segmentation...")
        summed_composite = self._create_summed_composite()

        try:
            ds = gdal.Open(str(summed_composite))
            cols, rows = ds.RasterXSize, ds.RasterYSize
            gt, proj = ds.GetGeoTransform(), ds.GetProjection()

            ds_foot = None
            if self.footprint_mask.exists():
                ds_foot = gdal.Open(str(self.footprint_mask))

            driver = gdal.GetDriverByName('GTiff')
            out_ds = driver.Create(str(self.seg_tif), cols, rows, 1, gdal.GDT_Int32,
                                   options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
            out_ds.SetGeoTransform(gt)
            out_ds.SetProjection(proj)
            out_band = out_ds.GetRasterBand(1)
            out_band.SetNoDataValue(0)

            # Load SAM model using samgeo
            from samgeo import SamGeo
            print(f"    Loading SAM model ({params['sam_model_type']}) on {params['sam_device']}...")
            sam_geo = SamGeo(
                model_type=params['sam_model_type'],
                checkpoint=params['sam_checkpoint'],
                device=params['sam_device'],
                sam_kwargs={
                    "points_per_side": 96,
                    "pred_iou_thresh": 0.55,
                    "stability_score_thresh": 0.55,
                    "crop_n_layers": 1,
                    "crop_n_points_downscale_factor": 2,
                    "min_mask_region_area": 10
                }
            )

            tile_size = params['tile_size']
            buffer = params['buffer']
            global_seg_id = 1

            for y in range(0, rows, tile_size):
                for x in range(0, cols, tile_size):
                    xsize_valid = min(tile_size, cols - x)
                    ysize_valid = min(tile_size, rows - y)

                    x_start_buf = max(0, x - buffer)
                    y_start_buf = max(0, y - buffer)
                    x_end_buf = min(cols, x + xsize_valid + buffer)
                    y_end_buf = min(rows, y + ysize_valid + buffer)

                    xsize_buf = x_end_buf - x_start_buf
                    ysize_buf = y_end_buf - y_start_buf

                    # Read single band composite
                    band = ds.GetRasterBand(1)
                    img_arr = band.ReadAsArray(x_start_buf, y_start_buf, xsize_buf, ysize_buf)
                    if img_arr is None:
                        continue

                    # Mask
                    if ds_foot:
                        valid_mask = ds_foot.GetRasterBand(1).ReadAsArray(x_start_buf, y_start_buf, xsize_buf, ysize_buf) > 0
                    else:
                        valid_mask = np.abs(img_arr) > 0

                    if not np.any(valid_mask):
                        continue

                    # Prep 8-bit image for SAM
                    img_8bit = np.zeros((ysize_buf, xsize_buf, 3), dtype=np.uint8)
                    p2, p98 = np.percentile(img_arr[valid_mask], (2, 98))
                    img_clip = np.clip(img_arr, p2, p98)

                    if p98 > p2:
                        val_scaled = ((img_clip - p2) / (p98 - p2) * 255).astype(np.uint8)
                        for b in range(3):
                            img_8bit[:, :, b] = val_scaled

                    import cv2
                    from scipy.ndimage import distance_transform_edt
                    
                    # Bilateral smooth to denoise radar speckle
                    img_smoothed = cv2.bilateralFilter(img_8bit[:, :, 0], d=9, sigmaColor=50, sigmaSpace=50)
                    for b in range(3):
                        img_8bit[:, :, b] = img_smoothed

                    # Generate masks
                    sam_geo.generate(
                        source=img_8bit,
                        output=None,
                        foreground=False,
                        unique=True,
                        min_size=10,
                        max_size=100000
                    )
                    segments_buf = sam_geo.objects.astype(np.int32)

                    # Fill empty spaces inside data footprint using Euclidean Distance Transform
                    zero_mask_buf = (segments_buf == 0) & valid_mask
                    if np.any(zero_mask_buf) and np.any(segments_buf > 0):
                        _, indices = distance_transform_edt(segments_buf == 0, return_indices=True)
                        segments_buf[zero_mask_buf] = segments_buf[tuple(indices)][zero_mask_buf]

                    y_offset = y - y_start_buf
                    x_offset = x - x_start_buf
                    segments_buf[~valid_mask] = 0

                    segments = segments_buf[y_offset : y_offset + ysize_valid, x_offset : x_offset + xsize_valid]
                    valid_mask_crop = valid_mask[y_offset : y_offset + ysize_valid, x_offset : x_offset + xsize_valid]

                    unique_segs = np.unique(segments[segments > 0])
                    if len(unique_segs) > 0:
                        max_seg = segments.max()
                        mapping = np.zeros(max_seg + 1, dtype=np.int32)
                        mapping[unique_segs] = np.arange(global_seg_id, global_seg_id + len(unique_segs))
                        
                        segments = mapping[segments]
                        segments[~valid_mask_crop] = 0
                        global_seg_id += len(unique_segs)
                    else:
                        segments[~valid_mask_crop] = 0

                    out_band.WriteArray(segments, x, y)

            out_ds.FlushCache()
            out_ds = None
            ds = None
            if ds_foot: ds_foot = None
            print(f"    Segmentation Raster saved to {self.seg_tif}\n")

        except Exception as e:
            print(f"Error in Python SAM segmentation: {e}")
            raise

    # --- Stage 2: Polygonize & Feature Extraction ---
    def stage_2_polygonize_and_extract_features(self):
        self._ensure_directories()
        stage = 2

        if self.seg_shp.exists():
            print(f"[Stage {stage}/{self.total_stages}] Segmentation Shapefile already exists, skipping.\n")
            return

        if not self.seg_tif.exists():
            print(f"ERROR: Segmentation raster {self.seg_tif} not found. Run Stage 1 first.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Converting SAM raster to Shapefile using gdal.Polygonize...")
        ds_seg = gdal.Open(str(self.seg_tif))
        seg_band = ds_seg.GetRasterBand(1)

        shp_drv = ogr.GetDriverByName('ESRI Shapefile')
        if os.path.exists(str(self.seg_shp)):
            shp_drv.DeleteDataSource(str(self.seg_shp))

        out_ds = shp_drv.CreateDataSource(str(self.seg_shp))
        srs = osr.SpatialReference()
        srs.ImportFromWkt(ds_seg.GetProjection())

        layer = out_ds.CreateLayer(self.file_prefix + "_segmentation_sam", srs=srs, geom_type=ogr.wkbPolygon)
        fd_dn = ogr.FieldDefn('DN', ogr.OFTInteger)
        layer.CreateField(fd_dn)

        print("    Running gdal.Polygonize...")
        gdal.Polygonize(seg_band, None, layer, 0, [], callback=None)
        out_ds = None  # Flush to disk

        print("    Calculating feature means per segment using fast numpy bincount binning...")
        ds_ras = gdal.Open(str(self.ras))
        nbands = ds_ras.RasterCount
        cols = ds_ras.RasterXSize
        rows = ds_ras.RasterYSize

        max_id = 0
        tile_size = 4096
        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)
                arr = seg_band.ReadAsArray(x, y, xsize, ysize)
                if arr is not None:
                    max_id = max(max_id, arr.max())

        print(f"    Maximum segment ID: {max_id}")

        sums = np.zeros((max_id + 1, nbands), dtype=np.float64)
        counts = np.zeros(max_id + 1, dtype=np.int64)

        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                seg_tile = seg_band.ReadAsArray(x, y, xsize, ysize)
                if seg_tile is None:
                    continue

                flat_seg = seg_tile.ravel()
                mask = flat_seg > 0
                if not np.any(mask):
                    continue

                flat_seg_masked = flat_seg[mask]
                counts += np.bincount(flat_seg_masked, minlength=max_id + 1)

                for b in range(1, nbands + 1):
                    b_band = ds_ras.GetRasterBand(b)
                    ras_tile = b_band.ReadAsArray(x, y, xsize, ysize)
                    if ras_tile is None:
                        continue
                    flat_ras = ras_tile.ravel()[mask]
                    flat_ras = np.nan_to_num(flat_ras)
                    sums[:, b - 1] += np.bincount(flat_seg_masked, weights=flat_ras, minlength=max_id + 1)

        means = sums / (counts[:, None] + 1e-9)

        print("    Loading Shapefile with geopandas (pyogrio)...")
        gdf = gpd.read_file(str(self.seg_shp), engine="pyogrio")
        
        print("    Populating attribute table with mean features...")
        dns = gdf['DN'].values
        dns_clamped = np.clip(dns, 0, max_id)
        
        for b in range(nbands):
            gdf[f'meanB{b}'] = means[dns_clamped, b]

        print("    Writing updated Shapefile back to disk...")
        gdf.to_file(str(self.seg_shp), engine="pyogrio")
        
        ds_seg = None
        ds_ras = None
        print(f"    Completed stage 2. Shapefile saved with {len(gdf)} segments.\n")

    # --- Stage 3: Split Samples ---
    def stage_3_split_samples(self, **kwargs):
        self._ensure_directories()
        params = self.stage3_params.copy()
        params.update(kwargs)
        stage = 3

        if not self.sample_shp.exists():
            print(f"ERROR: Input sample file not found: {self.sample_shp}")
            return

        gdf = gpd.read_file(str(self.sample_shp), engine="pyogrio")

        if self.country == 'NL':
            print("    Applying crop aggregation for Netherlands (Clover/Lucerne -> Grassland)...")
            gdf['crop_id'] = gdf['crop_id'].map(lambda cid: CROP_AGGREGATION_NL.get(cid, cid))

        if not self.seg_tif.exists():
            print("    [WARNING] Segmentation raster not found. Falling back to random sample split.")
            learn = gdf.sample(frac=params['learn_frac'], random_state=params['random_state'])
            control = gdf.drop(learn.index)
        else:
            print(f"    Aligning training samples with segmentation raster {self.seg_tif.name} to prevent spatial data leakage...")
            ds_seg = gdal.Open(str(self.seg_tif))
            gt = ds_seg.GetGeoTransform()
            inv_gt = gdal.InvGeoTransform(gt)
            raster_proj = ds_seg.GetProjection()
            cols = ds_seg.RasterXSize
            rows = ds_seg.RasterYSize
            seg_band = ds_seg.GetRasterBand(1)

            from pyproj import CRS
            if raster_proj and gdf.crs:
                target_crs = CRS.from_wkt(raster_proj)
                if gdf.crs != target_crs:
                    print("    Reprojecting points to match segmentation CRS...")
                    gdf = gdf.to_crs(target_crs)

            xs = gdf.geometry.x.values
            ys = gdf.geometry.y.values
            pxs = (inv_gt[0] + inv_gt[1] * xs + inv_gt[2] * ys).astype(int)
            pys = (inv_gt[3] + inv_gt[4] * xs + inv_gt[5] * ys).astype(int)

            seg_ids = []
            for px, py in zip(pxs, pys):
                if 0 <= px < cols and 0 <= py < rows:
                    try:
                        val = seg_band.ReadAsArray(int(px), int(py), 1, 1)[0, 0]
                        seg_ids.append(val)
                    except:
                        seg_ids.append(0)
                else:
                    seg_ids.append(0)

            gdf['seg_id'] = seg_ids
            ds_seg = None

            gdf_valid = gdf[gdf['seg_id'] > 0].copy()
            dropped = len(gdf) - len(gdf_valid)
            if dropped > 0:
                print(f"    Warning: Dropped {dropped} points that fell outside valid segments.")

            if len(gdf_valid) == 0:
                print("ERROR: No points fell within any valid segments. Falling back to random split.")
                learn = gdf.sample(frac=params['learn_frac'], random_state=params['random_state'])
                control = gdf.drop(learn.index)
            else:
                unique_segs = gdf_valid['seg_id'].unique()
                print(f"    Found {len(unique_segs)} unique segments for {len(gdf_valid)} valid points.")

                np.random.seed(params['random_state'])
                np.random.shuffle(unique_segs)
                split_idx = int(len(unique_segs) * params['learn_frac'])
                train_segs = set(unique_segs[:split_idx])

                learn = gdf_valid[gdf_valid['seg_id'].isin(train_segs)].copy()
                control = gdf_valid[~gdf_valid['seg_id'].isin(train_segs)].copy()

                learn = learn.drop(columns=['seg_id'])
                control = control.drop(columns=['seg_id'])

        learn.to_file(str(self.learn_shp), engine="pyogrio")
        control.to_file(str(self.control_shp), engine="pyogrio")
        print(f"Completed stage {stage}. Learn: {len(learn)}, Control: {len(control)}\n")

    # --- Stage 4: Select Polygons ---
    def stage_4_selection(self):
        self._ensure_directories()
        stage = 4
        if not self.sel_shp.exists():
            print(f"[Stage {stage}/{self.total_stages}] Running spatial intersection for sample selection")
            if not self.learn_shp.exists():
                print("ERROR: Learn samples not found. Run Stage 3 first.")
                return
            if not self.seg_shp.exists():
                print("ERROR: Segmentation shapefile not found. Run Stage 2 first.")
                return

            pts = gpd.read_file(self.learn_shp, engine="pyogrio")
            polys = gpd.read_file(self.seg_shp, engine="pyogrio")

            if polys.crs != pts.crs:
                print("Warning: CRS mismatch. Re-projecting segmentation to sample CRS.")
                polys = polys.to_crs(pts.crs)

            # Spatial join
            sel = gpd.sjoin(polys, pts, how='inner', predicate='intersects')

            # Apply Class Balancing (Oversampling) to prevent OTB classifier bias
            if not sel.empty:
                print("    Balancing classes in training polygons (Oversampling)...")
                max_size = sel['crop_id'].value_counts().max()
                balanced_dfs = []
                for crop_id, group in sel.groupby('crop_id'):
                    if len(group) < max_size:
                        resampled = group.sample(max_size, replace=True, random_state=42)
                        balanced_dfs.append(resampled)
                    else:
                        balanced_dfs.append(group)
                sel = pd.concat(balanced_dfs)
                sel = gpd.GeoDataFrame(sel, geometry='geometry', crs=polys.crs)

            sel.to_file(self.sel_shp, engine="pyogrio")
            print(f"[Stage {stage}/{self.total_stages}] Selected {len(sel)} features\n")
        else:
            print(f"[Stage {stage}/{self.total_stages}] Selection exists, skipping\n")

    # --- Stage 5: Train Classifier ---
    def stage_5_train_classifier(self, **kwargs):
        self._ensure_directories()
        force_retrain = kwargs.pop('force_retrain', False)
        params = self.stage5_params.copy()
        params.update(kwargs)
        stage = 5

        if not self.sel_shp.exists():
            print("ERROR: Selected samples file not found. Run Stage 4 first.")
            return

        df_sel = gpd.read_file(self.sel_shp, engine="pyogrio")
        feats = [c for c in df_sel.columns if c.startswith('meanB')]
        if not feats:
            print("ERROR: No features starting with 'meanB' found in selected shapefile.")
            return
        self.feat_str = ' '.join(feats)

        clf_name = params['classifier']
        model_fn = self.model_dir / f"{self.file_prefix}_model_otb_sam.{clf_name}"
        confmat_fn = self.model_dir / f"{self.file_prefix}_train_confmat_otb_sam.{clf_name}.csv"

        if force_retrain or not model_fn.exists() or os.path.getsize(model_fn) == 0:
            if force_retrain and model_fn.exists():
                print(f"[Stage {stage}/{self.total_stages}] Parameters changed. Forcing retrain.")
            elif model_fn.exists() and os.path.getsize(model_fn) == 0:
                print(f"[Stage {stage}/{self.total_stages}] Empty model, retrain.")

            otb_clf_name = 'libsvm' if clf_name == 'svm' else clf_name
            clf_str = f"-classifier {otb_clf_name}"

            if clf_name == 'rf':
                clf_str += (
                    f" -classifier.rf.max {params['rf_max']} -classifier.rf.min {params['rf_min']} "
                    f" -classifier.rf.var {params['rf_var']} -classifier.rf.cat {params['rf_cat']} "
                    f" -classifier.rf.acc {params['rf_acc']}"
                )
            elif clf_name == 'svm':
                clf_str += f" -classifier.libsvm.c {params.get('svm_c', 1.0)}"
                clf_str += f" -classifier.libsvm.k {params.get('svm_k', 'linear')}"
            elif clf_name == 'ann':
                clf_str += (
                    f" -classifier.ann.sizes {params['ann_sizes']} "
                    f" -classifier.ann.t {params['ann_t']} "
                    f" -classifier.ann.a {params['ann_a']}"
                )

            cmd = (
                f"otbcli_TrainVectorClassifier -io.vd {self.sel_shp} -io.out {model_fn} "
                f"-feat {self.feat_str} -cfield crop_id {clf_str} "
                f"-io.confmatout {confmat_fn}"
            )
            self._run_cmd(cmd, stage, f'Train OTB {clf_name.upper()}')

            try:
                if confmat_fn.exists():
                    print(f"\n--- Training Confusion Matrix ---")
                    df_cm = pd.read_csv(confmat_fn, skiprows=2, index_col=0)
                    df_cm = df_cm.drop(index=['Total', 'UA'], columns=['Total', 'PA'], errors='ignore')
                    df_cm = df_cm.dropna(how='all', axis=0).dropna(how='all', axis=1)
                    df_cm = df_cm.fillna(0)
                    print(df_cm.to_string())
            except Exception as e:
                print(f"Warning: Could not read training confusion matrix: {e}\n")
        else:
            print(f"[Stage {stage}/{self.total_stages}] Model exists, skipping\n")

    # --- Stage 6: Classify Vector ---
    def stage_6_classify_vector(self, force_recompute=False):
        self._ensure_directories()
        stage = 6

        clf_name = self.stage5_params['classifier']
        model_file = self.model_dir / f"{self.file_prefix}_model_otb_sam.{clf_name}"

        if not model_file.exists():
            print(f"ERROR: Model file {model_file} not found. Run Stage 5 first.")
            return

        if not self.feat_str:
            print("Feature string not set. Reading from selected samples...")
            try:
                if not self.sel_shp.exists():
                    raise FileNotFoundError("sel_shp not found, run stage 4")
                df_sel = gpd.read_file(self.sel_shp, engine="pyogrio")
                feats = [c for c in df_sel.columns if c.startswith('meanB')]
                self.feat_str = ' '.join(feats)
            except Exception as e:
                print(f"ERROR: Could not determine features: {e}.")
                return

        if not self.class_shp.exists() or force_recompute:
            if force_recompute and self.class_shp.exists():
                print(f"[Stage {stage}] Forcing recomputation of vector classification.")
                for ext in ['.shp', '.shx', '.dbf', '.prj', '.cpg']:
                    f = self.class_shp.with_suffix(ext)
                    if f.exists(): os.remove(f)

            cmd = (
                f"otbcli_VectorClassifier -in {self.seg_shp} -out {self.class_shp} "
                f"-model {model_file} -feat {self.feat_str} -cfield predicted -confmap true"
            )
            self._run_cmd(cmd, stage, 'Vector classification')
        else:
            print(f"[Stage {stage}/{self.total_stages}] Classification exists, skipping\n")

    # --- Stage 7: Rasterize Class ---
    def stage_7_rasterize_class(self, force_recompute=False):
        self._ensure_directories()
        stage = 7
        if not self.class_shp.exists():
            print(f"ERROR: Classified shapefile not found. Run Stage 6.")
            return
        if not self.class_tif.exists() or force_recompute:
            if force_recompute and self.class_tif.exists():
                os.remove(self.class_tif)
            cmd = (
                f"otbcli_Rasterization -in {self.class_shp} -out {self.class_tif} "
                f"-mode attribute -mode.attribute.field predicted -spx 10 -spy 10"
            )
            self._run_cmd(cmd, stage, 'Rasterize classification')
        else:
            print(f"[Stage {stage}/{self.total_stages}] Classified TIFF exists, skipping\n")

    # --- Stage 8: Rasterize Confidence ---
    def stage_8_rasterize_confidence(self, force_recompute=False):
        self._ensure_directories()
        stage = 8
        if not self.class_shp.exists():
            print(f"ERROR: Classified shapefile not found. Run Stage 6.")
            return
        if not self.conf_tif.exists() or force_recompute:
            if force_recompute and self.conf_tif.exists():
                os.remove(self.conf_tif)
            cmd = (
                f"otbcli_Rasterization -in {self.class_shp} -out {self.conf_tif} "
                f"-mode attribute -mode.attribute.field confidence -spx 10 -spy 10"
            )
            self._run_cmd(cmd, stage, 'Rasterize confidence')
        else:
            print(f"[Stage {stage}/{self.total_stages}] Confidence TIFF exists, skipping\n")

    # --- Helper: Apply Mask ---
    def _apply_mask(self, input_tif, mask_tif, out_tif, stage):
        print(f"[Stage {stage}/{self.total_stages}] Applying Arable & Data Footprint Mask...")

        ds_stack = gdal.Open(str(self.ras))
        if not ds_stack: raise RuntimeError(f"Could not open source raster {self.ras}")
        stack_band = ds_stack.GetRasterBand(1)

        ds_foot = None
        if self.footprint_mask.exists():
            ds_foot = gdal.Open(str(self.footprint_mask))
            foot_band = ds_foot.GetRasterBand(1)
        else:
            foot_band = None

        ds_in = gdal.Open(str(input_tif))
        gt = ds_in.GetGeoTransform()
        proj = ds_in.GetProjection()
        cols = ds_in.RasterXSize
        rows = ds_in.RasterYSize

        minx, maxy = gt[0], gt[3]
        maxx = minx + gt[1] * cols
        miny = maxy + gt[5] * rows

        if mask_tif.exists():
            temp_mask_vrt = str(out_tif).replace('.tif', '_mask_temp.vrt')
            mask_opts = gdal.WarpOptions(
                format='VRT',
                outputBounds=(minx, miny, maxx, maxy),
                width=cols,
                height=rows,
                dstSRS=proj,
                resampleAlg=gdal.GRA_NearestNeighbour
            )
            ds_mask = gdal.Warp(temp_mask_vrt, str(mask_tif), options=mask_opts)
            m_band = ds_mask.GetRasterBand(1)
        else:
            ds_mask = None
            m_band = None
            print(f"    WARNING: Arable mask not found at {mask_tif}. Will only apply data footprint mask.")

        in_band = ds_in.GetRasterBand(1)
        out_type = in_band.DataType

        driver = gdal.GetDriverByName('GTiff')
        out_ds = driver.Create(str(out_tif), cols, rows, 1, out_type,
                               options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        out_ds.SetGeoTransform(gt)
        out_ds.SetProjection(proj)
        out_band = out_ds.GetRasterBand(1)

        nodata = in_band.GetNoDataValue()
        if nodata is None: nodata = 0
        out_band.SetNoDataValue(nodata)

        tile_size = 4096

        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                arr = in_band.ReadAsArray(x, y, xsize, ysize)

                # Footprint Mask
                if foot_band:
                    f_arr = foot_band.ReadAsArray(x, y, xsize, ysize)
                    arr[f_arr == 0] = nodata
                elif stack_band:
                    try:
                        stack_arr = stack_band.ReadAsArray(x, y, xsize, ysize)
                        if stack_arr is not None:
                            arr[stack_arr == 0] = nodata
                    except:
                        pass

                # Arable Mask
                if m_band:
                    m_arr = m_band.ReadAsArray(x, y, xsize, ysize)
                    arr[m_arr < 0.5] = nodata

                out_band.WriteArray(arr, x, y)

        out_ds.FlushCache()
        out_ds = None
        ds_in = None
        ds_stack = None
        if ds_mask:
            ds_mask = None
            if os.path.exists(temp_mask_vrt): os.remove(temp_mask_vrt)

        print(f"Completed stage {stage}\n")

    # --- Stage 9: Mask Class ---
    def stage_9_mask_class(self, force_recompute=False):
        self._ensure_directories()
        stage = 9
        if not self.class_tif.exists():
            print("ERROR: Classified TIF not found. Run Stage 7.")
            return

        if not self.masked_class.exists() or force_recompute:
            self._apply_mask(self.class_tif, self.agri_mask, self.masked_class, stage)
        else:
            print(f"[Stage {stage}/{self.total_stages}] Masked classification exists, skipping\n")

    # --- Stage 10: Mask Confidence ---
    def stage_10_mask_confidence(self, force_recompute=False):
        self._ensure_directories()
        stage = 10
        if not self.conf_tif.exists():
            print("ERROR: Confidence TIF not found. Run Stage 8.")
            return

        if not self.masked_conf.exists() or force_recompute:
            self._apply_mask(self.conf_tif, self.agri_mask, self.masked_conf, stage)
        else:
            print(f"[Stage {stage}/{self.total_stages}] Masked confidence exists, skipping\n")

    # --- Stage 11: Calculate Metrics ---
    def stage_11_calculate_metrics(self):
        self._ensure_directories()
        stage = 11
        if not self.metrics_fp.exists():
            print(f"[Stage {stage}/{self.total_stages}] Computing metrics and exporting Excel...")

            if not self.control_shp.exists():
                print("ERROR: Control shapefile not found. Run Stage 3 first.")
                return
            if not self.masked_class.exists():
                print("ERROR: Masked classification raster not found. Run Stage 9.")
                return

            ctrl = gpd.read_file(str(self.control_shp), engine="pyogrio")

            ds = gdal.Open(str(self.masked_class))
            raster_proj = ds.GetProjection()
            raster_srs = osr.SpatialReference()
            raster_srs.ImportFromWkt(raster_proj)

            if ctrl.crs:
                if ctrl.crs.to_wkt() != raster_srs.ExportToWkt():
                    print("    Aligning control points CRS to match raster...")
                    try:
                        from pyproj import CRS
                        target_crs = CRS.from_wkt(raster_srs.ExportToWkt())
                        ctrl = ctrl.to_crs(target_crs)
                    except Exception as e:
                        print(f"    Could not auto-align CRS: {e}")

            band = ds.GetRasterBand(1)
            gt = ds.GetGeoTransform()
            inv = gdal.InvGeoTransform(gt)
            true_vals, pred_vals = [], []

            xs = ctrl.geometry.x.values
            ys = ctrl.geometry.y.values

            pxs = (inv[0] + inv[1] * xs + inv[2] * ys).astype(int)
            pys = (inv[3] + inv[4] * xs + inv[5] * ys).astype(int)
            crop_ids = ctrl['crop_id'].values

            if self.country == 'NL':
                print("    Applying crop aggregation for Netherlands validation labels...")
                crop_ids = np.array([CROP_AGGREGATION_NL.get(val, val) for val in crop_ids])

            for px, py, crop_id in zip(pxs, pys, crop_ids):
                try:
                    if 0 <= px < ds.RasterXSize and 0 <= py < ds.RasterYSize:
                        t = int(crop_id)
                        val_arr = band.ReadAsArray(px, py, 1, 1)
                        if val_arr is not None:
                            p = int(val_arr[0, 0])
                            if self.country == 'NL':
                                p = CROP_AGGREGATION_NL.get(p, p)
                            if t > 0 and p > 0 and p != -9999:
                                true_vals.append(t)
                                pred_vals.append(p)
                except Exception as e:
                    print(f"    [WARNING] Failed to extract point value: {e}")

            if not true_vals or not pred_vals:
                print("ERROR: No valid matching true/predicted values found.")
                return

            labels = sorted(list(set(true_vals + pred_vals)))
            cm = confusion_matrix(true_vals, pred_vals, labels=labels)
            precisions, recalls, f1s, _ = precision_recall_fscore_support(
                true_vals, pred_vals, labels=labels, average=None, zero_division=0
            )

            total = np.sum(cm)
            oa = np.trace(cm) / total
            sum_po = oa
            sum_pe = np.sum(np.sum(cm, axis=0) * np.sum(cm, axis=1)) / (total ** 2)
            kappa = (sum_po - sum_pe) / (1 - sum_pe) if (1 - sum_pe) != 0 else np.nan

            # calculate area per class
            resx, resy = abs(gt[1]), abs(gt[5])
            area_ha = resx * resy / 10000
            arr = band.ReadAsArray()
            unique_classes, counts = np.unique(arr[arr > 0], return_counts=True)
            class_areas = dict(zip(unique_classes, counts))
            areas = [{'Class': c, 'Area_ha': round(class_areas.get(c, 0) * area_ha, 2)} for c in labels]

            # Write Excel
            wb = openpyxl.Workbook()
            sh = wb.active
            sh.title = 'Results'

            # Confusion Matrix
            sh.cell(row=1, column=1, value='Confusion Matrix').font = Font(bold=True)
            sh.cell(row=2, column=1, value='True \\ Pred').font = Font(bold=True)
            for j, lbl in enumerate(labels, start=2):
                sh.cell(row=2, column=j, value=lbl).font = Font(bold=True)
            for i, lbl in enumerate(labels, start=3):
                sh.cell(row=i, column=1, value=lbl).font = Font(bold=True)
                for j, _ in enumerate(labels):
                    sh.cell(row=i, column=j + 2, value=int(cm[i - 3, j]))

            # OA & Kappa
            base = 4 + len(labels)
            sh.cell(row=base, column=1, value='Overall Accuracy').font = Font(bold=True)
            sh.cell(row=base, column=2, value=round(oa, 4))
            sh.cell(row=base + 1, column=1, value='Kappa').font = Font(bold=True)
            sh.cell(row=base + 1, column=2, value=round(kappa, 4))

            # Classification metrics
            start = base + 3
            headers = ['Class', 'Producer Acc (Recall)', 'User Acc (Precision)', 'F1-score']
            for j, h in enumerate(headers, start=1):
                sh.cell(row=start, column=j, value=h).font = Font(bold=True)
            for idx, c in enumerate(labels):
                row_idx = start + 1 + idx
                sh.cell(row=row_idx, column=1, value=c)
                sh.cell(row=row_idx, column=2, value=round(recalls[idx], 4))
                sh.cell(row=row_idx, column=3, value=round(precisions[idx], 4))
                sh.cell(row=row_idx, column=4, value=round(f1s[idx], 4))

            # Areas
            ar0 = start + 1 + len(labels) + 1
            sh.cell(row=ar0, column=1, value='Areas (ha)').font = Font(bold=True)
            sh.cell(row=ar0 + 1, column=1, value='Class').font = Font(bold=True)
            sh.cell(row=ar0 + 1, column=2, value='Area_ha').font = Font(bold=True)
            for idx, a in enumerate(areas, start=ar0 + 2):
                sh.cell(row=idx, column=1, value=a['Class'])
                sh.cell(row=idx, column=2, value=a['Area_ha'])

            wb.save(str(self.metrics_fp))
            print(f"Metrics saved to {self.metrics_fp}\n")
        else:
            print(f"[Stage 11] Metrics Excel exists, skipping")
        print(f"All done! Metrics available at {self.metrics_fp}")


# --- Interactive Menu Helpers ---

SAM_MODELS = {
    '1': {'name': 'vit_b  (Maly,  ~375 MB, SZYBKI,  ~2 GB VRAM - polecany do testow)',
           'model_type': 'vit_b',  'checkpoint': 'sam_vit_b_01ec64.pth',
           'url': 'https://dl.fbaipublicfiles.com/segment_anything/sam_vit_b_01ec64.pth'},
    '2': {'name': 'vit_l  (Sredni, ~1.2 GB, SREDNI,  ~6 GB VRAM)',
           'model_type': 'vit_l',  'checkpoint': 'sam_vit_l_0b3195.pth',
           'url': 'https://dl.fbaipublicfiles.com/segment_anything/sam_vit_l_0b3195.pth'},
    '3': {'name': 'vit_h  (Ogromny,~2.4 GB, WOLNY,  ~10 GB VRAM - najwyzsza dokladnosc)',
           'model_type': 'vit_h',  'checkpoint': 'sam_vit_h_4b8939.pth',
           'url': 'https://dl.fbaipublicfiles.com/segment_anything/sam_vit_h_4b8939.pth'},
}


def get_stage1_params_sam(param_dict):
    new_params = param_dict.copy()
    current_type = new_params.get('sam_model_type', 'vit_b')
    current_ckpt = new_params.get('sam_checkpoint', 'sam_vit_b_01ec64.pth')
    device = new_params.get('sam_device', 'cpu')
    tile_size = new_params.get('tile_size', 2048)
    buffer = new_params.get('buffer', 128)

    print("\n--- Stage 1: Segmentacja SAM ---")
    print(f"  Urzadzenie (device)  : {device}")
    print(f"  Rozmiar kafla (px)   : {tile_size}")
    print(f"  Bufor kafla (px)     : {buffer}")
    print(f"  Aktualny model       : {current_type}  [{current_ckpt}]")
    print()
    print("  Dostepne modele SAM:")
    for k, v in SAM_MODELS.items():
        marker = " <-- aktualny" if v['model_type'] == current_type else ""
        print(f"    [{k}] {v['name']}{marker}")
    print()

    choice = input("Wybierz model SAM (1/2/3) lub Enter aby zachowac aktualny: ").strip()
    if choice in SAM_MODELS:
        selected = SAM_MODELS[choice]
        new_params['sam_model_type'] = selected['model_type']
        
        sam_models_dir = aux_dir / 'SAM_models'
        ckpt_fn = selected['checkpoint']
        ckpt_path = sam_models_dir / ckpt_fn
        new_params['sam_checkpoint'] = str(ckpt_path)

        if not ckpt_path.exists():
            print(f"\n  [UWAGA] Plik wag '{ckpt_fn}' nie istnieje w katalogu {sam_models_dir}!")
            print(f"  Pobierz go z: {selected['url']}")
            proceed = input("  Kontynuowac mimo to? (y/n) [n]: ").strip().lower()
            if proceed != 'y':
                return None
        else:
            print(f"  [OK] Plik wag '{ckpt_fn}' znaleziony.")
    else:
        print("  Zachowuje aktualny model.")

    dev_choice = input(f"  Urzadzenie (cuda/cpu) [{device}]: ").strip().lower()
    if dev_choice in ('cuda', 'cpu'):
        new_params['sam_device'] = dev_choice

    ts_str = input(f"  Rozmiar kafla w px [{tile_size}]: ").strip()
    if ts_str:
        try:
            new_params['tile_size'] = int(ts_str)
        except ValueError:
            pass

    return new_params


def get_params(param_dict):
    new_params = param_dict.copy()
    print("--- Current Parameters ---")
    for key, val in new_params.items():
        print(f"  {key}: {val}")

    if input("Change parameters? (y/n) [n]: ").lower() != 'y':
        return new_params

    for key, val in new_params.items():
        new_val_str = input(f"Enter new value for '{key}' [{val}]: ")
        if not new_val_str:
            continue
        try:
            original_type = type(val)
            new_params[key] = original_type(new_val_str)
        except ValueError:
            print(f"Invalid value. Keeping default {val}.")
    return new_params


def get_classifier_params(param_dict):
    new_params = param_dict.copy()
    print("--- Current Parameters ---")
    for key, val in new_params.items():
        print(f"  {key}: {val}")

    if input("Change parameters? (y/n) [n]: ").lower() != 'y':
        return new_params

    clf = input(f"Enter classifier (rf, svm, ann) [{new_params['classifier']}]: ") or new_params['classifier']
    new_params['classifier'] = clf.lower()

    print(f"\n--- Setting parameters for {clf.upper()} ---")
    prefix = clf + '_'
    for key in [k for k in new_params if k.startswith(prefix)]:
        val = new_params[key]
        new_val_str = input(f"Enter new value for '{key}' [{val}]: ")
        if new_val_str:
            try:
                new_params[key] = type(val)(new_val_str)
            except ValueError:
                print(f"Invalid value.")
    return new_params


def main_menu(pipeline):
    menu = f"""
    --- OTB-SAM Hybrid OBIA Pipeline ---
    Track: {pipeline.track} ({pipeline.country})

    [0] Stage 0: Generate Data Footprint Mask
    [1] Stage 1: SAM Segmentation (Meta SAM) -> produces {pipeline.seg_tif.name}
    [2] Stage 2: Polygonize & Feature Extraction -> produces {pipeline.seg_shp.name}
    [3] Stage 3: Split Samples
    [4] Stage 4: Select Polygons -> produces {pipeline.sel_shp.name}
    [5] Stage 5: Train OTB Classifier ({pipeline.stage5_params['classifier'].upper()})
    [6] Stage 6: Classify Vector -> produces {pipeline.class_shp.name}
    [7] Stage 7: Rasterize Classification -> produces {pipeline.class_tif.name}
    [8] Stage 8: Rasterize Confidence -> produces {pipeline.conf_tif.name}
    [9] Stage 9: Mask Classification -> produces {pipeline.masked_class.name}
    [10] Stage 10: Mask Confidence -> produces {pipeline.masked_conf.name}
    [11] Stage 11: Calculate Metrics -> produces {pipeline.metrics_fp.name}

    [A] Run All Stages (forces rerun of inference, masking, metrics)
    [Q] Quit

    Enter your choice:
    """

    while True:
        choice = input(menu).strip().upper()
        try:
            if choice == '0':
                pipeline.stage_0_generate_footprint()
            elif choice == '1':
                new_params = get_stage1_params_sam(pipeline.stage1_params)
                if new_params is None:
                    continue
                pipeline.stage1_params.update(new_params)
                pipeline.stage_1_segmentation(**pipeline.stage1_params)
            elif choice == '2':
                pipeline.stage_2_polygonize_and_extract_features()
            elif choice == '3':
                new_params = get_params(pipeline.stage3_params)
                pipeline.stage3_params.update(new_params)
                pipeline.stage_3_split_samples(**pipeline.stage3_params)
            elif choice == '4':
                pipeline.stage_4_selection()
            elif choice == '5':
                new_params = get_classifier_params(pipeline.stage5_params)
                force = (pipeline.stage5_params != new_params)
                pipeline.stage5_params.update(new_params)
                pipeline.stage_5_train_classifier(force_retrain=force, **pipeline.stage5_params)
            elif choice == '6':
                pipeline.stage_6_classify_vector()
            elif choice == '7':
                pipeline.stage_7_rasterize_class()
            elif choice == '8':
                pipeline.stage_8_rasterize_confidence()
            elif choice == '9':
                pipeline.stage_9_mask_class(force_recompute=True)
            elif choice == '10':
                pipeline.stage_10_mask_confidence(force_recompute=True)
            elif choice == '11':
                pipeline.stage_11_calculate_metrics()
            elif choice == 'A':
                pipeline.stage_0_generate_footprint()
                pipeline.stage_1_segmentation(**pipeline.stage1_params)
                pipeline.stage_2_polygonize_and_extract_features()
                pipeline.stage_3_split_samples(**pipeline.stage3_params)
                pipeline.stage_4_selection()
                pipeline.stage_5_train_classifier(**pipeline.stage5_params)
                pipeline.stage_6_classify_vector(force_recompute=True)
                pipeline.stage_7_rasterize_class(force_recompute=True)
                pipeline.stage_8_rasterize_confidence(force_recompute=True)
                pipeline.stage_9_mask_class(force_recompute=True)
                pipeline.stage_10_mask_confidence(force_recompute=True)
                if pipeline.metrics_fp.exists(): pipeline.metrics_fp.unlink()
                pipeline.stage_11_calculate_metrics()
            elif choice == 'Q':
                break
        except Exception as e:
            print(f"\n--- ERROR ---: {e}")
            import traceback
            traceback.print_exc()


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Modular OTB-SAM Hybrid Pipeline")
    parser.add_argument('--track', required=True, help="Processing track name (e.g. NL/orbit_88 or PT/orbit_161)")
    parser.add_argument('--mask_variant', default='3class',
                        choices=['3class', 'allcrops'],
                        help="Agricultural mask variant: '3class' (jare/oziminy/rzepak, default) or 'allcrops'")
    args = parser.parse_args()

    try:
        pipeline = ProcessingPipeline(track=args.track, mask_variant=args.mask_variant)
        main_menu(pipeline)
    except Exception as e:
        print(f"Initialization Error: {e}")
        sys.exit(1)
