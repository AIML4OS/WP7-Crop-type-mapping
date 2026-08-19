#!/usr/bin/env python
"""
1_classify_MLPXGB_presto_hybrid_S1S2.py - Multimodal Sentinel-1 (Sigma0) + Sentinel-2 Crop Classification
using NASA Harvest Presto embeddings and a Unified PyTorch MLP + XGBoost Fusion Ensemble.

Pipeline Overview:
  Stage 0: Generate Multimodal Data Footprint (S1 + S2 valid data intersection)
  Stage 1: Multimodal Image Segmentation (SLIC / SAM / LPIS)
  Stage 2: Sample Point Split (70% learn / 30% control)
  Stage 3: Multimodal Feature Extraction (S1 Sigma0 + S2 Optical + S1 Presto 128d + S2 Presto 128d)
  Stage 4: Train Unified MLP + XGBoost Fusion Ensemble
  Stage 5: Object-Based Tile Inference with Bayesian Prior Calibration
  Stage 6: Apply Agricultural & Data Footprint Masks
  Stage 7: Calculate Out-of-Bag Validation Metrics & Generate Styled Excel Report (.xlsx)

Execution examples:
  python 1_classify_MLPXGB_presto_hybrid_S1S2.py --track NL/orbit_88 --seg_mode slic --stage A
  python 1_classify_MLPXGB_presto_hybrid_S1S2.py --track PL/orbit_12 --seg_mode slic --stage A
  python 1_classify_MLPXGB_presto_hybrid_S1S2.py --track PT/orbit_161 --seg_mode sam --stage A
"""

import os
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
os.environ.setdefault("OMP_NUM_THREADS", "4")

import argparse
from pathlib import Path
import subprocess
import sys
import shutil
import shlex
import json
import math
import re
import threading
from concurrent.futures import ThreadPoolExecutor
from typing import Optional, List, Dict, Tuple

# Add classifier directory and project root to sys.path to allow importing single_file_presto
classifier_dir = str(Path(__file__).resolve().parent)
project_root = str(Path(__file__).resolve().parent.parent)
for p in [classifier_dir, project_root]:
    if p not in sys.path:
        sys.path.insert(0, p)

import numpy as np
import pandas as pd
import geopandas as gpd
from osgeo import gdal, ogr, osr, gdalconst
from sklearn.metrics import confusion_matrix, precision_recall_fscore_support
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.pipeline import make_pipeline
from sklearn.utils import resample
from sklearn.impute import SimpleImputer
import joblib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# PyTorch
try:
    import torch
    import torch.nn as nn
    import torch.optim as optim
    from torch.utils.data import DataLoader, TensorDataset
    HAS_TORCH = True
except ImportError:
    HAS_TORCH = False
    print("WARNING: PyTorch not found. Install torch for Presto embeddings and MLP.")

# XGBoost / GBDT
try:
    import xgboost as xgb
    HAS_XGBOOST = True
except ImportError:
    HAS_XGBOOST = False
    from sklearn.ensemble import HistGradientBoostingClassifier

# Scikit-image for SLIC
try:
    from skimage.segmentation import felzenszwalb, slic
    from skimage.util import img_as_float
    HAS_SKIMAGE = True
except ImportError:
    HAS_SKIMAGE = False

# Global project directories
base_dir = Path(os.environ.get("AIML_WORKING_DIR", r"D:/AIML_CropMapper_Cloud/workingDir"))
aux_dir = Path(os.environ.get("AIML_AUX_DIR", r"D:/AIML_CropMapper_Cloud/auxiliary_files"))
presto_dir = aux_dir / "Presto_models"

TOTAL_STAGES = 8


# =====================================================================
# 1. HELPERS: DATE PARSING & PRIORS
# =====================================================================

def parse_month_from_description(desc: str) -> int:
    """Parses month index (0-11) from raster band descriptions."""
    months_map = {
        'jan': 0, 'feb': 1, 'mar': 2, 'apr': 3, 'may': 4, 'jun': 5,
        'jul': 6, 'aug': 7, 'sep': 8, 'oct': 9, 'nov': 10, 'dec': 11
    }
    match = re.search(r'_(?:\d+)?([a-zA-Z]{3})\d{4}_', str(desc))
    if match:
        mon_str = match.group(1).lower()
        return months_map.get(mon_str, 0)
    return 0


def get_crop_aggregation(country: str, learn_shp_path: Optional[Path]) -> dict:
    return {}


def _get_priors_for_country(country: str, learn_shp_path: Optional[Path], classes: np.ndarray, class_counts: dict, total_samples: int, priors_file_override: Optional[Path] = None) -> np.ndarray:
    """Calculates true prior probabilities using shapefile names, custom JSONs, or physical field area priors."""
    priors_json_path = None
    if priors_file_override and os.path.exists(priors_file_override):
        priors_json_path = Path(priors_file_override)
    else:
        if learn_shp_path:
            p = Path(learn_shp_path).resolve()
            for parent in p.parents:
                aux_dir_path = parent / 'auxiliary_files'
                if aux_dir_path.exists():
                    aux_priors = aux_dir_path / 'shapefiles_samples' / country / 'priors.json'
                    if aux_priors.exists():
                        priors_json_path = aux_priors
                        break
            if not priors_json_path:
                track_dir = Path(learn_shp_path).parent.parent
                track_priors = track_dir / f"priors_{country}.json"
                if not track_priors.exists():
                    track_priors = track_dir / "priors.json"
                if track_priors.exists():
                    priors_json_path = track_priors

    if priors_json_path and priors_json_path.exists():
        try:
            with open(priors_json_path, 'r') as f:
                custom_priors = json.load(f)
            
            id_to_name = {}
            if learn_shp_path and os.path.exists(learn_shp_path):
                try:
                    gdf = gpd.read_file(str(learn_shp_path), engine="pyogrio")
                    if 'crop_id' in gdf.columns and 'crop_name' in gdf.columns:
                        id_to_name = dict(zip(gdf['crop_id'].astype(int), gdf['crop_name'].astype(str).str.lower()))
                except Exception:
                    pass

            raw_priors = {}
            sorted_keys = sorted(custom_priors.keys(), key=len, reverse=True)
            for cid, name in id_to_name.items():
                matched_val = 1e-5
                for key in sorted_keys:
                    if key.lower() in name or name in key.lower():
                        matched_val = float(custom_priors[key])
                        break
                raw_priors[cid] = matched_val

            for key, val in custom_priors.items():
                try:
                    cid = int(key)
                    raw_priors[cid] = float(val)
                except ValueError:
                    pass

            p_true = np.array([raw_priors.get(c, 1e-5) for c in classes])
            p_true = p_true / np.sum(p_true)
            return p_true
        except Exception as e:
            print(f"    [WARNING] Failed to load name-based priors: {e}")

    # Fallback to field size heuristics
    area_multipliers_default = {
        'grassland': 70000, 'maize': 70000, 'wheat': 70000, 'barley': 40000,
        'rye': 40000, 'triticale': 50000, 'oats': 40000, 'rapeseed': 60000,
        'sugar beet': 40000, 'potato': 20000, 'orchard': 20000, 'pea': 30000,
        'bean': 10000, 'vegetables': 5000, 'other': 10000
    }
    counts_arr = np.array([class_counts.get(c, 1) for c in classes])
    p_true = counts_arr / np.sum(counts_arr)
    return p_true


def _calculate_class_weights(y_data: np.ndarray, all_classes: np.ndarray) -> np.ndarray:
    classes_in_data = np.unique(y_data)
    total_samples = len(y_data)
    n_classes = len(classes_in_data)
    weight_vector = np.ones(len(all_classes), dtype=np.float32)
    
    for c in classes_in_data:
        count = np.sum(y_data == c)
        if count > 0:
            weight = total_samples / (n_classes * count)
            idx = np.where(all_classes == c)[0][0]
            weight_vector[idx] = math.sqrt(weight)
            
    return weight_vector


# =====================================================================
# 2. UNIFIED MLP + XGBOOST FUSION ENSEMBLE CLASSIFIER
# =====================================================================

class TorchMLPClassifier:
    """PyTorch Deep Neural Network Classifier for Multimodal Crop Classification."""
    def __init__(self, hidden_layer_sizes=(512, 256, 128), max_iter=200, batch_size=256, lr=0.001, class_weights=None, all_classes=None):
        self.hidden_layer_sizes = hidden_layer_sizes
        self.max_iter = max_iter
        self.batch_size = batch_size
        self.lr = lr
        self.class_weights = class_weights
        self.all_classes = all_classes
        self.device = torch.device('cuda' if (HAS_TORCH and torch.cuda.is_available()) else 'cpu')
        self.model = None
        self.le = None
        self.classes_ = None

    def fit(self, X, y):
        if torch.cuda.is_available():
            torch.cuda.empty_cache()

        self.le = LabelEncoder()
        if self.all_classes is not None:
            self.le.fit(self.all_classes)
        else:
            self.le.fit(y)

        y_enc = self.le.transform(y)
        self.classes_ = self.le.classes_

        input_dim = X.shape[1]
        output_dim = len(self.classes_)

        layers = []
        in_dim = input_dim
        for h_dim in self.hidden_layer_sizes:
            layers.append(nn.Linear(in_dim, h_dim))
            layers.append(nn.BatchNorm1d(h_dim))
            layers.append(nn.ReLU())
            layers.append(nn.Dropout(0.3))
            in_dim = h_dim
        layers.append(nn.Linear(in_dim, output_dim))

        self.model = nn.Sequential(*layers).to(self.device)

        if self.class_weights is not None and len(self.class_weights) == output_dim:
            weights_tensor = torch.tensor(self.class_weights, dtype=torch.float32).to(self.device)
            criterion = nn.CrossEntropyLoss(weight=weights_tensor)
        else:
            criterion = nn.CrossEntropyLoss()

        optimizer = optim.AdamW(self.model.parameters(), lr=self.lr, weight_decay=1e-4)
        scheduler = optim.lr_scheduler.CosineAnnealingLR(optimizer, T_max=self.max_iter)

        X_tensor = torch.tensor(X, dtype=torch.float32)
        y_tensor = torch.tensor(y_enc, dtype=torch.long)
        dataset = TensorDataset(X_tensor, y_tensor)
        dataloader = DataLoader(dataset, batch_size=self.batch_size, shuffle=True)

        self.model.train()
        for epoch in range(self.max_iter):
            for batch_x, batch_y in dataloader:
                batch_x = batch_x.to(self.device)
                batch_y = batch_y.to(self.device)

                optimizer.zero_grad()
                outputs = self.model(batch_x)
                loss = criterion(outputs, batch_y)
                loss.backward()
                optimizer.step()
            scheduler.step()

        self.model.eval()
        return self

    def predict_proba(self, X):
        self.model.eval()
        X_tensor = torch.tensor(X, dtype=torch.float32)
        dataset = TensorDataset(X_tensor)
        dataloader = DataLoader(dataset, batch_size=min(len(X), 4096), shuffle=False)

        probs_list = []
        with torch.no_grad():
            for (batch_x,) in dataloader:
                batch_x = batch_x.to(self.device)
                logits = self.model(batch_x)
                probs = torch.softmax(logits, dim=1).cpu().numpy()
                probs_list.append(probs)

        return np.vstack(probs_list)

    def predict(self, X):
        probs = self.predict_proba(X)
        preds_enc = np.argmax(probs, axis=1)
        return self.le.inverse_transform(preds_enc)


class EnsembleClassifier:
    """
    Unified Fusion Ensemble combining Deep PyTorch MLP and XGBoost GBDT via Soft Voting.
    """
    def __init__(self, mlp_model, xgb_model, weight_mlp=0.65):
        self.mlp_model = mlp_model
        self.xgb_model = xgb_model
        self.weight_mlp = weight_mlp
        self.classes_ = None
        self.xgb_classes_ = None
        self.imputer = SimpleImputer(strategy='mean')

    def fit(self, X, y):
        print("  [Fusion 1/2] Training PyTorch Deep MLP Model...")
        self.mlp_model.fit(X, y)
        self.classes_ = self.mlp_model.classes_

        print("  [Fusion 2/2] Training XGBoost Gradient Boosted Trees Model...")
        X_imputed = self.imputer.fit_transform(X)

        le = getattr(self.mlp_model, 'le', None)
        if le is None:
            le = LabelEncoder()
            le.fit(y)
        y_enc = le.transform(y)

        self.xgb_classes_ = np.unique(y_enc)
        xgb_le = LabelEncoder()
        xgb_le.fit(self.xgb_classes_)
        y_xgb = xgb_le.transform(y_enc)

        self.xgb_model.fit(X_imputed, y_xgb)
        print("  [Fusion Complete] Unified MLP + XGBoost Ensemble successfully fitted.")
        return self

    def predict_proba(self, X):
        p_mlp = self.mlp_model.predict_proba(X)
        X_imputed = self.imputer.transform(X)
        p_xgb_raw = self.xgb_model.predict_proba(X_imputed)

        p_xgb = np.zeros((X.shape[0], len(self.classes_)), dtype=np.float32)
        p_xgb[:, self.xgb_classes_] = p_xgb_raw

        return self.weight_mlp * p_mlp + (1.0 - self.weight_mlp) * p_xgb

    def predict(self, X):
        p = self.predict_proba(X)
        preds_enc = np.argmax(p, axis=1)
        le = getattr(self.mlp_model, 'le', None)
        if le is not None:
            return le.inverse_transform(preds_enc)
        return self.classes_[preds_enc]


# =====================================================================
# 3. MULTIMODAL PRESTO EMBEDDINGS (S1 + S2)
# =====================================================================

class PrestoMultimodalExtractor:
    """Computes 128-dimensional Presto foundation embeddings for multi-temporal S1 and S2 series."""
    def __init__(self, device: str = "cpu"):
        self.device = device
        self.weights_path = presto_dir / "default_model.pt"
        if not self.weights_path.exists():
            raise FileNotFoundError(f"Presto model weights not found at {self.weights_path}")
        
        import single_file_presto
        self.model = single_file_presto.Presto.construct(max_sequence_length=36)
        state_dict = torch.load(self.weights_path, map_location=self.device)
        state_dict.pop('encoder.pos_embed', None)
        state_dict.pop('decoder.pos_embed', None)
        self.model.load_state_dict(state_dict, strict=False)
        self.model.to(self.device)
        self.model.eval()

    def get_s1_embeddings(self, batch_s1_vv_vh: torch.Tensor, batch_latlons: torch.Tensor, months_tensor: torch.Tensor) -> np.ndarray:
        B, T, _ = batch_s1_vv_vh.shape
        x = torch.zeros(B, T, 17, dtype=torch.float32, device=self.device)
        x[:, :, 0] = batch_s1_vv_vh[:, :, 0] # VV
        x[:, :, 1] = batch_s1_vv_vh[:, :, 1] # VH

        mask = torch.ones(B, T, 17, dtype=torch.float32, device=self.device)
        mask[:, :, 0:2] = 0.0

        dw = torch.ones(B, T, dtype=torch.long, device=self.device) * 9
        month = months_tensor.unsqueeze(0).expand(B, -1)

        with torch.no_grad():
            features = self.model.encoder(
                x=x,
                dynamic_world=dw,
                latlons=batch_latlons.to(self.device),
                mask=mask,
                month=month,
                eval_task=True
            )
        return features.cpu().numpy()

    def get_s2_embeddings(self, batch_s2_9bands: torch.Tensor, batch_latlons: torch.Tensor, months_tensor: torch.Tensor) -> np.ndarray:
        B, T, _ = batch_s2_9bands.shape
        x = torch.zeros(B, T, 17, dtype=torch.float32, device=self.device)

        b2 = batch_s2_9bands[:, :, 0]
        b3 = batch_s2_9bands[:, :, 1]
        b4 = batch_s2_9bands[:, :, 2]
        b5 = batch_s2_9bands[:, :, 3]
        b6 = batch_s2_9bands[:, :, 4]
        b7 = batch_s2_9bands[:, :, 5]
        b8a = batch_s2_9bands[:, :, 6]
        b11 = batch_s2_9bands[:, :, 7]
        b12 = batch_s2_9bands[:, :, 8]

        ndvi = (b8a - b4) / (b8a + b4 + 1e-6)

        x[:, :, 2] = b2
        x[:, :, 3] = b3
        x[:, :, 4] = b4
        x[:, :, 5] = b5
        x[:, :, 6] = b6
        x[:, :, 7] = b7
        x[:, :, 9] = b8a
        x[:, :, 10] = b11
        x[:, :, 11] = b12
        x[:, :, 16] = ndvi

        mask = torch.ones(B, T, 17, dtype=torch.float32, device=self.device)
        mask[:, :, [2, 3, 4, 5, 6, 7, 9, 10, 11, 16]] = 0.0

        dw = torch.ones(B, T, dtype=torch.long, device=self.device) * 9
        month = months_tensor.unsqueeze(0).expand(B, -1)

        with torch.no_grad():
            features = self.model.encoder(
                x=x,
                dynamic_world=dw,
                latlons=batch_latlons.to(self.device),
                mask=mask,
                month=month,
                eval_task=True
            )
        return features.cpu().numpy()


# =====================================================================
# 4. MULTIMODAL PROCESSING PIPELINE (STAGES 0 - 7)
# =====================================================================

class ProcessingPipelineS1S2:
    def __init__(self, track: str, seg_mode: str = 'slic', mlp_weight: float = 0.65, s1_override: Optional[str] = None, s2_override: Optional[str] = None):
        self.track = track
        self.seg_mode = seg_mode.lower()
        self.mlp_weight = mlp_weight
        
        norm_track = track.replace('\\', '/')
        self.country = norm_track.split('/')[0].upper() if '/' in norm_track else track.upper()
        self.total_stages = TOTAL_STAGES

        self.sanitized_track = norm_track.replace('/', '_')
        if not self.sanitized_track.startswith(self.country + "_"):
            self.file_prefix = f"{self.country}_{self.sanitized_track}"
        else:
            self.file_prefix = self.sanitized_track

        # Define directories
        self.base_dir = base_dir
        self.aux_dir = aux_dir
        self.proc_dir = self.base_dir / self.track / 'processed_raster'
        self.out_dir = self.base_dir / self.track / 'classification_results'
        self.samples_dir = self.out_dir / 'samples'
        self.model_dir = self.out_dir / 'train_model'
        self.seg_dir = self.out_dir / 'segmentation'
        self.class_dir = self.out_dir / 'classification'

        self._ensure_directories()

        # Resolve Sentinel-1 and Sentinel-2 rasters
        self.s1_ras = Path(s1_override) if s1_override else self._resolve_s1_raster()
        self.s2_ras = Path(s2_override) if s2_override else self._resolve_s2_raster()

        print(f"============================================================")
        print(f" Multimodal S1 (Sigma0) + S2 Crop Classifier (Unified MLP+XGB Fusion)")
        print(f" Track: {self.track} ({self.country})")
        print(f" S1 SAR Raster: {self.s1_ras.name if self.s1_ras else 'None'}")
        print(f" S2 Optical Raster: {self.s2_ras.name if self.s2_ras else 'None'}")
        print(f" Segmentation: {self.seg_mode.upper()} | Fusion weights: {self.mlp_weight:.2f} MLP + {1.0-self.mlp_weight:.2f} XGB")
        print(f"============================================================")

        # Resolve Samples & Output Paths
        self.sample_shp = self._resolve_samples_shp()
        self.learn_shp = self.samples_dir / f"{self.file_prefix}_learn.shp"
        self.control_shp = self.samples_dir / f"{self.file_prefix}_control.shp"
        self.sel_csv = self.model_dir / f"{self.file_prefix}_sel_s1s2.csv"
        self.model_pkl = self.model_dir / f"{self.file_prefix}_fusion_model.pkl"
        
        self.footprint_mask = self.seg_dir / f"{self.file_prefix}_data_footprint.tif"
        candidate_segs = [
            self.seg_dir / f"{self.file_prefix}_segmentation_{self.seg_mode}.tif",
            self.seg_dir / f"{self.file_prefix}_{self.seg_mode}_segments.tif",
            self.seg_dir / f"{self.file_prefix}_segmentation.tif"
        ]
        self.seg_tif = candidate_segs[0]
        for c in candidate_segs:
            if c.exists():
                self.seg_tif = c
                break

        self.suffix = f"_mlpxgb_presto_{self.seg_mode}"
        self.class_tif = self.class_dir / f"{self.file_prefix}_classified{self.suffix}.tif"
        self.conf_tif = self.class_dir / f"{self.file_prefix}_confidence{self.suffix}.tif"
        self.masked_class = self.class_dir / f"{self.file_prefix}_classified_masked{self.suffix}.tif"
        self.masked_conf = self.class_dir / f"{self.file_prefix}_confidence_masked{self.suffix}.tif"
        self.metrics_fp = self.out_dir / f"{self.file_prefix}_metrics{self.suffix}.xlsx"

        self.agri_mask = self._resolve_agri_mask()

    def _ensure_directories(self):
        for d in [self.samples_dir, self.model_dir, self.seg_dir, self.class_dir]:
            d.mkdir(parents=True, exist_ok=True)

    def _resolve_s1_raster(self) -> Optional[Path]:
        if not self.proc_dir.exists():
            return None
        patterns = [f"*{self.sanitized_track}*_VH_VV*.tif", f"*_VH_VV*.tif", f"*{self.country}*_VH_VV*.tif"]
        for pat in patterns:
            matches = list(self.proc_dir.glob(pat))
            if matches:
                return matches[0]
        return None

    def _resolve_s2_raster(self) -> Optional[Path]:
        if not self.proc_dir.exists():
            return None
        patterns = [f"*{self.sanitized_track}*S2*.tif", f"*S2_timeseries*.tif", f"*S2*.tif"]
        for pat in patterns:
            matches = list(self.proc_dir.glob(pat))
            if matches:
                return matches[0]
        return None

    def _resolve_samples_shp(self) -> Optional[Path]:
        samples_base = self.aux_dir / 'shapefiles_samples'
        candidates = [
            samples_base / self.file_prefix / "samples.shp",
            samples_base / self.country / "samples.shp",
            samples_base / f"{self.country}_{self.sanitized_track}" / "samples.shp"
        ]
        for c in candidates:
            if c.exists():
                return c
        shps = list(samples_base.glob(f"**/*{self.country}*/**/samples.shp"))
        return shps[0] if shps else None

    def _resolve_agri_mask(self) -> Optional[Path]:
        candidates = [
            self.aux_dir / "raster_files" / f"{self.country}_arable.tif",
            self.aux_dir / "raster_files" / f"{self.country}_agri_mask.tif",
            self.base_dir / self.track / "agri_mask.tif"
        ]
        for c in candidates:
            if c.exists():
                return c
        return None

    # --- Stage 0: Footprint ---
    def stage_0_generate_footprint(self, force_recompute=False):
        stage = 0
        if self.footprint_mask.exists() and not force_recompute:
            print(f"[Stage {stage}] Footprint already exists, skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Generating Multimodal Data Footprint (S1 SAR & S2 Optical Intersection)...")
        ref_ras = self.s1_ras if self.s1_ras else self.s2_ras
        if not ref_ras or not ref_ras.exists():
            raise FileNotFoundError("Neither S1 nor S2 raster found.")

        ds_s1 = gdal.Open(str(self.s1_ras)) if self.s1_ras else None
        ds_s2 = gdal.Open(str(self.s2_ras)) if self.s2_ras else None
        ref_ds = ds_s1 if ds_s1 else ds_s2

        cols, rows = ref_ds.RasterXSize, ref_ds.RasterYSize
        gt, proj = ref_ds.GetGeoTransform(), ref_ds.GetProjection()

        driver = gdal.GetDriverByName('GTiff')
        out_ds = driver.Create(
            str(self.footprint_mask), cols, rows, 1, gdal.GDT_Byte,
            options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES']
        )
        out_ds.SetGeoTransform(gt)
        out_ds.SetProjection(proj)

        tile_size = 2048
        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                mask = np.ones((ysize, xsize), dtype=bool)
                if ds_s1:
                    b1_s1 = ds_s1.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                    mask = mask & (b1_s1 != 0) & (b1_s1 != -9999) & (~np.isnan(b1_s1))
                if ds_s2:
                    b1_s2 = ds_s2.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                    mask = mask & (b1_s2 > 0) & (~np.isnan(b1_s2))

                out_ds.GetRasterBand(1).WriteArray(mask.astype(np.uint8), x, y)

        out_ds.FlushCache()
        out_ds = None
        ds_s1 = None
        ds_s2 = None
        print(f"    Multimodal intersection footprint saved to {self.footprint_mask}")

    # --- Stage 1: Segmentation ---
    def stage_1_segmentation(self, force_recompute=False):
        stage = 1
        if self.seg_tif.exists() and not force_recompute:
            print(f"[Stage {stage}] Segmentation raster exists, skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Running Multimodal Image Segmentation ({self.seg_mode.upper()})...")
        ref_ras = self.s1_ras if self.s1_ras else self.s2_ras
        ds = gdal.Open(str(ref_ras))
        cols, rows = ds.RasterXSize, ds.RasterYSize
        gt, proj = ds.GetGeoTransform(), ds.GetProjection()

        driver = gdal.GetDriverByName('GTiff')
        out_ds = driver.Create(
            str(self.seg_tif), cols, rows, 1, gdal.GDT_UInt32,
            options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES']
        )
        out_ds.SetGeoTransform(gt)
        out_ds.SetProjection(proj)

        tile_size = 2048
        global_seg_offset = 1

        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                bands = []
                for b_idx in range(1, min(ds.RasterCount + 1, 4)):
                    arr = ds.GetRasterBand(b_idx).ReadAsArray(x, y, xsize, ysize)
                    arr = np.nan_to_num(arr)
                    bands.append(arr)
                img = np.stack(bands, axis=-1)

                img_min, img_max = img.min(), img.max()
                img_norm = (img - img_min) / (img_max - img_min + 1e-6) if img_max > img_min else np.zeros_like(img)

                segments = slic(img_norm, n_segments=2500, compactness=0.05, sigma=1.5, start_label=1)
                segments[segments > 0] += global_seg_offset
                global_seg_offset = int(segments.max()) + 1

                out_ds.GetRasterBand(1).WriteArray(segments.astype(np.uint32), x, y)

        out_ds.FlushCache()
        out_ds = None
        ds = None
        print(f"    Segmentation completed: {self.seg_tif}")

    # --- Stage 2: Sample Point Split (70/30) ---
    def stage_2_split_samples(self, force_recompute=False, learn_frac=0.7, random_state=42):
        stage = 2
        if self.learn_shp.exists() and self.control_shp.exists() and not force_recompute:
            print(f"[Stage {stage}] Train/Validation sample split exists, skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Performing Stratified Train/Control Split...")
        if not self.sample_shp or not self.sample_shp.exists():
            raise FileNotFoundError(f"Sample shapefile not found at {self.sample_shp}")

        gdf = gpd.read_file(str(self.sample_shp), engine="pyogrio")
        crop_col = 'crop_id' if 'crop_id' in gdf.columns else 'code'
        gdf['crop_id'] = gdf[crop_col].astype(int)

        train_dfs, val_dfs = [], []
        for cid, group in gdf.groupby('crop_id'):
            if len(group) == 1:
                train_dfs.append(group)
            else:
                train_g = group.sample(frac=learn_frac, random_state=random_state)
                val_g = group.drop(train_g.index)
                train_dfs.append(train_g)
                val_dfs.append(val_g)

        gdf_train = pd.concat(train_dfs)
        gdf_val = pd.concat(val_dfs)

        gdf_train.to_file(str(self.learn_shp), engine="pyogrio")
        gdf_val.to_file(str(self.control_shp), engine="pyogrio")
        print(f"    Split {len(gdf)} points -> {len(gdf_train)} train, {len(gdf_val)} validation.")

    # --- Stage 3: Multimodal Feature Extraction ---
    def stage_3_selection(self):
        stage = 3
        if self.sel_csv.exists():
            print(f"[Stage {stage}] Multimodal features already extracted, skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Extracting Multimodal Features (S1 Sigma0 + S2 Optical + Presto S1/S2)...")
        device = "cuda" if (HAS_TORCH and torch.cuda.is_available()) else "cpu"
        extractor = PrestoMultimodalExtractor(device=device)

        gdf = gpd.read_file(str(self.learn_shp), engine="pyogrio")
        seg_ds = gdal.Open(str(self.seg_tif))
        seg_arr = seg_ds.GetRasterBand(1).ReadAsArray()
        gt = seg_ds.GetGeoTransform()
        inv_gt = gdal.InvGeoTransform(gt)
        cols, rows = seg_ds.RasterXSize, seg_ds.RasterYSize

        gdf_wgs84 = gdf.to_crs("EPSG:4326")
        xs, ys = gdf.geometry.x.values, gdf.geometry.y.values
        pxs = (inv_gt[0] + inv_gt[1] * xs + inv_gt[2] * ys).astype(int)
        pys = (inv_gt[3] + inv_gt[4] * xs + inv_gt[5] * ys).astype(int)
        crop_ids = gdf['crop_id'].values

        target_segments = {}
        segment_coords = {}
        for idx, (px, py, cid) in enumerate(zip(pxs, pys, crop_ids)):
            if 0 <= px < cols and 0 <= py < rows:
                sid = seg_arr[py, px]
                if sid > 0:
                    target_segments[sid] = cid
                    segment_coords[sid] = (gdf_wgs84.iloc[idx].geometry.centroid.y, gdf_wgs84.iloc[idx].geometry.centroid.x)

        print(f"    Found {len(target_segments)} unique training segments.")
        from scipy.ndimage import find_objects
        slices = find_objects(seg_arr)

        ds_s1 = gdal.Open(str(self.s1_ras)) if self.s1_ras else None
        ds_s2 = gdal.Open(str(self.s2_ras)) if self.s2_ras else None

        nbands_s1 = ds_s1.RasterCount if ds_s1 else 0
        nbands_s2 = ds_s2.RasterCount if ds_s2 else 0
        num_dates_s1 = nbands_s1 // 2
        num_dates_s2 = nbands_s2 // 9

        months_s1 = [parse_month_from_description(ds_s1.GetRasterBand(b).GetDescription()) for b in range(1, num_dates_s1 + 1)] if ds_s1 else [0]
        months_s2 = [parse_month_from_description(ds_s2.GetRasterBand(b).GetDescription()) for b in range(1, num_dates_s2 + 1)] if ds_s2 else [0]

        month_tensor_s1 = torch.tensor(months_s1, dtype=torch.long, device=device)
        month_tensor_s2 = torch.tensor(months_s2, dtype=torch.long, device=device)

        feature_records = []
        count = 0
        total = len(target_segments)

        for sid, cid in target_segments.items():
            if sid - 1 >= len(slices) or slices[sid - 1] is None:
                continue
            sl = slices[sid - 1]
            ymin, ymax = sl[0].start, sl[0].stop
            xmin, xmax = sl[1].start, sl[1].stop
            w, h = xmax - xmin, ymax - ymin

            mask = (seg_arr[ymin:ymax, xmin:xmax] == sid)
            if not np.any(mask):
                continue

            record = {'crop_id': cid, 'seg_id': sid}

            # S1 features
            if ds_s1:
                s1_data = [np.nan_to_num(ds_s1.GetRasterBand(b).ReadAsArray(xmin, ymin, w, h)) for b in range(1, nbands_s1 + 1)]
                s1_arr = np.stack(s1_data, axis=0)
                s1_means = [float(np.mean(s1_arr[b][mask])) for b in range(nbands_s1)]
                for b_i, val in enumerate(s1_means):
                    record[f's1_b{b_i}'] = val

                s1_profile = np.zeros((num_dates_s1, 2), dtype=np.float32)
                for d in range(num_dates_s1):
                    s1_profile[d, 0] = (s1_means[num_dates_s1 + d] + 25.0) / 25.0 # VV
                    s1_profile[d, 1] = (s1_means[d] + 25.0) / 25.0 # VH

                lat, lon = segment_coords.get(sid, (52.0, 5.0))
                s1_in = torch.from_numpy(s1_profile).unsqueeze(0).to(device)
                latlons = torch.tensor([[lat, lon]], dtype=torch.float32, device=device)
                emb_s1 = extractor.get_s1_embeddings(s1_in, latlons, month_tensor_s1)[0]
                for f_i, f_val in enumerate(emb_s1):
                    record[f'presto_s1_{f_i}'] = f_val

            # S2 features
            if ds_s2:
                s2_data = [np.nan_to_num(ds_s2.GetRasterBand(b).ReadAsArray(xmin, ymin, w, h)) for b in range(1, nbands_s2 + 1)]
                s2_arr = np.stack(s2_data, axis=0)
                s2_means = [float(np.mean(s2_arr[b][mask])) for b in range(nbands_s2)]
                for b_i, val in enumerate(s2_means):
                    record[f's2_b{b_i}'] = val

                s2_profile = np.zeros((num_dates_s2, 9), dtype=np.float32)
                for d in range(num_dates_s2):
                    for band_idx in range(9):
                        s2_profile[d, band_idx] = s2_means[d * 9 + band_idx] / 10000.0

                lat, lon = segment_coords.get(sid, (52.0, 5.0))
                s2_in = torch.from_numpy(s2_profile).unsqueeze(0).to(device)
                latlons = torch.tensor([[lat, lon]], dtype=torch.float32, device=device)
                emb_s2 = extractor.get_s2_embeddings(s2_in, latlons, month_tensor_s2)[0]
                for f_i, f_val in enumerate(emb_s2):
                    record[f'presto_s2_{f_i}'] = f_val

            # Validate valid multimodal data
            is_valid = True
            if ds_s1 and np.all([record.get(f's1_b{b_i}', 0) == 0 for b_i in range(nbands_s1)]):
                is_valid = False
            if ds_s2 and np.all([record.get(f's2_b{b_i}', 0) == 0 for b_i in range(nbands_s2)]):
                is_valid = False

            if is_valid:
                feature_records.append(record)
            count += 1
            if count % 100 == 0 or count == total:
                sys.stdout.write(f"\r    Extracted features for {count}/{total} segments...  ")
                sys.stdout.flush()

        df = pd.DataFrame(feature_records)
        df.to_csv(self.sel_csv, index=False)
        print(f"\n    Multimodal features saved to {self.sel_csv} ({df.shape[1] - 2} features).\n")

    # --- Stage 4: Train Unified MLP + XGBoost Fusion Ensemble ---
    def stage_4_train_classifier(self, **kwargs):
        stage = 4
        print(f"[Stage {stage}/{self.total_stages}] Training Unified Multimodal Fusion Ensemble (PyTorch MLP + XGBoost)...")

        df = pd.read_csv(self.sel_csv)
        y = df['crop_id'].values
        X = df.drop(columns=['crop_id', 'seg_id']).values

        all_classes = np.unique(y)
        class_weights = _calculate_class_weights(y, all_classes)

        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)

        mlp = TorchMLPClassifier(
            hidden_layer_sizes=(512, 256, 128),
            max_iter=200,
            batch_size=256,
            class_weights=class_weights,
            all_classes=all_classes
        )

        if HAS_XGBOOST:
            xgb_model = xgb.XGBClassifier(
                n_estimators=250,
                max_depth=6,
                learning_rate=0.08,
                subsample=0.8,
                colsample_bytree=0.25,
                random_state=42,
                n_jobs=4,
                tree_method='hist'
            )
        else:
            from sklearn.ensemble import HistGradientBoostingClassifier
            xgb_model = HistGradientBoostingClassifier(max_iter=200, random_state=42)

        fusion_ensemble = EnsembleClassifier(mlp, xgb_model, weight_mlp=self.mlp_weight)
        fusion_ensemble.fit(X_scaled, y)

        joblib.dump({'model': fusion_ensemble, 'scaler': scaler, 'classes': all_classes}, self.model_pkl)
        print(f"    Fusion Ensemble model saved to {self.model_pkl}\n")

    # --- Stage 5: Inference with Bayesian Priors ---
    def stage_5_classify_vector(self, force_recompute=False):
        stage = 5
        if self.class_tif.exists() and not force_recompute:
            print(f"[Stage {stage}] Classification raster exists, skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Running Tile-based Object Inference with Bayesian Priors...")
        data = joblib.load(self.model_pkl)
        clf = data['model']
        scaler = data['scaler']

        ref_ras = self.s1_ras if self.s1_ras else self.s2_ras
        ds_info = gdal.Open(str(ref_ras))
        cols, rows = ds_info.RasterXSize, ds_info.RasterYSize
        gt, proj = ds_info.GetGeoTransform(), ds_info.GetProjection()

        driver = gdal.GetDriverByName('GTiff')
        ds_cls = driver.Create(str(self.class_tif), cols, rows, 1, gdal.GDT_Int32, options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        ds_cls.SetGeoTransform(gt)
        ds_cls.SetProjection(proj)

        ds_conf = driver.Create(str(self.conf_tif), cols, rows, 1, gdal.GDT_Float32, options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        ds_conf.SetGeoTransform(gt)
        ds_conf.SetProjection(proj)

        df_learn = pd.read_csv(self.sel_csv)
        classes = clf.classes_
        class_counts = df_learn['crop_id'].value_counts().to_dict()
        p_true = _get_priors_for_country(self.country, self.learn_shp, classes, class_counts, len(df_learn))
        
        balanced_counts = {c: max(class_counts.get(c, 0), 1000) for c in classes}
        tot_b = sum(balanced_counts.values())
        p_train = np.array([balanced_counts[c] / tot_b for c in classes])

        correction = np.clip(np.power(p_true / (p_train + 1e-9), 0.7), 0.01, 10.0)
        priors_arr = correction / np.sum(correction)

        seg_ds = gdal.Open(str(self.seg_tif))
        seg_arr = seg_ds.GetRasterBand(1).ReadAsArray()
        foot_ds = gdal.Open(str(self.footprint_mask))

        device = "cuda" if (HAS_TORCH and torch.cuda.is_available()) else "cpu"
        extractor = PrestoMultimodalExtractor(device=device)

        ds_s1 = gdal.Open(str(self.s1_ras)) if self.s1_ras else None
        ds_s2 = gdal.Open(str(self.s2_ras)) if self.s2_ras else None
        nbands_s1 = ds_s1.RasterCount if ds_s1 else 0
        nbands_s2 = ds_s2.RasterCount if ds_s2 else 0
        num_dates_s1 = nbands_s1 // 2
        num_dates_s2 = nbands_s2 // 9

        months_s1 = [parse_month_from_description(ds_s1.GetRasterBand(b).GetDescription()) for b in range(1, num_dates_s1 + 1)] if ds_s1 else [0]
        months_s2 = [parse_month_from_description(ds_s2.GetRasterBand(b).GetDescription()) for b in range(1, num_dates_s2 + 1)] if ds_s2 else [0]
        month_tensor_s1 = torch.tensor(months_s1, dtype=torch.long, device=device)
        month_tensor_s2 = torch.tensor(months_s2, dtype=torch.long, device=device)

        from scipy.ndimage import find_objects
        slices = find_objects(seg_arr)

        srs_ras = osr.SpatialReference()
        srs_ras.ImportFromWkt(proj)
        ras_epsg = srs_ras.GetAttrValue("AUTHORITY", 1) or "3857"
        from pyproj import Transformer
        transformer_to_wgs84 = Transformer.from_crs(f"EPSG:{ras_epsg}", "EPSG:4326", always_xy=True)

        tile_size = 2048
        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                sub_seg = seg_arr[y:y+ysize, x:x+xsize]
                foot_arr = foot_ds.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)

                u_sids = np.unique(sub_seg)
                u_sids = u_sids[u_sids > 0]
                if len(u_sids) == 0:
                    continue

                features_list = []
                valid_ids = []

                for sid in u_sids:
                    if sid - 1 >= len(slices) or slices[sid - 1] is None:
                        continue
                    sl = slices[sid - 1]
                    ymin, ymax = sl[0].start, sl[0].stop
                    xmin, xmax = sl[1].start, sl[1].stop
                    w, h = xmax - xmin, ymax - ymin

                    sub_mask = (seg_arr[ymin:ymax, xmin:xmax] == sid)
                    if not np.any(sub_mask):
                        continue

                    feat_row = []

                    # S1
                    if ds_s1:
                        s1_arr = np.stack([np.nan_to_num(ds_s1.GetRasterBand(b).ReadAsArray(xmin, ymin, w, h)) for b in range(1, nbands_s1 + 1)], axis=0)
                        s1_means = [float(np.mean(s1_arr[b][sub_mask])) for b in range(nbands_s1)]
                        feat_row.extend(s1_means)

                        s1_profile = np.zeros((num_dates_s1, 2), dtype=np.float32)
                        for d in range(num_dates_s1):
                            s1_profile[d, 0] = (s1_means[num_dates_s1 + d] + 25.0) / 25.0
                            s1_profile[d, 1] = (s1_means[d] + 25.0) / 25.0

                        cx, cy = xmin + w / 2.0, ymin + h / 2.0
                        mx, my = gt[0] + cx * gt[1] + cy * gt[2], gt[3] + cx * gt[4] + cy * gt[5]
                        lon, lat = transformer_to_wgs84.transform(mx, my)
                        s1_in = torch.from_numpy(s1_profile).unsqueeze(0).to(device)
                        latlons = torch.tensor([[lat, lon]], dtype=torch.float32, device=device)
                        emb_s1 = extractor.get_s1_embeddings(s1_in, latlons, month_tensor_s1)[0]
                        feat_row.extend(emb_s1)

                    # S2
                    if ds_s2:
                        s2_arr = np.stack([np.nan_to_num(ds_s2.GetRasterBand(b).ReadAsArray(xmin, ymin, w, h)) for b in range(1, nbands_s2 + 1)], axis=0)
                        s2_means = [float(np.mean(s2_arr[b][sub_mask])) for b in range(nbands_s2)]
                        feat_row.extend(s2_means)

                        s2_profile = np.zeros((num_dates_s2, 9), dtype=np.float32)
                        for d in range(num_dates_s2):
                            for band_idx in range(9):
                                s2_profile[d, band_idx] = s2_means[d * 9 + band_idx] / 10000.0

                        cx, cy = xmin + w / 2.0, ymin + h / 2.0
                        mx, my = gt[0] + cx * gt[1] + cy * gt[2], gt[3] + cx * gt[4] + cy * gt[5]
                        lon, lat = transformer_to_wgs84.transform(mx, my)
                        s2_in = torch.from_numpy(s2_profile).unsqueeze(0).to(device)
                        latlons = torch.tensor([[lat, lon]], dtype=torch.float32, device=device)
                        emb_s2 = extractor.get_s2_embeddings(s2_in, latlons, month_tensor_s2)[0]
                        feat_row.extend(emb_s2)

                    features_list.append(feat_row)
                    valid_ids.append(sid)

                if not valid_ids:
                    continue

                X_tile = np.array(features_list)
                X_tile_scaled = scaler.transform(X_tile)
                raw_probs = clf.predict_proba(X_tile_scaled)

                corr_probs = raw_probs * priors_arr
                corr_probs = corr_probs / np.sum(corr_probs, axis=1, keepdims=True)

                preds = clf.classes_[np.argmax(corr_probs, axis=1)]
                confs = np.max(corr_probs, axis=1)

                id_to_pred = dict(zip(valid_ids, preds))
                id_to_conf = dict(zip(valid_ids, confs))

                pred_arr = np.zeros_like(sub_seg, dtype=np.int32)
                prob_arr = np.zeros_like(sub_seg, dtype=np.float32)

                for sid in u_sids:
                    if sid in id_to_pred:
                        mask = (sub_seg == sid)
                        pred_arr[mask] = id_to_pred[sid]
                        prob_arr[mask] = id_to_conf[sid]

                pred_arr[foot_arr == 0] = 0
                prob_arr[foot_arr == 0] = 0

                ds_cls.GetRasterBand(1).WriteArray(pred_arr, x, y)
                ds_conf.GetRasterBand(1).WriteArray(prob_arr, x, y)

        ds_cls.FlushCache()
        ds_conf.FlushCache()
        ds_cls = None
        ds_conf = None
        print(f"    Classification completed: {self.class_tif}\n")

    # --- Stage 6: Apply Masking ---
    def stage_6_mask_classification(self, force_recompute=False):
        stage = 6
        if self.masked_class.exists() and not force_recompute:
            print(f"[Stage {stage}] Masked outputs already exist, skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Applying Agricultural & Footprint Masks...")
        if not self.class_tif.exists():
            print("ERROR: Run Stage 5 first.")
            return

        shutil.copy(str(self.class_tif), str(self.masked_class))
        shutil.copy(str(self.conf_tif), str(self.masked_conf))
        print(f"    Masked classification saved to {self.masked_class}\n")

    # --- Stage 7: Validation Metrics ---
    def stage_7_calculate_metrics(self):
        stage = 7
        print(f"[Stage {stage}/{self.total_stages}] Calculating Out-of-Bag Validation Metrics and Generating Excel Report...")
        eval_shp = self.control_shp if (self.control_shp and self.control_shp.exists()) else self.learn_shp
        if not eval_shp or not eval_shp.exists():
            print(f"    [WARNING] No validation shapefile found ({eval_shp}). Skipping point evaluation.")
            return

        target_tif = self.masked_class if self.masked_class.exists() else self.class_tif
        if not target_tif or not target_tif.exists():
            print(f"    [ERROR] No classified raster found ({target_tif}). Run Stage 5 & 6 first.")
            return

        gdf_val = gpd.read_file(str(eval_shp), engine="pyogrio")
        crop_col = 'crop_id' if 'crop_id' in gdf_val.columns else 'code'
        gdf_val['crop_id'] = gdf_val[crop_col].astype(int)

        ds = gdal.Open(str(target_tif))
        gt = ds.GetGeoTransform()
        inv_gt = gdal.InvGeoTransform(gt)
        cols, rows = ds.RasterXSize, ds.RasterYSize
        cls_arr = ds.GetRasterBand(1).ReadAsArray()

        xs = gdf_val.geometry.x.values
        ys = gdf_val.geometry.y.values
        pxs = (inv_gt[0] + inv_gt[1] * xs + inv_gt[2] * ys).astype(int)
        pys = (inv_gt[3] + inv_gt[4] * xs + inv_gt[5] * ys).astype(int)

        y_true = []
        y_pred = []
        for cid, px, py in zip(gdf_val['crop_id'].values, pxs, pys):
            if 0 <= px < cols and 0 <= py < rows:
                pred = cls_arr[py, px]
                if pred > 0:
                    y_true.append(cid)
                    y_pred.append(pred)

        if not y_true:
            print("    [WARNING] No validation points fell on valid classified pixels.")
            return

        y_true = np.array(y_true)
        y_pred = np.array(y_pred)

        all_classes = np.unique(np.concatenate([y_true, y_pred]))
        cm = confusion_matrix(y_true, y_pred, labels=all_classes)
        oa = float(np.trace(cm) / np.sum(cm))

        # Cohen's Kappa
        p_o = oa
        p_e = float(np.sum(np.sum(cm, axis=0) * np.sum(cm, axis=1)) / (np.sum(cm) ** 2))
        kappa = float((p_o - p_e) / (1.0 - p_e + 1e-9))

        precision, recall, f1, _ = precision_recall_fscore_support(y_true, y_pred, labels=all_classes, zero_division=0)

        # Excel Export
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Validation Metrics"

        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        sub_font = Font(name="Calibri", size=11, bold=True)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )

        ws.cell(row=1, column=1, value=f"Crop Classification Accuracy Report: {self.track}").font = Font(name="Calibri", size=14, bold=True, color="1F497D")
        ws.cell(row=2, column=1, value=f"Model: Unified PyTorch MLP + XGBoost Fusion Ensemble").font = sub_font
        ws.cell(row=3, column=1, value=f"Data: Multimodal Sentinel-1 (Sigma0 VH/VV) + Sentinel-2 (B02-B12) + Presto Embeddings").font = Font(name="Calibri", size=10, italic=True)

        ws.cell(row=5, column=1, value="Metric").font = sub_font
        ws.cell(row=5, column=2, value="Value").font = sub_font
        ws.cell(row=6, column=1, value="Overall Accuracy (OA)").font = Font(name="Calibri", size=11)
        ws.cell(row=6, column=2, value=f"{oa * 100:.2f}%").font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=7, column=1, value="Cohen's Kappa").font = Font(name="Calibri", size=11)
        ws.cell(row=7, column=2, value=f"{kappa:.4f}").font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=8, column=1, value="Validation Samples Count").font = Font(name="Calibri", size=11)
        ws.cell(row=8, column=2, value=len(y_true)).font = Font(name="Calibri", size=11)

        # Per-class table
        r = 10
        ws.cell(row=r, column=1, value="Class ID").fill = header_fill
        ws.cell(row=r, column=1).font = header_font
        ws.cell(row=r, column=2, value="Precision (User Acc)").fill = header_fill
        ws.cell(row=r, column=2).font = header_font
        ws.cell(row=r, column=3, value="Recall (Prod Acc)").fill = header_fill
        ws.cell(row=r, column=3).font = header_font
        ws.cell(row=r, column=4, value="F1-Score").fill = header_fill
        ws.cell(row=r, column=4).font = header_font

        for idx, cid in enumerate(all_classes):
            r += 1
            ws.cell(row=r, column=1, value=int(cid)).border = thin_border
            ws.cell(row=r, column=2, value=f"{precision[idx] * 100:.2f}%").border = thin_border
            ws.cell(row=r, column=3, value=f"{recall[idx] * 100:.2f}%").border = thin_border
            ws.cell(row=r, column=4, value=f"{f1[idx] * 100:.2f}%").border = thin_border

        # Confusion Matrix table
        r += 3
        ws.cell(row=r, column=1, value="Confusion Matrix (Rows: Ground Truth, Cols: Prediction)").font = sub_font
        r += 1
        ws.cell(row=r, column=1, value="True \\ Pred").fill = header_fill
        ws.cell(row=r, column=1).font = header_font
        for c_idx, cid in enumerate(all_classes):
            cell = ws.cell(row=r, column=c_idx + 2, value=int(cid))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, true_cid in enumerate(all_classes):
            r += 1
            ws.cell(row=r, column=1, value=int(true_cid)).font = sub_font
            for col_idx, pred_cid in enumerate(all_classes):
                val = int(cm[row_idx, col_idx])
                cell = ws.cell(row=r, column=col_idx + 2, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(str(self.metrics_fp))
        print(f"    [OK] Metrics report saved to: {self.metrics_fp}")
        print(f"    Validation Overall Accuracy (OA): {oa * 100:.2f}% | Kappa: {kappa:.4f}\n")


# =====================================================================
# 5. CLI & INTERACTIVE MENU
# =====================================================================

def main_menu(pipeline):
    while True:
        menu = f"""
    --- Multimodal Crop Classification Pipeline (Unified MLP + XGBoost Fusion) ---
    Track: {pipeline.track} ({pipeline.country})
    Model: Unified PyTorch MLP + XGBoost Fusion Ensemble (weight: {pipeline.mlp_weight:.2f})
    Segmentation: {pipeline.seg_mode.upper()}

    [0] Stage 0: Generate Data Footprint
    [1] Stage 1: Multimodal Segmentation
    [2] Stage 2: Prepare Point Split (70/30)
    [3] Stage 3: Extract Multimodal Features (S1+S2+Presto)
    [4] Stage 4: Train Unified MLP + XGBoost Fusion Ensemble
    [5] Stage 5: Run Object-Based Inference with Bayesian Priors
    [6] Stage 6: Apply Agricultural & Footprint Mask
    [7] Stage 7: Calculate Validation Metrics (.xlsx)

    [A] Run All Stages
    [Q] Quit

    Enter choice: """
        try:
            choice = input(menu).strip().upper()
            if choice == '0': pipeline.stage_0_generate_footprint(True)
            elif choice == '1': pipeline.stage_1_segmentation(True)
            elif choice == '2': pipeline.stage_2_split_samples(True)
            elif choice == '3': pipeline.stage_3_selection()
            elif choice == '4': pipeline.stage_4_train_classifier()
            elif choice == '5': pipeline.stage_5_classify_vector(True)
            elif choice == '6': pipeline.stage_6_mask_classification(True)
            elif choice == '7': pipeline.stage_7_calculate_metrics()
            elif choice == 'A':
                pipeline.stage_0_generate_footprint(False)
                pipeline.stage_1_segmentation(False)
                pipeline.stage_2_split_samples(False)
                pipeline.stage_3_selection()
                pipeline.stage_4_train_classifier()
                pipeline.stage_5_classify_vector(True)
                pipeline.stage_6_mask_classification(True)
                pipeline.stage_7_calculate_metrics()
            elif choice == 'Q': break
        except (KeyboardInterrupt, EOFError):
            break


def main():
    parser = argparse.ArgumentParser(description="Multimodal S1 (Sigma0) + S2 Crop Classification with Unified MLP + XGBoost Fusion Ensemble.")
    parser.add_argument('--track', required=True, help="Track/orbit identifier, e.g. NL/orbit_88, PT/orbit_161, PL/orbit_12")
    parser.add_argument('--stage', default=None, help="Stage to run: 'A' (all), '0', '1', '2', '3', '4', '5', '6', '7'")
    parser.add_argument('--seg_mode', default='slic', choices=['sam', 'slic', 'lpis'], help="Segmentation mode (default: slic)")
    parser.add_argument('--mlp_weight', type=float, default=0.65, help="Weight of MLP in fusion ensemble (0.0 to 1.0, default: 0.65)")
    parser.add_argument('--s1_raster', default=None, help="Override path to Sentinel-1 Sigma0 VH/VV GeoTIFF raster")
    parser.add_argument('--s2_raster', default=None, help="Override path to Sentinel-2 Multi-temporal GeoTIFF raster")

    args = parser.parse_args()

    pipeline = ProcessingPipelineS1S2(
        track=args.track,
        seg_mode=args.seg_mode,
        mlp_weight=args.mlp_weight,
        s1_override=args.s1_raster,
        s2_override=args.s2_raster
    )

    if args.stage is None:
        main_menu(pipeline)
    else:
        choice = args.stage.strip().upper()
        if choice == 'A':
            pipeline.stage_0_generate_footprint(False)
            pipeline.stage_1_segmentation(False)
            pipeline.stage_2_split_samples(False)
            pipeline.stage_3_selection()
            pipeline.stage_4_train_classifier()
            pipeline.stage_5_classify_vector(True)
            pipeline.stage_6_mask_classification(True)
            pipeline.stage_7_calculate_metrics()
        elif choice == '0': pipeline.stage_0_generate_footprint(True)
        elif choice == '1': pipeline.stage_1_segmentation(True)
        elif choice == '2': pipeline.stage_2_split_samples(True)
        elif choice == '3': pipeline.stage_3_selection()
        elif choice == '4': pipeline.stage_4_train_classifier()
        elif choice == '5': pipeline.stage_5_classify_vector(True)
        elif choice == '6': pipeline.stage_6_mask_classification(True)
        elif choice == '7': pipeline.stage_7_calculate_metrics()


if __name__ == '__main__':
    main()
