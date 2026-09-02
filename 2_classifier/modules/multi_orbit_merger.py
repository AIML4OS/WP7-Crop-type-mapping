# How to run the script:
# ---------------------
# Country-Wide Merge (combines all processed orbits for a country e.g. NL, PL, PT, AT, FR)
#
# 1. Multimodal S1 + S2 Presto + MLP/XGBoost Fusion Ensemble (Recommended):
#    python 2_merge_classifications.py --track NL --suffix _mlpxgb_presto_slic
#    python 2_merge_classifications.py --track NL --suffix _mlpxgb_presto_lpis
#    python 2_merge_classifications.py --track NL --suffix _mlpxgb_presto_sam
#    python 2_merge_classifications.py --track PL --suffix _mlpxgb_presto_slic
#    python 2_merge_classifications.py --track PT --suffix _mlpxgb_presto_slic
#
# 2. NASA Harvest Presto + SAR Hybrid classification:
#    python 2_merge_classifications.py --track NL --suffix _presto_hybrid_lpis
#    python 2_merge_classifications.py --track NL --suffix _presto_hybrid_slic
#    python 2_merge_classifications.py --track NL --suffix _presto_hybrid_sam
#
# 3. Standard ANN classification (legacy):
#    python 2_merge_classifications.py --track NL --suffix _lpis
#    python 2_merge_classifications.py --track NL --suffix _slic
#    python 2_merge_classifications.py --track NL --suffix _sam
#
# 4. NASA Harvest Presto SAR-only classification (legacy):
#    python 2_merge_classifications.py --track NL --suffix _presto_lpis
#    python 2_merge_classifications.py --track NL --suffix _presto_slic
#    python 2_merge_classifications.py --track NL --suffix _presto_sam
#
# 5. Prithvi Foundation Model classification (legacy):
#    python 2_merge_classifications.py --track NL --suffix _prithvi_lpis
#    python 2_merge_classifications.py --track NL --suffix _prithvi_slic
#    python 2_merge_classifications.py --track NL --suffix _prithvi_sam

import argparse
import os
from pathlib import Path
from typing import List, Tuple
import json
import numpy as np
from osgeo import gdal
import geopandas as gpd
from sklearn.metrics import confusion_matrix, precision_recall_fscore_support
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def get_crop_aggregation(country, learn_shp_path):
    """
    Returns an empty dictionary because the grassland-related classes (Clover/Lucerne)
    have been directly integrated into Grassland in the shapefile itself.
    """
    return {}


# Discover tracks logic

def get_masked_filenames(track_prefix: str, suffix: str) -> List[Tuple[str, str]]:
    """Returns candidate (classification_masked, confidence_masked) filename pairs."""
    candidates = []

    # 1. Direct standard patterns
    candidates.append((f"{track_prefix}_classified_masked{suffix}.tif", f"{track_prefix}_confidence_masked{suffix}.tif"))
    candidates.append((f"{track_prefix}_classified{suffix}_masked.tif", f"{track_prefix}_confidence{suffix}_masked.tif"))

    # 2. Presto Hybrid specific
    if suffix.startswith('_presto_hybrid'):
        s_clean = suffix[14:]  # remove '_presto_hybrid'
        candidates.append((f"{track_prefix}_presto_hybrid_classified_masked{s_clean}.tif", f"{track_prefix}_presto_hybrid_confidence_masked{s_clean}.tif"))
        candidates.append((f"{track_prefix}_presto_hybrid_classified{s_clean}_masked.tif", f"{track_prefix}_presto_hybrid_confidence{s_clean}_masked.tif"))

    # 3. Presto SAR specific
    elif suffix.startswith('_presto'):
        s_clean = suffix[7:]  # remove '_presto'
        candidates.append((f"{track_prefix}_presto_classified_masked{s_clean}.tif", f"{track_prefix}_presto_confidence_masked{s_clean}.tif"))

    # 4. Prithvi specific
    elif suffix.startswith('_prithvi'):
        s_clean = suffix[8:]  # remove '_prithvi'
        candidates.append((f"{track_prefix}_prithvi_classified_masked{s_clean}.tif", f"{track_prefix}_prithvi_confidence_masked{s_clean}.tif"))

    # 5. MLPXGB / S1S2 specific
    if 'mlpxgb' in suffix or 's1s2' in suffix:
        candidates.append((f"{track_prefix}_classified_s1s2_masked.tif", f"{track_prefix}_confidence_s1s2_masked.tif"))
        candidates.append((f"{track_prefix}_classified_masked_s1s2.tif", f"{track_prefix}_confidence_masked_s1s2.tif"))

    return candidates


def find_masked_files(base_dir: Path, tr: str, country: str, suffix: str = ""):
    """
    Look for classification and confidence masked rasters in either:
      - classification_results/classification/
      - classification_results/
    Returns (cls_fp, conf_fp) or (None, None).
    """
    folders = [
        base_dir / tr / 'classification_results' / 'classification',
        base_dir / tr / 'classification_results'
    ]
    filename_pairs = get_masked_filenames(f"{country}_{tr}", suffix)
    for folder in folders:
        for cls_name, conf_name in filename_pairs:
            cls_fp = folder / cls_name
            conf_fp = folder / conf_name
            if cls_fp.exists() and conf_fp.exists():
                return cls_fp, conf_fp
    return None, None


def discover_tracks(base_dir: Path, prefix: str, suffix: str = ""):
    """
    Discovers track folders and their classification/confidence files.
    Supports:
      1. Direct specific orbit path: e.g. NL/orbit_161 or NL_orbit_161.
      2. Country-based orbits: Prefix is a country code (e.g. PL, NL, FR, PT, AT).
         We scan base_dir/prefix/ for orbit_* folders.
      3. Legacy tracks: Prefix is a track prefix (e.g. P1, P2).
         We scan base_dir/ for folders starting with prefix.
     Returns list of (tr, country, cls_fp, conf_fp).
    """
    tracks = []
    prefix_upper = prefix.upper()

    # Case 0: Prefix points directly to a specific orbit directory (e.g. NL/orbit_161 or NL_orbit_161)
    normalized_prefix = prefix.replace('\\', '/')
    specific_path = base_dir / normalized_prefix
    if not specific_path.is_dir() and '_' in normalized_prefix:
        alt_prefix = normalized_prefix.replace('_', '/', 1)
        alt_path = base_dir / alt_prefix
        if alt_path.is_dir():
            specific_path = alt_path
            normalized_prefix = alt_prefix

    if specific_path.is_dir() and (specific_path.name.startswith("orbit_") or "/" in normalized_prefix):
        tr = normalized_prefix
        parts = tr.split('/')
        country = parts[0].upper()
        sanitized = tr.replace('/', '_')
        track_prefix = sanitized if sanitized.upper().startswith(country.upper() + "_") else f"{country}_{sanitized}"
        filename_pairs = get_masked_filenames(track_prefix, suffix)
        
        candidates = [
            specific_path / '2_classification' / '3_maps',
            specific_path / 'classification_results' / 'classification',
            specific_path / 'classification_results'
        ]
        found = False
        for folder in candidates:
            for cls_name, conf_name in filename_pairs:
                cls_fp  = folder / cls_name
                conf_fp = folder / conf_name
                if cls_fp.exists() and conf_fp.exists():
                    tracks.append((tr, country, cls_fp, conf_fp))
                    found = True
                    break
            if found:
                break

    # Case 1: Prefix is a country code (2 letters)
    if not tracks and len(prefix) == 2 and prefix.isalpha():
        country = prefix_upper
        country_dir = base_dir / country
        if country_dir.exists():
            for sub in country_dir.iterdir():
                if not sub.is_dir() or not sub.name.startswith("orbit_"):
                    continue
                tr = f"{country}/{sub.name}"
                sanitized = tr.replace('/', '_').replace('\\', '_')
                
                track_prefix = sanitized if sanitized.upper().startswith(country.upper() + "_") else f"{country}_{sanitized}"
                filename_pairs = get_masked_filenames(track_prefix, suffix)
                
                candidates = [
                    sub / '2_classification' / '3_maps',
                    sub / 'classification_results' / 'classification',
                    sub / 'classification_results'
                ]
                found = False
                for folder in candidates:
                    for cls_name, conf_name in filename_pairs:
                        cls_fp  = folder / cls_name
                        conf_fp = folder / conf_name
                        if cls_fp.exists() and conf_fp.exists():
                            tracks.append((tr, country, cls_fp, conf_fp))
                            found = True
                            break
                    if found:
                        break
    
    # Case 2: Prefix is a legacy track prefix
    if not tracks:
        for sub in base_dir.iterdir():
            if not sub.is_dir():
                continue
            tr = sub.name
            if not tr.startswith(prefix):
                continue
            if '/' in tr or '\\' in tr:
                country = tr.replace('\\', '/').split('/')[0].upper()
            elif len(tr) == 2:
                country = tr.upper()
            else:
                continue
            
            sanitized = tr.replace('/', '_').replace('\\', '_')
            track_prefix = sanitized if sanitized.upper().startswith(country.upper() + "_") else f"{country}_{sanitized}"
            filename_pairs = get_masked_filenames(track_prefix, suffix)
            
            candidates = [
                base_dir / tr / 'classification_results' / 'classification',
                base_dir / tr / 'classification_results'
            ]
            found = False
            for folder in candidates:
                for cls_name, conf_name in filename_pairs:
                    cls_fp  = folder / cls_name
                    conf_fp = folder / conf_name
                    if cls_fp.exists() and conf_fp.exists():
                        tracks.append((tr, country, cls_fp, conf_fp))
                        found = True
                        break
                if found:
                    break
                    
    return tracks

def run_merge_for_country(country: str, seg_mode: str = 'slic', suffix: str = '', method: str = 'confidence'):
    prefix = country.upper()
    base_dir = Path(os.environ.get("AIML_WORKING_DIR", r"D:\AIML_CropMapper_Cloud\workingDirs"))
    if not base_dir.exists() or not (base_dir / prefix).exists():
        fallback_dir = Path(r"D:\AIML_CropMapper_Cloud\workingDir")
        if (fallback_dir / prefix).exists():
            base_dir = fallback_dir

    if not suffix:
        candidate_suffixes = [
            f"_mlpxgb_presto_{seg_mode}",
            f"_presto_hybrid_{seg_mode}",
            f"_{seg_mode}",
            f"_{seg_mode}_masked",
            ""
        ]
        for cand in candidate_suffixes:
            t_cand = discover_tracks(base_dir, prefix, suffix=cand)
            if t_cand:
                suffix = cand
                break

    tracks = discover_tracks(base_dir, prefix, suffix=suffix)
    if not tracks:
        raise FileNotFoundError(
            f"No valid classification/confidence files for tracks starting with {prefix} (suffix: '{suffix}')"
        )

    print(f"Discovered tracks for {prefix}: {[t for t,_,_,_ in tracks]}")

    # --- compute union extent & grid ---------------------------------------
    ds0   = gdal.Open(str(tracks[0][2]))
    proj  = ds0.GetProjection()
    gt0   = ds0.GetGeoTransform()
    resX, resY = gt0[1], abs(gt0[5])

    extents = []
    for _, _, cls_fp, _ in tracks:
        ds = gdal.Open(str(cls_fp))
        gt = ds.GetGeoTransform()
        c, r = ds.RasterXSize, ds.RasterYSize
        minX = gt[0]
        maxY = gt[3]
        maxX = gt[0] + c * gt[1]
        minY = gt[3] + r * gt[5]
        extents.append((minX, maxY, maxX, minY))

    minX = min(e[0] for e in extents)
    maxY = max(e[1] for e in extents)
    maxX = max(e[2] for e in extents)
    minY = min(e[3] for e in extents)

    cols = int(np.ceil((maxX - minX) / resX))
    rows = int(np.ceil((maxY - minY) / resY))
    gt_global = (minX, resX, 0, maxY, 0, -resY)

    print(f"Global mosaic: {cols} cols × {rows} rows")

    # --- warp and stack using VRTs to avoid OOM ------------------------------
    ds_cls_list = []
    ds_conf_list = []
    nodata_vals = []
    
    for tr, country, cls_fp, conf_fp in tracks:
        # classification VRT
        vrt_cls = gdal.Warp(
            '', str(cls_fp), format='VRT',
            width=cols, height=rows,
            outputBounds=(minX, minY, maxX, maxY),
            dstSRS=proj,
            resampleAlg=gdal.GRA_NearestNeighbour
        )
        ds_cls_list.append(vrt_cls)
        nodata_vals.append(vrt_cls.GetRasterBand(1).GetNoDataValue())

        # confidence VRT
        vrt_conf = gdal.Warp(
            '', str(conf_fp), format='VRT',
            width=cols, height=rows,
            outputBounds=(minX, minY, maxX, maxY),
            dstSRS=proj,
            resampleAlg=gdal.GRA_NearestNeighbour
        )
        ds_conf_list.append(vrt_conf)

    # --- Prepare output file ------------------------------------------------
    base_tr, base_country, _, _ = tracks[0]
    out_dir = base_dir / base_country / 'national_products'
    out_dir.mkdir(parents=True, exist_ok=True)
    out_tif = out_dir / f"{base_country}_national_crop_map{suffix}.tif"
    
    drv = gdal.GetDriverByName('GTiff')
    ds_out = drv.Create(str(out_tif), cols, rows, 1, gdal.GDT_Int32, 
                        options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
    ds_out.SetGeoTransform(gt_global)
    ds_out.SetProjection(proj)
    band_out = ds_out.GetRasterBand(1)
    band_out.SetNoDataValue(0)
    
    # --- Block by block processing ------------------------------------------
    gdal.SetCacheMax(4 * 1024 * 1024 * 1024)
    tile_size = 4096
    print("Merging tracks block-by-block to conserve memory...")
    
    for y in range(0, rows, tile_size):
        for x in range(0, cols, tile_size):
            xsize = min(tile_size, cols - x)
            ysize = min(tile_size, rows - y)
            
            c_stack = []
            cf_stack = []
            
            for i in range(len(tracks)):
                c_arr = ds_cls_list[i].GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                cf_arr = ds_conf_list[i].GetRasterBand(1).ReadAsArray(x, y, xsize, ysize).astype(np.float32)
                nod = nodata_vals[i]
                
                if nod is not None:
                    cf_arr[c_arr == nod] = np.nan
                    c_arr = np.where(c_arr == nod, 0, c_arr).astype(np.int32)
                    
                c_stack.append(c_arr)
                cf_stack.append(cf_arr)
                
            c_stack = np.stack(c_stack, axis=0)
            cf_stack = np.stack(cf_stack, axis=0)
            
            cf_stack[np.isnan(cf_stack)] = -np.inf
            idx = np.argmax(cf_stack, axis=0)
            final_block = np.take_along_axis(c_stack, idx[None,:,:], axis=0)[0]
            final_block[np.all(np.isneginf(cf_stack), axis=0)] = 0
            
            band_out.WriteArray(final_block, x, y)

    ds_out.FlushCache()
    
    # --- apply morphological sieve filter to restore objects -------------
    print("Applying Sieve Filter to remove slivers and isolated pixels...")
    # Using threshold of 50 pixels (approx field size depending on resolution)
    # 8-connectedness is generally preferred for diagonals
    gdal.SieveFilter(band_out, None, band_out, 50, 8, callback=None)
    ds_out.FlushCache()
    print(f"Merged classification saved: {out_tif}")

    # --- compute metrics & areas --------------------------------------------
    print("Calculating metrics and areas...")
    import pandas as pd
    ctrl_list = []
    for tr, country, _, _ in tracks:
        cand_ctrls = [
            base_dir / tr / '2_classification' / '1_samples_and_features' / f"{tr.replace('/', '_')}_control{suffix}.shp",
            base_dir / tr / '2_classification' / '1_samples_and_features' / f"control{suffix}.shp",
            base_dir / tr / '2_classification' / '1_samples_and_features' / f"{tr.replace('/', '_')}_control_{seg_mode}.shp",
            base_dir / tr / '2_classification' / '1_samples_and_features' / f"control_{seg_mode}.shp",
            base_dir / tr / 'classification_results' / 'samples' / f"control{suffix}.shp",
            base_dir / tr / 'classification_results' / 'samples' / f"control_{seg_mode}.shp",
            base_dir / tr / 'classification_results' / 'samples' / 'control.shp',
        ]
        track_ctrl_shp = None
        for c in cand_ctrls:
            if c.exists():
                track_ctrl_shp = c
                break

        if track_ctrl_shp and track_ctrl_shp.exists():
            try:
                gdf = gpd.read_file(str(track_ctrl_shp))
                if not gdf.empty:
                    ctrl_list.append(gdf)
                    print(f"  Loaded {len(gdf)} control points from {tr} ({track_ctrl_shp.name})")
            except Exception as e:
                print(f"  [WARNING] Failed to load control points from {tr}: {e}")
                
    if ctrl_list:
        ctrl = pd.concat(ctrl_list, ignore_index=True)
        # Drop duplicates based on geometry to avoid double-counting in overlap zones
        before_len = len(ctrl)
        ctrl = ctrl.drop_duplicates(subset=['geometry'])
        print(f"Merged control points: {len(ctrl)} total (dropped {before_len - len(ctrl)} duplicates)")
    else:
        # Fallback to the country-wide main samples database (if exists) or error
        national_samples_shp = base_dir.parent / 'auxiliary_files' / 'shapefiles_samples' / base_country / 'samples.shp'
        if national_samples_shp.exists():
            print(f"No track-specific control shapefiles found. Falling back to national main samples: {national_samples_shp}")
            ctrl = gpd.read_file(str(national_samples_shp))
        else:
            raise FileNotFoundError(f"No control shapefiles found and national fallback not found at {national_samples_shp}")
    inv      = gdal.InvGeoTransform(gt_global)

    x_coords = ctrl.geometry.x.values
    y_coords = ctrl.geometry.y.values

    # Crop aggregation for NL to reduce semantic confusion and match model classes
    if base_country == 'NL':
        print("Applying crop aggregation for Netherlands validation labels...")
        ref_shp = None
        for tr, _, _, _ in tracks:
            track_ctrl_shp = base_dir / tr / 'classification_results' / 'samples' / f"control{suffix}.shp"
            if not track_ctrl_shp.exists():
                track_ctrl_shp = base_dir / tr / 'classification_results' / 'samples' / 'control.shp'
            if track_ctrl_shp.exists():
                ref_shp = track_ctrl_shp
                break
        if not ref_shp:
            ref_shp = base_dir.parent / 'auxiliary_files' / 'shapefiles_samples' / base_country / 'samples.shp'
            
        crop_aggregation = get_crop_aggregation(base_country, ref_shp)
        ctrl['crop_id'] = ctrl['crop_id'].apply(lambda val: crop_aggregation.get(val, val))

    crop_ids = ctrl['crop_id'].values

    px_vals = (inv[0] + inv[1]*x_coords + inv[2]*y_coords).astype(int)
    py_vals = (inv[3] + inv[4]*x_coords + inv[5]*y_coords).astype(int)

    valid_mask = (px_vals >= 0) & (px_vals < cols) & (py_vals >= 0) & (py_vals < rows)

    valid_px = px_vals[valid_mask]
    valid_py = py_vals[valid_mask]
    valid_crop_ids = crop_ids[valid_mask].astype(int)
    
    # Read predictions locally per point to avoid full raster memory
    valid_preds = np.zeros(len(valid_px), dtype=int)
    for i in range(len(valid_px)):
        val = band_out.ReadAsArray(int(valid_px[i]), int(valid_py[i]), 1, 1)
        if val is not None:
            valid_preds[i] = val[0, 0]
            
    final_mask = (valid_crop_ids > 0) & (valid_preds > 0)

    true_vals = valid_crop_ids[final_mask].tolist()
    pred_vals = valid_preds[final_mask].tolist()
    # Guarantee all classes from control set are represented in the matrix, even if completely missed
    all_control_classes = set(ctrl['crop_id'].unique().astype(int))
    all_control_classes.discard(0)
    labels = sorted(list(all_control_classes.union(set(pred_vals))))

    cm     = confusion_matrix(true_vals, pred_vals, labels=labels)
    prec, rec, f1, _ = precision_recall_fscore_support(
        true_vals, pred_vals,
        labels=labels,
        average=None,
        zero_division=0
    )
    total = cm.sum()
    exp   = (cm.sum(axis=0) * cm.sum(axis=1)).sum() / (total**2)
    oa    = np.trace(cm) / total
    kappa = (oa - exp) / (1 - exp) if (1 - exp) else np.nan

    # Calculate areas block-by-block
    areas_counts = {}
    for y in range(0, rows, tile_size):
        for x in range(0, cols, tile_size):
            xsize = min(tile_size, cols - x)
            ysize = min(tile_size, rows - y)
            block = band_out.ReadAsArray(x, y, xsize, ysize)
            unique, counts = np.unique(block, return_counts=True)
            for u, c in zip(unique, counts):
                if u != 0:
                    areas_counts[u] = areas_counts.get(u, 0) + c

    resx, resy = abs(gt_global[1]), abs(gt_global[5])
    area_ha    = resx * resy / 10000
    areas      = [{
        'Class':   c,
        'Area_ha': round(areas_counts.get(c, 0) * area_ha, 2)
    } for c in labels]

    # --- Resolve English crop names mapping ---------------------------------
    crop_name_map = {}
    id_cols = [c for c in ['crop_id', 'crop_ids', 'code', 'id', 'class_id'] if c in ctrl.columns]
    name_cols = [c for c in ['crop_name', 'crop_names', 'crop_type', 'label', 'name', 'class_name', 'crop', 'nom'] if c in ctrl.columns]
    if id_cols and name_cols:
        for _, row in ctrl[[id_cols[0], name_cols[0]]].drop_duplicates().iterrows():
            try:
                cid = int(row[id_cols[0]])
                cname = str(row[name_cols[0]]).strip()
                if cname and cname.lower() not in ['none', 'nan', '']:
                    crop_name_map[cid] = cname
            except Exception:
                pass

    # Fallback to auxiliary samples shapefile
    aux_samples = Path(r"D:/AIML_CropMapper_Cloud/auxiliary_files/shapefiles_samples") / base_country / "samples.shp"
    if aux_samples.exists():
        try:
            gdf_s = gpd.read_file(str(aux_samples), engine="pyogrio")
            id_cols_s = [c for c in ['crop_id', 'crop_ids', 'code', 'id', 'class_id'] if c in gdf_s.columns]
            name_cols_s = [c for c in ['crop_name', 'crop_names', 'crop_type', 'label', 'name', 'class_name', 'crop', 'nom'] if c in gdf_s.columns]
            if id_cols_s and name_cols_s:
                for _, row in gdf_s[[id_cols_s[0], name_cols_s[0]]].drop_duplicates().iterrows():
                    try:
                        cid = int(row[id_cols_s[0]])
                        cname = str(row[name_cols_s[0]]).strip()
                        if cid not in crop_name_map and cname and cname.lower() not in ['none', 'nan', '']:
                            crop_name_map[cid] = cname
                    except Exception:
                        pass
        except Exception:
            pass

    # Fallback to priors.json
    aux_priors = Path(r"D:/AIML_CropMapper_Cloud/auxiliary_files/shapefiles_samples") / base_country / "priors.json"
    if aux_priors.exists():
        try:
            with open(aux_priors, 'r', encoding='utf-8') as pf:
                priors_data = json.load(pf)
                for p_idx, p_name in enumerate(priors_data.keys(), start=1):
                    if p_idx not in crop_name_map:
                        crop_name_map[p_idx] = p_name.title()
        except Exception:
            pass

    # --- write Excel report -------------------------------------------------
    xlsx = out_dir / f"{base_country}_final_metrics{suffix}.xlsx"
    wb   = openpyxl.Workbook()
    sh   = wb.active
    sh.title = 'Validation Metrics'

    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    sub_font = Font(name="Calibri", size=11, bold=True, color="1F497D")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    align_left = Alignment(horizontal='left', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    sh.cell(row=1, column=1, value=f"National Crop Classification Accuracy Report: {base_country}").font = Font(name="Calibri", size=14, bold=True, color="1F497D")
    sh.cell(row=2, column=1, value=f"Country: {base_country} | Segmentation: {seg_mode.upper()} | Model: Unified PyTorch MLP + XGBoost Fusion Ensemble").font = bold_font
    sh.cell(row=3, column=1, value="Data: Multimodal Sentinel-1 SAR (Sigma0 VH/VV) + Sentinel-2 MSI (B02-B12) + NASA Harvest Presto Embeddings").font = Font(name="Calibri", size=10, italic=True, color="595959")

    # Summary table
    sh.cell(row=5, column=1, value="Metric").fill = header_fill
    sh.cell(row=5, column=1).font = header_font
    sh.cell(row=5, column=1).alignment = align_left
    sh.cell(row=5, column=2, value="Value").fill = header_fill
    sh.cell(row=5, column=2).font = header_font
    sh.cell(row=5, column=2).alignment = align_left

    sh.cell(row=6, column=1, value="Overall Accuracy (OA)").font = regular_font
    sh.cell(row=6, column=1).border = thin_border
    sh.cell(row=6, column=2, value=f"{oa * 100:.1f}%").font = bold_font
    sh.cell(row=6, column=2).alignment = align_left
    sh.cell(row=6, column=2).border = thin_border

    sh.cell(row=7, column=1, value="Cohen's Kappa").font = regular_font
    sh.cell(row=7, column=1).border = thin_border
    sh.cell(row=7, column=2, value=f"{kappa:.4f}").font = bold_font
    sh.cell(row=7, column=2).alignment = align_left
    sh.cell(row=7, column=2).border = thin_border

    sh.cell(row=8, column=1, value="Total Validation Samples").font = regular_font
    sh.cell(row=8, column=1).border = thin_border
    sh.cell(row=8, column=2, value=f"{int(total):,}".replace(',', ' ')).font = regular_font
    sh.cell(row=8, column=2).alignment = align_left
    sh.cell(row=8, column=2).border = thin_border

    # Per‐class recall/precision/F1
    r1 = 10
    sh.cell(row=r1, column=1, value="Per-Class Classification Accuracy").font = sub_font
    r1 += 1
    headers_pc = ["Class ID", "Crop Name", "Producer Acc (Recall)", "User Acc (Precision)", "F1-Score", "Validation Samples"]
    for c_idx, h_text in enumerate(headers_pc, start=1):
        cell = sh.cell(row=r1, column=c_idx, value=h_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_left

    for idx, c in enumerate(labels):
        r_curr = r1 + 1 + idx
        c_name = crop_name_map.get(int(c), f"Class {c}")
        c_samples = int(np.sum(np.array(true_vals) == c))
        
        c1 = sh.cell(row=r_curr, column=1, value=int(c))
        c1.font = regular_font
        c1.alignment = align_left
        c1.border = thin_border
        
        c2 = sh.cell(row=r_curr, column=2, value=c_name)
        c2.font = regular_font
        c2.alignment = align_left
        c2.border = thin_border
        
        c3 = sh.cell(row=r_curr, column=3, value=f"{rec[idx] * 100:.1f}%")
        c3.font = regular_font
        c3.alignment = align_left
        c3.border = thin_border
        
        c4 = sh.cell(row=r_curr, column=4, value=f"{prec[idx] * 100:.1f}%")
        c4.font = regular_font
        c4.alignment = align_left
        c4.border = thin_border
        
        c5 = sh.cell(row=r_curr, column=5, value=f"{f1[idx] * 100:.1f}%")
        c5.font = regular_font
        c5.alignment = align_left
        c5.border = thin_border
        
        c6 = sh.cell(row=r_curr, column=6, value=f"{c_samples:,}".replace(',', ' '))
        c6.font = regular_font
        c6.alignment = align_left
        c6.border = thin_border

    # Confusion matrix table
    rc = r1 + 1 + len(labels) + 2
    sh.cell(row=rc, column=1, value="Confusion Matrix (Rows: Ground Truth, Cols: Prediction)").font = sub_font
    rc += 1
    c_top = sh.cell(row=rc, column=1, value="True \\ Pred")
    c_top.fill = header_fill
    c_top.font = header_font
    c_top.alignment = align_left
    for j, lbl in enumerate(labels):
        lbl_name = crop_name_map.get(int(lbl), str(lbl))
        cell = sh.cell(row=rc, column=j + 2, value=f"{int(lbl)}: {lbl_name}")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    for i, true_lbl in enumerate(labels):
        rc += 1
        t_name = crop_name_map.get(int(true_lbl), str(true_lbl))
        c_row_hdr = sh.cell(row=rc, column=1, value=f"{int(true_lbl)}: {t_name}")
        c_row_hdr.font = bold_font
        c_row_hdr.alignment = align_left
        c_row_hdr.border = thin_border
        for j in range(len(labels)):
            val = int(cm[i, j])
            cell = sh.cell(row=rc, column=j + 2, value=val)
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = align_center

    # Area per class
    ra = rc + 2
    total_area_ha = sum(a['Area_ha'] for a in areas)
    sh.cell(row=ra, column=1, value="Classified Agricultural Area Statistics").font = sub_font
    ra += 1
    headers_area = ["Class ID", "Crop Name", "Area (ha)", "Area (%)"]
    for c_idx, h_text in enumerate(headers_area, start=1):
        cell = sh.cell(row=ra, column=c_idx, value=h_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_left

    for a in areas:
        ra += 1
        c_id = int(a['Class'])
        c_name = crop_name_map.get(c_id, f"Class {c_id}")
        c_ha = int(round(a['Area_ha']))
        pct = (a['Area_ha'] / total_area_ha * 100.0) if total_area_ha > 0 else 0.0
        
        c1 = sh.cell(row=ra, column=1, value=c_id)
        c1.font = regular_font
        c1.alignment = align_left
        c1.border = thin_border
        
        c2 = sh.cell(row=ra, column=2, value=c_name)
        c2.font = regular_font
        c2.alignment = align_left
        c2.border = thin_border
        
        c3 = sh.cell(row=ra, column=3, value=f"{c_ha:,}".replace(',', ' '))
        c3.font = regular_font
        c3.alignment = align_left
        c3.border = thin_border
        
        c4 = sh.cell(row=ra, column=4, value=f"{pct:.1f}%")
        c4.font = regular_font
        c4.alignment = align_left
        c4.border = thin_border

    # Total row
    ra += 1
    tot_ha_int = int(round(total_area_ha))
    c1 = sh.cell(row=ra, column=1, value="Total")
    c1.font = bold_font
    c1.alignment = align_left
    c1.border = thin_border
    
    c2 = sh.cell(row=ra, column=2, value="All Agricultural Crops")
    c2.font = bold_font
    c2.alignment = align_left
    c2.border = thin_border
    
    c3 = sh.cell(row=ra, column=3, value=f"{tot_ha_int:,}".replace(',', ' '))
    c3.font = bold_font
    c3.alignment = align_left
    c3.border = thin_border
    
    c4 = sh.cell(row=ra, column=4, value="100.0%")
    c4.font = bold_font
    c4.alignment = align_left
    c4.border = thin_border

    for col in sh.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        sh.column_dimensions[col_letter].width = max(max_len + 4, 16)

    wb.save(str(xlsx))
    print(f"Final metrics saved: {xlsx}")


def main():
    parser = argparse.ArgumentParser(description="Multi-orbit national classification merger.")
    parser.add_argument('--track', required=True, help='Country code or base track prefix (e.g. NL, PL, FR)')
    parser.add_argument('--suffix', default='', help='Optional suffix of classification files')
    parser.add_argument('--seg_mode', default='slic', help='Segmentation mode (slic, sam, lpis)')
    args = parser.parse_args()

    run_merge_for_country(
        country=args.track,
        seg_mode=args.seg_mode,
        suffix=args.suffix
    )


if __name__ == '__main__':
    main()