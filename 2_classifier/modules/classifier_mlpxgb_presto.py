#!/usr/bin/env python
"""
1_classify_MLPXGB_presto_hybrid_S1S2.py - Multimodal Sentinel-1 (Sigma0) + Sentinel-2 Crop Classification
using NASA Harvest Presto embeddings and a Unified PyTorch MLP + XGBoost Fusion Ensemble.

Pipeline Overview:
  Stage 0: Generate Multimodal Data Footprint (S1 + S2 valid data intersection)
  Stage 1: Multimodal Image Segmentation (SLIC / SAM / LPIS)
  Stage 2: Sample Point Split (70% learn / 30% control)
  Stage 3: Multimodal Feature Extraction (S1 Sigma0 + S2 Optical + S1 Presto 128d + S2 Presto 128d)
  Stage 4: Train Unified MLP + XGBoost Fusion Ensemble
  Stage 5: Object-Based Tile Inference with Bayesian Prior Calibration
  Stage 6: Apply Agricultural & Data Footprint Masks
  Stage 7: Calculate Out-of-Bag Validation Metrics & Generate Styled Excel Report (.xlsx)

Execution examples:
  # Mode 1: SLIC Superpixel Segmentation (Fast, no external vector required):
  python 1_classify_MLPXGB_presto_hybrid_S1S2.py --track NL/orbit_88 --seg_mode slic --stage A

  # Mode 2: Official Cadastral LPIS Parcel Segmentation:
  python 1_classify_MLPXGB_presto_hybrid_S1S2.py --track NL/orbit_88 --seg_mode lpis --lpis_vector path/to/brp.gpkg --stage A
  python 1_classify_MLPXGB_presto_hybrid_S1S2.py --track PL/orbit_12 --seg_mode lpis --lpis_vector path/to/arimr.shp --stage A

  # Mode 3: Segment Anything (SAM) Deep Learning Segmentation:
  python 1_classify_MLPXGB_presto_hybrid_S1S2.py --track PT/orbit_161 --seg_mode sam --stage A
"""

import os
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
os.environ.setdefault("OMP_NUM_THREADS", "4")

import argparse
from pathlib import Path
import subprocess
import sys
import shutil
import shlex
import json
import math
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor
from typing import Optional, List, Dict, Tuple

# Add classifier directory and project root to sys.path to allow importing single_file_presto
classifier_dir = str(Path(__file__).resolve().parent)
project_root = str(Path(__file__).resolve().parent.parent)
for p in [classifier_dir, project_root]:
    if p not in sys.path:
        sys.path.insert(0, p)

import numpy as np
import pandas as pd
import geopandas as gpd
from osgeo import gdal, ogr, osr, gdalconst
from sklearn.metrics import confusion_matrix, precision_recall_fscore_support
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.pipeline import make_pipeline
from sklearn.utils import resample
from sklearn.impute import SimpleImputer
import joblib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# PyTorch
try:
    import torch
    import torch.nn as nn
    import torch.optim as optim
    from torch.utils.data import DataLoader, TensorDataset
    HAS_TORCH = True
except ImportError:
    HAS_TORCH = False
    print("WARNING: PyTorch not found. Install torch for Presto embeddings and MLP.")

# XGBoost / GBDT
try:
    import xgboost as xgb
    HAS_XGBOOST = True
except ImportError:
    HAS_XGBOOST = False
    from sklearn.ensemble import HistGradientBoostingClassifier

# Scikit-image for SLIC
try:
    from skimage.segmentation import felzenszwalb, slic
    from skimage.util import img_as_float
    HAS_SKIMAGE = True
except ImportError:
    HAS_SKIMAGE = False

# Global project directories
base_dir = Path(os.environ.get("AIML_WORKING_DIR", r"D:/AIML_CropMapper_Cloud/workingDirs"))
aux_dir = Path(os.environ.get("AIML_AUX_DIR", r"D:/AIML_CropMapper_Cloud/auxiliary_files"))
presto_dir = aux_dir / "Presto_models"

TOTAL_STAGES = 8


# =====================================================================
# 1. HELPERS: DATE PARSING & PRIORS
# =====================================================================

def parse_month_from_description(desc: str) -> int:
    """Parses month index (0-11) from raster band descriptions."""
    months_map = {
        'jan': 0, 'feb': 1, 'mar': 2, 'apr': 3, 'may': 4, 'jun': 5,
        'jul': 6, 'aug': 7, 'sep': 8, 'oct': 9, 'nov': 10, 'dec': 11
    }
    match = re.search(r'_(?:\d+)?([a-zA-Z]{3})\d{4}_', str(desc))
    if match:
        mon_str = match.group(1).lower()
        return months_map.get(mon_str, 0)
    return 0


def get_crop_aggregation(country: str, learn_shp_path: Optional[Path]) -> dict:
    return {}


def _get_priors_for_country(country: str, learn_shp_path: Optional[Path], classes: np.ndarray, class_counts: dict, total_samples: int, priors_file_override: Optional[Path] = None) -> np.ndarray:
    """Calculates true prior probabilities using shapefile names, custom JSONs, or physical field area priors."""
    priors_json_path = None
    if priors_file_override and os.path.exists(priors_file_override):
        priors_json_path = Path(priors_file_override)
    else:
        if learn_shp_path:
            p = Path(learn_shp_path).resolve()
            for parent in p.parents:
                aux_dir_path = parent / 'auxiliary_files'
                if aux_dir_path.exists():
                    aux_priors = aux_dir_path / 'shapefiles_samples' / country / 'priors.json'
                    if aux_priors.exists():
                        priors_json_path = aux_priors
                        break
            if not priors_json_path:
                track_dir = Path(learn_shp_path).parent.parent
                track_priors = track_dir / f"priors_{country}.json"
                if not track_priors.exists():
                    track_priors = track_dir / "priors.json"
                if track_priors.exists():
                    priors_json_path = track_priors

    if priors_json_path and priors_json_path.exists():
        try:
            with open(priors_json_path, 'r') as f:
                custom_priors = json.load(f)
            
            id_to_name = {}
            if learn_shp_path and os.path.exists(learn_shp_path):
                try:
                    gdf = gpd.read_file(str(learn_shp_path), engine="pyogrio")
                    if 'crop_id' in gdf.columns and 'crop_name' in gdf.columns:
                        id_to_name = dict(zip(gdf['crop_id'].astype(int), gdf['crop_name'].astype(str).str.lower()))
                except Exception:
                    pass

            raw_priors = {}
            sorted_keys = sorted(custom_priors.keys(), key=len, reverse=True)
            for cid, name in id_to_name.items():
                matched_val = 1e-5
                for key in sorted_keys:
                    if key.lower() in name or name in key.lower():
                        matched_val = float(custom_priors[key])
                        break
                raw_priors[cid] = matched_val

            for key, val in custom_priors.items():
                try:
                    cid = int(key)
                    raw_priors[cid] = float(val)
                except ValueError:
                    pass

            p_true = np.array([raw_priors.get(c, 1e-5) for c in classes])
            p_true = p_true / np.sum(p_true)
            return p_true
        except Exception as e:
            print(f"    [WARNING] Failed to load name-based priors: {e}")

    # Fallback to field size heuristics
    area_multipliers_default = {
        'grassland': 70000, 'maize': 70000, 'wheat': 70000, 'barley': 40000,
        'rye': 40000, 'triticale': 50000, 'oats': 40000, 'rapeseed': 60000,
        'sugar beet': 40000, 'potato': 20000, 'orchard': 20000, 'pea': 30000,
        'bean': 10000, 'vegetables': 5000, 'other': 10000
    }
    counts_arr = np.array([class_counts.get(c, 1) for c in classes])
    p_true = counts_arr / np.sum(counts_arr)
    return p_true


def _calculate_class_weights(y_data: np.ndarray, all_classes: np.ndarray) -> np.ndarray:
    classes_in_data = np.unique(y_data)
    total_samples = len(y_data)
    n_classes = len(classes_in_data)
    weight_vector = np.ones(len(all_classes), dtype=np.float32)
    
    for c in classes_in_data:
        count = np.sum(y_data == c)
        if count > 0:
            weight = total_samples / (n_classes * count)
            idx = np.where(all_classes == c)[0][0]
            weight_vector[idx] = math.sqrt(weight)
            
    return weight_vector


# =====================================================================
# 2. UNIFIED MLP + XGBOOST FUSION ENSEMBLE CLASSIFIER
# =====================================================================

class TorchMLPClassifier:
    """PyTorch Deep Neural Network Classifier for Multimodal Crop Classification."""
    def __init__(self, hidden_layer_sizes=(512, 256, 128), max_iter=200, batch_size=256, lr=0.001, class_weights=None, all_classes=None):
        self.hidden_layer_sizes = hidden_layer_sizes
        self.max_iter = max_iter
        self.batch_size = batch_size
        self.lr = lr
        self.class_weights = class_weights
        self.all_classes = all_classes
        self.device = torch.device('cuda' if (HAS_TORCH and torch.cuda.is_available()) else 'cpu')
        self.model = None
        self.le = None
        self.classes_ = None

    def fit(self, X, y):
        if torch.cuda.is_available():
            torch.cuda.empty_cache()

        self.le = LabelEncoder()
        if self.all_classes is not None:
            self.le.fit(self.all_classes)
        else:
            self.le.fit(y)

        y_enc = self.le.transform(y)
        self.classes_ = self.le.classes_

        input_dim = X.shape[1]
        output_dim = len(self.classes_)

        layers = []
        in_dim = input_dim
        for h_dim in self.hidden_layer_sizes:
            layers.append(nn.Linear(in_dim, h_dim))
            layers.append(nn.BatchNorm1d(h_dim))
            layers.append(nn.ReLU())
            layers.append(nn.Dropout(0.3))
            in_dim = h_dim
        layers.append(nn.Linear(in_dim, output_dim))

        self.model = nn.Sequential(*layers).to(self.device)

        if self.class_weights is not None and len(self.class_weights) == output_dim:
            weights_tensor = torch.tensor(self.class_weights, dtype=torch.float32).to(self.device)
            criterion = nn.CrossEntropyLoss(weight=weights_tensor)
        else:
            criterion = nn.CrossEntropyLoss()

        optimizer = optim.AdamW(self.model.parameters(), lr=self.lr, weight_decay=1e-4)
        scheduler = optim.lr_scheduler.CosineAnnealingLR(optimizer, T_max=self.max_iter)

        X_tensor = torch.tensor(X, dtype=torch.float32)
        y_tensor = torch.tensor(y_enc, dtype=torch.long)
        dataset = TensorDataset(X_tensor, y_tensor)
        dataloader = DataLoader(dataset, batch_size=self.batch_size, shuffle=True)

        self.model.train()
        for epoch in range(self.max_iter):
            for batch_x, batch_y in dataloader:
                batch_x = batch_x.to(self.device)
                batch_y = batch_y.to(self.device)

                optimizer.zero_grad()
                outputs = self.model(batch_x)
                loss = criterion(outputs, batch_y)
                loss.backward()
                optimizer.step()
            scheduler.step()

        self.model.eval()
        return self

    def predict_proba(self, X):
        self.model.eval()
        X_tensor = torch.tensor(X, dtype=torch.float32)
        dataset = TensorDataset(X_tensor)
        dataloader = DataLoader(dataset, batch_size=min(len(X), 4096), shuffle=False)

        probs_list = []
        with torch.no_grad():
            for (batch_x,) in dataloader:
                batch_x = batch_x.to(self.device)
                logits = self.model(batch_x)
                probs = torch.softmax(logits, dim=1).cpu().numpy()
                probs_list.append(probs)

        return np.vstack(probs_list)

    def predict(self, X):
        probs = self.predict_proba(X)
        preds_enc = np.argmax(probs, axis=1)
        return self.le.inverse_transform(preds_enc)


class EnsembleClassifier:
    """
    Unified Fusion Ensemble combining Deep PyTorch MLP and XGBoost GBDT via Soft Voting.
    """
    def __init__(self, mlp_model, xgb_model, weight_mlp=0.65):
        self.mlp_model = mlp_model
        self.xgb_model = xgb_model
        self.weight_mlp = weight_mlp
        self.classes_ = None
        self.xgb_classes_ = None
        self.imputer = SimpleImputer(strategy='mean')

    def fit(self, X, y):
        print("  [Fusion 1/2] Training PyTorch Deep MLP Model...")
        self.mlp_model.fit(X, y)
        self.classes_ = self.mlp_model.classes_

        print("  [Fusion 2/2] Training XGBoost Gradient Boosted Trees Model...")
        X_imputed = self.imputer.fit_transform(X)

        le = getattr(self.mlp_model, 'le', None)
        if le is None:
            le = LabelEncoder()
            le.fit(y)
        y_enc = le.transform(y)

        self.xgb_classes_ = np.unique(y_enc)
        xgb_le = LabelEncoder()
        xgb_le.fit(self.xgb_classes_)
        y_xgb = xgb_le.transform(y_enc)

        self.xgb_model.fit(X_imputed, y_xgb)
        print("  [Fusion Complete] Unified MLP + XGBoost Ensemble successfully fitted.")
        return self

    def predict_proba(self, X):
        p_mlp = self.mlp_model.predict_proba(X)
        X_imputed = self.imputer.transform(X)
        p_xgb_raw = self.xgb_model.predict_proba(X_imputed)

        p_xgb = np.zeros((X.shape[0], len(self.classes_)), dtype=np.float32)
        p_xgb[:, self.xgb_classes_] = p_xgb_raw

        return self.weight_mlp * p_mlp + (1.0 - self.weight_mlp) * p_xgb

    def predict(self, X):
        p = self.predict_proba(X)
        preds_enc = np.argmax(p, axis=1)
        le = getattr(self.mlp_model, 'le', None)
        if le is not None:
            return le.inverse_transform(preds_enc)
        return self.classes_[preds_enc]

# =====================================================================
# Register model classes in __main__ and legacy namespaces for unpickling
# =====================================================================
import sys
current_mod = sys.modules.get(__name__)
if current_mod:
    sys.modules['1_classify_MLPXGB_presto_hybrid_S1S2'] = current_mod
    sys.modules['classifier_mlpxgb_presto'] = current_mod

main_mod = sys.modules.get('__main__')
if main_mod:
    setattr(main_mod, 'EnsembleClassifier', EnsembleClassifier)
    setattr(main_mod, 'TorchMLPClassifier', TorchMLPClassifier)


# =====================================================================
# 3. MULTIMODAL PRESTO EMBEDDINGS (S1 + S2)
# =====================================================================

class PrestoMultimodalExtractor:
    """Computes 128-dimensional Presto foundation embeddings for multi-temporal S1 and S2 series."""
    def __init__(self, device: str = "cpu"):
        self.device = device
        self.weights_path = presto_dir / "default_model.pt"
        if not self.weights_path.exists():
            raise FileNotFoundError(f"Presto model weights not found at {self.weights_path}")
        
        try:
            import presto_model as single_file_presto
        except ImportError:
            import single_file_presto
        self.model = single_file_presto.Presto.construct(max_sequence_length=36)
        state_dict = torch.load(self.weights_path, map_location=self.device)
        state_dict.pop('encoder.pos_embed', None)
        state_dict.pop('decoder.pos_embed', None)
        self.model.load_state_dict(state_dict, strict=False)
        self.model.to(self.device)
        self.model.eval()

    def get_s1_embeddings(self, batch_s1_vv_vh: torch.Tensor, batch_latlons: torch.Tensor, months_tensor: torch.Tensor) -> np.ndarray:
        B, T, _ = batch_s1_vv_vh.shape
        x = torch.zeros(B, T, 17, dtype=torch.float32, device=self.device)
        x[:, :, 0] = batch_s1_vv_vh[:, :, 0] # VV
        x[:, :, 1] = batch_s1_vv_vh[:, :, 1] # VH

        mask = torch.ones(B, T, 17, dtype=torch.float32, device=self.device)
        mask[:, :, 0:2] = 0.0

        dw = torch.ones(B, T, dtype=torch.long, device=self.device) * 9
        month = months_tensor.unsqueeze(0).expand(B, -1)

        with torch.no_grad():
            features = self.model.encoder(
                x=x,
                dynamic_world=dw,
                latlons=batch_latlons.to(self.device),
                mask=mask,
                month=month,
                eval_task=True
            )
        return features.cpu().numpy()

    def get_s2_embeddings(self, batch_s2_9bands: torch.Tensor, batch_latlons: torch.Tensor, months_tensor: torch.Tensor) -> np.ndarray:
        B, T, _ = batch_s2_9bands.shape
        x = torch.zeros(B, T, 17, dtype=torch.float32, device=self.device)

        b2 = batch_s2_9bands[:, :, 0]
        b3 = batch_s2_9bands[:, :, 1]
        b4 = batch_s2_9bands[:, :, 2]
        b5 = batch_s2_9bands[:, :, 3]
        b6 = batch_s2_9bands[:, :, 4]
        b7 = batch_s2_9bands[:, :, 5]
        b8a = batch_s2_9bands[:, :, 6]
        b11 = batch_s2_9bands[:, :, 7]
        b12 = batch_s2_9bands[:, :, 8]

        ndvi = (b8a - b4) / (b8a + b4 + 1e-6)

        x[:, :, 2] = b2
        x[:, :, 3] = b3
        x[:, :, 4] = b4
        x[:, :, 5] = b5
        x[:, :, 6] = b6
        x[:, :, 7] = b7
        x[:, :, 9] = b8a
        x[:, :, 10] = b11
        x[:, :, 11] = b12
        x[:, :, 16] = ndvi

        mask = torch.ones(B, T, 17, dtype=torch.float32, device=self.device)
        mask[:, :, [2, 3, 4, 5, 6, 7, 9, 10, 11, 16]] = 0.0

        dw = torch.ones(B, T, dtype=torch.long, device=self.device) * 9
        month = months_tensor.unsqueeze(0).expand(B, -1)

        with torch.no_grad():
            features = self.model.encoder(
                x=x,
                dynamic_world=dw,
                latlons=batch_latlons.to(self.device),
                mask=mask,
                month=month,
                eval_task=True
            )
        return features.cpu().numpy()


# =====================================================================
# SAM WORKER (MULTIPROCESSING)
# =====================================================================

def sam_worker(tile_info, ras_path, footprint_path, params):
    try:
        import os
        import sys
        import time
        from osgeo import gdal
        import numpy as np
        import torch
        torch.set_num_threads(1)  # Force single thread in worker
        import cv2
        cv2.setNumThreads(0)
        
        from samgeo import SamGeo
        from skimage.util import img_as_float
        from scipy.ndimage import distance_transform_edt
        import scipy.ndimage as ndimage
        
        x, y, xsize_valid, ysize_valid, x_start_buf, y_start_buf, xsize_buf, ysize_buf, buffer = tile_info
        print(f"    [Worker x={x}, y={y}] Started reading tile data...", flush=True)
        
        ds = gdal.Open(ras_path, gdal.GA_ReadOnly)
        nbands = ds.RasterCount
        
        img_list = []
        for b in range(1, nbands + 1):
            band = ds.GetRasterBand(b)
            arr = band.ReadAsArray(x_start_buf, y_start_buf, xsize_buf, ysize_buf)
            if arr is None:
                print(f"    [Worker x={x}, y={y}] Failed to read band {b}!", flush=True)
                return x, y, None, None
            arr = np.nan_to_num(arr)
            img_list.append(arr)
            
        img = np.dstack(img_list)
        
        ds_foot = None
        if footprint_path and os.path.exists(footprint_path):
            ds_foot = gdal.Open(footprint_path, gdal.GA_ReadOnly)
            valid_mask_buf = ds_foot.GetRasterBand(1).ReadAsArray(x_start_buf, y_start_buf, xsize_buf, ysize_buf) > 0
            valid_mask = valid_mask_buf
        else:
            valid_mask = np.sum(np.abs(img), axis=2) > 0
            
        if not np.any(valid_mask):
            print(f"    [Worker x={x}, y={y}] Tile contains no active pixels.", flush=True)
            return x, y, None, None
            
        print(f"    [Worker x={x}, y={y}] Loading SAM model...", flush=True)
        t_model_start = time.time()
        sam_geo = SamGeo(
            model_type=params.get('sam_model_type', 'vit_h'),
            checkpoint=params.get('sam_checkpoint', None),
            device=params.get('sam_device', 'cuda' if torch.cuda.is_available() else 'cpu'),
            sam_kwargs={
                "points_per_side": params.get('points_per_side', 16),
                "pred_iou_thresh": params.get('pred_iou_thresh', 0.45),
                "stability_score_thresh": params.get('stability_score_thresh', 0.50),
                "crop_n_layers": params.get('crop_n_layers', 0),
                "crop_n_points_downscale_factor": params.get('crop_n_points_downscale_factor', 1),
                "min_mask_region_area": params.get('min_mask_region_area', 20),
                "box_nms_thresh": params.get('box_nms_thresh', 0.6)
            }
        )
        print(f"    [Worker x={x}, y={y}] SAM model loaded in {time.time() - t_model_start:.2f}s.", flush=True)
        
        img_8bit = np.zeros(img.shape, dtype=np.uint8)
        valid_pixels = valid_mask[:, :, np.newaxis]
        
        if np.any(valid_pixels):
            p2, p98 = np.percentile(img[valid_pixels], (2, 98))
            img_clip = np.clip(img, p2, p98)
            if p98 > p2:
                img_8bit[valid_pixels] = ((img_clip[valid_pixels] - p2) / (p98 - p2) * 255).astype(np.uint8)
                
            img_chan = np.ascontiguousarray(img_8bit[:, :, 0])
            img_smoothed = cv2.bilateralFilter(img_chan, d=9, sigmaColor=12, sigmaSpace=30)
            img_8bit[:, :, 0] = img_smoothed
            
            clahe_limit = params.get('clahe_limit', 0.0)
            if clahe_limit > 0.0:
                clahe = cv2.createCLAHE(clipLimit=clahe_limit, tileGridSize=(8,8))
                img_clahe = clahe.apply(img_8bit[:, :, 0])
                img_8bit[:, :, 0] = img_clahe
                
        if img_8bit.shape[2] == 1:
            img_rgb = np.repeat(img_8bit, 3, axis=2)
        else:
            img_rgb = img_8bit[:, :, :3]
            if img_rgb.shape[2] < 3:
                img_rgb = np.pad(img_rgb, ((0,0),(0,0),(0, 3-img_rgb.shape[2])), mode='constant')
                
        print(f"    [Worker x={x}, y={y}] Running SAM generate (points_per_side={params.get('points_per_side', 16)})...", flush=True)
        t_gen_start = time.time()
        sam_geo.generate(
            source=img_rgb,
            output=None,
            foreground=False,
            unique=True,
            min_size=10,
            max_size=100000
        )
        print(f"    [Worker x={x}, y={y}] SAM generate finished in {time.time() - t_gen_start:.2f}s.", flush=True)
        segments_buf = sam_geo.objects.astype(np.int32)
        
        zero_mask_buf = (segments_buf == 0) & valid_mask
        if np.any(zero_mask_buf) and np.any(segments_buf > 0):
            _, indices = distance_transform_edt(segments_buf == 0, return_indices=True)
            segments_buf[zero_mask_buf] = segments_buf[tuple(indices)][zero_mask_buf]
            
        segments_buf[~valid_mask] = 0
        
        median_size = params.get('median_size', 3)
        if median_size > 0:
            segments_buf = ndimage.median_filter(segments_buf, size=median_size)
            segments_buf[~valid_mask] = 0
            
        y_offset = y - y_start_buf
        x_offset = x - x_start_buf
        segments_buf[~valid_mask] = 0
        
        valid_mask_crop = valid_mask[y_offset : y_offset + ysize_valid, x_offset : x_offset + xsize_valid]
        segments = segments_buf[y_offset : y_offset + ysize_valid, x_offset : x_offset + xsize_valid]
        segments[~valid_mask_crop] = 0
        
        print(f"    [Worker x={x}, y={y}] Tile completed successfully.", flush=True)
        return x, y, segments, valid_mask_crop
    except Exception as e:
        print(f"Error in worker process for tile (x={tile_info[0]}, y={tile_info[1]}): {e}", flush=True)
        return tile_info[0], tile_info[1], None, None


# =====================================================================
# 4. MULTIMODAL PROCESSING PIPELINE (STAGES 0 - 7)
# =====================================================================

class ProcessingPipelineS1S2:
    def __init__(self, track: str, seg_mode: str = 'slic', mlp_weight: float = 0.65, s1_override: Optional[str] = None, s2_override: Optional[str] = None, lpis_vector: Optional[str] = None):
        self.track = track
        self.seg_mode = seg_mode.lower()
        self.mlp_weight = mlp_weight
        self.lpis_vector_override = lpis_vector
        
        norm_track = track.replace('\\', '/')
        self.country = norm_track.split('/')[0].upper() if '/' in norm_track else track.upper()
        self.total_stages = TOTAL_STAGES

        self.sanitized_track = norm_track.replace('/', '_')
        if not self.sanitized_track.startswith(self.country + "_"):
            self.file_prefix = f"{self.country}_{self.sanitized_track}"
        else:
            self.file_prefix = self.sanitized_track

        # Define directories
        # Define directories (New sequential structure in workingDirs/)
        self.base_dir = base_dir
        self.aux_dir = aux_dir
        self.proc_dir = self.base_dir / self.track / '1_input_stacks'
        self.out_dir = self.base_dir / self.track / '2_classification'
        self.seg_dir = self.out_dir / '0_segmentation'
        self.samples_dir = self.out_dir / '1_samples_and_features'
        self.model_dir = self.out_dir / '2_models'
        self.class_dir = self.out_dir / '3_maps'
        self.reports_dir = self.out_dir / '4_reports'

        self._ensure_directories()

        # Resolve Sentinel-1 and Sentinel-2 rasters
        self.s1_ras = Path(s1_override) if s1_override else self._resolve_s1_raster()
        self.s2_ras = Path(s2_override) if s2_override else self._resolve_s2_raster()

        print(f"============================================================")
        print(f" Multimodal S1 (Sigma0) + S2 Crop Classifier (Unified MLP+XGB Fusion)")
        print(f" Track: {self.track} ({self.country})")
        print(f" S1 SAR Raster: {self.s1_ras.name if self.s1_ras else 'None'}")
        print(f" S2 Optical Raster: {self.s2_ras.name if self.s2_ras else 'None'}")
        print(f" Segmentation: {self.seg_mode.upper()} | Fusion weights: {self.mlp_weight:.2f} MLP + {1.0-self.mlp_weight:.2f} XGB")
        print(f"============================================================")

        # Resolve Samples & Output Paths
        self.sample_shp = self._resolve_samples_shp()
        
        self.suffix = f"_mlpxgb_presto_{self.seg_mode}"

        # Standard canonical targets in workingDirs/
        self.footprint_mask = self.seg_dir / f"{self.file_prefix}_data_footprint.tif"
        self.seg_tif = self.seg_dir / f"{self.file_prefix}_segmentation_{self.seg_mode}.tif"
        self.learn_shp = self.samples_dir / f"{self.file_prefix}_learn_{self.seg_mode}.shp"
        self.control_shp = self.samples_dir / f"{self.file_prefix}_control_{self.seg_mode}.shp"
        self.sel_csv = self.samples_dir / f"{self.file_prefix}_mlpxgb_presto_learn_features_{self.seg_mode}.csv"
        self.model_pkl = self.model_dir / f"{self.file_prefix}_mlpxgb_presto_model_{self.seg_mode}.pkl"
        self.class_tif = self.class_dir / f"{self.file_prefix}_classified{self.suffix}.tif"
        self.conf_tif = self.class_dir / f"{self.file_prefix}_confidence{self.suffix}.tif"
        self.masked_class = self.class_dir / f"{self.file_prefix}_classified_masked{self.suffix}.tif"
        self.masked_conf = self.class_dir / f"{self.file_prefix}_confidence_masked{self.suffix}.tif"
        self.metrics_fp = self.reports_dir / f"report_{self.file_prefix}{self.suffix}.xlsx"

        # Fallback to legacy workingDir/ only if input stacks exist in legacy location and not in workingDirs
        if not self.proc_dir.exists():
            legacy_samples = [
                self.base_dir / self.track / 'classification_results' / 'samples' / f"{self.file_prefix}_learn_{self.seg_mode}.shp",
                self.base_dir / self.track / 'classification_results' / 'samples' / f"learn_{self.seg_mode}.shp",
                Path(r"D:/AIML_CropMapper_Cloud/workingDir") / self.track / 'classification_results' / 'samples' / f"{self.file_prefix}_learn_{self.seg_mode}.shp",
                Path(r"D:/AIML_CropMapper_Cloud/workingDir") / self.track / 'classification_results' / 'samples' / f"learn_{self.seg_mode}.shp",
            ]
            for c in legacy_samples:
                if c.exists():
                    self.learn_shp = c
                    break

            legacy_controls = [
                self.base_dir / self.track / 'classification_results' / 'samples' / f"{self.file_prefix}_control_{self.seg_mode}.shp",
                self.base_dir / self.track / 'classification_results' / 'samples' / f"control_{self.seg_mode}.shp",
                Path(r"D:/AIML_CropMapper_Cloud/workingDir") / self.track / 'classification_results' / 'samples' / f"{self.file_prefix}_control_{self.seg_mode}.shp",
                Path(r"D:/AIML_CropMapper_Cloud/workingDir") / self.track / 'classification_results' / 'samples' / f"control_{self.seg_mode}.shp",
            ]
            for c in legacy_controls:
                if c.exists():
                    self.control_shp = c
                    break

            legacy_footprints = [
                self.base_dir / self.track / 'classification_results' / 'segmentation' / f"{self.file_prefix}_data_footprint.tif",
                Path(r"D:/AIML_CropMapper_Cloud/workingDir") / self.track / 'classification_results' / 'segmentation' / f"{self.file_prefix}_data_footprint.tif",
            ]
            for c in legacy_footprints:
                if c.exists():
                    self.footprint_mask = c
                    break

            legacy_segs = [
                self.base_dir / self.track / 'classification_results' / 'segmentation' / f"{self.file_prefix}_segmentation_{self.seg_mode}.tif",
                Path(r"D:/AIML_CropMapper_Cloud/workingDir") / self.track / 'classification_results' / 'segmentation' / f"{self.file_prefix}_segmentation_{self.seg_mode}.tif",
            ]
            for c in legacy_segs:
                if c.exists():
                    self.seg_tif = c
                    break

            legacy_classes = [
                self.base_dir / self.track / 'classification_results' / 'classification' / f"{self.file_prefix}_classified{self.suffix}.tif",
                Path(r"D:/AIML_CropMapper_Cloud/workingDir") / self.track / 'classification_results' / 'classification' / f"{self.file_prefix}_classified{self.suffix}.tif",
            ]
            for c in legacy_classes:
                if c.exists():
                    self.class_tif = c
                    self.masked_class = c.parent / f"{self.file_prefix}_classified_masked{self.suffix}.tif"
                    break

            legacy_confs = [
                self.base_dir / self.track / 'classification_results' / 'classification' / f"{self.file_prefix}_confidence{self.suffix}.tif",
                Path(r"D:/AIML_CropMapper_Cloud/workingDir") / self.track / 'classification_results' / 'classification' / f"{self.file_prefix}_confidence{self.suffix}.tif",
            ]
            for c in legacy_confs:
                if c.exists():
                    self.conf_tif = c
                    self.masked_conf = c.parent / f"{self.file_prefix}_confidence_masked{self.suffix}.tif"
                    break

            if self.class_tif.parent != self.class_dir:
                self.metrics_fp = self.class_tif.parent.parent / f"{self.file_prefix}_metrics{self.suffix}.xlsx"

        self.agri_mask = self._resolve_agri_mask()

        # Segmentation params
        self.stage1_params = {
            'sam_model_type': 'vit_h',
            'sam_checkpoint': self._resolve_sam_checkpoint(),
            'sam_device': 'cuda' if (HAS_TORCH and torch.cuda.is_available()) else 'cpu',
            'tile_size': 2048,
            'buffer': 128,
            'points_per_side': 16,
            'pred_iou_thresh': 0.45,
            'stability_score_thresh': 0.50,
            'min_mask_region_area': 20,
            'box_nms_thresh': 0.6,
            'clahe_limit': 0.0,
            'median_size': 3
        }

    def _ensure_directories(self):
        for d in [self.seg_dir, self.samples_dir, self.model_dir, self.class_dir, self.reports_dir]:
            d.mkdir(parents=True, exist_ok=True)

    def _resolve_sam_checkpoint(self) -> Optional[str]:
        sam_dir = self.aux_dir / "SAM_models"
        for name in ['sam_vit_h_4b8939.pth', 'sam_vit_l_0b3195.pth', 'sam_vit_b_01ec64.pth']:
            p = sam_dir / name
            if p.exists():
                return str(p)
        return None

    def _resolve_s1_raster(self) -> Optional[Path]:
        candidate_dirs = [
            self.base_dir / self.track / '1_input_stacks',
            self.base_dir / self.track / 'processed_raster',
            Path(r"D:/AIML_CropMapper_Cloud/workingDir") / self.track / '1_input_stacks',
            Path(r"D:/AIML_CropMapper_Cloud/workingDir") / self.track / 'processed_raster'
        ]
        patterns = [f"*{self.sanitized_track}*_VH_VV*.tif", f"*_VH_VV*.tif", f"*{self.country}*_VH_VV*.tif", f"*{self.sanitized_track}*Sigma0*.tif", f"*Sigma0*.tif"]
        for c_dir in candidate_dirs:
            if c_dir.exists():
                for pat in patterns:
                    matches = list(c_dir.glob(pat))
                    if matches:
                        return matches[0]
        return None

    def _resolve_s2_raster(self) -> Optional[Path]:
        candidate_dirs = [
            self.base_dir / self.track / '1_input_stacks',
            self.base_dir / self.country / 'S2',
            self.base_dir / self.country / '1_input_stacks',
            self.base_dir / self.track / 'processed_raster',
            Path(r"D:/AIML_CropMapper_Cloud/workingDir") / self.track / '1_input_stacks',
            Path(r"D:/AIML_CropMapper_Cloud/workingDir") / self.country / 'S2',
            Path(r"D:/AIML_CropMapper_Cloud/workingDir") / self.track / 'processed_raster'
        ]
        patterns = [
            f"*{self.sanitized_track}*S2*.tif",
            f"{self.country}_S2_timeseries*.tif",
            f"*{self.country}*S2_timeseries*.tif",
            f"*S2_timeseries*.tif",
            f"*{self.country}*S2*.tif",
            f"*S2*.tif"
        ]
        for c_dir in candidate_dirs:
            if c_dir.exists():
                for pat in patterns:
                    matches = [f for f in c_dir.glob(pat) if f.is_file() and not f.name.endswith(".tmp.tif") and not f.name.endswith(".ovr")]
                    if matches:
                        return matches[0]

        # Cross-orbit discovery fallback: search other orbit folders of the same country
        country_dirs = [self.base_dir / self.country, Path(r"D:/AIML_CropMapper_Cloud/workingDir") / self.country]
        for c_root in country_dirs:
            if not c_root.exists():
                continue
            cross_candidates = list(c_root.glob("orbit_*/1_input_stacks/*S2*.tif"))
            for cand in cross_candidates:
                if cand.exists() and cand.stat().st_size > 100 * 1024 * 1024 and not cand.name.endswith(".tmp.tif") and not cand.name.endswith(".ovr"):
                    try:
                        ds_c = gdal.Open(str(cand))
                        if ds_c and ds_c.RasterCount >= 126:
                            dest_dir = self.base_dir / self.track / '1_input_stacks'
                            dest_dir.mkdir(parents=True, exist_ok=True)
                            dest_s2 = dest_dir / f"{self.file_prefix}_S2_timeseries.tif"
                            if not dest_s2.exists():
                                try:
                                    os.link(str(cand), str(dest_s2))
                                    print(f"    [OPTIMIZATION] Reused country S2 stack from {cand.name} via instant hardlink to {dest_s2.name}!")
                                except Exception:
                                    import shutil
                                    shutil.copy2(str(cand), str(dest_s2))
                                    print(f"    [OPTIMIZATION] Reused country S2 stack from {cand.name} via copy to {dest_s2.name}!")
                            return dest_s2 if dest_s2.exists() else cand
                    except Exception:
                        continue

        return None

    def _resolve_samples_shp(self) -> Optional[Path]:
        samples_base = self.aux_dir / 'shapefiles_samples'
        candidates = [
            samples_base / self.file_prefix / "samples.shp",
            samples_base / self.country / "samples.shp",
            samples_base / f"{self.country}_{self.sanitized_track}" / "samples.shp"
        ]
        for c in candidates:
            if c.exists():
                return c
        shps = list(samples_base.glob(f"**/*{self.country}*/**/samples.shp"))
        return shps[0] if shps else None

    def _resolve_agri_mask(self) -> Optional[Path]:
        agrimasks_dir = self.aux_dir / "raster_files" / "AgriMasks" / self.country
        candidates = [
            agrimasks_dir / f"{self.country}_agri_mask_allcrops_epsg3857.tif",
            agrimasks_dir / f"{self.country}_agri_mask_3class_epsg3857.tif",
            self.aux_dir / "raster_files" / f"{self.country}_arable.tif",
            self.aux_dir / "raster_files" / f"{self.country}_agri_mask.tif",
            self.aux_dir / "raster_files" / "EU_arable_areas_mask_3857.tif",
            self.base_dir / self.track / "agri_mask.tif"
        ]
        for c in candidates:
            if c.exists():
                return c
        if agrimasks_dir.exists():
            tifs = list(agrimasks_dir.glob("*.tif"))
            if tifs: return tifs[0]
        return None

    def _resolve_lpis_vector(self) -> Optional[Path]:
        if hasattr(self, 'lpis_vector_override') and self.lpis_vector_override:
            p = Path(self.lpis_vector_override)
            if p.exists():
                return p

        agrimasks_dir = self.aux_dir / "raster_files" / "AgriMasks" / self.country
        samples_dir = self.aux_dir / "shapefiles_samples" / self.country
        candidates = [
            agrimasks_dir / "brpgewaspercelen_definitief_2025.gpkg",
            agrimasks_dir / "lpis.gpkg",
            agrimasks_dir / "lpis.shp",
            samples_dir / "lpis.gpkg",
            samples_dir / "lpis.shp",
            samples_dir / "parcels.gpkg",
            samples_dir / "parcels.shp",
            samples_dir / "samples_all.shp",
            samples_dir / "samples_all.gpkg",
            samples_dir / "samples.shp",
            self.aux_dir / "shapefiles_samples" / f"{self.country}_{self.sanitized_track}" / "lpis.shp",
            self.base_dir / self.track / "lpis.shp",
            self.base_dir / self.track / "parcels.gpkg"
        ]
        for c in candidates:
            if c.exists():
                return c
        if agrimasks_dir.exists():
            gpkgs = list(agrimasks_dir.glob("*.gpkg"))
            if gpkgs: return gpkgs[0]
            shps = list(agrimasks_dir.glob("*.shp"))
            if shps: return shps[0]
        return None

    def _create_summed_composite(self, ref_ras: Path) -> Path:
        print("    [INFO] Creating high-SNR summed composite for segmentation...")
        composite_tif = self.seg_dir / f"{self.file_prefix}_summed_composite.tif"
        if composite_tif.exists():
            return composite_tif

        ds = gdal.Open(str(ref_ras))
        cols = ds.RasterXSize
        rows = ds.RasterYSize
        nbands = ds.RasterCount
        gt = ds.GetGeoTransform()
        proj = ds.GetProjection()

        driver = gdal.GetDriverByName('GTiff')
        out_ds = driver.Create(str(composite_tif), cols, rows, 1, gdal.GDT_Float32,
                               options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        out_ds.SetGeoTransform(gt)
        out_ds.SetProjection(proj)
        out_band = out_ds.GetRasterBand(1)
        out_band.SetNoDataValue(0)

        tile_size = 4096
        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                sum_arr = np.zeros((ysize, xsize), dtype=np.float32)
                valid_mask = np.zeros((ysize, xsize), dtype=bool)

                for b in range(1, nbands + 1):
                    band = ds.GetRasterBand(b)
                    arr = band.ReadAsArray(x, y, xsize, ysize)
                    nodata = band.GetNoDataValue()
                    if nodata is not None:
                        mask = (arr != nodata) & (~np.isnan(arr)) & (arr != 0)
                    else:
                        mask = (~np.isnan(arr)) & (arr != 0)
                    sum_arr[mask] += arr[mask]
                    valid_mask |= mask

                sum_arr[~valid_mask] = 0
                out_band.WriteArray(sum_arr, x, y)

        out_ds.FlushCache()
        out_ds = None
        ds = None
        return composite_tif

    # --- Stage 0: Footprint ---
    def stage_0_generate_footprint(self, force_recompute=False):
        stage = 0
        if self.footprint_mask.exists() and self.footprint_mask.stat().st_size > 1024 and not force_recompute:
            print(f"[Stage {stage}] Footprint already exists ({self.footprint_mask.name}), skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Generating Multimodal Data Footprint (S1 SAR & S2 Optical Intersection)...")
        ref_ras = self.s1_ras if self.s1_ras else self.s2_ras
        if not ref_ras or not ref_ras.exists():
            raise FileNotFoundError("Neither S1 nor S2 raster found.")

        ds_s1 = gdal.Open(str(self.s1_ras)) if self.s1_ras else None
        ds_s2 = gdal.Open(str(self.s2_ras)) if self.s2_ras else None
        ref_ds = ds_s1 if ds_s1 else ds_s2

        cols, rows = ref_ds.RasterXSize, ref_ds.RasterYSize
        gt, proj = ref_ds.GetGeoTransform(), ref_ds.GetProjection()

        driver = gdal.GetDriverByName('GTiff')
        out_ds = driver.Create(
            str(self.footprint_mask), cols, rows, 1, gdal.GDT_Byte,
            options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES']
        )
        out_ds.SetGeoTransform(gt)
        out_ds.SetProjection(proj)

        tile_size = 4096
        total_blocks = math.ceil(cols / tile_size) * math.ceil(rows / tile_size)
        done_blocks = 0

        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                mask = np.ones((ysize, xsize), dtype=bool)
                if ds_s1:
                    b1_s1 = ds_s1.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                    mask = mask & (b1_s1 != 0) & (b1_s1 != -9999) & (~np.isnan(b1_s1))
                if ds_s2:
                    b1_s2 = ds_s2.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                    mask = mask & (b1_s2 > 0) & (~np.isnan(b1_s2))

                out_ds.GetRasterBand(1).WriteArray(mask.astype(np.uint8), x, y)
                done_blocks += 1
                if done_blocks % 10 == 0 or done_blocks == total_blocks:
                    pct = (done_blocks / total_blocks) * 100.0
                    print(f"    [FOOTPRINT PROGRESS] {done_blocks}/{total_blocks} blocks completed ({pct:.1f}%)", flush=True)

        out_ds.FlushCache()
        out_ds = None
        ds_s1 = None
        ds_s2 = None
        print(f"    Multimodal intersection footprint saved to {self.footprint_mask}")

    # --- Stage 1: Segmentation (LPIS / SAM / SLIC) ---
    def stage_1_segmentation(self, force_recompute=False):
        stage = 1
        if self.seg_tif.exists() and not force_recompute:
            print(f"[Stage {stage}] Segmentation raster exists ({self.seg_tif.name}), skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Running Multimodal Image Segmentation ({self.seg_mode.upper()})...")
        ref_ras = self.s1_ras if self.s1_ras else self.s2_ras
        if not ref_ras or not ref_ras.exists():
            raise FileNotFoundError("Reference raster not found for segmentation.")

        ds = gdal.Open(str(ref_ras))
        cols, rows = ds.RasterXSize, ds.RasterYSize
        gt, proj = ds.GetGeoTransform(), ds.GetProjection()

        if self.seg_mode == 'lpis':
            lpis_file = self._resolve_lpis_vector()
            if lpis_file and lpis_file.exists():
                print(f"    Loading official LPIS parcel vector from: {lpis_file}...")
                minx = gt[0]
                maxy = gt[3]
                maxx = minx + cols * gt[1]
                miny = maxy + rows * gt[5]

                try:
                    import pyogrio
                    info = pyogrio.read_info(str(lpis_file))
                    lpis_crs = info.get('crs')
                except Exception:
                    gdf_temp = gpd.read_file(str(lpis_file), rows=1)
                    lpis_crs = gdf_temp.crs.to_string() if gdf_temp.crs else None
                    info = {'fid_column': None}

                srs_target = osr.SpatialReference()
                srs_target.ImportFromWkt(proj)
                target_epsg = srs_target.GetAttrValue("AUTHORITY", 1) or "3857"

                from pyproj import Transformer
                transformer = Transformer.from_crs(f"EPSG:{target_epsg}", lpis_crs, always_xy=True)
                p1 = transformer.transform(minx, miny)
                p2 = transformer.transform(maxx, maxy)
                lpis_bbox = (min(p1[0], p2[0]), min(p1[1], p2[1]), max(p1[0], p2[0]), max(p1[1], p2[1]))

                print(f"    Querying LPIS with spatial filter bbox: {lpis_bbox}")
                try:
                    import pyogrio
                    gdf = pyogrio.read_dataframe(str(lpis_file), bbox=lpis_bbox)
                except Exception:
                    gdf = gpd.read_file(str(lpis_file), bbox=lpis_bbox)

                print(f"    Loaded {len(gdf)} intersecting parcels. Reprojecting to EPSG:{target_epsg}...")
                gdf_target = gdf.to_crs(f"EPSG:{target_epsg}")

                fid_col = info.get('fid_column') if isinstance(info, dict) else None
                if fid_col and fid_col in gdf_target.columns:
                    id_col = fid_col
                elif 'id' in gdf_target.columns:
                    id_col = 'id'
                elif 'id_0' in gdf_target.columns:
                    id_col = 'id_0'
                else:
                    id_col = None

                if id_col is None:
                    gdf_target['lpis_id'] = np.arange(1, len(gdf_target) + 1)
                    id_col = 'lpis_id'
                else:
                    gdf_target[id_col] = pd.to_numeric(gdf_target[id_col], errors='coerce').fillna(0).astype(np.int32)
                    if gdf_target[id_col].sum() == 0 or gdf_target[id_col].nunique() < len(gdf_target):
                        gdf_target['lpis_id'] = np.arange(1, len(gdf_target) + 1)
                        id_col = 'lpis_id'

                temp_gpkg = self.seg_dir / f"temp_lpis_{self.sanitized_track}.gpkg"
                gdf_target.geometry = gdf_target.geometry.force_2d()
                gdf_target.to_file(str(temp_gpkg), driver="GPKG", engine="pyogrio")

                driver = gdal.GetDriverByName("GTiff")
                ds_out = driver.Create(str(self.seg_tif), cols, rows, 1, gdal.GDT_Int32,
                                       options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
                ds_out.SetGeoTransform(gt)
                ds_out.SetProjection(proj)

                band = ds_out.GetRasterBand(1)
                band.SetNoDataValue(0)
                band.Fill(0)

                print(f"    Rasterizing parcels to {self.seg_tif.name} (burning column '{id_col}')...")
                gdal.Rasterize(ds_out, str(temp_gpkg), attribute=id_col, callback=gdal.TermProgress_nocb)

                # Strictly mask with multimodal footprint (S1 + S2 intersection)
                if self.footprint_mask.exists():
                    print(f"    Masking LPIS segmentation with multimodal footprint ({self.footprint_mask.name})...")
                    ds_foot = gdal.Open(str(self.footprint_mask))
                    tile_size = 4096
                    for y in range(0, rows, tile_size):
                        for x in range(0, cols, tile_size):
                            xsize = min(tile_size, cols - x)
                            ysize = min(tile_size, rows - y)
                            seg_block = band.ReadAsArray(x, y, xsize, ysize)
                            foot_block = ds_foot.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                            if np.any(foot_block == 0):
                                seg_block[foot_block == 0] = 0
                                band.WriteArray(seg_block, x, y)
                    ds_foot = None

                ds_out.FlushCache()
                ds_out = None
                if os.path.exists(temp_gpkg):
                    try: os.remove(temp_gpkg)
                    except: pass
                print(f"    [OK] LPIS segmentation raster created and masked with footprint: {self.seg_tif}")
                return
            else:
                print(f"    [WARNING] LPIS vector dataset not found. Falling back to SLIC.")

        # Create composite if needed for SLIC/SAM
        try:
            comp_ras = self._create_summed_composite(ref_ras)
        except Exception:
            comp_ras = ref_ras

        if self.seg_mode == 'sam':
            self._run_python_segmentation_tiled(comp_ras, self.stage1_params, 'python_sam')
        else:
            slic_params = {
                'tile_size': 2048,
                'buffer': 64,
                'n_segments': 32000,
                'compactness': 0.05,
                'slic_sigma': 1.5
            }
            self._run_python_segmentation_tiled(comp_ras, slic_params, 'python_slic')

    def _run_python_segmentation_tiled(self, ras_path: Path, params: dict, method: str):
        print(f"    Running Tiled Python Segmentation ({method})...")
        ds = gdal.Open(str(ras_path))
        ds_foot = gdal.Open(str(self.footprint_mask)) if self.footprint_mask.exists() else None

        cols = ds.RasterXSize
        rows = ds.RasterYSize
        nbands = ds.RasterCount
        gt = ds.GetGeoTransform()
        proj = ds.GetProjection()

        driver = gdal.GetDriverByName('GTiff')
        out_ds = driver.Create(str(self.seg_tif), cols, rows, 1, gdal.GDT_Int32,
                               options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        out_ds.SetGeoTransform(gt)
        out_ds.SetProjection(proj)
        out_band = out_ds.GetRasterBand(1)
        out_band.SetNoDataValue(0)

        tile_size = params.get('tile_size', 2048)
        buffer = params.get('buffer', 128)
        global_seg_id = 1

        if method == 'python_sam':
            from concurrent.futures import ProcessPoolExecutor, as_completed
            tile_tasks = []
            for y in range(0, rows, tile_size):
                for x in range(0, cols, tile_size):
                    xsize_valid = min(tile_size, cols - x)
                    ysize_valid = min(tile_size, rows - y)
                    x_start_buf = max(0, x - buffer)
                    y_start_buf = max(0, y - buffer)
                    x_end_buf = min(cols, x + xsize_valid + buffer)
                    y_end_buf = min(rows, y + ysize_valid + buffer)
                    xsize_buf = x_end_buf - x_start_buf
                    ysize_buf = y_end_buf - y_start_buf

                    if ds_foot:
                        band_foot = ds_foot.GetRasterBand(1)
                        arr_foot = band_foot.ReadAsArray(x_start_buf, y_start_buf, xsize_buf, ysize_buf)
                        valid_mask = arr_foot > 0 if arr_foot is not None else None
                    else:
                        band = ds.GetRasterBand(1)
                        arr = band.ReadAsArray(x_start_buf, y_start_buf, xsize_buf, ysize_buf)
                        valid_mask = np.nan_to_num(arr) > 0 if arr is not None else None

                    if valid_mask is not None and np.any(valid_mask):
                        tile_tasks.append((x, y, xsize_valid, ysize_valid, x_start_buf, y_start_buf, xsize_buf, ysize_buf, buffer))

            total_tasks = len(tile_tasks)
            print(f"    Total active tiles to process with SAM: {total_tasks}")
            max_workers = min(8, os.cpu_count() or 4)

            with ProcessPoolExecutor(max_workers=max_workers) as executor:
                futures = {
                    executor.submit(sam_worker, task, str(ras_path), str(self.footprint_mask) if self.footprint_mask.exists() else None, params): task
                    for task in tile_tasks
                }
                completed_count = 0
                for future in as_completed(futures):
                    x, y, segments, valid_mask_crop = future.result()
                    completed_count += 1
                    pct = (completed_count / total_tasks) * 100.0
                    print(f"    [SAM PROGRESS] Tile {completed_count}/{total_tasks} ({pct:.1f}%) finished (x={x}, y={y}) | Total segments: {global_seg_id-1:,}", flush=True)

                    if segments is not None:
                        seg_valid_mask = segments > 0
                        unique_segs = np.unique(segments[seg_valid_mask])
                        if len(unique_segs) > 0:
                            max_seg = segments.max()
                            mapping = np.zeros(max_seg + 1, dtype=np.int32)
                            mapping[unique_segs] = np.arange(global_seg_id, global_seg_id + len(unique_segs))
                            segments = mapping[segments]
                            segments[~valid_mask_crop] = 0
                            global_seg_id += len(unique_segs)
                        else:
                            segments[~valid_mask_crop] = 0
                        out_band.WriteArray(segments.astype(np.int32), x, y)
                    else:
                        for task in tile_tasks:
                            if task[0] == x and task[1] == y:
                                xsize_valid, ysize_valid = task[2], task[3]
                                out_band.WriteArray(np.zeros((ysize_valid, xsize_valid), dtype=np.int32), x, y)
                                break
        else:
            # Tiled SLIC
            total_tiles = math.ceil(rows / tile_size) * math.ceil(cols / tile_size)
            tile_idx = 0
            for y in range(0, rows, tile_size):
                for x in range(0, cols, tile_size):
                    tile_idx += 1
                    xsize_valid = min(tile_size, cols - x)
                    ysize_valid = min(tile_size, rows - y)
                    x_start_buf = max(0, x - buffer)
                    y_start_buf = max(0, y - buffer)
                    x_end_buf = min(cols, x + xsize_valid + buffer)
                    y_end_buf = min(rows, y + ysize_valid + buffer)
                    xsize_buf = x_end_buf - x_start_buf
                    ysize_buf = y_end_buf - y_start_buf

                    img_list = []
                    for b in range(1, nbands + 1):
                        band = ds.GetRasterBand(b)
                        arr = band.ReadAsArray(x_start_buf, y_start_buf, xsize_buf, ysize_buf)
                        if arr is None:
                            img_list = None
                            break
                        arr = np.nan_to_num(arr)
                        img_list.append(arr)

                    if img_list is None:
                        continue

                    img = np.dstack(img_list)
                    if ds_foot:
                        valid_mask_buf = ds_foot.GetRasterBand(1).ReadAsArray(x_start_buf, y_start_buf, xsize_buf, ysize_buf)
                        if valid_mask_buf is not None:
                            valid_mask = valid_mask_buf > 0
                        else:
                            valid_mask = np.sum(np.abs(img), axis=2) > 0
                    else:
                        valid_mask = np.sum(np.abs(img), axis=2) > 0

                    if not np.any(valid_mask):
                        pct = (tile_idx / total_tiles) * 100.0
                        if tile_idx % 10 == 0 or tile_idx == total_tiles:
                            print(f"    [SLIC PROGRESS] Tile {tile_idx}/{total_tiles} ({pct:.1f}%) [NoData tile skipped] | Total segments: {global_seg_id-1:,}", flush=True)
                        continue

                    img_norm = img_as_float(img)
                    from skimage.segmentation import slic
                    max_tile_pixels = (tile_size + 2 * buffer) ** 2
                    pixels_per_segment = max_tile_pixels / params.get('n_segments', 32000)
                    active_pixels = np.sum(valid_mask)
                    n_segments_dynamic = max(1, int(active_pixels / pixels_per_segment))
                    segments_buf = slic(img_norm, n_segments=n_segments_dynamic, compactness=params.get('compactness', 0.05),
                                        sigma=params.get('slic_sigma', 1.5), start_label=1, mask=valid_mask)

                    y_offset = y - y_start_buf
                    x_offset = x - x_start_buf
                    segments_buf[~valid_mask] = 0

                    valid_mask_crop = valid_mask[y_offset : y_offset + ysize_valid, x_offset : x_offset + xsize_valid]
                    segments = segments_buf[y_offset : y_offset + ysize_valid, x_offset : x_offset + xsize_valid]

                    seg_valid_mask = segments > 0
                    unique_segs = np.unique(segments[seg_valid_mask])
                    if len(unique_segs) > 0:
                        max_seg = segments.max()
                        mapping = np.zeros(max_seg + 1, dtype=np.int32)
                        mapping[unique_segs] = np.arange(global_seg_id, global_seg_id + len(unique_segs))
                        segments = mapping[segments]
                        segments[~valid_mask_crop] = 0
                        global_seg_id += len(unique_segs)
                    else:
                        segments[~valid_mask_crop] = 0

                    out_band.WriteArray(segments.astype(np.int32), x, y)
                    pct = (tile_idx / total_tiles) * 100.0
                    print(f"    [SLIC PROGRESS] Tile {tile_idx}/{total_tiles} ({pct:.1f}%) completed | Total segments: {global_seg_id-1:,}", flush=True)

        out_band.FlushCache()
        out_ds = None
        ds = None
        print(f"    Segmentation completed: {self.seg_tif}")

    # --- Stage 2: Sample Point Split (70/30) ---
    def stage_2_split_samples(self, force_recompute=False, learn_frac=0.7, random_state=42):
        stage = 2
        if self.learn_shp.exists() and self.control_shp.exists() and not force_recompute:
            print(f"[Stage {stage}] Train/Validation sample split exists, skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Performing Stratified Train/Control Split...")
        if not self.sample_shp or not self.sample_shp.exists():
            raise FileNotFoundError(f"Sample shapefile not found at {self.sample_shp}")

        gdf = gpd.read_file(str(self.sample_shp), engine="pyogrio")
        crop_col = 'crop_id' if 'crop_id' in gdf.columns else 'code'
        gdf['crop_id'] = gdf[crop_col].astype(int)

        train_dfs, val_dfs = [], []
        for cid, group in gdf.groupby('crop_id'):
            if len(group) == 1:
                train_dfs.append(group)
            else:
                train_g = group.sample(frac=learn_frac, random_state=random_state)
                val_g = group.drop(train_g.index)
                train_dfs.append(train_g)
                val_dfs.append(val_g)

        gdf_train = pd.concat(train_dfs)
        gdf_val = pd.concat(val_dfs)

        gdf_train.to_file(str(self.learn_shp), engine="pyogrio")
        gdf_val.to_file(str(self.control_shp), engine="pyogrio")
        print(f"    Split {len(gdf)} points -> {len(gdf_train)} train, {len(gdf_val)} validation.")

    # --- Stage 3: Multimodal Feature Extraction ---
    def stage_3_selection(self, force_recompute=False):
        stage = 3
        if self.sel_csv.exists() and not force_recompute:
            try:
                df_test = pd.read_csv(self.sel_csv, nrows=5)
                n_feats = df_test.shape[1] - 2
                if self.s1_ras and self.s2_ras and n_feats < 400:
                    print(f"[Stage {stage}] Existing CSV has only {n_feats} features (expected ~426 multimodal features). Re-extracting...")
                else:
                    print(f"[Stage {stage}] Multimodal features already extracted ({n_feats} features in {self.sel_csv.name}), skipping.")
                    return
            except Exception:
                pass

        print(f"[Stage {stage}/{self.total_stages}] Extracting Multimodal Features (S1 Sigma0 + S2 Optical + Presto S1/S2)...")
        device = "cuda" if (HAS_TORCH and torch.cuda.is_available()) else "cpu"
        extractor = PrestoMultimodalExtractor(device=device)

        gdf = gpd.read_file(str(self.learn_shp), engine="pyogrio")
        seg_ds = gdal.Open(str(self.seg_tif))
        seg_arr = seg_ds.GetRasterBand(1).ReadAsArray()
        gt = seg_ds.GetGeoTransform()
        inv_gt = gdal.InvGeoTransform(gt)
        cols, rows = seg_ds.RasterXSize, seg_ds.RasterYSize

        gdf_wgs84 = gdf.to_crs("EPSG:4326")
        seg_proj = seg_ds.GetProjection()
        if seg_proj and gdf.crs:
            from pyproj import CRS
            target_crs = CRS.from_wkt(seg_proj)
            if gdf.crs != target_crs:
                gdf_proj = gdf.to_crs(target_crs)
            else:
                gdf_proj = gdf
        else:
            gdf_proj = gdf

        if hasattr(gdf_proj.geometry, 'x'):
            xs = gdf_proj.geometry.x.values
            ys = gdf_proj.geometry.y.values
        else:
            xs = gdf_proj.geometry.centroid.x.values
            ys = gdf_proj.geometry.centroid.y.values

        pxs = (inv_gt[0] + inv_gt[1] * xs + inv_gt[2] * ys).astype(int)
        pys = (inv_gt[3] + inv_gt[4] * xs + inv_gt[5] * ys).astype(int)
        crop_ids = gdf['crop_id'].values

        target_segments = {}
        segment_coords = {}
        for idx, (px, py, cid) in enumerate(zip(pxs, pys, crop_ids)):
            if 0 <= px < cols and 0 <= py < rows:
                sid = seg_arr[py, px]
                if sid > 0:
                    target_segments[sid] = cid
                    geom_wgs = gdf_wgs84.iloc[idx].geometry
                    centroid_wgs = geom_wgs.centroid if hasattr(geom_wgs, 'centroid') else geom_wgs
                    segment_coords[sid] = (centroid_wgs.y, centroid_wgs.x)

        print(f"    Found {len(target_segments)} unique training segments.")
        from scipy.ndimage import find_objects
        slices = find_objects(seg_arr)

        ds_s1 = gdal.Open(str(self.s1_ras)) if self.s1_ras else None
        ds_s2 = gdal.Open(str(self.s2_ras)) if self.s2_ras else None

        nbands_s1 = ds_s1.RasterCount if ds_s1 else 0
        nbands_s2 = ds_s2.RasterCount if ds_s2 else 0
        num_dates_s1 = nbands_s1 // 2
        num_dates_s2 = nbands_s2 // 9

        months_s1 = [parse_month_from_description(ds_s1.GetRasterBand(b).GetDescription()) for b in range(1, num_dates_s1 + 1)] if ds_s1 else [0]
        months_s2 = [parse_month_from_description(ds_s2.GetRasterBand(b).GetDescription()) for b in range(1, num_dates_s2 + 1)] if ds_s2 else [0]

        month_tensor_s1 = torch.tensor(months_s1, dtype=torch.long, device=device)
        month_tensor_s2 = torch.tensor(months_s2, dtype=torch.long, device=device)

        feature_records = []
        batch_records = []
        batch_s1_profiles = []
        batch_s2_profiles = []
        batch_latlons = []
        count = 0
        total = len(target_segments)

        for sid, cid in target_segments.items():
            if sid - 1 >= len(slices) or slices[sid - 1] is None:
                continue
            sl = slices[sid - 1]
            ymin, ymax = sl[0].start, sl[0].stop
            xmin, xmax = sl[1].start, sl[1].stop
            w, h = xmax - xmin, ymax - ymin

            mask = (seg_arr[ymin:ymax, xmin:xmax] == sid)
            if not np.any(mask):
                continue

            record = {'crop_id': cid, 'seg_id': sid}
            lat, lon = segment_coords.get(sid, (52.0, 5.0))

            # S1 features
            s1_prof = None
            if ds_s1:
                s1_data = [np.nan_to_num(ds_s1.GetRasterBand(b).ReadAsArray(xmin, ymin, w, h)) for b in range(1, nbands_s1 + 1)]
                s1_arr = np.stack(s1_data, axis=0)
                s1_means = [float(np.mean(s1_arr[b][mask])) for b in range(nbands_s1)]
                for b_i, val in enumerate(s1_means):
                    record[f's1_b{b_i}'] = val

                s1_prof = np.zeros((num_dates_s1, 2), dtype=np.float32)
                for d in range(num_dates_s1):
                    s1_prof[d, 0] = (s1_means[num_dates_s1 + d] + 25.0) / 25.0 # VV
                    s1_prof[d, 1] = (s1_means[d] + 25.0) / 25.0 # VH

            # S2 features
            s2_prof = None
            if ds_s2:
                s2_data = [np.nan_to_num(ds_s2.GetRasterBand(b).ReadAsArray(xmin, ymin, w, h)) for b in range(1, nbands_s2 + 1)]
                s2_arr = np.stack(s2_data, axis=0)
                s2_means = [float(np.mean(s2_arr[b][mask])) for b in range(nbands_s2)]
                for b_i, val in enumerate(s2_means):
                    record[f's2_b{b_i}'] = val

                s2_prof = np.zeros((num_dates_s2, 9), dtype=np.float32)
                for d in range(num_dates_s2):
                    for band_idx in range(9):
                        s2_prof[d, band_idx] = s2_means[d * 9 + band_idx] / 10000.0

            # Validate valid multimodal data
            is_valid = True
            if ds_s1 and np.all([record.get(f's1_b{b_i}', 0) == 0 for b_i in range(nbands_s1)]):
                is_valid = False
            if ds_s2 and np.all([record.get(f's2_b{b_i}', 0) == 0 for b_i in range(nbands_s2)]):
                is_valid = False

            if is_valid:
                batch_records.append(record)
                if s1_prof is not None: batch_s1_profiles.append(s1_prof)
                if s2_prof is not None: batch_s2_profiles.append(s2_prof)
                batch_latlons.append([lat, lon])

            count += 1
            if count % 200 == 0 or count == total:
                sys.stdout.write(f"\r    Extracting raster stats: {count}/{total} segments...  ")
                sys.stdout.flush()

        print(f"\n    Computing Presto embeddings in parallel batches (total {len(batch_records)} valid segments)...")
        batch_size = 256
        n_batches = math.ceil(len(batch_records) / batch_size) if batch_records else 0
        for b_idx in range(n_batches):
            start_i = b_idx * batch_size
            end_i = min(len(batch_records), (b_idx + 1) * batch_size)
            b_ll = torch.tensor(batch_latlons[start_i:end_i], dtype=torch.float32, device=device)

            if ds_s1 and batch_s1_profiles:
                b_s1_tensor = torch.from_numpy(np.stack(batch_s1_profiles[start_i:end_i], axis=0)).to(device)
                embs_s1 = extractor.get_s1_embeddings(b_s1_tensor, b_ll, month_tensor_s1)
                for rec_i, emb in enumerate(embs_s1):
                    for f_i, f_val in enumerate(emb):
                        batch_records[start_i + rec_i][f'presto_s1_{f_i}'] = f_val

            if ds_s2 and batch_s2_profiles:
                b_s2_tensor = torch.from_numpy(np.stack(batch_s2_profiles[start_i:end_i], axis=0)).to(device)
                embs_s2 = extractor.get_s2_embeddings(b_s2_tensor, b_ll, month_tensor_s2)
                for rec_i, emb in enumerate(embs_s2):
                    for f_i, f_val in enumerate(emb):
                        batch_records[start_i + rec_i][f'presto_s2_{f_i}'] = f_val

            sys.stdout.write(f"\r    Presto batch progress: {end_i}/{len(batch_records)} segments ({(end_i/len(batch_records)*100):.1f}%)...  ")
            sys.stdout.flush()

        df = pd.DataFrame(batch_records)
        df.to_csv(self.sel_csv, index=False)
        print(f"\n    Multimodal features saved to {self.sel_csv} ({df.shape[1] - 2} features).\n")

    # --- Stage 4: Train Unified MLP + XGBoost Fusion Ensemble ---
    def stage_4_train_classifier(self, force_recompute=False, **kwargs):
        stage = 4
        if self.model_pkl.exists() and not force_recompute:
            print(f"[Stage {stage}] Fusion Ensemble model exists ({self.model_pkl.name}), skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Training Unified Multimodal Fusion Ensemble (PyTorch MLP + XGBoost)...")

        df = pd.read_csv(self.sel_csv)
        y = df['crop_id'].values
        X = df.drop(columns=['crop_id', 'seg_id']).values

        all_classes = np.unique(y)
        class_weights = _calculate_class_weights(y, all_classes)

        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)

        mlp = TorchMLPClassifier(
            hidden_layer_sizes=(512, 256, 128),
            max_iter=200,
            batch_size=256,
            class_weights=class_weights,
            all_classes=all_classes
        )

        if HAS_XGBOOST:
            xgb_model = xgb.XGBClassifier(
                n_estimators=250,
                max_depth=6,
                learning_rate=0.08,
                subsample=0.8,
                colsample_bytree=0.25,
                random_state=42,
                n_jobs=-1,
                tree_method='hist'
            )
        else:
            from sklearn.ensemble import HistGradientBoostingClassifier
            xgb_model = HistGradientBoostingClassifier(max_iter=200, random_state=42)

        fusion_ensemble = EnsembleClassifier(mlp, xgb_model, weight_mlp=self.mlp_weight)
        fusion_ensemble.fit(X_scaled, y)

        joblib.dump({'model': fusion_ensemble, 'scaler': scaler, 'classes': all_classes}, self.model_pkl)
        print(f"    Fusion Ensemble model saved to {self.model_pkl}\n")

    # --- Stage 5: Inference with Bayesian Priors ---
    def stage_5_classify_vector(self, force_recompute=False):
        stage = 5
        if self.class_tif.exists() and not force_recompute:
            print(f"[Stage {stage}] Classification raster exists, skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Running Vectorized Tile-based Object Inference with Bayesian Priors...")
        main_mod = sys.modules.get('__main__')
        if main_mod:
            setattr(main_mod, 'EnsembleClassifier', EnsembleClassifier)
            setattr(main_mod, 'TorchMLPClassifier', TorchMLPClassifier)
        sys.modules['1_classify_MLPXGB_presto_hybrid_S1S2'] = sys.modules[__name__]
        sys.modules['classifier_mlpxgb_presto'] = sys.modules[__name__]

        data = joblib.load(self.model_pkl)
        clf = data['model']
        scaler = data['scaler']

        ref_ras = self.s1_ras if self.s1_ras else self.s2_ras
        ds_info = gdal.Open(str(ref_ras))
        cols, rows = ds_info.RasterXSize, ds_info.RasterYSize
        gt, proj = ds_info.GetGeoTransform(), ds_info.GetProjection()

        driver = gdal.GetDriverByName('GTiff')
        ds_cls = driver.Create(str(self.class_tif), cols, rows, 1, gdal.GDT_Int32, options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        ds_cls.SetGeoTransform(gt)
        ds_cls.SetProjection(proj)

        ds_conf = driver.Create(str(self.conf_tif), cols, rows, 1, gdal.GDT_Float32, options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        ds_conf.SetGeoTransform(gt)
        ds_conf.SetProjection(proj)

        df_learn = pd.read_csv(self.sel_csv)
        classes = clf.classes_
        class_counts = df_learn['crop_id'].value_counts().to_dict()
        p_true = _get_priors_for_country(self.country, self.learn_shp, classes, class_counts, len(df_learn))
        
        balanced_counts = {c: max(class_counts.get(c, 0), 1000) for c in classes}
        tot_b = sum(balanced_counts.values())
        p_train = np.array([balanced_counts[c] / tot_b for c in classes])

        correction = np.clip(np.power(p_true / (p_train + 1e-9), 0.7), 0.01, 10.0)
        priors_arr = correction / np.sum(correction)

        seg_ds = gdal.Open(str(self.seg_tif))
        foot_ds = gdal.Open(str(self.footprint_mask))

        device = "cuda" if (HAS_TORCH and torch.cuda.is_available()) else "cpu"
        extractor = PrestoMultimodalExtractor(device=device)

        ds_s1 = gdal.Open(str(self.s1_ras)) if self.s1_ras else None
        ds_s2 = gdal.Open(str(self.s2_ras)) if self.s2_ras else None
        nbands_s1 = ds_s1.RasterCount if ds_s1 else 0
        nbands_s2 = ds_s2.RasterCount if ds_s2 else 0
        num_dates_s1 = nbands_s1 // 2
        num_dates_s2 = nbands_s2 // 9

        months_s1 = [parse_month_from_description(ds_s1.GetRasterBand(b).GetDescription()) for b in range(1, num_dates_s1 + 1)] if ds_s1 else [0]
        months_s2 = [parse_month_from_description(ds_s2.GetRasterBand(b).GetDescription()) for b in range(1, num_dates_s2 + 1)] if ds_s2 else [0]
        month_tensor_s1 = torch.tensor(months_s1, dtype=torch.long, device=device)
        month_tensor_s2 = torch.tensor(months_s2, dtype=torch.long, device=device)

        from scipy import ndimage
        srs_ras = osr.SpatialReference()
        srs_ras.ImportFromWkt(proj)
        ras_epsg = srs_ras.GetAttrValue("AUTHORITY", 1) or "3857"
        from pyproj import Transformer
        transformer_to_wgs84 = Transformer.from_crs(f"EPSG:{ras_epsg}", "EPSG:4326", always_xy=True)

        tile_size = 2048
        total_tiles = math.ceil(cols / tile_size) * math.ceil(rows / tile_size)
        tile_cnt = 0
        total_segments_classified = 0
        t_infer_start = time.time()

        print(f"    Starting vectorized tile-based inference ({total_tiles} tiles of {tile_size}x{tile_size} px)...")

        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                tile_cnt += 1
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                sub_seg = seg_ds.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                foot_arr = foot_ds.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)

                u_sids = np.unique(sub_seg)
                u_sids = u_sids[u_sids > 0]
                if len(u_sids) == 0:
                    if tile_cnt % 25 == 0 or tile_cnt == total_tiles:
                        elapsed = time.time() - t_infer_start
                        rate = tile_cnt / elapsed if elapsed > 0 else 0
                        eta_sec = (total_tiles - tile_cnt) / rate if rate > 0 else 0
                        eta_str = f"{int(eta_sec//60)}m {int(eta_sec%60):02d}s"
                        sys.stdout.write(
                            f"\r    [INFERENCE] Tile {tile_cnt}/{total_tiles} ({(tile_cnt/total_tiles*100):.1f}%) | "
                            f"Objects: {total_segments_classified:,} | Time: {int(elapsed//60)}m {int(elapsed%60):02d}s | ETA: {eta_str}  "
                        )
                        sys.stdout.flush()
                    continue

                flat_labels = sub_seg.ravel()
                counts = np.bincount(flat_labels)
                valid_counts = np.maximum(counts[u_sids], 1)

                feat_blocks = []

                # Centroids calculation
                centers = ndimage.center_of_mass(np.ones_like(sub_seg), labels=sub_seg, index=u_sids)
                cy_arr = np.array([c[0] for c in centers], dtype=np.float64) + y
                cx_arr = np.array([c[1] for c in centers], dtype=np.float64) + x
                mx_arr = gt[0] + cx_arr * gt[1] + cy_arr * gt[2]
                my_arr = gt[3] + cx_arr * gt[4] + cy_arr * gt[5]
                lons, lats = transformer_to_wgs84.transform(mx_arr, my_arr)
                latlons = np.column_stack([lats, lons]).astype(np.float32)
                b_ll = torch.from_numpy(latlons).to(device)

                # Sentinel-1 Block Read & Vectorized Zonal Means
                embs_s1 = None
                s1_means = None
                if ds_s1:
                    s1_tile = np.nan_to_num(ds_s1.ReadAsArray(x, y, xsize, ysize).astype(np.float32))
                    if s1_tile.ndim == 2:
                        s1_tile = s1_tile[np.newaxis, ...]
                    s1_means = np.zeros((len(u_sids), nbands_s1), dtype=np.float32)
                    for b in range(nbands_s1):
                        sums = np.bincount(flat_labels, weights=s1_tile[b].ravel())
                        s1_means[:, b] = sums[u_sids] / valid_counts

                    s1_profiles = np.zeros((len(u_sids), num_dates_s1, 2), dtype=np.float32)
                    for d in range(num_dates_s1):
                        s1_profiles[:, d, 0] = (s1_means[:, num_dates_s1 + d] + 25.0) / 25.0
                        s1_profiles[:, d, 1] = (s1_means[:, d] + 25.0) / 25.0

                    b_s1 = torch.from_numpy(s1_profiles).to(device)
                    embs_s1 = extractor.get_s1_embeddings(b_s1, b_ll, month_tensor_s1)

                # Sentinel-2 Block Read & Vectorized Zonal Means
                embs_s2 = None
                s2_means = None
                if ds_s2:
                    s2_tile = np.nan_to_num(ds_s2.ReadAsArray(x, y, xsize, ysize).astype(np.float32))
                    if s2_tile.ndim == 2:
                        s2_tile = s2_tile[np.newaxis, ...]
                    s2_means = np.zeros((len(u_sids), nbands_s2), dtype=np.float32)
                    for b in range(nbands_s2):
                        sums = np.bincount(flat_labels, weights=s2_tile[b].ravel())
                        s2_means[:, b] = sums[u_sids] / valid_counts

                    s2_profiles = np.zeros((len(u_sids), num_dates_s2, 9), dtype=np.float32)
                    for d in range(num_dates_s2):
                        for band_idx in range(9):
                            s2_profiles[:, d, band_idx] = s2_means[:, d * 9 + band_idx] / 10000.0

                    b_s2 = torch.from_numpy(s2_profiles).to(device)
                    embs_s2 = extractor.get_s2_embeddings(b_s2, b_ll, month_tensor_s2)

                # Exact column order matching Stage 3: [s1_means, s2_means, embs_s1, embs_s2]
                feat_blocks = []
                if s1_means is not None:
                    feat_blocks.append(s1_means)
                if s2_means is not None:
                    feat_blocks.append(s2_means)
                if embs_s1 is not None:
                    feat_blocks.append(embs_s1)
                if embs_s2 is not None:
                    feat_blocks.append(embs_s2)

                X_tile = np.hstack(feat_blocks)
                X_tile_scaled = scaler.transform(X_tile)
                raw_probs = clf.predict_proba(X_tile_scaled)

                corr_probs = raw_probs * priors_arr
                corr_probs = corr_probs / np.sum(corr_probs, axis=1, keepdims=True)

                preds = clf.classes_[np.argmax(corr_probs, axis=1)]
                confs = np.max(corr_probs, axis=1)

                # O(1) Fast Vectorized LUT Remapping
                max_sid = int(np.max(u_sids))
                lut_pred = np.zeros(max_sid + 1, dtype=np.int32)
                lut_conf = np.zeros(max_sid + 1, dtype=np.float32)
                lut_pred[u_sids] = preds
                lut_conf[u_sids] = confs

                pred_arr = lut_pred[sub_seg]
                prob_arr = lut_conf[sub_seg]

                pred_arr[foot_arr == 0] = 0
                prob_arr[foot_arr == 0] = 0

                ds_cls.GetRasterBand(1).WriteArray(pred_arr, x, y)
                ds_conf.GetRasterBand(1).WriteArray(prob_arr, x, y)

                total_segments_classified += len(u_sids)

                elapsed = time.time() - t_infer_start
                rate = tile_cnt / elapsed if elapsed > 0 else 0
                eta_sec = (total_tiles - tile_cnt) / rate if rate > 0 else 0
                eta_str = f"{int(eta_sec//60)}m {int(eta_sec%60):02d}s"
                sys.stdout.write(
                    f"\r    [INFERENCE] Tile {tile_cnt}/{total_tiles} ({(tile_cnt/total_tiles*100):.1f}%) | "
                    f"Objects: {total_segments_classified:,} | Time: {int(elapsed//60)}m {int(elapsed%60):02d}s | ETA: {eta_str}  "
                )
                sys.stdout.flush()

        ds_cls.FlushCache()
        ds_conf.FlushCache()
        ds_cls = None
        ds_conf = None
        total_time_min = (time.time() - t_infer_start) / 60.0
        print(f"\n    [INFERENCE COMPLETE] Successfully classified {total_segments_classified:,} objects across {total_tiles} tiles in {total_time_min:.1f} minutes.")
        print(f"    Raw classification saved: {self.class_tif}\n")

    # --- Stage 6: Apply Masking ---
    def stage_6_mask_classification(self, force_recompute=False):
        stage = 6
        if self.masked_class.exists() and self.masked_conf.exists() and not force_recompute:
            print(f"[Stage {stage}] Masked outputs already exist, skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Applying Agricultural & Footprint Masks...")
        ref_ras = self.class_tif if self.class_tif.exists() else (self.s1_ras if self.s1_ras else self.s2_ras)
        if not ref_ras or not ref_ras.exists():
            print("ERROR: Reference raster or classification output not found.")
            return

        if not self.class_tif.exists() or not self.conf_tif.exists():
            print("ERROR: Classification outputs not found. Run Stage 5 first.")
            return

        ds_ref = gdal.Open(str(ref_ras))
        cols = ds_ref.RasterXSize
        rows = ds_ref.RasterYSize
        gt = ds_ref.GetGeoTransform()
        proj = ds_ref.GetProjection()
        ds_ref = None

        mask_tif = self.agri_mask
        ds_mask = None
        temp_mask_vrt = None
        if mask_tif and mask_tif.exists():
            print(f"    Warping country agricultural mask to match classification raster bounds...")
            minx = gt[0]
            maxy = gt[3]
            maxx = minx + gt[1] * cols
            miny = maxy + gt[5] * rows

            temp_mask_vrt = str(self.masked_class).replace('.tif', '_mask_temp.vrt')
            mask_opts = gdal.WarpOptions(
                format='VRT',
                outputBounds=(minx, miny, maxx, maxy),
                width=cols,
                height=rows,
                dstSRS=proj,
                resampleAlg=gdal.GRA_NearestNeighbour
            )
            ds_mask = gdal.Warp(temp_mask_vrt, str(mask_tif), options=mask_opts)
            if not ds_mask:
                print(f"    [WARNING] Failed to warp agricultural mask ({mask_tif}). Continuing without it.")
            else:
                print(f"    Applying country agricultural mask: {mask_tif.name}")
        else:
            print("    [INFO] Arable mask not found or not configured. Masking with data footprint only.")

        ds_foot = gdal.Open(str(self.footprint_mask)) if (self.footprint_mask and self.footprint_mask.exists()) else None
        ds_cls = gdal.Open(str(self.class_tif))
        ds_conf = gdal.Open(str(self.conf_tif))

        driver = gdal.GetDriverByName('GTiff')
        out_cls = driver.Create(str(self.masked_class), cols, rows, 1, gdal.GDT_Int32,
                                options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        out_cls.SetGeoTransform(gt)
        out_cls.SetProjection(proj)
        out_cls.GetRasterBand(1).SetNoDataValue(0)

        out_conf = driver.Create(str(self.masked_conf), cols, rows, 1, gdal.GDT_Float32,
                                 options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        out_conf.SetGeoTransform(gt)
        out_conf.SetProjection(proj)
        out_conf.GetRasterBand(1).SetNoDataValue(0)

        tile_size = 4096
        total_tiles = math.ceil(cols / tile_size) * math.ceil(rows / tile_size)
        tile_cnt = 0

        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                cls_arr = ds_cls.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                conf_arr = ds_conf.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)

                combined_mask = np.ones((ysize, xsize), dtype=bool)
                if ds_foot:
                    foot_arr = ds_foot.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                    combined_mask = combined_mask & (foot_arr > 0)

                if ds_mask:
                    mask_arr = ds_mask.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                    combined_mask = combined_mask & (mask_arr > 0)

                cls_arr[~combined_mask] = 0
                conf_arr[~combined_mask] = 0.0

                out_cls.GetRasterBand(1).WriteArray(cls_arr, x, y)
                out_conf.GetRasterBand(1).WriteArray(conf_arr, x, y)

                tile_cnt += 1
                if tile_cnt % 20 == 0 or tile_cnt == total_tiles:
                    pct = (tile_cnt / total_tiles) * 100.0
                    print(f"    [MASKING PROGRESS] {tile_cnt}/{total_tiles} tiles ({pct:.1f}%)", flush=True)

        out_cls.GetRasterBand(1).FlushCache()
        out_conf.GetRasterBand(1).FlushCache()

        # Build pyramids
        try:
            out_cls.BuildOverviews('NEAREST', [2, 4, 8, 16, 32, 64])
            out_conf.BuildOverviews('AVERAGE', [2, 4, 8, 16, 32, 64])
        except Exception:
            pass

        out_cls = None
        out_conf = None
        ds_mask = None
        ds_foot = None
        ds_cls = None
        ds_conf = None

        if temp_mask_vrt and os.path.exists(temp_mask_vrt):
            try:
                os.remove(temp_mask_vrt)
            except Exception:
                pass

        print(f"    Masked classification saved to {self.masked_class}")
        print(f"    Masked confidence saved to {self.masked_conf}\n")

    # --- Stage 7: Validation Metrics ---
    def stage_7_calculate_metrics(self):
        stage = 7
        print(f"[Stage {stage}/{self.total_stages}] Calculating Out-of-Bag Validation Metrics and Generating Excel Report...")
        eval_shp = self.control_shp if (self.control_shp and self.control_shp.exists()) else self.learn_shp
        if not eval_shp or not eval_shp.exists():
            print(f"    [WARNING] No validation shapefile found ({eval_shp}). Skipping point evaluation.")
            return

        target_tif = self.masked_class if self.masked_class.exists() else self.class_tif
        if not target_tif or not target_tif.exists():
            print(f"    [ERROR] No classified raster found ({target_tif}). Run Stage 5 & 6 first.")
            return

        gdf_val = gpd.read_file(str(eval_shp), engine="pyogrio")
        crop_col = 'crop_id' if 'crop_id' in gdf_val.columns else 'code'
        gdf_val['crop_id'] = gdf_val[crop_col].astype(int)

        ds = gdal.Open(str(target_tif))
        gt = ds.GetGeoTransform()
        inv_gt = gdal.InvGeoTransform(gt)
        cols, rows = ds.RasterXSize, ds.RasterYSize
        cls_arr = ds.GetRasterBand(1).ReadAsArray()
        ras_proj = ds.GetProjection()

        if ras_proj and gdf_val.crs:
            from pyproj import CRS
            target_crs = CRS.from_wkt(ras_proj)
            if gdf_val.crs != target_crs:
                gdf_val_proj = gdf_val.to_crs(target_crs)
            else:
                gdf_val_proj = gdf_val
        else:
            gdf_val_proj = gdf_val

        if hasattr(gdf_val_proj.geometry, 'x'):
            xs = gdf_val_proj.geometry.x.values
            ys = gdf_val_proj.geometry.y.values
        else:
            xs = gdf_val_proj.geometry.centroid.x.values
            ys = gdf_val_proj.geometry.centroid.y.values

        pxs = (inv_gt[0] + inv_gt[1] * xs + inv_gt[2] * ys).astype(int)
        pys = (inv_gt[3] + inv_gt[4] * xs + inv_gt[5] * ys).astype(int)

        y_true = []
        y_pred = []
        for cid, px, py in zip(gdf_val['crop_id'].values, pxs, pys):
            if 0 <= px < cols and 0 <= py < rows:
                pred = cls_arr[py, px]
                if pred > 0:
                    y_true.append(cid)
                    y_pred.append(pred)

        if not y_true:
            print("    [WARNING] No validation points fell on valid classified pixels.")
            return

        y_true = np.array(y_true)
        y_pred = np.array(y_pred)

        all_classes = np.unique(np.concatenate([y_true, y_pred]))
        cm = confusion_matrix(y_true, y_pred, labels=all_classes)
        oa = float(np.trace(cm) / np.sum(cm))

        # Cohen's Kappa
        p_o = oa
        p_e = float(np.sum(np.sum(cm, axis=0) * np.sum(cm, axis=1)) / (np.sum(cm) ** 2))
        kappa = float((p_o - p_e) / (1.0 - p_e + 1e-9))

        precision, recall, f1, _ = precision_recall_fscore_support(y_true, y_pred, labels=all_classes, zero_division=0)

        # Resolve English crop names mapping
        crop_name_map = {}
        id_cols = [c for c in ['crop_id', 'crop_ids', 'code', 'id', 'class_id'] if c in gdf_val.columns]
        name_cols = [c for c in ['crop_name', 'crop_names', 'crop_type', 'label', 'name', 'class_name', 'crop', 'nom'] if c in gdf_val.columns]
        if id_cols and name_cols:
            for _, row in gdf_val[[id_cols[0], name_cols[0]]].drop_duplicates().iterrows():
                try:
                    cid = int(row[id_cols[0]])
                    cname = str(row[name_cols[0]]).strip()
                    if cname and cname.lower() not in ['none', 'nan', '']:
                        crop_name_map[cid] = cname
                except Exception:
                    pass

        # Fallback to master sample_shp if any crop name missing
        if self.sample_shp and self.sample_shp.exists():
            try:
                gdf_samp = gpd.read_file(str(self.sample_shp), engine="pyogrio")
                id_cols_s = [c for c in ['crop_id', 'crop_ids', 'code', 'id', 'class_id'] if c in gdf_samp.columns]
                name_cols_s = [c for c in ['crop_name', 'crop_names', 'crop_type', 'label', 'name', 'class_name', 'crop', 'nom'] if c in gdf_samp.columns]
                if id_cols_s and name_cols_s:
                    for _, row in gdf_samp[[id_cols_s[0], name_cols_s[0]]].drop_duplicates().iterrows():
                        try:
                            cid = int(row[id_cols_s[0]])
                            cname = str(row[name_cols_s[0]]).strip()
                            if cid not in crop_name_map and cname and cname.lower() not in ['none', 'nan', '']:
                                crop_name_map[cid] = cname
                        except Exception:
                            pass
            except Exception:
                pass

        # Fallback to country priors.json if available
        country_priors_file = self.aux_dir / "shapefiles_samples" / self.country / "priors.json"
        if country_priors_file.exists():
            try:
                with open(country_priors_file, 'r', encoding='utf-8') as pf:
                    priors_data = json.load(pf)
                    for p_idx, p_name in enumerate(priors_data.keys(), start=1):
                        if p_idx not in crop_name_map:
                            crop_name_map[p_idx] = p_name.title()
            except Exception:
                pass

        # Excel Export
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Validation Metrics"

        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        sub_font = Font(name="Calibri", size=11, bold=True)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )

        ws.cell(row=1, column=1, value=f"Crop Classification Accuracy Report: {self.track}").font = Font(name="Calibri", size=14, bold=True, color="1F497D")
        ws.cell(row=2, column=1, value=f"Model: Unified PyTorch MLP + XGBoost Fusion Ensemble").font = sub_font
        ws.cell(row=3, column=1, value=f"Data: Multimodal Sentinel-1 (Sigma0 VH/VV) + Sentinel-2 (B02-B12) + Presto Embeddings").font = Font(name="Calibri", size=10, italic=True)

        ws.cell(row=5, column=1, value="Metric").font = sub_font
        ws.cell(row=5, column=2, value="Value").font = sub_font
        ws.cell(row=6, column=1, value="Overall Accuracy (OA)").font = Font(name="Calibri", size=11)
        ws.cell(row=6, column=2, value=f"{oa * 100:.2f}%").font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=7, column=1, value="Cohen's Kappa").font = Font(name="Calibri", size=11)
        ws.cell(row=7, column=2, value=f"{kappa:.4f}").font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=8, column=1, value="Validation Samples Count").font = Font(name="Calibri", size=11)
        ws.cell(row=8, column=2, value=len(y_true)).font = Font(name="Calibri", size=11)

        # Per-class table
        r = 10
        ws.cell(row=r, column=1, value="Class ID").fill = header_fill
        ws.cell(row=r, column=1).font = header_font
        ws.cell(row=r, column=2, value="Crop Name").fill = header_fill
        ws.cell(row=r, column=2).font = header_font
        ws.cell(row=r, column=3, value="Precision (User Acc)").fill = header_fill
        ws.cell(row=r, column=3).font = header_font
        ws.cell(row=r, column=4, value="Recall (Prod Acc)").fill = header_fill
        ws.cell(row=r, column=4).font = header_font
        ws.cell(row=r, column=5, value="F1-Score").fill = header_fill
        ws.cell(row=r, column=5).font = header_font
        ws.cell(row=r, column=6, value="Validation Samples").fill = header_fill
        ws.cell(row=r, column=6).font = header_font

        for idx, cid in enumerate(all_classes):
            r += 1
            c_name = crop_name_map.get(int(cid), f"Class {cid}")
            sample_cnt = int(np.sum(y_true == cid))
            ws.cell(row=r, column=1, value=int(cid)).border = thin_border
            ws.cell(row=r, column=2, value=c_name).border = thin_border
            ws.cell(row=r, column=3, value=f"{precision[idx] * 100:.2f}%").border = thin_border
            ws.cell(row=r, column=4, value=f"{recall[idx] * 100:.2f}%").border = thin_border
            ws.cell(row=r, column=5, value=f"{f1[idx] * 100:.2f}%").border = thin_border
            ws.cell(row=r, column=6, value=sample_cnt).border = thin_border

        # Confusion Matrix table
        r += 3
        ws.cell(row=r, column=1, value="Confusion Matrix (Rows: Ground Truth, Cols: Prediction)").font = sub_font
        r += 1
        ws.cell(row=r, column=1, value="True \\ Pred").fill = header_fill
        ws.cell(row=r, column=1).font = header_font
        for c_idx, cid in enumerate(all_classes):
            c_name = crop_name_map.get(int(cid), str(cid))
            cell = ws.cell(row=r, column=c_idx + 2, value=f"{int(cid)}: {c_name}")
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, true_cid in enumerate(all_classes):
            r += 1
            t_name = crop_name_map.get(int(true_cid), str(true_cid))
            ws.cell(row=r, column=1, value=f"{int(true_cid)}: {t_name}").font = sub_font
            for col_idx, pred_cid in enumerate(all_classes):
                val = int(cm[row_idx, col_idx])
                cell = ws.cell(row=r, column=col_idx + 2, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

        # Area estimation table
        unique_cls, counts = np.unique(cls_arr, return_counts=True)
        valid_cls_mask = unique_cls > 0
        if np.any(valid_cls_mask):
            unique_cls = unique_cls[valid_cls_mask]
            counts = counts[valid_cls_mask]
            
            resx = abs(gt[1])
            resy = abs(gt[5])
            pixel_ha = (resx * resy) / 10000.0
            total_ha = float(np.sum(counts) * pixel_ha)

            r += 3
            ws.cell(row=r, column=1, value="Classified Agricultural Area Statistics").font = sub_font
            r += 1
            ws.cell(row=r, column=1, value="Class ID").fill = header_fill
            ws.cell(row=r, column=1).font = header_font
            ws.cell(row=r, column=2, value="Crop Name").fill = header_fill
            ws.cell(row=r, column=2).font = header_font
            ws.cell(row=r, column=3, value="Area (ha)").fill = header_fill
            ws.cell(row=r, column=3).font = header_font
            ws.cell(row=r, column=4, value="Area (%)").fill = header_fill
            ws.cell(row=r, column=4).font = header_font

            for u_id, u_count in zip(unique_cls, counts):
                r += 1
                c_name = crop_name_map.get(int(u_id), f"Class {u_id}")
                c_ha = float(u_count * pixel_ha)
                c_pct = (c_ha / total_ha * 100.0) if total_ha > 0 else 0.0
                ws.cell(row=r, column=1, value=int(u_id)).border = thin_border
                ws.cell(row=r, column=2, value=c_name).border = thin_border
                ws.cell(row=r, column=3, value=round(c_ha, 2)).border = thin_border
                ws.cell(row=r, column=4, value=f"{c_pct:.2f}%").border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

        wb.save(str(self.metrics_fp))
        print(f"    [OK] Metrics report saved to: {self.metrics_fp}")
        print(f"    Validation Overall Accuracy (OA): {oa * 100:.2f}% | Kappa: {kappa:.4f}\n")


# =====================================================================
# 5. CLI & INTERACTIVE MENU
# =====================================================================

def main_menu(pipeline):
    while True:
        menu = f"""
    --- Multimodal Crop Classification Pipeline (Unified MLP + XGBoost Fusion) ---
    Track: {pipeline.track} ({pipeline.country})
    Model: Unified PyTorch MLP + XGBoost Fusion Ensemble (weight: {pipeline.mlp_weight:.2f})
    Segmentation: {pipeline.seg_mode.upper()}

    [0] Stage 0: Generate Data Footprint
    [1] Stage 1: Multimodal Segmentation
    [2] Stage 2: Prepare Point Split (70/30)
    [3] Stage 3: Extract Multimodal Features (S1+S2+Presto)
    [4] Stage 4: Train Unified MLP + XGBoost Fusion Ensemble
    [5] Stage 5: Run Object-Based Inference with Bayesian Priors
    [6] Stage 6: Apply Agricultural & Footprint Mask
    [7] Stage 7: Calculate Validation Metrics (.xlsx)

    [A] Run All Stages
    [Q] Quit

    Enter choice: """
        try:
            choice = input(menu).strip().upper()
            if choice == '0': pipeline.stage_0_generate_footprint(True)
            elif choice == '1': pipeline.stage_1_segmentation(True)
            elif choice == '2': pipeline.stage_2_split_samples(True)
            elif choice == '3': pipeline.stage_3_selection(True)
            elif choice == '4': pipeline.stage_4_train_classifier(True)
            elif choice == '5': pipeline.stage_5_classify_vector(True)
            elif choice == '6': pipeline.stage_6_mask_classification(True)
            elif choice == '7': pipeline.stage_7_calculate_metrics()
            elif choice == 'A':
                pipeline.stage_0_generate_footprint(False)
                pipeline.stage_1_segmentation(False)
                pipeline.stage_2_split_samples(False)
                pipeline.stage_3_selection(False)
                pipeline.stage_4_train_classifier(False)
                pipeline.stage_5_classify_vector(True)
                pipeline.stage_6_mask_classification(True)
                pipeline.stage_7_calculate_metrics()
            elif choice == 'Q': break
        except (KeyboardInterrupt, EOFError):
            break


def main():
    parser = argparse.ArgumentParser(description="Multimodal S1 (Sigma0) + S2 Crop Classification with Unified MLP + XGBoost Fusion Ensemble.")
    parser.add_argument('--track', required=True, help="Track/orbit identifier, e.g. NL/orbit_88, PT/orbit_161, PL/orbit_12")
    parser.add_argument('--stage', default=None, help="Stage to run: 'A' (all), '0', '1', '2', '3', '4', '5', '6', '7'")
    parser.add_argument('--seg_mode', default='slic', choices=['sam', 'slic', 'lpis'], help="Segmentation mode (default: slic)")
    parser.add_argument('--mlp_weight', type=float, default=0.65, help="Weight of MLP in fusion ensemble (0.0 to 1.0, default: 0.65)")
    parser.add_argument('--s1_raster', default=None, help="Override path to Sentinel-1 Sigma0 VH/VV GeoTIFF raster")
    parser.add_argument('--s2_raster', default=None, help="Override path to Sentinel-2 Multi-temporal GeoTIFF raster")
    parser.add_argument('--lpis_vector', default=None, help="Path to official LPIS cadastral parcel vector file (.shp, .gpkg) for --seg_mode lpis")

    args = parser.parse_args()

    pipeline = ProcessingPipelineS1S2(
        track=args.track,
        seg_mode=args.seg_mode,
        mlp_weight=args.mlp_weight,
        s1_override=args.s1_raster,
        s2_override=args.s2_raster,
        lpis_vector=args.lpis_vector
    )

    if args.stage is None:
        main_menu(pipeline)
    else:
        choice = args.stage.strip().upper()
        if choice == 'A':
            pipeline.stage_0_generate_footprint(False)
            pipeline.stage_1_segmentation(False)
            pipeline.stage_2_split_samples(False)
            pipeline.stage_3_selection(False)
            pipeline.stage_4_train_classifier(False)
            pipeline.stage_5_classify_vector(True)
            pipeline.stage_6_mask_classification(True)
            pipeline.stage_7_calculate_metrics()
        elif choice == '0': pipeline.stage_0_generate_footprint(True)
        elif choice == '1': pipeline.stage_1_segmentation(True)
        elif choice == '2': pipeline.stage_2_split_samples(True)
        elif choice == '3': pipeline.stage_3_selection(True)
        elif choice == '4': pipeline.stage_4_train_classifier(True)
        elif choice == '5': pipeline.stage_5_classify_vector(True)
        elif choice == '6': pipeline.stage_6_mask_classification(True)
        elif choice == '7': pipeline.stage_7_calculate_metrics()


if __name__ == '__main__':
    main()
