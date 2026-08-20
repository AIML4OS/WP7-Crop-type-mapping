# How to run the script:
# ---------------------
# Country-Wide Merge (combines all processed orbits for a country e.g. NL, PL, PT, AT, FR)
#
# 1. Multimodal S1 + S2 Presto + MLP/XGBoost Fusion Ensemble (Recommended):
#    python 2_merge_classifications.py --track NL --suffix _mlpxgb_presto_slic
#    python 2_merge_classifications.py --track NL --suffix _mlpxgb_presto_lpis
#    python 2_merge_classifications.py --track NL --suffix _mlpxgb_presto_sam
#    python 2_merge_classifications.py --track PL --suffix _mlpxgb_presto_slic
#    python 2_merge_classifications.py --track PT --suffix _mlpxgb_presto_slic
#
# 2. NASA Harvest Presto + SAR Hybrid classification:
#    python 2_merge_classifications.py --track NL --suffix _presto_hybrid_lpis
#    python 2_merge_classifications.py --track NL --suffix _presto_hybrid_slic
#    python 2_merge_classifications.py --track NL --suffix _presto_hybrid_sam
#
# 3. Standard ANN classification (legacy):
#    python 2_merge_classifications.py --track NL --suffix _lpis
#    python 2_merge_classifications.py --track NL --suffix _slic
#    python 2_merge_classifications.py --track NL --suffix _sam
#
# 4. NASA Harvest Presto SAR-only classification (legacy):
#    python 2_merge_classifications.py --track NL --suffix _presto_lpis
#    python 2_merge_classifications.py --track NL --suffix _presto_slic
#    python 2_merge_classifications.py --track NL --suffix _presto_sam
#
# 5. Prithvi Foundation Model classification (legacy):
#    python 2_merge_classifications.py --track NL --suffix _prithvi_lpis
#    python 2_merge_classifications.py --track NL --suffix _prithvi_slic
#    python 2_merge_classifications.py --track NL --suffix _prithvi_sam

import argparse
import os
from pathlib import Path
from typing import List, Tuple
import numpy as np
from osgeo import gdal
import geopandas as gpd
from sklearn.metrics import confusion_matrix, precision_recall_fscore_support
import openpyxl
from openpyxl.styles import Font

def get_crop_aggregation(country, learn_shp_path):
    """
    Returns an empty dictionary because the grassland-related classes (Clover/Lucerne)
    have been directly integrated into Grassland in the shapefile itself.
    """
    return {}


# Discover tracks logic

def get_masked_filenames(track_prefix: str, suffix: str) -> List[Tuple[str, str]]:
    """Returns candidate (classification_masked, confidence_masked) filename pairs."""
    candidates = []

    # 1. Direct standard patterns
    candidates.append((f"{track_prefix}_classified_masked{suffix}.tif", f"{track_prefix}_confidence_masked{suffix}.tif"))
    candidates.append((f"{track_prefix}_classified{suffix}_masked.tif", f"{track_prefix}_confidence{suffix}_masked.tif"))

    # 2. Presto Hybrid specific
    if suffix.startswith('_presto_hybrid'):
        s_clean = suffix[14:]  # remove '_presto_hybrid'
        candidates.append((f"{track_prefix}_presto_hybrid_classified_masked{s_clean}.tif", f"{track_prefix}_presto_hybrid_confidence_masked{s_clean}.tif"))
        candidates.append((f"{track_prefix}_presto_hybrid_classified{s_clean}_masked.tif", f"{track_prefix}_presto_hybrid_confidence{s_clean}_masked.tif"))

    # 3. Presto SAR specific
    elif suffix.startswith('_presto'):
        s_clean = suffix[7:]  # remove '_presto'
        candidates.append((f"{track_prefix}_presto_classified_masked{s_clean}.tif", f"{track_prefix}_presto_confidence_masked{s_clean}.tif"))

    # 4. Prithvi specific
    elif suffix.startswith('_prithvi'):
        s_clean = suffix[8:]  # remove '_prithvi'
        candidates.append((f"{track_prefix}_prithvi_classified_masked{s_clean}.tif", f"{track_prefix}_prithvi_confidence_masked{s_clean}.tif"))

    # 5. MLPXGB / S1S2 specific
    if 'mlpxgb' in suffix or 's1s2' in suffix:
        candidates.append((f"{track_prefix}_classified_s1s2_masked.tif", f"{track_prefix}_confidence_s1s2_masked.tif"))
        candidates.append((f"{track_prefix}_classified_masked_s1s2.tif", f"{track_prefix}_confidence_masked_s1s2.tif"))

    return candidates


def find_masked_files(base_dir: Path, tr: str, country: str, suffix: str = ""):
    """
    Look for classification and confidence masked rasters in either:
      - classification_results/classification/
      - classification_results/
    Returns (cls_fp, conf_fp) or (None, None).
    """
    folders = [
        base_dir / tr / 'classification_results' / 'classification',
        base_dir / tr / 'classification_results'
    ]
    filename_pairs = get_masked_filenames(f"{country}_{tr}", suffix)
    for folder in folders:
        for cls_name, conf_name in filename_pairs:
            cls_fp = folder / cls_name
            conf_fp = folder / conf_name
            if cls_fp.exists() and conf_fp.exists():
                return cls_fp, conf_fp
    return None, None


def discover_tracks(base_dir: Path, prefix: str, suffix: str = ""):
    """
    Discovers track folders and their classification/confidence files.
    Supports:
      1. Direct specific orbit path: e.g. NL/orbit_161 or NL_orbit_161.
      2. Country-based orbits: Prefix is a country code (e.g. PL, NL, FR, PT, AT).
         We scan base_dir/prefix/ for orbit_* folders.
      3. Legacy tracks: Prefix is a track prefix (e.g. P1, P2).
         We scan base_dir/ for folders starting with prefix.
     Returns list of (tr, country, cls_fp, conf_fp).
    """
    tracks = []
    prefix_upper = prefix.upper()

    # Case 0: Prefix points directly to a specific orbit directory (e.g. NL/orbit_161 or NL_orbit_161)
    normalized_prefix = prefix.replace('\\', '/')
    specific_path = base_dir / normalized_prefix
    if not specific_path.is_dir() and '_' in normalized_prefix:
        alt_prefix = normalized_prefix.replace('_', '/', 1)
        alt_path = base_dir / alt_prefix
        if alt_path.is_dir():
            specific_path = alt_path
            normalized_prefix = alt_prefix

    if specific_path.is_dir():
        tr = normalized_prefix
        parts = tr.split('/')
        country = parts[0].upper()
        sanitized = tr.replace('/', '_')
        
        track_prefix = sanitized if sanitized.upper().startswith(country.upper() + "_") else f"{country}_{sanitized}"
        filename_pairs = get_masked_filenames(track_prefix, suffix)
        
        candidates = [
            specific_path / 'classification_results' / 'classification',
            specific_path / 'classification_results'
        ]
        found = False
        for folder in candidates:
            for cls_name, conf_name in filename_pairs:
                cls_fp  = folder / cls_name
                conf_fp = folder / conf_name
                if cls_fp.exists() and conf_fp.exists():
                    tracks.append((tr, country, cls_fp, conf_fp))
                    found = True
                    break
            if found:
                break

    # Case 1: Prefix is a country code (2 letters)
    if not tracks and len(prefix) == 2 and prefix.isalpha():
        country = prefix_upper
        country_dir = base_dir / country
        if country_dir.exists():
            for sub in country_dir.iterdir():
                if not sub.is_dir() or not sub.name.startswith("orbit_"):
                    continue
                tr = f"{country}/{sub.name}"
                sanitized = tr.replace('/', '_').replace('\\', '_')
                
                track_prefix = sanitized if sanitized.upper().startswith(country.upper() + "_") else f"{country}_{sanitized}"
                filename_pairs = get_masked_filenames(track_prefix, suffix)
                
                candidates = [
                    sub / 'classification_results' / 'classification',
                    sub / 'classification_results'
                ]
                found = False
                for folder in candidates:
                    for cls_name, conf_name in filename_pairs:
                        cls_fp  = folder / cls_name
                        conf_fp = folder / conf_name
                        if cls_fp.exists() and conf_fp.exists():
                            tracks.append((tr, country, cls_fp, conf_fp))
                            found = True
                            break
                    if found:
                        break
    
    # Case 2: Prefix is a legacy track prefix
    if not tracks:
        for sub in base_dir.iterdir():
            if not sub.is_dir():
                continue
            tr = sub.name
            if not tr.startswith(prefix):
                continue
            if '/' in tr or '\\' in tr:
                country = tr.replace('\\', '/').split('/')[0].upper()
            elif len(tr) == 2:
                country = tr.upper()
            else:
                continue
            
            sanitized = tr.replace('/', '_').replace('\\', '_')
            track_prefix = sanitized if sanitized.upper().startswith(country.upper() + "_") else f"{country}_{sanitized}"
            filename_pairs = get_masked_filenames(track_prefix, suffix)
            
            candidates = [
                base_dir / tr / 'classification_results' / 'classification',
                base_dir / tr / 'classification_results'
            ]
            found = False
            for folder in candidates:
                for cls_name, conf_name in filename_pairs:
                    cls_fp  = folder / cls_name
                    conf_fp = folder / conf_name
                    if cls_fp.exists() and conf_fp.exists():
                        tracks.append((tr, country, cls_fp, conf_fp))
                        found = True
                        break
                if found:
                    break
                    
    return tracks

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--track', required=True,
                        help='Base track prefix or country code (e.g. P1, PL, FR)')
    parser.add_argument('--suffix', default='',
                        help='Optional suffix of classification files (e.g. _prithvi)')
    args = parser.parse_args()
    prefix = args.track
    suffix = args.suffix

    base_dir = Path(os.environ.get("AIML_WORKING_DIR", r"D:\AIML_CropMapper_Cloud\workingDir"))
    tracks   = discover_tracks(base_dir, prefix, suffix=suffix)
    if not tracks:
        raise FileNotFoundError(
            f"No valid classification/confidence files for tracks starting with {prefix}"
        )

    print(f"Discovered tracks: {[t for t,_,_,_ in tracks]}")

    # --- compute union extent & grid ---------------------------------------
    ds0   = gdal.Open(str(tracks[0][2]))
    proj  = ds0.GetProjection()
    gt0   = ds0.GetGeoTransform()
    resX, resY = gt0[1], abs(gt0[5])

    extents = []
    for _, _, cls_fp, _ in tracks:
        ds = gdal.Open(str(cls_fp))
        gt = ds.GetGeoTransform()
        c, r = ds.RasterXSize, ds.RasterYSize
        minX = gt[0]
        maxY = gt[3]
        maxX = gt[0] + c * gt[1]
        minY = gt[3] + r * gt[5]
        extents.append((minX, maxY, maxX, minY))

    minX = min(e[0] for e in extents)
    maxY = max(e[1] for e in extents)
    maxX = max(e[2] for e in extents)
    minY = min(e[3] for e in extents)

    cols = int(np.ceil((maxX - minX) / resX))
    rows = int(np.ceil((maxY - minY) / resY))
    gt_global = (minX, resX, 0, maxY, 0, -resY)

    print(f"Global mosaic: {cols} cols × {rows} rows")

    # --- warp and stack using VRTs to avoid OOM ------------------------------
    ds_cls_list = []
    ds_conf_list = []
    nodata_vals = []
    
    for tr, country, cls_fp, conf_fp in tracks:
        # classification VRT
        vrt_cls = gdal.Warp(
            '', str(cls_fp), format='VRT',
            width=cols, height=rows,
            outputBounds=(minX, minY, maxX, maxY),
            dstSRS=proj,
            resampleAlg=gdal.GRA_NearestNeighbour
        )
        ds_cls_list.append(vrt_cls)
        nodata_vals.append(vrt_cls.GetRasterBand(1).GetNoDataValue())

        # confidence VRT
        vrt_conf = gdal.Warp(
            '', str(conf_fp), format='VRT',
            width=cols, height=rows,
            outputBounds=(minX, minY, maxX, maxY),
            dstSRS=proj,
            resampleAlg=gdal.GRA_NearestNeighbour
        )
        ds_conf_list.append(vrt_conf)

    # --- Prepare output file ------------------------------------------------
    base_tr, base_country, _, _ = tracks[0]
    if '/' in base_tr or '\\' in base_tr:
        out_dir = base_dir / base_country / 'classification_results'
    else:
        out_dir = base_dir / base_tr / 'classification_results'
    out_dir.mkdir(parents=True, exist_ok=True)
    out_tif = out_dir / f"{base_country}_final_classification{suffix}.tif"
    
    drv = gdal.GetDriverByName('GTiff')
    ds_out = drv.Create(str(out_tif), cols, rows, 1, gdal.GDT_Int32, 
                        options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
    ds_out.SetGeoTransform(gt_global)
    ds_out.SetProjection(proj)
    band_out = ds_out.GetRasterBand(1)
    band_out.SetNoDataValue(0)
    
    # --- Block by block processing ------------------------------------------
    gdal.SetCacheMax(4 * 1024 * 1024 * 1024)
    tile_size = 4096
    print("Merging tracks block-by-block to conserve memory...")
    
    for y in range(0, rows, tile_size):
        for x in range(0, cols, tile_size):
            xsize = min(tile_size, cols - x)
            ysize = min(tile_size, rows - y)
            
            c_stack = []
            cf_stack = []
            
            for i in range(len(tracks)):
                c_arr = ds_cls_list[i].GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                cf_arr = ds_conf_list[i].GetRasterBand(1).ReadAsArray(x, y, xsize, ysize).astype(np.float32)
                nod = nodata_vals[i]
                
                if nod is not None:
                    cf_arr[c_arr == nod] = np.nan
                    c_arr = np.where(c_arr == nod, 0, c_arr).astype(np.int32)
                    
                c_stack.append(c_arr)
                cf_stack.append(cf_arr)
                
            c_stack = np.stack(c_stack, axis=0)
            cf_stack = np.stack(cf_stack, axis=0)
            
            cf_stack[np.isnan(cf_stack)] = -np.inf
            idx = np.argmax(cf_stack, axis=0)
            final_block = np.take_along_axis(c_stack, idx[None,:,:], axis=0)[0]
            final_block[np.all(np.isneginf(cf_stack), axis=0)] = 0
            
            band_out.WriteArray(final_block, x, y)

    ds_out.FlushCache()
    
    # --- apply morphological sieve filter to restore objects -------------
    print("Applying Sieve Filter to remove slivers and isolated pixels...")
    # Using threshold of 50 pixels (approx field size depending on resolution)
    # 8-connectedness is generally preferred for diagonals
    gdal.SieveFilter(band_out, None, band_out, 50, 8, callback=None)
    ds_out.FlushCache()
    print(f"Merged classification saved: {out_tif}")

    # --- compute metrics & areas --------------------------------------------
    print("Calculating metrics and areas...")
    import pandas as pd
    ctrl_list = []
    for tr, country, _, _ in tracks:
        track_ctrl_shp = base_dir / tr / 'classification_results' / 'samples' / f"control{suffix}.shp"
        if not track_ctrl_shp.exists():
            track_ctrl_shp = base_dir / tr / 'classification_results' / 'samples' / 'control.shp'
        if track_ctrl_shp.exists():
            try:
                gdf = gpd.read_file(str(track_ctrl_shp))
                if not gdf.empty:
                    ctrl_list.append(gdf)
                    print(f"  Loaded {len(gdf)} control points from {tr}")
            except Exception as e:
                print(f"  [WARNING] Failed to load control points from {tr}: {e}")
                
    if ctrl_list:
        ctrl = pd.concat(ctrl_list, ignore_index=True)
        # Drop duplicates based on geometry to avoid double-counting in overlap zones
        before_len = len(ctrl)
        ctrl = ctrl.drop_duplicates(subset=['geometry'])
        print(f"Merged control points: {len(ctrl)} total (dropped {before_len - len(ctrl)} duplicates)")
    else:
        # Fallback to the country-wide main samples database (if exists) or error
        national_samples_shp = base_dir.parent / 'auxiliary_files' / 'shapefiles_samples' / base_country / 'samples.shp'
        if national_samples_shp.exists():
            print(f"No track-specific control shapefiles found. Falling back to national main samples: {national_samples_shp}")
            ctrl = gpd.read_file(str(national_samples_shp))
        else:
            raise FileNotFoundError(f"No control shapefiles found and national fallback not found at {national_samples_shp}")
    inv      = gdal.InvGeoTransform(gt_global)

    x_coords = ctrl.geometry.x.values
    y_coords = ctrl.geometry.y.values

    # Crop aggregation for NL to reduce semantic confusion and match model classes
    if base_country == 'NL':
        print("Applying crop aggregation for Netherlands validation labels...")
        ref_shp = None
        for tr, _, _, _ in tracks:
            track_ctrl_shp = base_dir / tr / 'classification_results' / 'samples' / f"control{suffix}.shp"
            if not track_ctrl_shp.exists():
                track_ctrl_shp = base_dir / tr / 'classification_results' / 'samples' / 'control.shp'
            if track_ctrl_shp.exists():
                ref_shp = track_ctrl_shp
                break
        if not ref_shp:
            ref_shp = base_dir.parent / 'auxiliary_files' / 'shapefiles_samples' / base_country / 'samples.shp'
            
        crop_aggregation = get_crop_aggregation(base_country, ref_shp)
        ctrl['crop_id'] = ctrl['crop_id'].apply(lambda val: crop_aggregation.get(val, val))

    crop_ids = ctrl['crop_id'].values

    px_vals = (inv[0] + inv[1]*x_coords + inv[2]*y_coords).astype(int)
    py_vals = (inv[3] + inv[4]*x_coords + inv[5]*y_coords).astype(int)

    valid_mask = (px_vals >= 0) & (px_vals < cols) & (py_vals >= 0) & (py_vals < rows)

    valid_px = px_vals[valid_mask]
    valid_py = py_vals[valid_mask]
    valid_crop_ids = crop_ids[valid_mask].astype(int)
    
    # Read predictions locally per point to avoid full raster memory
    valid_preds = np.zeros(len(valid_px), dtype=int)
    for i in range(len(valid_px)):
        val = band_out.ReadAsArray(int(valid_px[i]), int(valid_py[i]), 1, 1)
        if val is not None:
            valid_preds[i] = val[0, 0]
            
    final_mask = (valid_crop_ids > 0) & (valid_preds > 0)

    true_vals = valid_crop_ids[final_mask].tolist()
    pred_vals = valid_preds[final_mask].tolist()
    # Guarantee all classes from control set are represented in the matrix, even if completely missed
    all_control_classes = set(ctrl['crop_id'].unique().astype(int))
    all_control_classes.discard(0)
    labels = sorted(list(all_control_classes.union(set(pred_vals))))

    cm     = confusion_matrix(true_vals, pred_vals, labels=labels)
    prec, rec, f1, _ = precision_recall_fscore_support(
        true_vals, pred_vals,
        labels=labels,
        average=None,
        zero_division=0
    )
    total = cm.sum()
    exp   = (cm.sum(axis=0) * cm.sum(axis=1)).sum() / (total**2)
    oa    = np.trace(cm) / total
    kappa = (oa - exp) / (1 - exp) if (1 - exp) else np.nan

    # Calculate areas block-by-block
    areas_counts = {}
    for y in range(0, rows, tile_size):
        for x in range(0, cols, tile_size):
            xsize = min(tile_size, cols - x)
            ysize = min(tile_size, rows - y)
            block = band_out.ReadAsArray(x, y, xsize, ysize)
            unique, counts = np.unique(block, return_counts=True)
            for u, c in zip(unique, counts):
                if u != 0:
                    areas_counts[u] = areas_counts.get(u, 0) + c

    resx, resy = abs(gt_global[1]), abs(gt_global[5])
    area_ha    = resx * resy / 10000
    areas      = [{
        'Class':   c,
        'Area_ha': round(areas_counts.get(c, 0) * area_ha, 2)
    } for c in labels]

    # --- write Excel report -------------------------------------------------
    xlsx = out_dir / f"{base_country}_final_metrics{suffix}.xlsx"
    wb   = openpyxl.Workbook()
    sh   = wb.active
    sh.title = 'Results'

    # Confusion matrix table
    sh.cell(1,1,'Confusion Matrix').font = Font(bold=True)
    for j, lbl in enumerate(labels, start=2):
        sh.cell(2,j,lbl).font = Font(bold=True)
    for i, lbl in enumerate(labels, start=3):
        sh.cell(i,1,lbl).font = Font(bold=True)
        for j in range(len(labels)):
            sh.cell(i,j+2,int(cm[i-3,j]))

    # Overall accuracy & kappa
    r0 = 3 + len(labels)
    sh.cell(r0,1,'Overall Accuracy').font = Font(bold=True)
    sh.cell(r0,2,round(oa,2))
    sh.cell(r0+1,1,'Kappa').font          = Font(bold=True)
    sh.cell(r0+1,2,round(kappa,2))

    # Per‐class recall/precision/F1
    r1 = r0 + 3
    hdrs = ['Class','Producer Acc','User Acc','F1-score']
    for j, h in enumerate(hdrs, start=1):
        sh.cell(r1,j,h).font = Font(bold=True)
    for idx, c in enumerate(labels, start=r1+1):
        sh.cell(idx,1,c)
        sh.cell(idx,2,round(rec[idx-r1-1],2))
        sh.cell(idx,3,round(prec[idx-r1-1],2))
        sh.cell(idx,4,round(f1[idx-r1-1],2))

    # Area per class
    ra = r1 + 1 + len(labels) + 1
    sh.cell(ra,1,'Areas (ha)').font = Font(bold=True)
    sh.cell(ra+1,1,'Class').font   = Font(bold=True)
    sh.cell(ra+1,2,'Area_ha').font = Font(bold=True)
    for i, a in enumerate(areas, start=ra+2):
        sh.cell(i,1,a['Class'])
        sh.cell(i,2,a['Area_ha'])

    wb.save(str(xlsx))
    print(f"Final metrics saved: {xlsx}")

if __name__ == '__main__':
    main()