import os
import argparse
from pathlib import Path
import subprocess
import sys
import shutil
import shlex
import geopandas as gpd
import numpy as np
import pandas as pd
from osgeo import gdal, ogr, osr, gdalconst
from sklearn.metrics import confusion_matrix, precision_recall_fscore_support
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.utils import resample
import joblib
import openpyxl
from openpyxl.styles import Font
from pyogrio import read_info, read_dataframe

from concurrent.futures import ThreadPoolExecutor
import threading
import math
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader, TensorDataset

# Limit PyTorch to 1 thread to prevent thread explosion / CPU thrashing when run inside ThreadPoolExecutor
torch.set_num_threads(1)

def compute_210_features_dict(mean_matrix, std_matrix, nbands=44):
    feature_data = {}
    for b in range(nbands):
        feature_data[f'meanB{b}'] = mean_matrix[:, b]
    return feature_data

def _calculate_class_weights(y_data, all_classes):
    classes_in_data = np.unique(y_data)
    total_samples = len(y_data)
    n_classes = len(classes_in_data)
    weight_vector = np.ones(len(all_classes))
    
    for c in classes_in_data:
        count = np.sum(y_data == c)
        if count > 0:
            weight = total_samples / (n_classes * count)
            idx = np.where(all_classes == c)[0][0]
            weight_vector[idx] = math.sqrt(weight)
            
    return weight_vector

# Crop aggregation mapping for the Netherlands (NL) to reduce semantic confusion
CROP_AGGREGATION_NL = {
    2: 5,   # Clover -> Grassland
    7: 5,   # Lucerne -> Grassland
}

def get_crop_aggregation(country, learn_shp_path):
    aggregation = {}
    if country == 'NL' and learn_shp_path and os.path.exists(learn_shp_path):
        try:
            import geopandas as gpd
            gdf = gpd.read_file(str(learn_shp_path), engine="pyogrio")
            if 'crop_id' in gdf.columns and 'crop_name' in gdf.columns:
                id_to_name = dict(zip(gdf['crop_id'].astype(int), gdf['crop_name'].astype(str).str.lower()))
                
                # Find ID of grassland
                grassland_id = None
                for cid, name in id_to_name.items():
                    if "grassland" in name or "grass" in name:
                        grassland_id = cid
                        break
                
                if grassland_id is not None:
                    for cid, name in id_to_name.items():
                        if any(k in name for k in ["clover", "klaver", "lucerne", "luzerne"]):
                            aggregation[cid] = grassland_id
        except Exception as e:
            print(f"    [WARNING] Failed to dynamically construct NL crop aggregation: {e}")
            
    if not aggregation and country == 'NL':
        aggregation = CROP_AGGREGATION_NL
        
    return aggregation

def _get_priors_for_country(country, learn_shp_path, classes, class_counts, total_samples, priors_file_override=None):
    # Try custom JSON override first
    priors_json_path = None
    if priors_file_override and os.path.exists(priors_file_override):
        priors_json_path = Path(priors_file_override)
    else:
        # Resolve project root relative to learn_shp_path
        if learn_shp_path:
            p = Path(learn_shp_path).resolve()
            for parent in p.parents:
                aux_dir = parent / 'auxiliary_files'
                if aux_dir.exists():
                    aux_priors = aux_dir / 'shapefiles_samples' / country / 'priors.json'
                    if aux_priors.exists():
                        priors_json_path = aux_priors
                        break
            
            # Fallback to track_dir/priors.json
            if not priors_json_path:
                track_dir = Path(learn_shp_path).parent.parent
                track_priors = track_dir / f"priors_{country}.json"
                if not track_priors.exists():
                    track_priors = track_dir / "priors.json"
                if track_priors.exists():
                    priors_json_path = track_priors

    if priors_json_path and priors_json_path.exists():
        try:
            import json
            with open(priors_json_path, 'r') as f:
                custom_priors = json.load(f)
            print(f"    Loaded name-based priors from {priors_json_path}")
            
            # Read crop names from shapefile
            id_to_name = {}
            if learn_shp_path and os.path.exists(learn_shp_path):
                try:
                    import geopandas as gpd
                    gdf = gpd.read_file(str(learn_shp_path), engine="pyogrio")
                    if 'crop_id' in gdf.columns and 'crop_name' in gdf.columns:
                        id_to_name = dict(zip(gdf['crop_id'].astype(int), gdf['crop_name'].astype(str).str.lower()))
                except Exception as e:
                    print(f"    [WARNING] Could not read crop names from shapefile: {e}")
            
            # Build priors map dynamically
            raw_priors = {}
            sorted_keys = sorted(custom_priors.keys(), key=len, reverse=True)
            for cid, name in id_to_name.items():
                matched_val = 1e-5
                for key in sorted_keys:
                    if key.lower() in name or name in key.lower():
                        matched_val = float(custom_priors[key])
                        break
                raw_priors[cid] = matched_val
                
            for key, val in custom_priors.items():
                try:
                    cid = int(key)
                    raw_priors[cid] = float(val)
                except ValueError:
                    pass
            
            # Apply aggregation
            crop_aggregation = get_crop_aggregation(country, learn_shp_path)
            aggregated_priors = {}
            for cid, val in raw_priors.items():
                mapped_cid = crop_aggregation.get(cid, cid)
                aggregated_priors[mapped_cid] = aggregated_priors.get(mapped_cid, 0.0) + val
                
            p_true = np.array([aggregated_priors.get(c, 1e-5) for c in classes])
            p_true = p_true / np.sum(p_true)
            return p_true
        except Exception as e:
            print(f"    [WARNING] Failed to load name-based priors: {e}")

    # NL specific priors fallback
    if country == 'NL':
        real_priors_nl = {
            1: 0.0030, 2: 0.0015, 3: 0.0775, 4: 0.0057, 5: 0.7214,
            6: 0.0033, 7: 0.0056, 8: 0.0847, 9: 0.0060, 10: 0.0007,
            11: 0.0073, 12: 0.0087, 13: 0.0255, 14: 0.0023, 15: 0.0149,
            16: 0.0058, 17: 0.0044, 18: 0.0006, 19: 0.0034, 20: 0.0178
        }
        crop_aggregation = get_crop_aggregation(country, learn_shp_path)
        aggregated_priors = {}
        for cid, val in real_priors_nl.items():
            mapped_cid = crop_aggregation.get(cid, cid)
            aggregated_priors[mapped_cid] = aggregated_priors.get(mapped_cid, 0.0) + val
        p_true = np.array([aggregated_priors.get(c, 1e-5) for c in classes])
        p_true = p_true / np.sum(p_true)
        return p_true


    # PL & any other country - dynamic estimation using keyword-based field sizes
    id_to_name = {}
    if learn_shp_path and os.path.exists(learn_shp_path):
        try:
            import geopandas as gpd
            gdf = gpd.read_file(str(learn_shp_path), engine="pyogrio")
            if 'crop_id' in gdf.columns and 'crop_name' in gdf.columns:
                id_to_name = dict(zip(gdf['crop_id'].astype(int), gdf['crop_name'].astype(str)))
        except Exception as e:
            print(f"    [WARNING] Could not read crop names from shapefile: {e}")

    # Area threshold keyword matcher
    def get_area_multiplier(cid):
        name = id_to_name.get(cid, '').lower()
        if not name:
            # Fallback to standard 37 classes PL hardcoded values if crop_id maps to them (1 to 37)
            area_thresholds_pl = {
                1: 3000, 2: 5000, 3: 5000, 4: 40000, 5: 5000, 6: 10000, 7: 5000, 8: 30000, 9: 15000,
                10: 20000, 11: 30000, 12: 40000, 13: 5000, 14: 70000, 15: 2000, 16: 30000, 17: 5000,
                18: 3000, 19: 40000, 20: 2000, 21: 40000, 22: 2000, 23: 10000, 24: 25000, 25: 70000,
                26: 10000, 27: 50000, 28: 2000, 29: 60000, 30: 3000, 31: 15000, 32: 70000, 33: 5000,
                34: 3000, 35: 5000, 36: 20000, 37: 40000
            }
            return area_thresholds_pl.get(cid, 10000)
            
        # Large field crops (~40k - 100k m2)
        if any(k in name for k in ['grass', 'tiuz', 'grassland', 'pasture', 'trawa', 'trawiast', 'blijvend', 'tijdelijk', 'permanent', 'clover', 'klaver', 'lucerne', 'luzerne']):
            return 70000
        if any(k in name for k in ['maize', 'mais', 'kukurydza', 'corn']):
            return 70000
        if any(k in name for k in ['wheat', 'pszenica', 'tarwe']):
            return 70000
        if any(k in name for k in ['barley', 'jeczmien', 'gerst']):
            return 40000
        if any(k in name for k in ['rye', 'zyto', 'rogge']):
            return 40000
        if any(k in name for k in ['triticale', 'pszenzyto', 'koorn']):
            return 50000
        if any(k in name for k in ['oats', 'owies', 'haver']):
            return 40000
        if any(k in name for k in ['rapeseed', 'rzepak', 'koolzaad']):
            return 60000
        if any(k in name for k in ['sugar beet', 'burak', 'suikerbiet']):
            return 40000
        if any(k in name for k in ['fallow', 'braak', 'ugor']):
            return 40000
            
        # Medium/small (~10k - 30k m2)
        if any(k in name for k in ['potato', 'ziemniak', 'aardappel']):
            return 20000
        if any(k in name for k in ['orchard', 'fruit', 'sad', 'appel', 'peer', 'jablon', 'sliwa', 'wisnia']):
            return 20000
        if any(k in name for k in ['pea', 'groch', 'erwt']):
            return 30000
        if any(k in name for k in ['bean', 'fasola', 'boon']):
            return 10000
        if any(k in name for k in ['aronia', 'blueberry', 'borowka', 'currant', 'porzeczka', 'bessen']):
            return 10000
        if any(k in name for k in ['nursery', 'ornamental', 'szkolka', 'sier', 'boomkwekerij']):
            return 10000
            
        # Small / vegetable crops (<10k m2)
        if any(k in name for k in ['onion', 'cebula', 'ui']):
            return 5000
        if any(k in name for k in ['strawberry', 'truskawka', 'aardbei']):
            return 5000
        if any(k in name for k in ['cabbage', 'brassica', 'kapusta', 'kool']):
            return 5000
        if any(k in name for k in ['carrot', 'marchew', 'peen']):
            return 3000
        if any(k in name for k in ['tomato', 'pomidor', 'tomaat']):
            return 2000
        if any(k in name for k in ['cucumber', 'ogorek', 'komkommer']):
            return 2000
        if any(k in name for k in ['tobacco', 'tyton', 'tabak']):
            return 3000
        if any(k in name for k in ['raspberry', 'blackberry', 'malina', 'jezyna', 'framboos']):
            return 5000
        return 10000

    area_multipliers = np.array([get_area_multiplier(c) for c in classes])
    counts_arr = np.array([class_counts.get(c, 0) for c in classes])
    true_area_dist = counts_arr * area_multipliers
    p_true = true_area_dist / (np.sum(true_area_dist) + 1e-9)
    return p_true

class TorchMLPClassifier:
    def __init__(self, hidden_layer_sizes=(256, 128, 64), max_iter=120, batch_size=256, lr=0.001, class_weights=None, all_classes=None):
        self.hidden_layer_sizes = hidden_layer_sizes
        self.max_iter = max_iter
        self.batch_size = batch_size
        self.lr = lr
        self.class_weights = class_weights
        self.all_classes = all_classes
        self.device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
        self.model = None
        self.le = None
        self.classes_ = None
        self.input_dim = None
        self.output_dim = None

    def fit(self, X, y):
        self.le = LabelEncoder()
        if self.all_classes is not None:
            self.le.fit(self.all_classes)
        else:
            self.le.fit(y)
        
        y_enc = self.le.transform(y)
        self.classes_ = self.le.classes_
        
        self.input_dim = X.shape[1]
        self.output_dim = len(self.classes_)

        layers = []
        in_dim = self.input_dim
        for h_dim in self.hidden_layer_sizes:
            layers.append(nn.Linear(in_dim, h_dim))
            layers.append(nn.ReLU())
            layers.append(nn.BatchNorm1d(h_dim))
            layers.append(nn.Dropout(0.3))
            in_dim = h_dim
        layers.append(nn.Linear(in_dim, self.output_dim))

        self.model = nn.Sequential(*layers).to(self.device)
        
        if self.class_weights is not None:
            weights_tensor = torch.tensor(self.class_weights, dtype=torch.float32).to(self.device)
            criterion = nn.CrossEntropyLoss(weight=weights_tensor)
        else:
            criterion = nn.CrossEntropyLoss()
            
        optimizer = optim.Adam(self.model.parameters(), lr=self.lr, weight_decay=1e-4)

        X_tensor = torch.tensor(X, dtype=torch.float32)
        y_tensor = torch.tensor(y_enc, dtype=torch.long)

        dataset = TensorDataset(X_tensor, y_tensor)
        loader = DataLoader(dataset, batch_size=self.batch_size, shuffle=True)

        self.model.train()
        for epoch in range(self.max_iter):
            for batch_X, batch_y in loader:
                batch_X = batch_X.to(self.device)
                batch_y = batch_y.to(self.device)

                optimizer.zero_grad()
                outputs = self.model(batch_X)
                loss = criterion(outputs, batch_y)
                loss.backward()
                optimizer.step()

        return self

    def predict(self, X):
        self.model.eval()
        X_tensor = torch.tensor(X, dtype=torch.float32)
        dataset = TensorDataset(X_tensor)
        loader = DataLoader(dataset, batch_size=1024, shuffle=False)
        predictions = []

        with torch.no_grad():
            for batch_X in loader:
                batch_X = batch_X[0].to(self.device)
                outputs = self.model(batch_X)
                _, predicted = torch.max(outputs.data, 1)
                predictions.append(predicted.cpu().numpy())

        return self.le.inverse_transform(np.concatenate(predictions))

    def predict_proba(self, X):
        self.model.eval()
        X_tensor = torch.tensor(X, dtype=torch.float32)
        dataset = TensorDataset(X_tensor)
        loader = DataLoader(dataset, batch_size=1024, shuffle=False)
        probabilities = []

        with torch.no_grad():
            for batch_X in loader:
                batch_X = batch_X[0].to(self.device)
                outputs = self.model(batch_X)
                probs = torch.softmax(outputs, dim=1)
                probabilities.append(probs.cpu().numpy())

        return np.concatenate(probabilities, axis=0)

    def __getstate__(self):
        state = self.__dict__.copy()
        if self.model is not None:
            self.model.cpu()
            state['model_state_dict'] = self.model.state_dict()
            state['model'] = None
        return state

    def __setstate__(self, state):
        self.__dict__.update(state)
        if 'model_state_dict' in state and state['model_state_dict'] is not None:
            layers = []
            in_dim = self.input_dim
            for h_dim in self.hidden_layer_sizes:
                layers.append(nn.Linear(in_dim, h_dim))
                layers.append(nn.ReLU())
                layers.append(nn.BatchNorm1d(h_dim))
                layers.append(nn.Dropout(0.3))
                in_dim = h_dim
            layers.append(nn.Linear(in_dim, self.output_dim))
            self.model = nn.Sequential(*layers)
            self.model.load_state_dict(state['model_state_dict'])
            self.model.to(self.device)

# Try importing scikit-image
try:
    from skimage.segmentation import felzenszwalb, slic, watershed
    from skimage.filters import sobel
    from skimage.graph import rag_mean_color, merge_hierarchical
    from skimage.util import img_as_float
    from skimage.measure import regionprops_table
    import scipy.ndimage as ndi

    HAS_SKIMAGE = True
except ImportError:
    HAS_SKIMAGE = False
    print("WARNING: scikit-image not found. This script requires scikit-image for raster segmentation.")

# Try importing SAM
try:
    import torch
    from segment_anything import sam_model_registry, SamAutomaticMaskGenerator
    HAS_SAM = True
except ImportError:
    HAS_SAM = False
    print("WARNING: segment-anything or torch not found.")

# --- Configuration (Global) ---

# ==============================================================================
# HOW TO RUN THE SCRIPT (INSTRUCTIONS):
#
# Available CLI arguments:
#   --track          : Name of the processing track (e.g. NL/orbit_88, PL/orbit_12)
#   --seg_mode       : Initial segmentation mode. Choose:
#                      * 'sam'  (default, uses Meta AI SAM or other dynamically computed segmentation)
#                      * 'lpis' (rasterization of actual cadastral parcel boundaries from vector GPKG/SHP databases)
#                      Note: In the interactive menu, you can dynamically switch between any of the 5 supported
#                            segmentation methods. The output filenames, models, and metric spreadsheets
#                            will automatically update to include a suffix matching the selected method
#                            (e.g., '_sam', '_lpis', '_otb_meanshift',
#                            '_felzenszwalb', '_slic'). This prevents file name collisions.
#   --mask_variant   : Agricultural crop mask variant. Choose:
#                      * 'allcrops' (full cropland mask including all 18+ classes, recommended for NL)
#                      * '3class'   (aggregated/masked for the 3 main PL crop types: spring crops, winter crops, rapeseed)
#
# LAUNCH EXAMPLES:
#
# 1. Netherlands (NL) - Segmentation using actual LPIS boundaries (BRP) and allcrops mask:
#    python 1_classify_ann.py --track NL/orbit_88 --seg_mode lpis --mask_variant allcrops
#
# 2. Netherlands (NL) - Segmentation using Meta SAM model and allcrops mask:
#    python 1_classify_ann.py --track NL/orbit_88 --seg_mode sam --mask_variant allcrops
#
# 3. Poland (PL) - Segmentation using Meta SAM model and 3-class crop mask:
#    python 1_classify_ann.py --track PL/orbit_12 --seg_mode sam --mask_variant 3class
#
# 4. Poland (PL) - Segmentation using LPIS boundaries (ARiMR) and 3-class crop mask:
#    python 1_classify_ann.py --track PL/orbit_12 --seg_mode lpis --mask_variant 3class
#
# 5. Netherlands (NL) - Segmentation using SLIC algorithm and allcrops mask:
#    python 1_classify_ann.py --track NL/orbit_88 --seg_mode slic --mask_variant allcrops
#
# 6. Netherlands (NL) - Segmentation using Felzenszwalb algorithm and allcrops mask:
#    python 1_classify_ann.py --track NL/orbit_88 --seg_mode felzenszwalb --mask_variant allcrops
#
# 7. Netherlands (NL) - Segmentation using OTB Mean-Shift and allcrops mask:
#    python 1_classify_ann.py --track NL/orbit_88 --seg_mode otb_meanshift --mask_variant allcrops
#
# RUNNING THE ENTIRE PIPELINE (ALL STAGES):
# To process the entire pipeline (Stages 0-8) sequentially for a specific segmentation method:
#   1. Run the command with the desired --seg_mode (e.g., --seg_mode slic)
#   2. In the interactive terminal menu, choose option '[A] Run All Stages'
#   This will execute the entire footprint, segmentation, training, inference, and accuracy calculation
#   automatically for the chosen segmentation method.
# ==============================================================================
# Base Paths provided by user
base_dir = Path("D:/AIML_CropMapper_Cloud/workingDir")
aux_dir = Path("D:/AIML_CropMapper_Cloud/auxiliary_files")

# OTB Installation Path (Still used for some auxiliary tasks if needed, but main flow is Python)
otb_dir = Path("D:/AIML_CropMapper_Cloud/bin/OTB-6.2.0-Win64")

TOTAL_STAGES = 8


# --- Main Pipeline Class ---

class ProcessingPipeline:
    def __init__(self, track, mask_variant='3class', seg_mode='sam'):
        self.track = track
        self.mask_variant = mask_variant  # '3class' or 'allcrops'
        self.seg_mode = seg_mode          # 'sam' or 'lpis'
        
        # Dedicated support for dynamic tracks per country (e.g. PL/orbit_12 or 2-letter country code)
        if '/' in track or '\\' in track:
            normalized_track = track.replace('\\', '/')
            self.country = normalized_track.split('/')[0].upper()
        elif len(track) == 2:
            self.country = track.upper()
        else:
            print(f"Error: Track '{track}' does not contain country code and is not a 2-letter country code.")
            sys.exit(1)

        self.total_stages = TOTAL_STAGES
        print(f"Initializing pipeline for Track: {self.track}, Country: {self.country}, Segmentation: {self.seg_mode.upper()}")

        # Sanitized track name for filenames (no slashes)
        self.sanitized_track = self.track.replace('/', '_').replace('\\', '_')
        if self.sanitized_track.upper().startswith(self.country.upper() + "_"):
            self.file_prefix = self.sanitized_track
        else:
            self.file_prefix = f"{self.country}_{self.sanitized_track}"

        # --- 1. Define all paths ---
        self.base_dir = base_dir
        self.aux_dir = aux_dir
        self.proc_dir = self.base_dir / self.track / 'processed_raster'
        self.out_dir = self.base_dir / self.track / 'classification_results'
        self.samples_dir = self.out_dir / 'samples'
        self.model_dir = self.out_dir / 'train_model'
        self.seg_dir = self.out_dir / 'segmentation'
        self.class_dir = self.out_dir / 'classification'

        self._ensure_directories()

        # --- 2. Resolve input raster ---
        search_patterns = [
            f"{self.sanitized_track}_*_VH_VV*.tif",
            f"*_{self.sanitized_track}_*_VH_VV*.tif",
            f"*{self.sanitized_track}*.tif",
            f"{self.track}_*_VH_VV*.tif",
            f"*_{self.track}_*_VH_VV*.tif",
            f"*{self.track}*.tif",
            f"{self.sanitized_track}_*_VH_VV*.hdr",
            f"*{self.sanitized_track}*.hdr",
        ]

        self.hdr = None
        if self.proc_dir.exists():
            for pattern in search_patterns:
                self.hdr = next(self.proc_dir.glob(pattern), None)
                if self.hdr:
                    break

            if not self.hdr:
                raise FileNotFoundError(f"No raster file (TIF/HDR) found for track {self.track} in {self.proc_dir}")

            self.ras = self._resolve_raster(self.hdr)
            print(f"Input raster found: {self.ras}")
        else:
            raise FileNotFoundError(f"Processing directory does not exist: {self.proc_dir}")

        # Samples
        samples_base = self.aux_dir / 'shapefiles_samples'
        candidate_paths = [
            samples_base / self.file_prefix / "samples.shp",
            samples_base / f"{self.country}_{self.sanitized_track}" / "samples.shp",
            samples_base / f"{self.country}_{self.track}" / "samples.shp",
            samples_base / self.sanitized_track / "samples.shp",
            samples_base / self.track / "samples.shp",
            samples_base / self.country / "samples.shp",
            samples_base / "samples.shp"
        ]

        self.sample_shp = None
        for p in candidate_paths:
            if p.exists():
                self.sample_shp = p
                print(f"Training samples found at: {self.sample_shp}")
                break

        if not self.sample_shp:
            print(f"\nCRITICAL WARNING: Could not find 'samples.shp' inside {samples_base}")
            self.sample_shp = samples_base / self.file_prefix / "samples.shp"

        # --- 3. Initialize seg_mode and output file paths ---
        method_mapping_inv = {
            'sam': 'python_sam',
            'lpis': 'lpis',
            'otb_meanshift': 'otb_meanshift_summed',
            'felzenszwalb': 'python_felzenszwalb',
            'slic': 'python_slic'
        }
        initial_method = method_mapping_inv.get(seg_mode, 'python_sam')

        # Initialize parameters dictionary first so update_paths can populate method-specific defaults
        self.stage1_params = {
            'method': initial_method
        }
        self.update_paths(initial_method)

        # Legacy compatibility migration for SAM segmentation files
        if self.seg_mode == 'sam':
            legacy_tif = self.seg_dir / f"{self.file_prefix}_segmentation.tif"
            legacy_sqlite = self.seg_dir / f"{self.file_prefix}_segmentation.sqlite"
            if legacy_tif.exists() and not self.seg_tif.exists():
                print(f"    [INFO] Migrating legacy SAM segmentation raster to new naming format...")
                try:
                    os.rename(str(legacy_tif), str(self.seg_tif))
                    if legacy_sqlite.exists() and not self.seg_shp.exists():
                        os.rename(str(legacy_sqlite), str(self.seg_shp))
                except Exception as e:
                    print(f"    [WARNING] Failed to migrate legacy segmentation files: {e}")

        self.footprint_mask = self.seg_dir / f"{self.file_prefix}_data_footprint.tif"
        
        # Enable footprint masking for all tracks. The footprint mask generation has been
        # updated to be robust (checking all bands and removing the noise sieve filter),
        # which automatically makes it cover 100% of land area for full-swath orbits (like Orbit 88)
        # without dropping points, while correctly masking out the empty 75% for narrow-swath orbits (like Orbit 161).
        self.use_footprint_mask = True
        
        # Agricultural mask - resolved per country (set in _resolve_agri_mask)
        self.agri_mask = self._resolve_agri_mask()

        # --- 4. Parameters ---
        self.stage2_params = {
            'learn_frac': 0.7, 'random_state': 42
        }
        self.stage4_params = {
            'classifier': 'ann_sklearn',
            'sk_hidden_sizes': '256,128,64',
            'sk_activation': 'relu',
            'sk_solver': 'adam',
            'sk_alpha': 0.0001,
            'sk_max_iter': 120,
            'balance_threshold': 1000
        }

        self.feat_cols = []

    # --- Utility Methods ---

    def update_paths(self, method_name):
        """
        Dynamically updates the output file paths based on the selected segmentation method
        (e.g., 'sam', 'lpis', 'otb_meanshift', 'felzenszwalb', 'slic').
        """
        suffix_mapping = {
            'python_sam': 'sam',
            'lpis': 'lpis',
            'otb_meanshift_summed': 'otb_meanshift',
            'python_felzenszwalb': 'felzenszwalb',
            'python_slic': 'slic'
        }
        
        self.seg_suffix = suffix_mapping.get(method_name, 'sam')
        
        # 'lpis' mode does vector parcel boundary rasterization, others do image-based segmentation
        if method_name == 'lpis':
            self.seg_mode = 'lpis'
        else:
            self.seg_mode = 'sam'
        
        self.seg_tif = self.seg_dir / f"{self.file_prefix}_segmentation_{self.seg_suffix}.tif"
        self.seg_shp = self.seg_dir / f"{self.file_prefix}_segmentation_{self.seg_suffix}.sqlite"
        self.learn_shp = self.samples_dir / f"learn_{self.seg_suffix}.shp"
        self.control_shp = self.samples_dir / f"control_{self.seg_suffix}.shp"
        self.sel_csv = self.samples_dir / f"{self.file_prefix}_learn_features_{self.seg_suffix}.csv"
        self.class_tif = self.class_dir / f"{self.file_prefix}_classified_{self.seg_suffix}.tif"
        self.conf_tif = self.class_dir / f"{self.file_prefix}_confidence_map_{self.seg_suffix}.tif"
        self.masked_class = self.class_dir / f"{self.file_prefix}_classified_masked_{self.seg_suffix}.tif"
        self.masked_conf = self.class_dir / f"{self.file_prefix}_confidence_masked_{self.seg_suffix}.tif"
        self.metrics_fp = self.class_dir / f"{self.file_prefix}_metrics_{self.seg_suffix}.xlsx"
        self.model_pkl = self.model_dir / f"{self.file_prefix}_model_{self.seg_suffix}.pkl"
        
        print(f"    [INFO] Switched segmentation mode to: {self.seg_mode.upper()} (Suffix: {self.seg_suffix})")
        print(f"    Updated output paths accordingly (e.g. classification output: {self.class_tif.name})")

        # Keep stage1_params['method'] consistent with mode if it has been initialized
        if hasattr(self, 'stage1_params'):
            self.stage1_params['method'] = method_name
            # Initialize default values for the selected method in stage1_params if they don't exist
            if method_name == 'python_sam':
                self.stage1_params.setdefault('tile_size', 2048)
                self.stage1_params.setdefault('buffer', 128)
                self.stage1_params.setdefault('sam_checkpoint', str(self.aux_dir / 'SAM_models' / 'sam_vit_h_4b8939.pth'))
                self.stage1_params.setdefault('sam_model_type', 'vit_h')
                self.stage1_params.setdefault('sam_device', 'cuda' if (HAS_SAM and torch.cuda.is_available()) else 'cpu')
            elif method_name in ['otb_meanshift_summed', 'otb_meanshift']:
                self.stage1_params.setdefault('spatialr', 4)
                self.stage1_params.setdefault('ranger', 0.3)
                self.stage1_params.setdefault('minsize', 20)
                self.stage1_params.setdefault('tilesizex', 4096)
                self.stage1_params.setdefault('tilesizey', 4096)
                self.stage1_params.setdefault('ram', 4096)
            elif method_name == 'python_felzenszwalb':
                self.stage1_params.setdefault('tile_size', 4096)
                self.stage1_params.setdefault('buffer', 256)
                self.stage1_params.setdefault('scale', 50.0)
                self.stage1_params.setdefault('sigma', 0.8)
                self.stage1_params.setdefault('min_size', 15)
            elif method_name == 'python_slic':
                self.stage1_params.setdefault('tile_size', 4096)
                self.stage1_params.setdefault('buffer', 256)
                self.stage1_params.setdefault('n_segments', 35000)  # Tuned 1: 35000, Original: 20000
                self.stage1_params.setdefault('compactness', 0.03)  # Tuned 1: 0.03, Original: 0.1
                self.stage1_params.setdefault('slic_sigma', 1.0)    # Tuned 1: 1.0

    def _resolve_agri_mask(self) -> Path:
        """
        Smart-selects the binary agricultural mask per country.
        Variant selected by self.mask_variant ('3class' or 'allcrops').

        Search order (for mask_variant='3class'):
          1. AgriMasks/<COUNTRY>/<COUNTRY>_agri_mask_3class_epsg3857.tif
          2. AgriMasks/<COUNTRY>/<COUNTRY>_agri_mask_allcrops_epsg3857.tif
          3. EU_arable_areas_mask_3857.tif (fallback)

        To generate the mask, run: python build_agri_mask.py --country {self.country}
        """
        raster_dir  = self.aux_dir / 'raster_files'
        country_dir = raster_dir / 'AgriMasks' / self.country

        mask_3class   = country_dir / f"{self.country}_agri_mask_3class_epsg3857.tif"
        mask_allcrops = country_dir / f"{self.country}_agri_mask_allcrops_epsg3857.tif"
        mask_eu       = raster_dir / 'EU_arable_areas_mask_3857.tif'

        # Candidates order depends on the selected variant
        if self.mask_variant == 'allcrops':
            candidates = [mask_allcrops, mask_3class, mask_eu]
            print(f"  Mask variant: allcrops (all crops including permanent ones)")
        else:
            candidates = [mask_3class, mask_allcrops, mask_eu]
            print(f"  Mask variant: 3class (spring/winter/rapeseed)")

        for p in candidates:
            if p.exists():
                print(f"Agricultural mask selected: {p}")
                return p

        print(f"[WARNING] No agricultural mask found for country '{self.country}'.")
        print(f"  Generate mask with: python build_agri_mask.py --country {self.country}")
        return mask_eu


    def _ensure_directories(self):
        for d in [self.samples_dir, self.model_dir, self.seg_dir, self.class_dir]:
            d.mkdir(parents=True, exist_ok=True)

    def _run_cmd(self, cmd, stage, desc, ram=None):
        print(f"[Stage {stage}/{self.total_stages}] {desc}")
        env = os.environ.copy()
        otb_bin = str(otb_dir / "bin")
        otb_lib = str(otb_dir / "lib")
        otb_apps = str(otb_dir / "lib" / "otb" / "applications")
        env["PATH"] = f"{otb_bin};{otb_lib};{env['PATH']}"
        env["OTB_APPLICATION_PATH"] = otb_apps
        env["GEOTIFF_CSV"] = str(otb_dir / "share" / "epsg_csv")
        env["GDAL_DATA"] = str(otb_dir / "share" / "gdal")

        if ram:
            env["OTB_MAX_RAM_HINT"] = str(ram)

        if isinstance(cmd, str):
            cmd = shlex.split(cmd, posix=os.name != 'nt')

        # Resolve full path to executable, crucial for Windows .bat execution without shell=True
        executable = shutil.which(cmd[0], path=env.get("PATH"))
        if executable:
            cmd[0] = executable

        proc = subprocess.Popen(cmd, shell=False, stdout=sys.stdout, stderr=sys.stderr, env=env)
        proc.communicate()
        if proc.returncode != 0:
            raise RuntimeError(f"Stage {stage} failed with return code {proc.returncode}: {cmd}")
        print(f"Completed stage {stage}/{self.total_stages}\n")

    def _resolve_raster(self, hdr):
        if hdr.suffix.lower() in ['.tif', '.tiff']: return hdr
        for ext in ['.img', '.tif', '.TIF']:
            p = hdr.with_suffix(ext)
            if p.exists(): return p
        p_no_ext = hdr.with_suffix('')
        if p_no_ext.exists() and p_no_ext.is_file(): return p_no_ext
        raise FileNotFoundError(f"No raster image (.img/.tif) found matching header {hdr.stem}")

    def _apply_mask(self, input_tif, mask_tif, out_tif, stage):
        print(f"[Stage {stage}/{self.total_stages}] Applying Arable & Data Footprint Mask...")

        # Open the original radar stack to check for data footprint (blank areas)
        ds_stack = gdal.Open(str(self.ras))
        if not ds_stack: raise RuntimeError(f"Could not open source raster {self.ras} for footprint masking.")
        stack_band = ds_stack.GetRasterBand(1)

        # 0. Open Data Footprint Mask (Stage 0 output)
        if self.use_footprint_mask and self.footprint_mask.exists():
            ds_foot = gdal.Open(str(self.footprint_mask))
            foot_band = ds_foot.GetRasterBand(1)
        else:
            ds_foot = None
            foot_band = None

        if not mask_tif.exists():
            print(f"    WARNING: Arable mask not found at {mask_tif}. Will only apply data footprint mask.")
            has_arable_mask = False
        else:
            has_arable_mask = True

        ds_in = gdal.Open(str(input_tif))
        gt = ds_in.GetGeoTransform()
        proj = ds_in.GetProjection()
        cols = ds_in.RasterXSize
        rows = ds_in.RasterYSize

        minx = gt[0]
        maxy = gt[3]
        maxx = minx + gt[1] * cols
        miny = maxy + gt[5] * rows

        if has_arable_mask:
            # Warp the mask to match the input raster exactly
            temp_mask_vrt = str(out_tif).replace('.tif', '_mask_temp.vrt')
            mask_opts = gdal.WarpOptions(
                format='VRT',
                outputBounds=(minx, miny, maxx, maxy),
                width=cols,
                height=rows,
                dstSRS=proj,
                resampleAlg=gdal.GRA_NearestNeighbour
            )
            ds_mask = gdal.Warp(temp_mask_vrt, str(mask_tif), options=mask_opts)
            if not ds_mask: raise RuntimeError("Failed to warp the arable mask.")
            m_band = ds_mask.GetRasterBand(1)
        else:
            ds_mask = None

        in_band = ds_in.GetRasterBand(1)
        out_type = in_band.DataType

        driver = gdal.GetDriverByName('GTiff')
        out_ds = driver.Create(str(out_tif), cols, rows, 1, out_type,
                               options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        out_ds.SetGeoTransform(gt)
        out_ds.SetProjection(proj)
        out_band = out_ds.GetRasterBand(1)

        nodata = in_band.GetNoDataValue()
        if nodata is None: nodata = 0
        out_band.SetNoDataValue(nodata)

        tile_size = 4096

        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                # 1. Read the Classification/Confidence Data
                arr = in_band.ReadAsArray(x, y, xsize, ysize)

                # 2. Apply Footprint Mask
                if foot_band:
                    f_arr = foot_band.ReadAsArray(x, y, xsize, ysize)
                    arr[f_arr == 0] = nodata
                elif stack_band:
                    try:
                        stack_arr = stack_band.ReadAsArray(x, y, xsize, ysize)
                        # Fallback to simple zero-check if Stage 0 mask is missing
                        if stack_arr is not None:
                            arr[stack_arr == 0] = nodata
                    except Exception as e:
                        print(f"Warning: Failed to read stack band for fallback masking: {e}")

                # 3. Read and Apply the Arable Mask (if it exists)
                if ds_mask:
                    m_arr = m_band.ReadAsArray(x, y, xsize, ysize)
                    arr[m_arr < 0.5] = nodata

                out_band.WriteArray(arr, x, y)

        out_ds.FlushCache()

        # Build pyramids (overviews) for faster loading in GIS software
        # Using NEAREST for classification (integers) and AVERAGE for confidence (floats)
        resampling = "NEAREST" if out_type in [gdal.GDT_Byte, gdal.GDT_Int16, gdal.GDT_Int32] else "AVERAGE"
        print(f"    [INFO] Building pyramids for masked raster with resampling: {resampling}...")
        out_ds.BuildOverviews(resampling=resampling, overviewlist=[2, 4, 8, 16, 32])

        ds_mask = None
        out_ds = None
        ds_in = None
        ds_stack = None
        if has_arable_mask and os.path.exists(temp_mask_vrt): os.remove(temp_mask_vrt)

        print(f"Completed stage {stage}\n")

    def stage_0_generate_footprint(self, force_recompute=False):
        """Generates a precise data footprint mask from the input raster."""
        self._ensure_directories()
        if not self.use_footprint_mask:
            print("[Stage 0] Footprint mask is bypassed for this track, skipping.")
            return
        if self.footprint_mask.exists() and not force_recompute:
            print("[Stage 0] Data footprint mask already exists, skipping.")
            return

        print(f"[Stage 0/{self.total_stages}] Generating robust data footprint mask from radar stack...")
        ds = gdal.Open(str(self.ras))
        cols, rows = ds.RasterXSize, ds.RasterYSize
        gt, proj = ds.GetGeoTransform(), ds.GetProjection()

        driver = gdal.GetDriverByName('GTiff')
        out_ds = driver.Create(str(self.footprint_mask), cols, rows, 1, gdal.GDT_Byte,
                               options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        out_ds.SetGeoTransform(gt)
        out_ds.SetProjection(proj)
        out_band = out_ds.GetRasterBand(1)

        tile_size = 4096
        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                # Check a subset of bands (every 8th band, representing different acquisition dates)
                # to see if the pixel has valid data. This is 8x faster and highly robust.
                # Loop through bands one by one to keep memory usage extremely low (~67 MB per tile)
                combined_mask = np.zeros((ysize, xsize), dtype=bool)
                for b in range(1, ds.RasterCount + 1, 8):
                    data = ds.GetRasterBand(b).ReadAsArray(x, y, xsize, ysize)
                    if data is not None:
                        combined_mask |= (np.abs(data) > 1e-7) & (~np.isnan(data))

                out_band.WriteArray(combined_mask.astype(np.uint8), x, y)

        out_ds.FlushCache()
        # Sieve filter is removed to prevent dropping valid small agricultural parcels
        out_ds = None
        ds = None
        print(f"    Footprint mask saved to {self.footprint_mask}\n")


    def _create_summed_composite(self):
        """Creates a single-band composite by summing the log-domain (dB) values of all SAR bands to reduce speckle while preserving low-backscatter crop contrast."""
        print("    [INFO] Creating a log-domain (dB) summed composite of all SAR bands...")

        gdal.SetCacheMax(4 * 1024 * 1024 * 1024)  # Increase GDAL cache to 4GB to prevent allocation errors

        composite_tif = self.seg_dir / f"{self.file_prefix}_summed_composite.tif"

        if composite_tif.exists():
            print(f"    [INFO] Summed composite already exists at {composite_tif}.")
            return composite_tif

        ds = gdal.Open(str(self.ras))
        if not ds:
            raise RuntimeError(f"Could not open source raster {self.ras}")

        cols = ds.RasterXSize
        rows = ds.RasterYSize
        nbands = ds.RasterCount
        gt = ds.GetGeoTransform()
        proj = ds.GetProjection()

        driver = gdal.GetDriverByName('GTiff')
        out_ds = driver.Create(str(composite_tif), cols, rows, 1, gdal.GDT_Float32,
                               options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        out_ds.SetGeoTransform(gt)
        out_ds.SetProjection(proj)
        out_band = out_ds.GetRasterBand(1)
        out_band.SetNoDataValue(0)

        tile_size = 4096
        
        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                sum_arr = np.zeros((ysize, xsize), dtype=np.float32)
                valid_mask = np.zeros((ysize, xsize), dtype=bool)

                for b in range(1, nbands + 1):
                    band = ds.GetRasterBand(b)
                    arr = band.ReadAsArray(x, y, xsize, ysize)
                    
                    if arr is None:
                        raise RuntimeError(f"Failed to read block at x={x}, y={y} for band {b}. Possible memory exhaustion.")
                        
                    nodata = band.GetNoDataValue()
                    
                    if nodata is not None:
                        mask = (arr != nodata) & (~np.isnan(arr)) & (arr != 0)
                    else:
                        mask = (~np.isnan(arr)) & (arr != 0)
                        
                    sum_arr[mask] += arr[mask]
                    valid_mask |= mask

                # Set areas where all bands were nodata to 0
                sum_arr[~valid_mask] = 0

                out_band.WriteArray(sum_arr, x, y)

        out_ds.FlushCache()
        out_ds = None
        ds = None
        print(f"    [INFO] Summed composite saved to {composite_tif}")

        return composite_tif

    # --- Stage 1: OTB Segmentation (Direct Raster Pipeline) ---
    def stage_1_segmentation(self, **kwargs):
        self._ensure_directories()
        params = self.stage1_params.copy()
        params.update(kwargs)
        stage = 1

        if self.seg_tif.exists():
            print(f"[Stage {stage}/{self.total_stages}] Segmentation Raster exists, skipping\n")
            return

        if self.seg_mode == 'lpis':
            print(f"[Stage {stage}/{self.total_stages}] Running LPIS (Parcel Boundary) Rasterization...")
            lpis_dir = self.aux_dir / 'shapefiles_samples' / self.country
            lpis_candidates = list(lpis_dir.glob("*.gpkg")) + list(lpis_dir.glob("*.shp"))
            lpis_candidates = [p for p in lpis_candidates if p.name not in ['samples.shp', 'learn.shp', 'control.shp']]
            
            if not lpis_candidates:
                raise FileNotFoundError(f"No LPIS vector file (.gpkg/.shp) found in {lpis_dir}")
            
            lpis_file = lpis_candidates[0]
            print(f"    LPIS file selected: {lpis_file}")
            
            ds_ras = gdal.Open(str(self.ras))
            cols = ds_ras.RasterXSize
            rows = ds_ras.RasterYSize
            gt = ds_ras.GetGeoTransform()
            proj = ds_ras.GetProjection()
            
            minx = gt[0]
            maxy = gt[3]
            maxx = minx + gt[1] * cols
            miny = maxy + gt[5] * rows
            
            info = read_info(str(lpis_file))
            
            srs_target = osr.SpatialReference()
            srs_target.ImportFromWkt(proj)
            target_epsg = srs_target.GetAttrValue("AUTHORITY", 1)
            
            from pyproj import Transformer
            lpis_crs = info.get('crs')
            print(f"    LPIS CRS: {lpis_crs}")
            
            transformer = Transformer.from_crs(f"EPSG:{target_epsg}", lpis_crs, always_xy=True)
            p1 = transformer.transform(minx, miny)
            p2 = transformer.transform(maxx, maxy)
            lpis_bbox = (min(p1[0], p2[0]), min(p1[1], p2[1]), max(p1[0], p2[0]), max(p1[1], p2[1]))
            
            print(f"    Querying LPIS with spatial filter bbox: {lpis_bbox}")
            gdf = read_dataframe(str(lpis_file), bbox=lpis_bbox)
            print(f"    Loaded {len(gdf)} intersecting parcels. Reprojecting to EPSG:{target_epsg}...")
            gdf_target = gdf.to_crs(f"EPSG:{target_epsg}")
            
            # Resolve unique ID column
            fid_col = info.get('fid_column')
            if fid_col and fid_col in gdf_target.columns:
                id_col = fid_col
            elif 'id' in gdf_target.columns:
                id_col = 'id'
            elif 'id_0' in gdf_target.columns:
                id_col = 'id_0'
            else:
                id_col = None

            if id_col is None:
                gdf_target['lpis_id'] = np.arange(1, len(gdf_target) + 1)
                id_col = 'lpis_id'
            else:
                gdf_target[id_col] = pd.to_numeric(gdf_target[id_col], errors='coerce').fillna(0).astype(np.int32)
                if gdf_target[id_col].sum() == 0 or gdf_target[id_col].nunique() < len(gdf_target):
                    gdf_target['lpis_id'] = np.arange(1, len(gdf_target) + 1)
                    id_col = 'lpis_id'
            
            temp_gpkg = self.seg_dir / f"temp_lpis_{self.sanitized_track}.gpkg"
            gdf_target.geometry = gdf_target.geometry.force_2d()
            gdf_target.to_file(str(temp_gpkg), driver="GPKG", engine="pyogrio")
            
            driver = gdal.GetDriverByName("GTiff")
            ds_out = driver.Create(
                str(self.seg_tif),
                cols, rows, 1,
                gdal.GDT_Int32,
                options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES']
            )
            ds_out.SetGeoTransform(gt)
            ds_out.SetProjection(proj)
            
            band = ds_out.GetRasterBand(1)
            band.SetNoDataValue(0)
            band.Fill(0)
            
            print(f"    Rasterizing parcels to {self.seg_tif.name} (burning column '{id_col}')...")
            gdal.Rasterize(ds_out, str(temp_gpkg), attribute=id_col)
            
            ds_out.FlushCache()
            ds_out = None
            ds_ras = None
            
            if os.path.exists(temp_gpkg):
                os.remove(temp_gpkg)
                    
            print(f"Completed stage {stage}: LPIS rasterized.\n")
            return

        method = params.get('method', 'otb_meanshift')

        if method in ['otb_meanshift', 'otb_meanshift_summed']:
            print(f"[Stage {stage}/{self.total_stages}] Running OTB Large-Scale Mean-Shift (Vector Mode) [{method}]...")

            input_raster_for_seg = self.ras
            if method == 'otb_meanshift_summed':
                try:
                    input_raster_for_seg = self._create_summed_composite()
                except Exception as e:
                    print(f"    [WARNING] Failed to create summed composite: {e}. Falling back to full stack.")
                    input_raster_for_seg = self.ras

            if not self.seg_shp.exists():
                cmd = (
                    f"otbcli_LargeScaleMeanShift -in {input_raster_for_seg} -spatialr {params['spatialr']} "
                    f"-ranger {params['ranger']} -minsize {params['minsize']} "
                    f"-tilesizex {params['tilesizex']} -tilesizey {params['tilesizey']} "
                    f"-mode vector -mode.vector.out {self.seg_shp} "
                    f"-cleanup false -ram {params['ram']}"
                )
                self._run_cmd(cmd, stage, 'OTB LargeScaleMeanShift (Vector)')
                if not self.seg_shp.exists():
                    raise RuntimeError("OTB Vector Segmentation failed: output shapefile not found.")
            else:
                print(f"    [INFO] Segmentation vector already exists at {self.seg_shp}")

            # --- Rasterize the vector shapefile so ANN processing can use it as a tiled grid ---
            if not self.seg_tif.exists():
                print(f"    [INFO] Rasterizing segmentation for ANN feature extraction...")

                ds_stack = gdal.Open(str(input_raster_for_seg))
                gt = ds_stack.GetGeoTransform()
                proj = ds_stack.GetProjection()
                cols = ds_stack.RasterXSize
                rows = ds_stack.RasterYSize

                # OTB generates a column 'DN' or 'label' for region IDs. Typically 'DN' for vector mode in old versions, or 'label'.
                # gdal.Rasterize will burn the object ID into the pixels.

                driver = gdal.GetDriverByName('GTiff')
                out_ds = driver.Create(str(self.seg_tif), cols, rows, 1, gdal.GDT_Int32,
                                       options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
                out_ds.SetGeoTransform(gt)
                out_ds.SetProjection(proj)

                # Check what field OTB generated
                shp_ds = ogr.Open(str(self.seg_shp))
                layer = shp_ds.GetLayer()
                field_names = [field.name for field in layer.schema]
                id_field = 'DN' if 'DN' in field_names else 'label' if 'label' in field_names else field_names[0]

                print(f"    [INFO] Burning field '{id_field}' into raster...")
                gdal.RasterizeLayer(out_ds, [1], layer, options=[f"ATTRIBUTE={id_field}"])

                out_ds.FlushCache()
                out_ds = None
                shp_ds = None
                ds_stack = None
                print(f"    [INFO] Rasterization complete: {self.seg_tif}")
            else:
                print(f"    [INFO] Rasterized segmentation already exists.")

            return

        if method in ['python_felzenszwalb', 'python_slic', 'python_sam']:
            if method != 'python_sam' and not HAS_SKIMAGE:
                print("Error: scikit-image not installed.")
                return
            if method == 'python_sam' and not HAS_SAM:
                print("Error: segment-anything or torch not installed. Install via: pip install git+https://github.com/facebookresearch/segment-anything.git")
                return
                
            original_ras = self.ras
            if method in ['python_sam', 'python_felzenszwalb', 'python_slic']:
                try:
                    self.ras = self._create_summed_composite()
                except Exception as e:
                    print(f"    [WARNING] Failed to create summed composite: {e}. Falling back to full stack.")
            
            # Use appropriate internal method logic
            self._run_python_segmentation_tiled(params, stage, method)
            
            self.ras = original_ras
            return

        print(f"Error: Unknown segmentation method '{method}'")

    def _run_python_segmentation_tiled(self, params, stage, method):
        print(f"[Stage {stage}/{self.total_stages}] Running Tiled Python Segmentation ({method})...")

        try:
            gdal.SetCacheMax(4 * 1024 * 1024 * 1024)  # Increase GDAL cache to 4GB
            
            ds = gdal.Open(str(self.ras))
            if not ds: raise RuntimeError("Could not open raster")

            # Open Data Footprint Mask (Stage 0 output) if it exists
            ds_foot = None
            if self.use_footprint_mask and self.footprint_mask.exists():
                ds_foot = gdal.Open(str(self.footprint_mask))

            cols = ds.RasterXSize
            rows = ds.RasterYSize
            nbands = ds.RasterCount
            gt = ds.GetGeoTransform()
            proj = ds.GetProjection()

            driver = gdal.GetDriverByName('GTiff')
            out_ds = driver.Create(str(self.seg_tif), cols, rows, 1, gdal.GDT_Int32,
                                   options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
            out_ds.SetGeoTransform(gt)
            out_ds.SetProjection(proj)
            out_band = out_ds.GetRasterBand(1)
            out_band.SetNoDataValue(0)

            tile_size = params.get('tile_size', 2048 if method == 'python_sam' else 4096)
            buffer = params.get('buffer', 128 if method == 'python_sam' else 256)
            global_seg_id = 1
            
            sam_geo = None
            if method == 'python_sam':
                print(f"    Loading SAM-Geo model ({params['sam_model_type']}) to {params['sam_device']}...")
                try:
                    from samgeo import SamGeo
                    sam_geo = SamGeo(
                        model_type=params['sam_model_type'],
                        checkpoint=params['sam_checkpoint'],
                        device=params['sam_device'],
                        sam_kwargs={
                            "points_per_side": 128,
                            "pred_iou_thresh": 0.45,
                            "stability_score_thresh": 0.45,
                            "crop_n_layers": 1,
                            "crop_n_points_downscale_factor": 2,
                            "min_mask_region_area": 5
                        }
                    )
                except Exception as e:
                    print(f"    [ERROR] Failed to load SAM-Geo model: {e}")
                    print("    Please ensure you have installed segment-geospatial and have the proper checkpoint.")
                    return

            for y in range(0, rows, tile_size):
                for x in range(0, cols, tile_size):
                    xsize_valid = min(tile_size, cols - x)
                    ysize_valid = min(tile_size, rows - y)

                    x_start_buf = max(0, x - buffer)
                    y_start_buf = max(0, y - buffer)
                    x_end_buf = min(cols, x + xsize_valid + buffer)
                    y_end_buf = min(rows, y + ysize_valid + buffer)

                    xsize_buf = x_end_buf - x_start_buf
                    ysize_buf = y_end_buf - y_start_buf

                    print(f"    Processing Tile: x={x}, y={y} (buffered {xsize_buf}x{ysize_buf})")

                    img_list = []
                    for b in range(1, nbands + 1):
                        band = ds.GetRasterBand(b)
                        arr = band.ReadAsArray(x_start_buf, y_start_buf, xsize_buf, ysize_buf)
                        if arr is None:
                            print(f"    [WARNING] Failed to read block at x={x_start_buf}, y={y_start_buf}. Skipping tile.")
                            img_list = None
                            break
                        arr = np.nan_to_num(arr)
                        img_list.append(arr)

                    if img_list is None: continue

                    img = np.dstack(img_list)
                    
                    # Use footprint mask if available, otherwise fallback to sum > 0
                    if ds_foot:
                        valid_mask_buf = ds_foot.GetRasterBand(1).ReadAsArray(x_start_buf, y_start_buf, xsize_buf, ysize_buf) > 0
                        valid_mask = valid_mask_buf
                    else:
                        valid_mask = np.sum(np.abs(img), axis=2) > 0
                        
                    if not np.any(valid_mask): continue

                    # Normalize inputs to [0.0, 1.0] using 2%-98% percentiles for robust traditional segmentations
                    img_norm = np.zeros(img.shape, dtype=np.float32)
                    for b in range(img.shape[2]):
                        band_data = img[:, :, b]
                        p2, p98 = np.percentile(band_data[valid_mask], (2, 98))
                        if p98 > p2:
                            img_norm[:, :, b] = np.clip((band_data - p2) / (p98 - p2), 0.0, 1.0)
                        else:
                            img_norm[:, :, b] = 0.0

                    if method == 'python_sam':
                        import cv2
                        from scipy.ndimage import distance_transform_edt
                        
                        # Convert float32 1-band to 8-bit RGB for SAM
                        img_8bit = np.zeros(img.shape, dtype=np.uint8)
                        valid_pixels = valid_mask[:, :, np.newaxis]
                        
                        if np.any(valid_pixels):
                            p2, p98 = np.percentile(img[valid_pixels], (2, 98))
                            img_clip = np.clip(img, p2, p98)
                            
                            # Avoid division by zero
                            if p98 > p2:
                                img_8bit[valid_pixels] = ((img_clip[valid_pixels] - p2) / (p98 - p2) * 255).astype(np.uint8)
                                
                            # Apply CLAHE to enhance contrast in darker regions (SAR data is very skewed)
                            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                            img_clahe = clahe.apply(img_8bit[:, :, 0])
                            img_8bit[:, :, 0] = img_clahe
                            
                            # Apply bilateral filter to smooth out speckle noise while preserving sharp boundaries
                            print("    [SAM-Geo] Applying bilateral filter for edge-preserving speckle smoothing...")
                            img_chan = np.ascontiguousarray(img_8bit[:, :, 0])
                            img_smoothed = cv2.bilateralFilter(img_chan, d=9, sigmaColor=50, sigmaSpace=50)
                            img_8bit[:, :, 0] = img_smoothed
                            
                        # SAM requires 3 channel RGB
                        if img_8bit.shape[2] == 1:
                            img_rgb = np.repeat(img_8bit, 3, axis=2)
                        else:
                            img_rgb = img_8bit[:, :, :3]
                            if img_rgb.shape[2] < 3:
                                img_rgb = np.pad(img_rgb, ((0,0),(0,0),(0, 3-img_rgb.shape[2])), mode='constant')
                                
                        sam_geo.generate(
                            source=img_rgb,
                            output=None,
                            foreground=False,
                            unique=True,
                            min_size=10,
                            max_size=100000
                        )
                        segments_buf = sam_geo.objects.astype(np.int32)
                                    
                        # Fill empty spaces (NoData) with nearest segment (distance transform)
                        zero_mask_buf = (segments_buf == 0) & valid_mask
                        if np.any(zero_mask_buf) and np.any(segments_buf > 0):
                            _, indices = distance_transform_edt(segments_buf == 0, return_indices=True)
                            segments_buf[zero_mask_buf] = segments_buf[tuple(indices)][zero_mask_buf]
                    elif method == 'python_felzenszwalb':
                        # Add a tiny amount of random noise to break flat zero regions (fixes scikit-image Cython infinite loop)
                        noise = np.random.normal(0, 1e-6, img_norm.shape).astype(np.float32)
                        img_norm_noise = img_norm + noise
                        segments_buf = felzenszwalb(img_norm_noise, scale=params['scale'], sigma=params['sigma'],
                                                min_size=params['min_size'])
                    elif method == 'python_slic':
                        max_tile_pixels = (tile_size + 2 * buffer) ** 2
                        pixels_per_segment = max_tile_pixels / params['n_segments']
                        active_pixels = np.sum(valid_mask)
                        n_segments_dynamic = max(1, int(active_pixels / pixels_per_segment))
                        segments_buf = slic(img_norm, n_segments=n_segments_dynamic, compactness=params['compactness'],
                                        sigma=params['slic_sigma'], start_label=1, mask=valid_mask)

                    y_offset = y - y_start_buf
                    x_offset = x - x_start_buf
                    
                    # Mask out segments outside footprint
                    segments_buf[~valid_mask] = 0

                    # Get the valid mask for the unbuffered tile
                    valid_mask_crop = valid_mask[y_offset : y_offset + ysize_valid, x_offset : x_offset + xsize_valid]

                    segments = segments_buf[y_offset : y_offset + ysize_valid, x_offset : x_offset + xsize_valid]
                    
                    seg_valid_mask = segments > 0
                    
                    unique_segs = np.unique(segments[seg_valid_mask])
                    if len(unique_segs) > 0:
                        max_seg = segments.max()
                        mapping = np.zeros(max_seg + 1, dtype=np.int32)
                        mapping[unique_segs] = np.arange(global_seg_id, global_seg_id + len(unique_segs))
                        
                        segments = mapping[segments]
                        segments[~valid_mask_crop] = 0
                        global_seg_id += len(unique_segs)
                    else:
                        segments[~valid_mask_crop] = 0

                    out_band.WriteArray(segments.astype(np.int32), x, y)

            out_ds.FlushCache()
            out_ds = None
            ds = None
            ds_foot = None
            print(f"    Segmentation Raster saved to {self.seg_tif}\n")

        except Exception as e:
            print(f"ERROR in Python segmentation: {e}")
            raise

    # --- Stage 2: Sample Split ---
    def stage_2_split_samples(self, **kwargs):
        self._ensure_directories()
        params = self.stage2_params.copy()
        params.update(kwargs)
        stage = 2

        if not self.sample_shp.exists():
            print("ERROR: Input sample file not found.")
            return

        gdf = gpd.read_file(str(self.sample_shp), engine="pyogrio")

        # Prevent spatial data leakage by splitting on unique segment IDs
        if not self.seg_tif.exists():
            print("ERROR: Segmentation raster not found. Run Stage 1 first to do spatial-leakage-free split.")
            return

        print(f"    Aligning training samples with segmentation raster {self.seg_tif.name} to prevent spatial data leakage...")
        ds_seg = gdal.Open(str(self.seg_tif))
        gt = ds_seg.GetGeoTransform()
        inv_gt = gdal.InvGeoTransform(gt)
        raster_proj = ds_seg.GetProjection()
        cols = ds_seg.RasterXSize
        rows = ds_seg.RasterYSize
        seg_band = ds_seg.GetRasterBand(1)

        # Open footprint mask to filter out samples outside the radar swath
        foot_band = None
        ds_foot = None
        if self.use_footprint_mask and self.footprint_mask.exists():
            print(f"    Filtering points by footprint mask {self.footprint_mask.name} to discard NoData areas...")
            ds_foot = gdal.Open(str(self.footprint_mask))
            foot_band = ds_foot.GetRasterBand(1)

        # Reproject points if CRS differs
        from pyproj import CRS
        if raster_proj and gdf.crs:
            target_crs = CRS.from_wkt(raster_proj)
            if gdf.crs != target_crs:
                print("    Reprojecting points to match segmentation CRS...")
                gdf = gdf.to_crs(target_crs)

        xs = gdf.geometry.x.values
        ys = gdf.geometry.y.values
        pxs = (inv_gt[0] + inv_gt[1] * xs + inv_gt[2] * ys).astype(int)
        pys = (inv_gt[3] + inv_gt[4] * xs + inv_gt[5] * ys).astype(int)

        seg_ids = []
        for px, py in zip(pxs, pys):
            if 0 <= px < cols and 0 <= py < rows:
                try:
                    # Filter out points that fall outside the active radar footprint
                    if foot_band is not None:
                        is_active = foot_band.ReadAsArray(int(px), int(py), 1, 1)[0, 0] > 0
                        if not is_active:
                            seg_ids.append(0)
                            continue
                    
                    val = seg_band.ReadAsArray(int(px), int(py), 1, 1)[0, 0]
                    seg_ids.append(val)
                except:
                    seg_ids.append(0)
            else:
                seg_ids.append(0)

        gdf['seg_id'] = seg_ids
        ds_seg = None  # Close dataset
        ds_foot = None  # Close dataset

        # Keep only points that fall into a valid segment (seg_id > 0)
        gdf_valid = gdf[gdf['seg_id'] > 0].copy()
        dropped = len(gdf) - len(gdf_valid)
        if dropped > 0:
            print(f"    Warning: Dropped {dropped} points that fell outside valid segments.")

        if len(gdf_valid) == 0:
            print("ERROR: No points fell within any valid segments.")
            return

        unique_segs = gdf_valid['seg_id'].unique()
        print(f"    Found {len(unique_segs)} unique segments for {len(gdf_valid)} valid points.")

        # Split unique segment IDs
        np.random.seed(params['random_state'])
        np.random.shuffle(unique_segs)
        split_idx = int(len(unique_segs) * params['learn_frac'])
        train_segs = set(unique_segs[:split_idx])

        learn = gdf_valid[gdf_valid['seg_id'].isin(train_segs)].copy()
        control = gdf_valid[~gdf_valid['seg_id'].isin(train_segs)].copy()

        # Remove temporary column to avoid shapefile writing issues
        learn = learn.drop(columns=['seg_id'])
        control = control.drop(columns=['seg_id'])

        learn.to_file(str(self.learn_shp), engine="pyogrio")
        control.to_file(str(self.control_shp), engine="pyogrio")
        print(f"Completed stage {stage}. Total valid: {len(gdf_valid)}, Learn: {len(learn)}, Control: {len(control)}\n")

    # --- Stage 3: Feature Extraction (Object-Based) ---
    def stage_3_selection(self):
        self._ensure_directories()
        stage = 3

        if self.sel_csv.exists():
            print(f"[Stage {stage}] Features already extracted, skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Extracting OBJECT-BASED features for Training Points...")

        if not self.learn_shp.exists():
            print("ERROR: Learn samples not found.")
            return

        gdf = gpd.read_file(str(self.learn_shp), engine="pyogrio")

        ds = gdal.Open(str(self.ras))
        gt = ds.GetGeoTransform()
        inv_gt = gdal.InvGeoTransform(gt)
        raster_proj = ds.GetProjection()
        nbands = ds.RasterCount
        cols = ds.RasterXSize
        rows = ds.RasterYSize

        # --- FIX: ALIGN CRS FOR TRAINING SAMPLES ---
        from pyproj import CRS
        if raster_proj and gdf.crs:
            target_crs = CRS.from_wkt(raster_proj)
            if gdf.crs != target_crs:
                print(f"    Warning: Reprojecting samples from {gdf.crs.name} to Match Raster CRS...")
                gdf = gdf.to_crs(target_crs)

        print(f"    Finding target segments for {len(gdf)} points...")

        seg_ds = gdal.Open(str(self.seg_tif))
        seg_band = seg_ds.GetRasterBand(1)

        # Open footprint mask to verify points during segment mapping
        foot_ds = None
        foot_band = None
        if self.use_footprint_mask and self.footprint_mask.exists():
            foot_ds = gdal.Open(str(self.footprint_mask))
            foot_band = foot_ds.GetRasterBand(1)

        target_segments = {}

        xs = gdf.geometry.x.values
        ys = gdf.geometry.y.values

        pxs = (inv_gt[0] + inv_gt[1] * xs + inv_gt[2] * ys).astype(int)
        pys = (inv_gt[3] + inv_gt[4] * xs + inv_gt[5] * ys).astype(int)
        crop_ids = gdf['crop_id'].values

        for px, py, crop_id in zip(pxs, pys, crop_ids):
            if 0 <= px < cols and 0 <= py < rows:
                try:
                    # Skip points outside the active footprint
                    if foot_band is not None:
                        is_active = foot_band.ReadAsArray(int(px), int(py), 1, 1)[0, 0] > 0
                        if not is_active:
                            continue
                    
                    seg_id = seg_band.ReadAsArray(int(px), int(py), 1, 1)[0, 0]
                    if seg_id > 0:
                        target_segments[seg_id] = crop_id
                except:
                    pass

        if not target_segments:
            print("ERROR: No valid samples found overlapping the raster.")
            return

        print(f"    Found {len(target_segments)} unique segments for training.")
        print("    Calculating true segment means (Optimized Tiled Read)...")

        target_ids_set = set(target_segments.keys())

        sums = {tid: np.zeros(nbands, dtype=np.float64) for tid in target_ids_set}
        sums_sq = {tid: np.zeros(nbands, dtype=np.float64) for tid in target_ids_set}
        counts = {tid: 0 for tid in target_ids_set}

        # Open footprint mask for tiled read masking
        footprint_ds = None
        footprint_band = None
        if self.use_footprint_mask and self.footprint_mask.exists():
            footprint_ds = gdal.Open(str(self.footprint_mask))
            footprint_band = footprint_ds.GetRasterBand(1)

        tile_size = 2048

        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                seg_arr = seg_band.ReadAsArray(x, y, xsize, ysize)
                tile_ids = np.unique(seg_arr)

                intersect_ids = target_ids_set.intersection(tile_ids)

                if not intersect_ids:
                    continue

                sys.stdout.write(f"\r      Reading required data from Tile: x={x}, y={y}    ")
                sys.stdout.flush()

                stack_tile = ds.ReadAsArray(x, y, xsize, ysize)
                stack_tile = np.nan_to_num(stack_tile, copy=False)

                foot_arr = None
                if footprint_band is not None:
                    foot_arr = footprint_band.ReadAsArray(x, y, xsize, ysize) > 0

                for tid in intersect_ids:
                    mask = (seg_arr == tid)
                    if foot_arr is not None:
                        mask = mask & foot_arr
                    
                    pixel_count = np.sum(mask)
                    if pixel_count == 0:
                        continue
                    counts[tid] += pixel_count
                    for b in range(nbands):
                        vals = stack_tile[b][mask]
                        sums[tid][b] += np.sum(vals)
                        sums_sq[tid][b] += np.sum(vals ** 2)

        footprint_ds = None
        print("\n    Aggregation complete. Formatting features...")

        valid_tids = [tid for tid in target_ids_set if counts[tid] > 0]

        crop_ids = [target_segments[tid] for tid in valid_tids]
        seg_ids = valid_tids

        if len(valid_tids) > 0:
            n_vals = np.array([counts[tid] for tid in valid_tids])[:, None]
            sums_arr = np.array([sums[tid] for tid in valid_tids])
            sums_sq_arr = np.array([sums_sq[tid] for tid in valid_tids])

            mean_matrix = sums_arr / n_vals
            var_matrix = (sums_sq_arr / n_vals) - (mean_matrix ** 2)
            var_matrix = np.maximum(var_matrix, 0)
            std_matrix = np.sqrt(var_matrix)
        else:
            mean_matrix = np.empty((0, nbands))
            std_matrix = np.empty((0, nbands))

        feature_data = compute_210_features_dict(mean_matrix, std_matrix, nbands=nbands)
        feature_data['crop_id'] = crop_ids
        feature_data['seg_id'] = seg_ids

        df_final = pd.DataFrame(feature_data)
        df_final.to_csv(self.sel_csv, index=False)
        
        # Close GDAL datasets to release file locks and native handles
        ds = None
        seg_ds = None
        foot_ds = None
        
        print(f"    Object-Based Features saved to {self.sel_csv}\n")

    # --- Stage 4: Train Classifier ---
    def stage_4_train_classifier(self, **kwargs):
        self._ensure_directories()
        params = self.stage4_params.copy()
        params.update(kwargs)
        stage = 4

        if not self.sel_csv.exists():
            print("ERROR: Feature CSV not found.")
            return

        model_fn = self.model_pkl

        print(f"[Stage {stage}/{self.total_stages}] Training ANN...")

        df = pd.read_csv(self.sel_csv)
        local_samples = len(df)
        local_classes = df['crop_id'].nunique()

        # Check if local dataset is too small to train a viable model
        if local_samples < 500 or local_classes < 5:
            print(f"\n[Fallback System] WARNING: Local training dataset is too small (Samples: {local_samples}, Classes: {local_classes}).")
            print(f"[Fallback System] Local model training is bypassed. Stage 5 (Inference) will automatically use the best national pre-trained model.")
            # Delete any existing local model file to ensure fallback triggers in Stage 5
            if self.model_pkl.exists():
                try:
                    self.model_pkl.unlink()
                except Exception as e:
                    print(f"    [WARNING] Failed to remove old local model file: {e}")
            return

        feat_cols = [c for c in df.columns if c not in ['crop_id', 'seg_id']]
        self.feat_cols = feat_cols

        print(f"    Original samples: {local_samples}")

        X = df[feat_cols].values
        y = df['crop_id'].values

        # Apply crop aggregation for NL if country is NL
        if self.country == 'NL':
            print("    Applying crop aggregation for Netherlands to reduce semantic confusion...")
            crop_aggregation = get_crop_aggregation(self.country, self.learn_shp)
            y = np.array([crop_aggregation.get(val, val) for val in y])

        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)

        all_classes = np.unique(y)
        class_weights = _calculate_class_weights(y, all_classes)

        hidden_sizes = tuple(map(int, str(params.get('sk_hidden_sizes', '256,128,64')).split(',')))
        max_iter = params.get('sk_max_iter', 120)

        print(f"    Training TorchMLPClassifier on {X.shape[0]} samples with input dim {X.shape[1]}, hidden sizes {hidden_sizes}...")
        clf = TorchMLPClassifier(
            hidden_layer_sizes=hidden_sizes,
            max_iter=max_iter,
            batch_size=256,
            class_weights=class_weights,
            all_classes=all_classes
        )

        clf.fit(X_scaled, y)

        joblib.dump({'model': clf, 'scaler': scaler, 'feats': feat_cols}, model_fn)
        print(f"Model saved to {model_fn}")

        y_pred = clf.predict(X_scaled)
        labels = sorted(list(set(y)))
        cm = confusion_matrix(y, y_pred, labels=labels)
        print("\n--- Training Confusion Matrix ---")
        print(pd.DataFrame(cm, index=labels, columns=labels).to_string())
        print("\n")

    def find_best_fallback_model(self):
        """
        Scans all orbit directories for the same country to locate the best available pre-trained model
        for the current segmentation suffix. Ranks candidates by number of classes, then geographic distance
        from the local orbit (descending proximity), then training sample size, and finally classification overall accuracy.
        """
        print(f"\n[Fallback System] Scanning for pre-trained models for country '{self.country}' (Suffix: {self.seg_suffix})...")
        country_dir = self.base_dir / self.country
        if not country_dir.exists():
            print(f"[Fallback System] Country directory {country_dir} does not exist.")
            return None, None

        # Get center coordinate of local raster
        local_center = None
        try:
            ds = gdal.Open(str(self.ras))
            if ds:
                gt = ds.GetGeoTransform()
                cols = ds.RasterXSize
                rows = ds.RasterYSize
                local_center = (gt[0] + (cols * gt[1]) / 2.0, gt[3] + (rows * gt[5]) / 2.0)
                ds = None
        except Exception as e:
            print(f"  [WARNING] Failed to get local raster center: {e}")

        candidates = []
        # Find all track/orbit directories
        for orbit_path in country_dir.iterdir():
            if not orbit_path.is_dir():
                continue
            
            # Skip current track
            current_track_name = Path(self.track).name
            if orbit_path.name == current_track_name:
                continue

            # Check if model pkl exists
            model_pkl_path = orbit_path / 'classification_results' / 'train_model' / f"{self.country}_{orbit_path.name}_model_{self.seg_suffix}.pkl"
            if not model_pkl_path.exists():
                continue

            # Load model to check number of classes
            n_classes = 0
            try:
                model_data = joblib.load(model_pkl_path)
                clf = model_data.get('model')
                if clf and hasattr(clf, 'classes_'):
                    n_classes = len(clf.classes_)
            except Exception as e:
                print(f"  [WARNING] Failed to load model for {orbit_path.name}: {e}")
                continue

            # Check training sample size from CSV
            n_samples = 0
            csv_path = orbit_path / 'classification_results' / 'samples' / f"{self.country}_{orbit_path.name}_learn_features_{self.seg_suffix}.csv"
            if csv_path.exists():
                try:
                    n_samples = sum(1 for _ in open(csv_path, encoding='utf-8')) - 1
                except Exception as e:
                    print(f"  [WARNING] Failed to read sample features CSV for {orbit_path.name}: {e}")

            # Check accuracy from metrics spreadsheet
            oa = 0.0
            metrics_path = orbit_path / 'classification_results' / 'classification' / f"{self.country}_{orbit_path.name}_metrics_{self.seg_suffix}.xlsx"
            if metrics_path.exists():
                try:
                    wb = openpyxl.load_workbook(str(metrics_path), read_only=True)
                    if 'Results' in wb.sheetnames:
                        sh = wb['Results']
                        for row in range(1, 40):
                            cell_val = sh.cell(row=row, column=1).value
                            if cell_val == 'Overall Accuracy':
                                oa = float(sh.cell(row=row, column=2).value or 0.0)
                                break
                    wb.close()
                except Exception as e:
                    pass

            # Calculate geographic distance between candidate and local raster centers
            distance_km = float('inf')
            try:
                cand_proc_dir = orbit_path / 'processed_raster'
                cand_ras = None
                if cand_proc_dir.exists():
                    search_patterns = [
                        f"{self.country}_{orbit_path.name}_*_VH_VV*.tif",
                        f"*{orbit_path.name}*.tif",
                    ]
                    for pattern in search_patterns:
                        cand_hdr = next(cand_proc_dir.glob(pattern), None)
                        if cand_hdr:
                            if cand_hdr.suffix.lower() in ['.tif', '.tiff']:
                                cand_ras = cand_hdr
                            else:
                                for ext in ['.img', '.tif', '.TIF']:
                                    p = cand_hdr.with_suffix(ext)
                                    if p.exists():
                                        cand_ras = p
                                        break
                            if cand_ras:
                                break
                
                if cand_ras and cand_ras.exists():
                    ds_cand = gdal.Open(str(cand_ras))
                    if ds_cand:
                        gt_cand = ds_cand.GetGeoTransform()
                        cols_cand = ds_cand.RasterXSize
                        rows_cand = ds_cand.RasterYSize
                        cand_center = (gt_cand[0] + (cols_cand * gt_cand[1]) / 2.0, gt_cand[3] + (rows_cand * gt_cand[5]) / 2.0)
                        ds_cand = None
                        
                        if local_center and cand_center:
                            # Distance in kilometers (assuming coordinates are projected meters)
                            distance_km = math.sqrt((local_center[0] - cand_center[0])**2 + (local_center[1] - cand_center[1])**2) / 1000.0
            except Exception as e:
                pass

            candidates.append({
                'track': orbit_path.name,
                'model_path': model_pkl_path,
                'samples_csv_path': csv_path,
                'learn_shp_path': orbit_path / 'classification_results' / 'samples' / f"learn_{self.seg_suffix}.shp",
                'n_classes': n_classes,
                'n_samples': n_samples,
                'accuracy': oa,
                'distance': distance_km
            })

        if not candidates:
            print("[Fallback System] No other model files found.")
            return None, None

        # Sort: first by classes (descending), then distance (ascending), then sample size (descending), then accuracy (descending)
        # We use -x['distance'] to sort distance ascending while using reverse=True
        candidates.sort(key=lambda x: (x['n_classes'], -x['distance'], x['n_samples'], x['accuracy']), reverse=True)

        print("Found candidates:")
        for cand in candidates:
            dist_str = f"{cand['distance']:.1f} km" if cand['distance'] != float('inf') else "unknown"
            print(f"  - Track: {cand['track']}, Distance: {dist_str}, Classes: {cand['n_classes']}, Samples: {cand['n_samples']}, OA: {cand['accuracy']:.4f}")

        best = candidates[0]
        print(f"[Fallback System] Selected best model from track: {best['track']}\n")
        return best['model_path'], best

    # --- Stage 5: Tiled Inference (Object-Based) ---
    def stage_5_classify_vector(self, force_recompute=False):
        # Renamed logic, kept name for compatibility
        self._ensure_directories()
        stage = 5

        # Check if we should use fallback model
        model_file = self.model_pkl
        fallback_path, fallback_info = self.find_best_fallback_model()
        
        use_fallback = False
        reason = ""
        
        if not model_file.exists():
            use_fallback = True
            reason = "Local model file does not exist"
        else:
            # Check local stats
            local_samples = 0
            if self.sel_csv.exists():
                try:
                    local_samples = sum(1 for _ in open(self.sel_csv, encoding='utf-8')) - 1
                except:
                    pass
            
            local_classes = 0
            try:
                local_data = joblib.load(model_file)
                local_clf = local_data.get('model')
                if local_clf and hasattr(local_clf, 'classes_'):
                    local_classes = len(local_clf.classes_)
            except:
                pass

            if fallback_info:
                # If local model has fewer classes than fallback, or local model has very few samples (e.g. < 500)
                # and fallback has more, fallback is preferred.
                if local_classes < fallback_info['n_classes']:
                    use_fallback = True
                    reason = f"Local model has fewer classes ({local_classes}) than the best national fallback model ({fallback_info['n_classes']})"
                elif local_samples < 500 and fallback_info['n_samples'] > local_samples * 2:
                    use_fallback = True
                    reason = f"Local model trained on very few samples ({local_samples}) compared to fallback model ({fallback_info['n_samples']})"
        
        used_fallback = False
        if use_fallback and fallback_path:
            print(f"\n[Fallback System] TRIGGERED: {reason}")
            print(f"[Fallback System] Swapping model to fallback: {fallback_path.name} (from track: {fallback_info['track']})")
            model_file = fallback_path
            used_fallback = True
        else:
            print(f"\n[Fallback System] Using local model: {model_file.name}")
            
        if not model_file.exists():
            print("ERROR: Model not found.")
            return

        if self.class_tif.exists() and not force_recompute:
            print(f"[Stage {stage}] Classification Raster exists, skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Running Tiled Object-Based Inference (Parallelized)...")

        data = joblib.load(model_file)
        clf = data['model']
        scaler = data['scaler']
        feat_cols = data['feats']

        # Calculate priors from training samples
        train_csv_to_use = fallback_info['samples_csv_path'] if used_fallback else self.sel_csv
        learn_shp_to_use = fallback_info['learn_shp_path'] if used_fallback else self.learn_shp

        if not train_csv_to_use.exists():
            print(f"ERROR: Training samples CSV not found: {train_csv_to_use}")
            return

        df_train = pd.read_csv(train_csv_to_use)
        y_train = df_train['crop_id'].values
        if self.country == 'NL':
            crop_aggregation = get_crop_aggregation(self.country, learn_shp_to_use)
            y_train = np.array([crop_aggregation.get(val, val) for val in y_train])
            
        classes = clf.classes_
        n_classes = len(classes)
        total_samples = len(y_train)
        class_counts = pd.Series(y_train).value_counts()
        
        # Calculate true prior probabilities P_true
        p_true = _get_priors_for_country(
            country=self.country,
            learn_shp_path=learn_shp_to_use,
            classes=classes,
            class_counts=class_counts,
            total_samples=total_samples
        )
        # Calculate training bias P_train introduced by class weights:
        # weight = sqrt(total / (n_classes * count))
        train_bias = np.array([math.sqrt(total_samples / (n_classes * class_counts.get(c, 1))) for c in classes])
        p_train = train_bias / np.sum(train_bias)
        
        # Exact Bayesian correction factor: P_true / P_train
        correction = p_true / (p_train + 1e-9)
        
        # Apply SATMIROL power smoothing factor (0.7) to prevent over-correction
        correction = np.power(correction, 0.7)
        
        # [NEW] Cap extreme multipliers to prevent "black hole" effect for dominant classes (like Grassland)
        correction = np.clip(correction, 0.3, 1.5)
        
        priors_arr = correction / np.sum(correction)

        ds_stack_info = gdal.Open(str(self.ras))
        cols = ds_stack_info.RasterXSize
        rows = ds_stack_info.RasterYSize
        nbands = ds_stack_info.RasterCount
        gt = ds_stack_info.GetGeoTransform()
        proj = ds_stack_info.GetProjection()
        ds_stack_info = None

        driver = gdal.GetDriverByName('GTiff')
        ds_cls = driver.Create(str(self.class_tif), cols, rows, 1, gdal.GDT_Int32,
                               options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        ds_cls.SetGeoTransform(gt)
        ds_cls.SetProjection(proj)
        ds_cls.GetRasterBand(1).SetNoDataValue(0)

        ds_conf = driver.Create(str(self.conf_tif), cols, rows, 1, gdal.GDT_Float32,
                                options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        ds_conf.SetGeoTransform(gt)
        ds_conf.SetProjection(proj)
        ds_conf.GetRasterBand(1).SetNoDataValue(0)

        tile_size = 2048

        # We need a lock when writing to the output datasets
        write_lock = threading.Lock()

        def process_tile(x, y):
            xsize = min(tile_size, cols - x)
            ysize = min(tile_size, rows - y)

            # Each thread needs its own GDAL dataset handles to be thread-safe
            ds_stack = gdal.Open(str(self.ras))
            ds_seg = gdal.Open(str(self.seg_tif))
            ds_foot = gdal.Open(str(self.footprint_mask)) if self.use_footprint_mask and self.footprint_mask.exists() else None

            try:
                seg_arr = ds_seg.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                if ds_foot is not None:
                    foot_arr = ds_foot.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                    # Apply footprint to segmentation tile
                    seg_arr[foot_arr == 0] = 0

                if np.all(seg_arr == 0):
                    return

                img_list = []
                for b in range(1, nbands + 1):
                    band = ds_stack.GetRasterBand(b)
                    arr = band.ReadAsArray(x, y, xsize, ysize)
                    arr = np.nan_to_num(arr)
                    img_list.append(arr)
                img = np.dstack(img_list)

                flat_seg = seg_arr.ravel()
                mask = flat_seg > 0
                valid_seg = flat_seg[mask]

                if len(valid_seg) == 0:
                    return

                # Get unique segment IDs in this tile (excluding 0)
                unique_ids = np.unique(seg_arr)
                unique_ids = unique_ids[unique_ids > 0]

                if len(unique_ids) == 0:
                    return

                # Calculate means and standard deviations for all bands using scipy.ndimage (extremely fast and memory efficient)
                means = np.zeros((len(unique_ids), nbands), dtype=np.float32)
                stds = np.zeros((len(unique_ids), nbands), dtype=np.float32)

                for b in range(nbands):
                    means[:, b] = ndi.mean(img[:, :, b], labels=seg_arr, index=unique_ids)
                    stds[:, b] = ndi.standard_deviation(img[:, :, b], labels=seg_arr, index=unique_ids)

                stds = np.nan_to_num(stds)

                # Compute the 210 advanced features using our helper function
                feature_data = compute_210_features_dict(means, stds, nbands=nbands)

                X_tile = np.zeros((len(unique_ids), len(feat_cols)), dtype=np.float32)
                for idx, col in enumerate(feat_cols):
                    X_tile[:, idx] = feature_data[col]
                X_scaled = scaler.transform(X_tile)

                # Get raw probabilities and apply Bayesian Prior Correction
                raw_probs = clf.predict_proba(X_scaled)
                corrected_probs = raw_probs * priors_arr
                row_sums = np.sum(corrected_probs, axis=1, keepdims=True) + 1e-9
                corrected_probs = corrected_probs / row_sums

                preds = clf.classes_[np.argmax(corrected_probs, axis=1)]
                probs = np.max(corrected_probs, axis=1)

                sort_idx = np.argsort(unique_ids)
                sorted_ids = unique_ids[sort_idx]
                sorted_preds = preds[sort_idx]
                sorted_probs = probs[sort_idx]

                idx_map = np.searchsorted(sorted_ids, valid_seg)
                idx_map = np.clip(idx_map, 0, len(sorted_ids) - 1)
                valid_match = sorted_ids[idx_map] == valid_seg

                flat_cls = np.zeros_like(flat_seg, dtype=np.int32)
                flat_conf = np.zeros_like(flat_seg, dtype=np.float32)

                global_mask = np.zeros(len(flat_seg), dtype=bool)
                global_mask[mask] = valid_match

                flat_cls[global_mask] = sorted_preds[idx_map[valid_match]]
                flat_conf[global_mask] = sorted_probs[idx_map[valid_match]]

                cls_tile = flat_cls.reshape(ysize, xsize)
                conf_tile = flat_conf.reshape(ysize, xsize)

                with write_lock:
                    print(f"    Writing Tile: x={x}, y={y}")
                    ds_cls.GetRasterBand(1).WriteArray(cls_tile, x, y)
                    ds_conf.GetRasterBand(1).WriteArray(conf_tile, x, y)
            finally:
                ds_stack = None
                ds_seg = None
                ds_foot = None


        tiles_to_process = []
        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                tiles_to_process.append((x, y))

        print(f"    Dispatching {len(tiles_to_process)} tiles to 20 workers...")
        with ThreadPoolExecutor(max_workers=20) as executor:
            futures = [executor.submit(process_tile, x, y) for x, y in tiles_to_process]
            for future in futures:
                future.result() # Wait for completion and raise any exceptions

        ds_cls.FlushCache()
        ds_conf.FlushCache()

        # Build pyramids (overviews) for faster loading in GIS software
        print("    [INFO] Building pyramids for raw classification and confidence rasters...")
        ds_cls.BuildOverviews(resampling="NEAREST", overviewlist=[2, 4, 8, 16, 32])
        ds_conf.BuildOverviews(resampling="AVERAGE", overviewlist=[2, 4, 8, 16, 32])

        ds_cls = None
        ds_conf = None
        print(f"    Classification saved to {self.class_tif}\n")

    # --- Stage 6: Mask Class ---
    def stage_6_mask_class(self, force_recompute=False):
        self._ensure_directories()
        stage = 6
        mask_file = self.agri_mask
        if not mask_file.exists():
            print(f"[Stage {stage}] ERROR: Agricultural mask not found: {mask_file}")
            print(f"  Run: python auxiliary_files/raster_files/AgriMasks/build_agri_mask.py --country {self.country}")
            return
        if not self.class_tif.exists():
            print(f"ERROR: Classified TIF not found.")
            return

        if not self.masked_class.exists() or force_recompute:
            self._apply_mask(self.class_tif, mask_file, self.masked_class, stage)
        else:
            print(f"[Stage {stage}/{self.total_stages}] Masked classification exists, skipping\n")

    # --- Stage 7: Mask Confidence ---
    def stage_7_mask_confidence(self, force_recompute=False):
        self._ensure_directories()
        stage = 7
        mask_file = self.agri_mask
        if not mask_file.exists():
            print(f"[Stage {stage}] ERROR: Agricultural mask not found: {mask_file}")
            print(f"  Run: python auxiliary_files/raster_files/AgriMasks/build_agri_mask.py --country {self.country}")
            return
        if not self.conf_tif.exists():
            print(f"ERROR: Confidence TIF not found.")
            return

        if not self.masked_conf.exists() or force_recompute:
            self._apply_mask(self.conf_tif, mask_file, self.masked_conf, stage)
        else:
            print(f"[Stage {stage}/{self.total_stages}] Masked confidence exists, skipping\n")

    # --- Stage 8: Metrics ---
    def stage_8_calculate_metrics(self):
        self._ensure_directories()
        stage = 8
        if not self.metrics_fp.exists():
            print(f"[Stage {stage}/{self.total_stages}] Computing metrics...")

            if not self.control_shp.exists():
                print(f"ERROR: Control shapefile not found.")
                return
            if not self.masked_class.exists():
                print(f"ERROR: Masked classification not found.")
                return

            ctrl = gpd.read_file(str(self.control_shp), engine="pyogrio")

            ds = gdal.Open(str(self.masked_class))
            raster_proj = ds.GetProjection()
            raster_srs = osr.SpatialReference()
            raster_srs.ImportFromWkt(raster_proj)

            if ctrl.crs:
                if ctrl.crs.to_wkt() != raster_srs.ExportToWkt():
                    print("    Aligning control points CRS to match raster...")
                    try:
                        from pyproj import CRS
                        target_crs = CRS.from_wkt(raster_srs.ExportToWkt())
                        ctrl = ctrl.to_crs(target_crs)
                    except Exception as e:
                        print(f"    Could not auto-align CRS: {e}")

            band = ds.GetRasterBand(1)
            gt = ds.GetGeoTransform()
            inv = gdal.InvGeoTransform(gt)
            true_vals, pred_vals = [], []

            xs = ctrl.geometry.x.values
            ys = ctrl.geometry.y.values

            pxs = (inv[0] + inv[1] * xs + inv[2] * ys).astype(int)
            pys = (inv[3] + inv[4] * xs + inv[5] * ys).astype(int)
            crop_ids = ctrl['crop_id'].values

            # Apply crop aggregation for NL
            if self.country == 'NL':
                print("    Applying crop aggregation for Netherlands validation labels...")
                crop_aggregation = get_crop_aggregation(self.country, self.control_shp)
                crop_ids = np.array([crop_aggregation.get(val, val) for val in crop_ids])

            for px, py, crop_id in zip(pxs, pys, crop_ids):
                try:
                    if 0 <= px < ds.RasterXSize and 0 <= py < ds.RasterYSize:
                        t = int(crop_id)
                        val_arr = band.ReadAsArray(px, py, 1, 1)
                        if val_arr is not None:
                            p = int(val_arr[0, 0])
                            if t > 0 and p > 0 and p != -9999:
                                true_vals.append(t)
                                pred_vals.append(p)
                except Exception as e:
                    print(f"    [WARNING] Failed to extract point value: {e}")

            if not true_vals or not pred_vals:
                print("ERROR: No valid matching true/predicted values found.")
                print("HINT: Ensure your test points intersect valid data areas in the masked raster.")
                return

            labels = sorted(list(set(true_vals + pred_vals)))
            cm = confusion_matrix(true_vals, pred_vals, labels=labels)
            precisions, recalls, f1s, _ = precision_recall_fscore_support(
                true_vals, pred_vals, labels=labels, average=None, zero_division=0
            )

            total = np.sum(cm)
            oa = np.trace(cm) / total
            sum_po = oa
            sum_pe = np.sum(np.sum(cm, axis=0) * np.sum(cm, axis=1)) / (total ** 2)
            kappa = (sum_po - sum_pe) / (1 - sum_pe) if (1 - sum_pe) != 0 else np.nan

            resx, resy = abs(gt[1]), abs(gt[5])
            area_ha = resx * resy / 10000

            arr = band.ReadAsArray()
            if arr is not None:
                unique_classes, counts = np.unique(arr[arr > 0], return_counts=True)
                class_areas = dict(zip(unique_classes, counts))
                areas = [{'Class': c, 'Area_ha': round(class_areas.get(c, 0) * area_ha, 2)} for c in labels]
            else:
                areas = [{'Class': c, 'Area_ha': 0} for c in labels]

            wb = openpyxl.Workbook()
            sh = wb.active
            sh.title = 'Results'

            sh.cell(row=1, column=1, value='Confusion Matrix').font = Font(bold=True)
            sh.cell(row=2, column=1, value='True \\ Pred').font = Font(bold=True)
            for j, lbl in enumerate(labels, start=2):
                sh.cell(row=2, column=j, value=lbl).font = Font(bold=True)
            for i, lbl in enumerate(labels, start=3):
                sh.cell(row=i, column=1, value=lbl).font = Font(bold=True)
                for j, _ in enumerate(labels):
                    sh.cell(row=i, column=j + 2, value=int(cm[i - 3, j]))

            # Compute weighted overall accuracy
            weighted_oa = None
            try:
                df_train = pd.read_csv(self.sel_csv)
                y_train = df_train['crop_id'].values
                if self.country == 'NL':
                    crop_aggregation = get_crop_aggregation(self.country, self.learn_shp)
                    y_train = np.array([crop_aggregation.get(val, val) for val in y_train])
                
                train_classes = sorted(list(set(y_train)))
                train_class_counts = pd.Series(y_train).value_counts()
                train_total_samples = len(y_train)
                
                p_true_all = _get_priors_for_country(
                    country=self.country,
                    learn_shp_path=self.learn_shp,
                    classes=train_classes,
                    class_counts=train_class_counts,
                    total_samples=train_total_samples
                )
                p_true_dict = dict(zip(train_classes, p_true_all))
                
                active_priors = np.array([p_true_dict.get(lbl, 1e-5) for lbl in labels])
                active_priors = active_priors / np.sum(active_priors)
                
                weighted_oa = np.sum(active_priors * recalls)
                print(f"    Weighted Overall Accuracy ({self.country} Area-Adjusted): {weighted_oa:.4f}")
            except Exception as e:
                print(f"    [WARNING] Could not compute Weighted Overall Accuracy: {e}")

            base = 4 + len(labels)
            sh.cell(row=base, column=1, value='Overall Accuracy').font = Font(bold=True)
            sh.cell(row=base, column=2, value=round(oa, 4))
            sh.cell(row=base + 1, column=1, value='Kappa').font = Font(bold=True)
            sh.cell(row=base + 1, column=2, value=round(kappa, 4))

            if weighted_oa is not None:
                sh.cell(row=base + 2, column=1, value='Weighted Overall Accuracy').font = Font(bold=True)
                sh.cell(row=base + 2, column=2, value=round(weighted_oa, 4))
                start = base + 4
            else:
                start = base + 3
            headers = ['Class', 'Producer Acc (Recall)', 'User Acc (Precision)', 'F1-score']
            for j, h in enumerate(headers, start=1):
                sh.cell(row=start, column=j, value=h).font = Font(bold=True)
            for idx, c in enumerate(labels):
                row_idx = start + 1 + idx
                sh.cell(row=row_idx, column=1, value=c)
                sh.cell(row=row_idx, column=2, value=round(recalls[idx], 4))
                sh.cell(row=row_idx, column=3, value=round(precisions[idx], 4))
                sh.cell(row=row_idx, column=4, value=round(f1s[idx], 4))

            ar0 = start + 1 + len(labels) + 1
            sh.cell(row=ar0, column=1, value='Areas (ha)').font = Font(bold=True)
            sh.cell(row=ar0 + 1, column=1, value='Class').font = Font(bold=True)
            sh.cell(row=ar0 + 1, column=2, value='Area_ha').font = Font(bold=True)
            for idx, a in enumerate(areas, start=ar0 + 2):
                sh.cell(row=idx, column=1, value=a['Class'])
                sh.cell(row=idx, column=2, value=a['Area_ha'])

            wb.save(str(self.metrics_fp))
            print(f"Metrics saved to {self.metrics_fp}\n")
        else:
            print(f"[Stage 8] Metrics Excel exists, skipping")


# --- Interactive Menu Helpers ---

SAM_MODELS = {
    '1': {'name': 'vit_b  (Small, ~375 MB, FAST,   ~2 GB VRAM - recommended for testing)',
           'model_type': 'vit_b',  'checkpoint': 'sam_vit_b_01ec64.pth',
           'url': 'https://dl.fbaipublicfiles.com/segment_anything/sam_vit_b_01ec64.pth'},
    '2': {'name': 'vit_l  (Medium, ~1.2 GB, MEDIUM, ~6 GB VRAM)',
           'model_type': 'vit_l',  'checkpoint': 'sam_vit_l_0b3195.pth',
           'url': 'https://dl.fbaipublicfiles.com/segment_anything/sam_vit_l_0b3195.pth'},
    '3': {'name': 'vit_h  (Huge,  ~2.4 GB, SLOW,   ~10 GB VRAM - highest accuracy)',
           'model_type': 'vit_h',  'checkpoint': 'sam_vit_h_4b8939.pth',
           'url': 'https://dl.fbaipublicfiles.com/segment_anything/sam_vit_h_4b8939.pth'},
}


def get_stage1_params_sam(param_dict):
    """Interactive menu for selecting the segmentation method and Stage 1 parameters."""
    new_params = param_dict.copy()
    current_method = new_params.get('method', 'python_sam')
    
    print("\n=== SELECT SEGMENTATION METHOD ===")
    print(f"  Current method: {current_method.upper()}")
    print()
    print("  [1] Meta SAM (Deep learning, default) [python_sam]")
    print("  [2] OTB Mean-Shift on summed dB (Fast C++ engine) [otb_meanshift_summed]")
    print("  [3] Felzenszwalb algorithm on full raster [python_felzenszwalb]")
    print("  [4] SLIC algorithm on full raster [python_slic]")
    print("  [5] LPIS boundary rasterization (Cadastral vector data) [lpis]")
    print("  [Enter] Keep current method")
    
    choice = input("Choose option (1-5): ").strip()
    
    method_mapping = {
        '1': 'python_sam',
        '2': 'otb_meanshift_summed',
        '3': 'python_felzenszwalb',
        '4': 'python_slic',
        '5': 'lpis'
    }
    
    if choice in method_mapping:
        new_params['method'] = method_mapping[choice]
        method = new_params['method']
        print(f"  Selected method: {method.upper()}")
        
        # Initialize default values for the selected method if they don't exist
        if method == 'python_sam':
            new_params.setdefault('tile_size', 2048)
            new_params.setdefault('buffer', 128)
            new_params.setdefault('sam_checkpoint', str(aux_dir / 'SAM_models' / 'sam_vit_h_4b8939.pth'))
            new_params.setdefault('sam_model_type', 'vit_h')
            new_params.setdefault('sam_device', 'cuda' if torch.cuda.is_available() else 'cpu')
        elif method in ['otb_meanshift_summed', 'otb_meanshift']:
            new_params.setdefault('spatialr', 4)
            new_params.setdefault('ranger', 0.3)
            new_params.setdefault('minsize', 20)
            new_params.setdefault('tilesizex', 4096)
            new_params.setdefault('tilesizey', 4096)
            new_params.setdefault('ram', 4096)
        elif method == 'python_felzenszwalb':
            new_params.setdefault('tile_size', 4096)
            new_params.setdefault('buffer', 256)
            new_params.setdefault('scale', 50.0)
            new_params.setdefault('sigma', 0.8)
            new_params.setdefault('min_size', 15)
        elif method == 'python_slic':
            new_params.setdefault('tile_size', 4096)
            new_params.setdefault('buffer', 256)
            new_params.setdefault('n_segments', 35000)  # Tuned 1: 35000, Original: 20000
            new_params.setdefault('compactness', 0.03)  # Tuned 1: 0.03, Original: 0.1
            new_params.setdefault('slic_sigma', 1.0)    # Tuned 1: 1.0
        elif method == 'lpis':
            pass
            
    method = new_params.get('method', 'python_sam')
    
    # Specific interactive choice for SAM models
    if method == 'python_sam':
        current_type = new_params.get('sam_model_type', 'vit_h')
        current_ckpt = new_params.get('sam_checkpoint', 'sam_vit_h_4b8939.pth')
        print("\n  Available SAM models:")
        for k, v in SAM_MODELS.items():
            marker = " <-- current" if v['model_type'] == current_type else ""
            print(f"    [{k}] {v['name']}{marker}")
        print()
        
        sam_choice = input("Choose SAM model (1/2/3) or Enter to keep current: ").strip()
        if sam_choice in SAM_MODELS:
            selected = SAM_MODELS[sam_choice]
            new_params['sam_model_type'] = selected['model_type']
            
            sam_models_dir = aux_dir / 'SAM_models'
            ckpt_fn = selected['checkpoint']
            ckpt_path = sam_models_dir / ckpt_fn
            new_params['sam_checkpoint'] = str(ckpt_path)

            if not ckpt_path.exists():
                print(f"\n  [WARNING] Checkpoint weight file '{ckpt_fn}' does not exist in {sam_models_dir}!")
                print(f"  Download it from: {selected['url']}")
                proceed = input("  Continue anyway? (y/n) [n]: ").strip().lower()
                if proceed != 'y':
                    return None
            else:
                print(f"  [OK] Checkpoint weight file '{ckpt_fn}' found.")
                
    # Filter parameters to show and edit based on chosen method
    show_keys = []
    if method == 'python_sam':
        show_keys = ['tile_size', 'buffer', 'sam_device']
    elif method in ['otb_meanshift_summed', 'otb_meanshift']:
        show_keys = ['spatialr', 'ranger', 'minsize', 'tilesizex', 'tilesizey', 'ram']
    elif method == 'python_felzenszwalb':
        show_keys = ['tile_size', 'buffer', 'scale', 'sigma', 'min_size']
    elif method == 'python_slic':
        show_keys = ['tile_size', 'buffer', 'n_segments', 'compactness', 'slic_sigma']
    elif method == 'lpis':
        show_keys = []

    if method == 'lpis':
        print("\n  No configurable parameters for LPIS rasterization.")
        print("  It will use the cadastral vector files in the country folder.")
    else:
        print(f"\n--- Edit parameters for: {method.upper()} ---")
        for key in show_keys:
            val = new_params.get(key)
            new_val_str = input(f"  {key} [{val}]: ").strip()
            if new_val_str:
                try:
                    if isinstance(val, bool):
                        new_params[key] = new_val_str.lower() in ('true', '1', 'y', 'yes')
                    else:
                        new_params[key] = type(val)(new_val_str) if val is not None else new_val_str
                except ValueError:
                    print("    Invalid value. Keeping default.")

    print("\n--- APPROVED SEGMENTATION PARAMETERS ---")
    print(f"  Method: {method.upper()}")
    for key in show_keys:
        print(f"  {key}: {new_params.get(key)}")
    print("==========================================\n")
    return new_params


def get_params(param_dict):
    new_params = param_dict.copy()
    print("--- Current Parameters ---")
    for key, val in new_params.items():
        print(f"  {key}: {val}")

    if input("Change parameters? (y/n) [n]: ").lower() != 'y':
        return new_params

    for key, val in new_params.items():
        new_val_str = input(f"Enter new value for '{key}' [{val}]: ")
        if not new_val_str:
            continue
        try:
            original_type = type(val)
            new_params[key] = original_type(new_val_str)
        except ValueError:
            print(f"Invalid value. Keeping default {val}.")
    return new_params


def get_classifier_params(param_dict):
    new_params = param_dict.copy()
    print("--- Current Parameters ---")
    for key, val in new_params.items():
        print(f"  {key}: {val}")

    if input("Change parameters? (y/n) [n]: ").lower() != 'y':
        return new_params

    clf = input(f"Enter classifier (ann_sklearn) [{new_params['classifier']}]: ") or new_params['classifier']
    new_params['classifier'] = clf.lower()

    print(f"\n--- Setting parameters for {clf.upper()} ---")
    prefix = 'sk_'
    for key in [k for k in new_params if k.startswith(prefix)]:
        val = new_params[key]
        new_val_str = input(f"Enter new value for '{key}' [{val}]: ")
        if new_val_str:
            try:
                new_params[key] = type(val)(new_val_str)
            except ValueError:
                print(f"Invalid value.")
    return new_params


# --- Main Execution ---

def main_menu(pipeline):
    while True:
        seg_method = pipeline.stage1_params.get('method', 'python_sam')
        seg_desc = "LPIS Cadastral Vector Rasterization" if pipeline.seg_mode == 'lpis' else f"SAR Segmentation ({seg_method})"
        
        menu = f"""
    --- Raster-Based OBIA Pipeline (ANN) ---
    Track: {pipeline.track} ({pipeline.country})
    Segmentation Mode: {pipeline.seg_mode.upper()} ({seg_method})

    [0] Stage 0: Generate Data Footprint Mask
    [1] Stage 1: {seg_desc}
    [2] Stage 2: Split Samples
    [3] Stage 3: Extract Features (Object-based Training)
    [4] Stage 4: Train ANN Classifier
    [5] Stage 5: Tiled Object-Based Inference
    [6] Stage 6: Mask Classification
    [7] Stage 7: Mask Confidence
    [8] Stage 8: Calculate Metrics

    [A] Run All Stages (Forces overwrite of Stages 5-8 to clear old bugs)
    [Q] Quit

    Enter your choice: 
        """
        try:
            choice = input(menu).strip().upper()
        except (EOFError, KeyboardInterrupt):
            print("\nExiting interactive menu due to standard input disconnection or interruption.")
            break
        try:
            if choice == '0':
                pipeline.stage_0_generate_footprint()
            elif choice == '1':
                new_params = get_stage1_params_sam(pipeline.stage1_params)
                if new_params is None:
                    print("  Segmentation parameter setup cancelled.")
                    continue
                pipeline.stage1_params.update(new_params)
                pipeline.update_paths(pipeline.stage1_params['method'])
                pipeline.stage_1_segmentation(**pipeline.stage1_params)
            elif choice == '2':
                new_params = get_params(pipeline.stage2_params)
                pipeline.stage2_params.update(new_params)
                pipeline.stage_2_split_samples(**pipeline.stage2_params)
            elif choice == '3':
                pipeline.stage_3_selection()
            elif choice == '4':
                new_params = get_classifier_params(pipeline.stage4_params)
                force = (pipeline.stage4_params != new_params)
                pipeline.stage4_params.update(new_params)
                pipeline.stage_4_train_classifier(force_retrain=force, **pipeline.stage4_params)
            elif choice == '5':
                pipeline.stage_5_classify_vector()
            elif choice == '6':
                pipeline.stage_6_mask_class(force_recompute=True)
            elif choice == '7':
                pipeline.stage_7_mask_confidence(force_recompute=True)
            elif choice == '8':
                pipeline.stage_8_calculate_metrics()
            elif choice == 'A':
                print(
                    "\nNOTE: Running 'A' will automatically force recomputation of Stages 5-8 to clear any corrupted old files.")
                pipeline.stage_0_generate_footprint()
                pipeline.stage_1_segmentation(**pipeline.stage1_params)
                pipeline.stage_2_split_samples(**pipeline.stage2_params)
                pipeline.stage_3_selection()
                pipeline.stage_4_train_classifier(**pipeline.stage4_params)
                # Force recompute inference and masking
                pipeline.stage_5_classify_vector(force_recompute=True)
                pipeline.stage_6_mask_class(force_recompute=True)
                pipeline.stage_7_mask_confidence(force_recompute=True)
                if pipeline.metrics_fp.exists(): pipeline.metrics_fp.unlink()
                pipeline.stage_8_calculate_metrics()
            elif choice == 'Q':
                break
        except Exception as e:
            print(f"\n--- ERROR ---: {e}")
            import traceback
            traceback.print_exc()


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Modular OBIA Pipeline (True Object-Based Training)")
    parser.add_argument('--track', required=True, help="Processing track name (e.g. NL/orbit_88 or PT/orbit_161)")
    parser.add_argument('--mask_variant', default='3class',
                        choices=['3class', 'allcrops'],
                        help="Agricultural mask variant: '3class' (spring/winter/rapeseed, default) "
                             "or 'allcrops' (all crops including permanent ones)")
    parser.add_argument('--seg_mode', default='sam',
                        choices=['sam', 'lpis', 'otb_meanshift', 'felzenszwalb', 'slic'],
                        help="Initial segmentation mode/suffix (default: 'sam'). "
                             "Determines which files (features, model, results) are processed.")
    args = parser.parse_args()

    try:
        pipeline = ProcessingPipeline(track=args.track, mask_variant=args.mask_variant, seg_mode=args.seg_mode)
        main_menu(pipeline)
    except Exception as e:
        print(f"Initialization Error: {e}")
        sys.exit(1)