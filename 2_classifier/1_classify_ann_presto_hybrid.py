# How to run the script:
# python 1_classify_ann_presto_hybrid.py --track NL/orbit_88 --seg_mode slic
# python 1_classify_ann_presto_hybrid.py --track PT/orbit_161 --seg_mode slic

import os
import argparse
from pathlib import Path
import subprocess
import sys
# Add project root to sys.path to allow importing single_file_presto
project_root = str(Path(__file__).resolve().parent.parent)
if project_root not in sys.path:
    sys.path.insert(0, project_root)
import shutil
import shlex
import geopandas as gpd
import numpy as np
import pandas as pd
from osgeo import gdal, ogr, osr, gdalconst
from sklearn.metrics import confusion_matrix, precision_recall_fscore_support
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.utils import resample
import joblib
import openpyxl
from openpyxl.styles import Font
from pyogrio import read_info, read_dataframe
import threading
from concurrent.futures import ThreadPoolExecutor
import json
import math
import re

# Try importing PyTorch
try:
    import torch
    import torch.nn as nn
    import torch.optim as optim
    from torch.utils.data import DataLoader, TensorDataset
    HAS_TORCH = True
except ImportError:
    HAS_TORCH = False
    print("WARNING: PyTorch not found. Please install it to use Presto-SAR.")

# Try importing scikit-image
try:
    from skimage.segmentation import felzenszwalb, slic
    from skimage.util import img_as_float
    HAS_SKIMAGE = True
except ImportError:
    HAS_SKIMAGE = False
    print("WARNING: scikit-image not found.")

# Global Paths
base_dir = Path("D:/AIML_CropMapper_Cloud/workingDir")
aux_dir = Path("D:/AIML_CropMapper_Cloud/auxiliary_files")
presto_dir = aux_dir / "Presto_models"

TOTAL_STAGES = 8


def parse_month_from_description(desc):
    """Parses the month from GDAL band descriptions (e.g. 'Sigma0_VH_18Oct2024_db') to 0-11 index."""
    months_map = {
        'jan': 0, 'feb': 1, 'mar': 2, 'apr': 3, 'may': 4, 'jun': 5,
        'jul': 6, 'aug': 7, 'sep': 8, 'oct': 9, 'nov': 10, 'dec': 11
    }
    match = re.search(r'_(?:\d+)?([a-zA-Z]{3})\d{4}_', desc)
    if match:
        mon_str = match.group(1).lower()
        return months_map.get(mon_str, 0)
    return 0


def get_crop_aggregation(country, learn_shp_path):
    """
    Returns an empty dictionary because the grassland-related classes (Clover/Lucerne)
    have been directly integrated into Grassland in the shapefile itself.
    """
    return {}


def _get_priors_for_country(country, learn_shp_path, classes, class_counts, total_samples, priors_file_override=None):
    # Try custom JSON override first
    priors_json_path = None
    if priors_file_override and os.path.exists(priors_file_override):
        priors_json_path = Path(priors_file_override)
    else:
        # Resolve project root relative to learn_shp_path
        if learn_shp_path:
            p = Path(learn_shp_path).resolve()
            for parent in p.parents:
                aux_dir_path = parent / 'auxiliary_files'
                if aux_dir_path.exists():
                    aux_priors = aux_dir_path / 'shapefiles_samples' / country / 'priors.json'
                    if aux_priors.exists():
                        priors_json_path = aux_priors
                        break
            
            # Fallback to track_dir/priors.json
            if not priors_json_path:
                track_dir = Path(learn_shp_path).parent.parent
                track_priors = track_dir / f"priors_{country}.json"
                if not track_priors.exists():
                    track_priors = track_dir / "priors.json"
                if track_priors.exists():
                    priors_json_path = track_priors

    if priors_json_path and priors_json_path.exists():
        try:
            with open(priors_json_path, 'r') as f:
                custom_priors = json.load(f)
            print(f"    Loaded name-based priors from {priors_json_path}")
            
            # Read crop names from shapefile
            id_to_name = {}
            if learn_shp_path and os.path.exists(learn_shp_path):
                try:
                    import geopandas as gpd
                    gdf = gpd.read_file(str(learn_shp_path), engine="pyogrio")
                    if 'crop_id' in gdf.columns and 'crop_name' in gdf.columns:
                        id_to_name = dict(zip(gdf['crop_id'].astype(int), gdf['crop_name'].astype(str).str.lower()))
                except Exception as e:
                    print(f"    [WARNING] Could not read crop names from shapefile: {e}")
            
            # Build priors map dynamically
            raw_priors = {}
            sorted_keys = sorted(custom_priors.keys(), key=len, reverse=True)
            for cid, name in id_to_name.items():
                matched_val = 1e-5
                for key in sorted_keys:
                    if key.lower() in name or name in key.lower():
                        matched_val = float(custom_priors[key])
                        break
                raw_priors[cid] = matched_val
                
            for key, val in custom_priors.items():
                try:
                    cid = int(key)
                    raw_priors[cid] = float(val)
                except ValueError:
                    pass
            
            # Apply aggregation
            crop_aggregation = get_crop_aggregation(country, learn_shp_path)
            aggregated_priors = {}
            for cid, val in raw_priors.items():
                mapped_cid = crop_aggregation.get(cid, cid)
                aggregated_priors[mapped_cid] = aggregated_priors.get(mapped_cid, 0.0) + val
                
            p_true = np.array([aggregated_priors.get(c, 1e-5) for c in classes])
            p_true = p_true / np.sum(p_true)
            return p_true
        except Exception as e:
            print(f"    [WARNING] Failed to load name-based priors: {e}")

    # NL specific priors fallback
    if country == 'NL':
        real_priors_nl = {
            1: 0.0030,   # Asparagus
            2: 0.0034,   # Beans
            3: 0.0087,   # Beets
            4: 0.0058,   # Carrots
            5: 0.0033,   # Chicory
            6: 0.0023,   # Flax
            7: 0.0060,   # Flower Bulbs
            8: 0.7270,   # Grassland
            9: 0.0847,   # Maize
            10: 0.0015,  # Oats
            11: 0.0073,  # Onions
            12: 0.0044,  # Peas
            13: 0.0775,  # Potatoes
            14: 0.0149,  # Spring Barley
            15: 0.0007,  # Spring Wheat
            16: 0.0178,  # Winter Barley
            17: 0.0255   # Winter Wheat
        }
        p_true = np.array([real_priors_nl.get(c, 1e-5) for c in classes])
        p_true = p_true / np.sum(p_true)
        return p_true

    # PL & any other country - dynamic estimation using keyword-based field sizes
    id_to_name = {}
    if learn_shp_path and os.path.exists(learn_shp_path):
        try:
            import geopandas as gpd
            gdf = gpd.read_file(str(learn_shp_path), engine="pyogrio")
            if 'crop_id' in gdf.columns and 'crop_name' in gdf.columns:
                id_to_name = dict(zip(gdf['crop_id'].astype(int), gdf['crop_name'].astype(str)))
        except Exception as e:
            print(f"    [WARNING] Could not read crop names from shapefile: {e}")

    def get_area_multiplier(cid):
        name = id_to_name.get(cid, '').lower()
        if not name:
            area_thresholds_pl = {
                1: 3000, 2: 5000, 3: 5000, 4: 40000, 5: 5000, 6: 10000, 7: 5000, 8: 30000, 9: 15000,
                10: 20000, 11: 30000, 12: 40000, 13: 5000, 14: 70000, 15: 2000, 16: 30000, 17: 5000,
                18: 3000, 19: 40000, 20: 2000, 21: 40000, 22: 2000, 23: 10000, 24: 25000, 25: 70000,
                26: 10000, 27: 50000, 28: 2000, 29: 60000, 30: 3000, 31: 15000, 32: 70000, 33: 5000,
                34: 3000, 35: 5000, 36: 20000, 37: 40000
            }
            return area_thresholds_pl.get(cid, 10000)
            
        if any(k in name for k in ['grass', 'tiuz', 'grassland', 'pasture', 'trawa', 'trawiast', 'blijvend', 'tijdelijk', 'permanent', 'clover', 'klaver', 'lucerne', 'luzerne']):
            return 70000
        if any(k in name for k in ['maize', 'mais', 'kukurydza', 'corn']):
            return 70000
        if any(k in name for k in ['wheat', 'pszenica', 'tarwe']):
            return 70000
        if any(k in name for k in ['barley', 'jeczmien', 'gerst']):
            return 40000
        if any(k in name for k in ['rye', 'zyto', 'rogge']):
            return 40000
        if any(k in name for k in ['triticale', 'pszenzyto', 'koorn']):
            return 50000
        if any(k in name for k in ['oats', 'owies', 'haver']):
            return 40000
        if any(k in name for k in ['rapeseed', 'rzepak', 'koolzaad']):
            return 60000
        if any(k in name for k in ['sugar beet', 'burak', 'suikerbiet']):
            return 40000
        if any(k in name for k in ['fallow', 'braak', 'ugor']):
            return 40000
            
        if any(k in name for k in ['potato', 'ziemniak', 'aardappel']):
            return 20000
        if any(k in name for k in ['orchard', 'fruit', 'sad', 'appel', 'peer', 'jablon', 'sliwa', 'wisnia']):
            return 20000
        if any(k in name for k in ['pea', 'groch', 'erwt']):
            return 30000
        if any(k in name for k in ['bean', 'fasola', 'boon']):
            return 10000
        if any(k in name for k in ['aronia', 'blueberry', 'borowka', 'currant', 'porzeczka', 'bessen']):
            return 10000
        if any(k in name for k in ['nursery', 'ornamental', 'szkolka', 'sier', 'boomkwekerij']):
            return 10000
            
        if any(k in name for k in ['onion', 'cebula', 'ui']):
            return 5000
        if any(k in name for k in ['strawberry', 'truskawka', 'aardbei']):
            return 5000
        if any(k in name for k in ['cabbage', 'brassica', 'kapusta', 'kool']):
            return 5000
        if any(k in name for k in ['carrot', 'marchew', 'peen']):
            return 3000
        if any(k in name for k in ['tomato', 'pomidor', 'tomaat']):
            return 2000
        if any(k in name for k in ['cucumber', 'ogorek', 'komkommer']):
            return 2000
        if any(k in name for k in ['tobacco', 'tyton', 'tabak']):
            return 3000
        if any(k in name for k in ['raspberry', 'blackberry', 'malina', 'jezyna', 'framboos']):
            return 5000
        return 10000

    area_multipliers = np.array([get_area_multiplier(c) for c in classes])
    counts_arr = np.array([class_counts.get(c, 0) for c in classes])
    true_area_dist = counts_arr * area_multipliers
    p_true = true_area_dist / (np.sum(true_area_dist) + 1e-9)
    return p_true


def _calculate_class_weights(y_data, all_classes):
    classes_in_data = np.unique(y_data)
    total_samples = len(y_data)
    n_classes = len(classes_in_data)
    weight_vector = np.ones(len(all_classes))
    
    for c in classes_in_data:
        count = np.sum(y_data == c)
        if count > 0:
            weight = total_samples / (n_classes * count)
            idx = np.where(all_classes == c)[0][0]
            weight_vector[idx] = math.sqrt(weight)
            
    return weight_vector


class TorchMLPClassifier:
    def __init__(self, hidden_layer_sizes=(256, 128, 64), max_iter=120, batch_size=256, lr=0.001, class_weights=None, all_classes=None):
        self.hidden_layer_sizes = hidden_layer_sizes
        self.max_iter = max_iter
        self.batch_size = batch_size
        self.lr = lr
        self.class_weights = class_weights
        self.all_classes = all_classes
        self.device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
        self.model = None
        self.le = None
        self.classes_ = None
        self.input_dim = None
        self.output_dim = None

    def fit(self, X, y):
        self.le = LabelEncoder()
        if self.all_classes is not None:
            self.le.fit(self.all_classes)
        else:
            self.le.fit(y)
        
        y_enc = self.le.transform(y)
        self.classes_ = self.le.classes_
        
        self.input_dim = X.shape[1]
        self.output_dim = len(self.classes_)

        layers = []
        in_dim = self.input_dim
        for h_dim in self.hidden_layer_sizes:
            layers.append(nn.Linear(in_dim, h_dim))
            layers.append(nn.ReLU())
            layers.append(nn.BatchNorm1d(h_dim))
            layers.append(nn.Dropout(0.3))
            in_dim = h_dim
        layers.append(nn.Linear(in_dim, self.output_dim))

        self.model = nn.Sequential(*layers).to(self.device)
        
        if self.class_weights is not None:
            weights_tensor = torch.tensor(self.class_weights, dtype=torch.float32).to(self.device)
            criterion = nn.CrossEntropyLoss(weight=weights_tensor)
        else:
            criterion = nn.CrossEntropyLoss()
            
        optimizer = optim.Adam(self.model.parameters(), lr=self.lr, weight_decay=1e-4)

        X_tensor = torch.tensor(X, dtype=torch.float32)
        y_tensor = torch.tensor(y_enc, dtype=torch.long)

        dataset = TensorDataset(X_tensor, y_tensor)
        loader = DataLoader(dataset, batch_size=self.batch_size, shuffle=True)

        self.model.train()
        for epoch in range(self.max_iter):
            for batch_X, batch_y in loader:
                batch_X = batch_X.to(self.device)
                batch_y = batch_y.to(self.device)

                optimizer.zero_grad()
                outputs = self.model(batch_X)
                loss = criterion(outputs, batch_y)
                loss.backward()
                optimizer.step()

        return self

    def predict(self, X):
        self.model.eval()
        X_tensor = torch.tensor(X, dtype=torch.float32)
        dataset = TensorDataset(X_tensor)
        loader = DataLoader(dataset, batch_size=1024, shuffle=False)
        predictions = []

        with torch.no_grad():
            for batch_X in loader:
                batch_X = batch_X[0].to(self.device)
                outputs = self.model(batch_X)
                _, predicted = torch.max(outputs.data, 1)
                predictions.append(predicted.cpu().numpy())

        return self.le.inverse_transform(np.concatenate(predictions))

    def predict_proba(self, X):
        self.model.eval()
        X_tensor = torch.tensor(X, dtype=torch.float32)
        dataset = TensorDataset(X_tensor)
        loader = DataLoader(dataset, batch_size=1024, shuffle=False)
        probabilities = []

        with torch.no_grad():
            for batch_X in loader:
                batch_X = batch_X[0].to(self.device)
                outputs = self.model(batch_X)
                probs = torch.softmax(outputs, dim=1)
                probabilities.append(probs.cpu().numpy())

        return np.concatenate(probabilities, axis=0)

    def __getstate__(self):
        state = self.__dict__.copy()
        if self.model is not None:
            self.model.cpu()
            state['model_state_dict'] = self.model.state_dict()
            state['model'] = None
        return state

    def __setstate__(self, state):
        self.__dict__.update(state)
        if 'model_state_dict' in state and state['model_state_dict'] is not None:
            layers = []
            in_dim = self.input_dim
            for h_dim in self.hidden_layer_sizes:
                layers.append(nn.Linear(in_dim, h_dim))
                layers.append(nn.ReLU())
                layers.append(nn.BatchNorm1d(h_dim))
                layers.append(nn.Dropout(0.3))
                in_dim = h_dim
            layers.append(nn.Linear(in_dim, self.output_dim))
            self.model = nn.Sequential(*layers)
            self.model.load_state_dict(state['model_state_dict'])
            self.model.to(self.device)


def sam_worker(tile_info, ras_path, footprint_path, params):
    try:
        import os
        import sys
        import time
        from osgeo import gdal
        import numpy as np
        import torch
        torch.set_num_threads(1)  # Force single thread in worker to prevent OpenMP collisions
        import cv2
        cv2.setNumThreads(0)      # Disable cv2 multi-threading
        
        from samgeo import SamGeo
        from skimage.util import img_as_float
        from scipy.ndimage import distance_transform_edt
        import scipy.ndimage as ndimage
        
        x, y, xsize_valid, ysize_valid, x_start_buf, y_start_buf, xsize_buf, ysize_buf, buffer = tile_info
        print(f"    [Worker x={x}, y={y}] Started reading tile data...", flush=True)
        
        ds = gdal.Open(ras_path, gdal.GA_ReadOnly)
        nbands = ds.RasterCount
        
        img_list = []
        for b in range(1, nbands + 1):
            band = ds.GetRasterBand(b)
            arr = band.ReadAsArray(x_start_buf, y_start_buf, xsize_buf, ysize_buf)
            if arr is None:
                print(f"    [Worker x={x}, y={y}] Failed to read band {b}!", flush=True)
                return x, y, None, None
            arr = np.nan_to_num(arr)
            img_list.append(arr)
            
        img = np.dstack(img_list)
        
        ds_foot = None
        if footprint_path and os.path.exists(footprint_path):
            ds_foot = gdal.Open(footprint_path, gdal.GA_ReadOnly)
            valid_mask_buf = ds_foot.GetRasterBand(1).ReadAsArray(x_start_buf, y_start_buf, xsize_buf, ysize_buf) > 0
            valid_mask = valid_mask_buf
        else:
            valid_mask = np.sum(np.abs(img), axis=2) > 0
            
        if not np.any(valid_mask):
            print(f"    [Worker x={x}, y={y}] Tile contains no active pixels.", flush=True)
            return x, y, None, None
            
        print(f"    [Worker x={x}, y={y}] Loading SAM model...", flush=True)
        t_model_start = time.time()
        sam_geo = SamGeo(
            model_type=params['sam_model_type'],
            checkpoint=params['sam_checkpoint'],
            device=params['sam_device'],
            sam_kwargs={
                "points_per_side": params.get('points_per_side', 16),
                "pred_iou_thresh": params.get('pred_iou_thresh', 0.45),
                "stability_score_thresh": params.get('stability_score_thresh', 0.50),
                "crop_n_layers": params.get('crop_n_layers', 0),
                "crop_n_points_downscale_factor": params.get('crop_n_points_downscale_factor', 1),
                "min_mask_region_area": params.get('min_mask_region_area', 20),
                "box_nms_thresh": params.get('box_nms_thresh', 0.6)
            }
        )
        print(f"    [Worker x={x}, y={y}] SAM model loaded in {time.time() - t_model_start:.2f}s.", flush=True)
        
        img_8bit = np.zeros(img.shape, dtype=np.uint8)
        valid_pixels = valid_mask[:, :, np.newaxis]
        
        if np.any(valid_pixels):
            p2, p98 = np.percentile(img[valid_pixels], (2, 98))
            img_clip = np.clip(img, p2, p98)
            if p98 > p2:
                img_8bit[valid_pixels] = ((img_clip[valid_pixels] - p2) / (p98 - p2) * 255).astype(np.uint8)
                
            img_chan = np.ascontiguousarray(img_8bit[:, :, 0])
            img_smoothed = cv2.bilateralFilter(img_chan, d=9, sigmaColor=12, sigmaSpace=30)
            img_8bit[:, :, 0] = img_smoothed
            
            clahe_limit = params.get('clahe_limit', 0.0)
            if clahe_limit > 0.0:
                clahe = cv2.createCLAHE(clipLimit=clahe_limit, tileGridSize=(8,8))
                img_clahe = clahe.apply(img_8bit[:, :, 0])
                img_8bit[:, :, 0] = img_clahe
                
        if img_8bit.shape[2] == 1:
            img_rgb = np.repeat(img_8bit, 3, axis=2)
        else:
            img_rgb = img_8bit[:, :, :3]
            if img_rgb.shape[2] < 3:
                img_rgb = np.pad(img_rgb, ((0,0),(0,0),(0, 3-img_rgb.shape[2])), mode='constant')
                
        print(f"    [Worker x={x}, y={y}] Running SAM generate (points_per_side={params.get('points_per_side', 16)})...", flush=True)
        t_gen_start = time.time()
        sam_geo.generate(
            source=img_rgb,
            output=None,
            foreground=False,
            unique=True,
            min_size=10,
            max_size=100000
        )
        print(f"    [Worker x={x}, y={y}] SAM generate finished in {time.time() - t_gen_start:.2f}s.", flush=True)
        segments_buf = sam_geo.objects.astype(np.int32)
        
        zero_mask_buf = (segments_buf == 0) & valid_mask
        if np.any(zero_mask_buf) and np.any(segments_buf > 0):
            _, indices = distance_transform_edt(segments_buf == 0, return_indices=True)
            segments_buf[zero_mask_buf] = segments_buf[tuple(indices)][zero_mask_buf]
            
        segments_buf[~valid_mask] = 0
        
        median_size = params.get('median_size', 3)
        if median_size > 0:
            segments_buf = ndimage.median_filter(segments_buf, size=median_size)
            segments_buf[~valid_mask] = 0
            
        y_offset = y - y_start_buf
        x_offset = x - x_start_buf
        segments_buf[~valid_mask] = 0
        
        valid_mask_crop = valid_mask[y_offset : y_offset + ysize_valid, x_offset : x_offset + xsize_valid]
        segments = segments_buf[y_offset : y_offset + ysize_valid, x_offset : x_offset + xsize_valid]
        segments[~valid_mask_crop] = 0
        
        print(f"    [Worker x={x}, y={y}] Tile completed successfully.", flush=True)
        return x, y, segments, valid_mask_crop
    except Exception as e:
        print(f"Error in worker process for tile (x={tile_info[0]}, y={tile_info[1]}): {e}", flush=True)
        return tile_info[0], tile_info[1], None, None


class ProcessingPipeline:
    def __init__(self, track, seg_mode='sam'):
        self.track = track
        self.seg_mode = seg_mode.lower()
        normalized_track = track.replace('\\', '/')
        if '/' in normalized_track:
            self.country = normalized_track.split('/')[0].upper()
        elif len(track) == 2:
            self.country = track.upper()
        else:
            print(f"Error: Track '{track}' does not contain country code.")
            sys.exit(1)

        self.total_stages = TOTAL_STAGES
        print(f"Initializing Presto-SAR Hybrid pipeline for Track: {self.track}, Country: {self.country}, Mode: {self.seg_mode}")

        self.sanitized_track = self.track.replace('/', '_').replace('\\', '_')
        if self.sanitized_track.upper().startswith(self.country.upper() + "_"):
            self.file_prefix = self.sanitized_track
        else:
            self.file_prefix = f"{self.country}_{self.sanitized_track}"

        # --- 1. Define paths ---
        self.base_dir = base_dir
        self.aux_dir = aux_dir
        self.proc_dir = self.base_dir / self.track / 'processed_raster'
        self.out_dir = self.base_dir / self.track / 'classification_results'
        self.samples_dir = self.out_dir / 'samples'
        self.model_dir = self.out_dir / 'train_model'
        self.seg_dir = self.out_dir / 'segmentation'
        self.class_dir = self.out_dir / 'classification'

        self._ensure_directories()

        # --- 2. Resolve input raster ---
        search_patterns = [
            f"{self.sanitized_track}_*_VH_VV*.tif",
            f"*_{self.sanitized_track}_*_VH_VV*.tif",
            f"*{self.sanitized_track}*.tif",
            f"{self.track}_*_VH_VV*.tif",
            f"*{self.track}*.tif",
        ]

        self.hdr = None
        if self.proc_dir.exists():
            for pattern in search_patterns:
                self.hdr = next(self.proc_dir.glob(pattern), None)
                if self.hdr:
                    break

            if not self.hdr:
                raise FileNotFoundError(f"No raster file (TIF) found for track {self.track} in {self.proc_dir}")

            self.ras = self.hdr
            print(f"Input raster found: {self.ras}")
        else:
            raise FileNotFoundError(f"Processing directory does not exist: {self.proc_dir}")

        # Samples
        samples_base = self.aux_dir / 'shapefiles_samples'
        candidate_paths = [
            samples_base / self.file_prefix / "samples.shp",
            samples_base / f"{self.country}_{self.sanitized_track}" / "samples.shp",
            samples_base / f"{self.country}_{self.track}" / "samples.shp",
            samples_base / self.sanitized_track / "samples.shp",
            samples_base / self.track / "samples.shp",
            samples_base / self.country / "samples.shp",
            samples_base / "samples.shp"
        ]

        self.sample_shp = None
        for p in candidate_paths:
            if p.exists():
                self.sample_shp = p
                print(f"Training samples found at: {self.sample_shp}")
                break

        if not self.sample_shp:
            print(f"\nCRITICAL WARNING: Could not find 'samples.shp' inside {samples_base}")
            self.sample_shp = samples_base / self.file_prefix / "samples.shp"

        self.learn_shp = None
        self.control_shp = None
        self.sel_csv = None
        self.class_tif = None
        self.conf_tif = None
        self.masked_class = None
        self.masked_conf = None
        self.metrics_fp = None
        self.model_pkl = None

        self.stage1_params = {}
        self.stage2_params = {
            'learn_frac': 0.7,
            'random_state': 42
        }
        self.stage4_params = {
            'sk_hidden_sizes': '256,128,64',
            'sk_max_iter': 120,
            'balance_threshold': 1000
        }

        # Map seg_mode to method_name
        mode_to_method = {
            'sam': 'python_sam',
            'slic': 'python_slic',
            'lpis': 'lpis'
        }
        method_name = mode_to_method.get(self.seg_mode, 'python_sam')

        self.footprint_mask = self.seg_dir / f"{self.file_prefix}_data_footprint.tif"
        self.use_footprint_mask = True
        self.update_paths(method_name)
        self.agri_mask = self._resolve_agri_mask()

    def _ensure_directories(self):
        for d in [self.out_dir, self.samples_dir, self.model_dir, self.seg_dir, self.class_dir]:
            d.mkdir(parents=True, exist_ok=True)

    def _resolve_agri_mask(self):
        raster_dir = self.aux_dir / 'raster_files'
        country_dir = raster_dir / 'AgriMasks' / self.country
        mask_3class = country_dir / f"{self.country}_agri_mask_3class_epsg3857.tif"
        mask_allcrops = country_dir / f"{self.country}_agri_mask_allcrops_epsg3857.tif"

        if mask_allcrops.exists():
            print(f"    [OK] Agricultural Mask selected: {mask_allcrops.name}")
            return mask_allcrops
        elif mask_3class.exists():
            print(f"    [OK] Agricultural Mask selected: {mask_3class.name}")
            return mask_3class
        else:
            print(f"    [WARNING] No agricultural mask found for country '{self.country}'.")
            return None

    def update_paths(self, method_name):
        suffix_mapping = {
            'python_sam': 'sam',
            'lpis': 'lpis',
            'python_slic': 'slic'
        }
        
        self.seg_suffix = suffix_mapping.get(method_name, 'sam')
        self.seg_mode = 'lpis' if method_name == 'lpis' else 'sam'
        
        suffix = f"_{self.seg_suffix}"
        
        self.seg_tif = self.seg_dir / f"{self.file_prefix}_segmentation{suffix}.tif"
        self.seg_shp = self.seg_dir / f"{self.file_prefix}_segmentation{suffix}.sqlite"
        self.learn_shp = self.samples_dir / f"learn_{self.seg_suffix}.shp"
        self.control_shp = self.samples_dir / f"control_{self.seg_suffix}.shp"
        self.sel_csv = self.samples_dir / f"{self.file_prefix}_presto_hybrid_learn_features{suffix}.csv"
        self.class_tif = self.class_dir / f"{self.file_prefix}_presto_hybrid_classified{suffix}.tif"
        self.conf_tif = self.class_dir / f"{self.file_prefix}_presto_hybrid_confidence_map{suffix}.tif"
        self.masked_class = self.class_dir / f"{self.file_prefix}_presto_hybrid_classified_masked{suffix}.tif"
        self.masked_conf = self.class_dir / f"{self.file_prefix}_presto_hybrid_confidence_masked{suffix}.tif"
        self.metrics_fp = self.class_dir / f"{self.file_prefix}_presto_hybrid_metrics{suffix}.xlsx"
        self.model_pkl = self.model_dir / f"{self.file_prefix}_presto_hybrid_model{suffix}.pkl"
        
        print(f"    [INFO] Switched segmentation mode to: {self.seg_mode.upper()} (Suffix: {self.seg_suffix})")
        print(f"    Updated output paths accordingly (e.g. classification output: {self.class_tif.name})")

        self.stage1_params['method'] = method_name
        if method_name == 'python_sam':
            self.stage1_params.setdefault('tile_size', 1024)
            self.stage1_params.setdefault('buffer', 64)
            self.stage1_params.setdefault('sam_checkpoint', str(self.aux_dir / 'SAM_models' / 'sam_vit_b_01ec64.pth'))
            self.stage1_params.setdefault('sam_model_type', 'vit_b')
            self.stage1_params.setdefault('sam_device', 'cuda' if (HAS_TORCH and torch.cuda.is_available()) else 'cpu')
            self.stage1_params.setdefault('points_per_side', 16)
            self.stage1_params.setdefault('crop_n_layers', 0)
        elif method_name == 'python_slic':
            self.stage1_params.setdefault('tile_size', 2048)
            self.stage1_params.setdefault('buffer', 64)
            self.stage1_params.setdefault('n_segments', 32000)  # Original: 15000, Tuned: 32000
            self.stage1_params.setdefault('compactness', 0.05)  # Original: 0.1, Tuned: 0.05
            self.stage1_params.setdefault('slic_sigma', 1.5)    # Original: 1.5, Tuned: 1.5

    # --- Stage 0: Robust Data Footprint ---
    def stage_0_generate_footprint(self, force_recompute=False):
        stage = 0
        if self.footprint_mask.exists() and not force_recompute:
            print("[Stage 0] Data footprint mask already exists, skipping.")
            return

        print(f"[Stage 0/{self.total_stages}] Generating robust data footprint mask from radar stack...")
        ds = gdal.Open(str(self.ras))
        cols = ds.RasterXSize
        rows = ds.RasterYSize
        nbands = ds.RasterCount
        gt = ds.GetGeoTransform()
        proj = ds.GetProjection()

        driver = gdal.GetDriverByName('GTiff')
        out_ds = driver.Create(str(self.footprint_mask), cols, rows, 1, gdal.GDT_Byte,
                               options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        out_ds.SetGeoTransform(gt)
        out_ds.SetProjection(proj)
        out_band = out_ds.GetRasterBand(1)

        tile_size = 2048
        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                valid_mask = np.ones((ysize, xsize), dtype=bool)
                for b in range(1, min(nbands + 1, 5)):
                    arr = ds.GetRasterBand(b).ReadAsArray(x, y, xsize, ysize)
                    arr = np.nan_to_num(arr)
                    valid_mask &= (arr != 0) & (~np.isnan(arr))

                out_arr = np.where(valid_mask, 1, 0).astype(np.uint8)
                out_band.WriteArray(out_arr, x, y)

        out_band.FlushCache()
        out_ds = None
        ds = None
        print(f"    Footprint mask saved to {self.footprint_mask}\n")

    def _create_summed_composite(self):
        print("    [INFO] Creating a log-domain (dB) summed composite of all SAR bands...")
        gdal.SetCacheMax(4 * 1024 * 1024 * 1024)

        composite_tif = self.seg_dir / f"{self.file_prefix}_summed_composite.tif"
        if composite_tif.exists():
            print(f"    [INFO] Summed composite already exists at {composite_tif}.")
            return composite_tif

        ds = gdal.Open(str(self.ras))
        cols = ds.RasterXSize
        rows = ds.RasterYSize
        nbands = ds.RasterCount
        gt = ds.GetGeoTransform()
        proj = ds.GetProjection()

        driver = gdal.GetDriverByName('GTiff')
        out_ds = driver.Create(str(composite_tif), cols, rows, 1, gdal.GDT_Float32,
                               options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        out_ds.SetGeoTransform(gt)
        out_ds.SetProjection(proj)
        out_band = out_ds.GetRasterBand(1)
        out_band.SetNoDataValue(0)

        tile_size = 4096
        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                sum_arr = np.zeros((ysize, xsize), dtype=np.float32)
                valid_mask = np.zeros((ysize, xsize), dtype=bool)

                for b in range(1, nbands + 1):
                    band = ds.GetRasterBand(b)
                    arr = band.ReadAsArray(x, y, xsize, ysize)
                    nodata = band.GetNoDataValue()
                    
                    if nodata is not None:
                        mask = (arr != nodata) & (~np.isnan(arr)) & (arr != 0)
                    else:
                        mask = (~np.isnan(arr)) & (arr != 0)
                        
                    sum_arr[mask] += arr[mask]
                    valid_mask |= mask

                sum_arr[~valid_mask] = 0
                out_band.WriteArray(sum_arr, x, y)

        out_ds.FlushCache()
        out_ds = None
        ds = None
        print(f"    [INFO] Summed composite saved to {composite_tif}")
        return composite_tif

    # --- Stage 1: Segmentation (LPIS, SLIC, or SAM) ---
    def stage_1_segmentation(self, force_recompute=False, **kwargs):
        stage = 1
        if self.seg_tif.exists() and not force_recompute:
            print(f"[Stage {stage}/{self.total_stages}] Segmentation Raster already exists, skipping.\n")
            return

        self._ensure_directories()

        if self.seg_mode == 'lpis':
            print(f"[Stage {stage}/{self.total_stages}] Running LPIS (Parcel Boundary) Rasterization...")
            lpis_dir = self.aux_dir / 'shapefiles_samples' / self.country
            lpis_candidates = list(lpis_dir.glob("*.gpkg")) + list(lpis_dir.glob("*.shp"))
            lpis_candidates = [p for p in lpis_candidates if p.name not in ['samples.shp', 'learn.shp', 'control.shp', 'samples_all.shp']]
            
            if not lpis_candidates:
                raise FileNotFoundError(f"No LPIS vector file (.gpkg/.shp) found in {lpis_dir}")
            
            lpis_file = lpis_candidates[0]
            print(f"    LPIS file selected: {lpis_file}")
            
            ds_ras = gdal.Open(str(self.ras))
            cols = ds_ras.RasterXSize
            rows = ds_ras.RasterYSize
            gt = ds_ras.GetGeoTransform()
            proj = ds_ras.GetProjection()
            
            minx = gt[0]
            maxy = gt[3]
            maxx = minx + gt[1] * cols
            miny = maxy + gt[5] * rows
            
            try:
                from pyogrio import read_info, read_dataframe
                info = read_info(str(lpis_file))
                lpis_crs = info.get('crs')
            except ImportError:
                import geopandas as gpd
                gdf_temp = gpd.read_file(str(lpis_file), rows=1)
                lpis_crs = gdf_temp.crs.to_string() if gdf_temp.crs else None
                info = {'fid_column': None}
                
            srs_target = osr.SpatialReference()
            srs_target.ImportFromWkt(proj)
            target_epsg = srs_target.GetAttrValue("AUTHORITY", 1) or "3857"
            
            from pyproj import Transformer
            print(f"    LPIS CRS: {lpis_crs}")
            
            transformer = Transformer.from_crs(f"EPSG:{target_epsg}", lpis_crs, always_xy=True)
            p1 = transformer.transform(minx, miny)
            p2 = transformer.transform(maxx, maxy)
            lpis_bbox = (min(p1[0], p2[0]), min(p1[1], p2[1]), max(p1[0], p2[0]), max(p1[1], p2[1]))
            
            print(f"    Querying LPIS with spatial filter bbox: {lpis_bbox}")
            try:
                gdf = read_dataframe(str(lpis_file), bbox=lpis_bbox)
            except ImportError:
                import geopandas as gpd
                gdf = gpd.read_file(str(lpis_file), bbox=lpis_bbox)
                
            print(f"    Loaded {len(gdf)} intersecting parcels. Reprojecting to EPSG:{target_epsg}...")
            gdf_target = gdf.to_crs(f"EPSG:{target_epsg}")
            
            fid_col = info.get('fid_column') if isinstance(info, dict) else None
            if fid_col and fid_col in gdf_target.columns:
                id_col = fid_col
            elif 'id' in gdf_target.columns:
                id_col = 'id'
            elif 'id_0' in gdf_target.columns:
                id_col = 'id_0'
            else:
                id_col = None

            if id_col is None:
                gdf_target['lpis_id'] = np.arange(1, len(gdf_target) + 1)
                id_col = 'lpis_id'
            else:
                gdf_target[id_col] = pd.to_numeric(gdf_target[id_col], errors='coerce').fillna(0).astype(np.int32)
                if gdf_target[id_col].sum() == 0 or gdf_target[id_col].nunique() < len(gdf_target):
                    gdf_target['lpis_id'] = np.arange(1, len(gdf_target) + 1)
                    id_col = 'lpis_id'
            
            temp_gpkg = self.seg_dir / f"temp_lpis_{self.sanitized_track}.gpkg"
            gdf_target.geometry = gdf_target.geometry.force_2d()
            gdf_target.to_file(str(temp_gpkg), driver="GPKG", engine="pyogrio")
            
            driver = gdal.GetDriverByName("GTiff")
            ds_out = driver.Create(str(self.seg_tif), cols, rows, 1, gdal.GDT_Int32,
                                   options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
            ds_out.SetGeoTransform(gt)
            ds_out.SetProjection(proj)
            
            band = ds_out.GetRasterBand(1)
            band.SetNoDataValue(0)
            band.Fill(0)
            
            print(f"    Rasterizing parcels to {self.seg_tif.name} (burning column '{id_col}')...")
            gdal.Rasterize(ds_out, str(temp_gpkg), attribute=id_col)
            
            ds_out.FlushCache()
            ds_out = None
            ds_ras = None
            
            if os.path.exists(temp_gpkg):
                os.remove(temp_gpkg)
            print(f"Completed stage {stage}: LPIS rasterized.\n")
            return

        original_ras = self.ras
        try:
            self.ras = self._create_summed_composite()
        except Exception as e:
            print(f"    [WARNING] Failed to create summed composite: {e}. Falling back to full stack.")
            
        if self.seg_suffix == 'slic':
            print(f"[Stage {stage}/{self.total_stages}] Running SLIC superpixels segmentation...")
            slic_params = {
                'method': 'python_slic',
                'tile_size': 2048,
                'buffer': 64,
                'n_segments': 32000,  # Original: 15000, Tuned: 32000
                'compactness': 0.05,  # Original: 0.1, Tuned: 0.05
                'slic_sigma': 1.5     # Original: 1.5, Tuned: 1.5
            }
            self._run_python_segmentation_tiled(slic_params, stage, 'python_slic')
        else:
            print(f"[Stage {stage}/{self.total_stages}] Running SAM-Geo segmentation...")
            self._run_python_segmentation_tiled(self.stage1_params, stage, 'python_sam')
            
        self.ras = original_ras

    def _run_python_segmentation_tiled(self, params, stage, method):
        print(f"    Running Tiled Python Segmentation ({method})...")
        try:
            ds = gdal.Open(str(self.ras))
            ds_foot = None
            if self.footprint_mask.exists():
                ds_foot = gdal.Open(str(self.footprint_mask))

            cols = ds.RasterXSize
            rows = ds.RasterYSize
            nbands = ds.RasterCount
            gt = ds.GetGeoTransform()
            proj = ds.GetProjection()

            driver = gdal.GetDriverByName('GTiff')
            out_ds = driver.Create(str(self.seg_tif), cols, rows, 1, gdal.GDT_Int32,
                                   options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
            out_ds.SetGeoTransform(gt)
            out_ds.SetProjection(proj)
            out_band = out_ds.GetRasterBand(1)
            out_band.SetNoDataValue(0)

            tile_size = params.get('tile_size', 2048)
            buffer = params.get('buffer', 128)
            global_seg_id = 1

            if method == 'python_sam':
                # Run in parallel using ProcessPoolExecutor
                from concurrent.futures import ProcessPoolExecutor, as_completed
                
                # Gather active tiles first
                tile_tasks = []
                for y in range(0, rows, tile_size):
                    for x in range(0, cols, tile_size):
                        xsize_valid = min(tile_size, cols - x)
                        ysize_valid = min(tile_size, rows - y)
                        x_start_buf = max(0, x - buffer)
                        y_start_buf = max(0, y - buffer)
                        x_end_buf = min(cols, x + xsize_valid + buffer)
                        y_end_buf = min(rows, y + ysize_valid + buffer)
                        xsize_buf = x_end_buf - x_start_buf
                        ysize_buf = y_end_buf - y_start_buf
                        
                        # Read footprint chunk to see if active
                        if ds_foot:
                            band_foot = ds_foot.GetRasterBand(1)
                            arr_foot = band_foot.ReadAsArray(x_start_buf, y_start_buf, xsize_buf, ysize_buf)
                            valid_mask = arr_foot > 0 if arr_foot is not None else None
                        else:
                            band = ds.GetRasterBand(1)
                            arr = band.ReadAsArray(x_start_buf, y_start_buf, xsize_buf, ysize_buf)
                            valid_mask = np.nan_to_num(arr) > 0 if arr is not None else None
                            
                        if valid_mask is not None and np.any(valid_mask):
                            tile_tasks.append((x, y, xsize_valid, ysize_valid, x_start_buf, y_start_buf, xsize_buf, ysize_buf, buffer))
                            
                total_tasks = len(tile_tasks)
                print(f"    Total active tiles to process with SAM: {total_tasks}")
                
                max_workers = 8  # Use 8 processes on 16-core CPU
                print(f"    Processing tiles in parallel using ProcessPoolExecutor (Workers: {max_workers})...")
                
                with ProcessPoolExecutor(max_workers=max_workers) as executor:
                    futures = {
                        executor.submit(sam_worker, task, str(self.ras), str(self.footprint_mask) if self.footprint_mask.exists() else None, params): task
                        for task in tile_tasks
                    }
                    
                    completed_count = 0
                    for future in as_completed(futures):
                        x, y, segments, valid_mask_crop = future.result()
                        completed_count += 1
                        print(f"    [Tile Finished] x={x}, y={y} | Progress: {completed_count}/{total_tasks} tiles ({completed_count/total_tasks*100:.1f}%)")
                            
                        if segments is not None:
                            seg_valid_mask = segments > 0
                            unique_segs = np.unique(segments[seg_valid_mask])
                            if len(unique_segs) > 0:
                                max_seg = segments.max()
                                mapping = np.zeros(max_seg + 1, dtype=np.int32)
                                mapping[unique_segs] = np.arange(global_seg_id, global_seg_id + len(unique_segs))
                                segments = mapping[segments]
                                segments[~valid_mask_crop] = 0
                                global_seg_id += len(unique_segs)
                            else:
                                segments[~valid_mask_crop] = 0
                            out_band.WriteArray(segments.astype(np.int32), x, y)
                        else:
                            # Write empty array for inactive/failed tiles
                            for task in tile_tasks:
                                if task[0] == x and task[1] == y:
                                    xsize_valid, ysize_valid = task[2], task[3]
                                    out_band.WriteArray(np.zeros((ysize_valid, xsize_valid), dtype=np.int32), x, y)
                                    break
            else:
                # Original sequential logic for SLIC
                for y in range(0, rows, tile_size):
                    for x in range(0, cols, tile_size):
                        xsize_valid = min(tile_size, cols - x)
                        ysize_valid = min(tile_size, rows - y)
                        x_start_buf = max(0, x - buffer)
                        y_start_buf = max(0, y - buffer)
                        x_end_buf = min(cols, x + xsize_valid + buffer)
                        y_end_buf = min(rows, y + ysize_valid + buffer)
                        xsize_buf = x_end_buf - x_start_buf
                        ysize_buf = y_end_buf - y_start_buf

                        img_list = []
                        for b in range(1, nbands + 1):
                            band = ds.GetRasterBand(b)
                            arr = band.ReadAsArray(x_start_buf, y_start_buf, xsize_buf, ysize_buf)
                            if arr is None:
                                img_list = None
                                break
                            arr = np.nan_to_num(arr)
                            img_list.append(arr)

                        if img_list is None:
                            continue

                        img = np.dstack(img_list)

                        if ds_foot:
                            valid_mask_buf = ds_foot.GetRasterBand(1).ReadAsArray(x_start_buf, y_start_buf, xsize_buf, ysize_buf) > 0
                            valid_mask = valid_mask_buf
                        else:
                            valid_mask = np.sum(np.abs(img), axis=2) > 0

                        if not np.any(valid_mask):
                            continue

                        img_norm = img_as_float(img)

                        from skimage.segmentation import slic
                        max_tile_pixels = (tile_size + 2 * buffer) ** 2
                        pixels_per_segment = max_tile_pixels / params['n_segments']
                        active_pixels = np.sum(valid_mask)
                        n_segments_dynamic = max(1, int(active_pixels / pixels_per_segment))
                        segments_buf = slic(img_norm, n_segments=n_segments_dynamic, compactness=params['compactness'],
                                            sigma=params['slic_sigma'], start_label=1, mask=valid_mask)

                        y_offset = y - y_start_buf
                        x_offset = x - x_start_buf
                        segments_buf[~valid_mask] = 0

                        valid_mask_crop = valid_mask[y_offset : y_offset + ysize_valid, x_offset : x_offset + xsize_valid]
                        segments = segments_buf[y_offset : y_offset + ysize_valid, x_offset : x_offset + xsize_valid]
                        
                        seg_valid_mask = segments > 0
                        unique_segs = np.unique(segments[seg_valid_mask])
                        if len(unique_segs) > 0:
                            max_seg = segments.max()
                            mapping = np.zeros(max_seg + 1, dtype=np.int32)
                            mapping[unique_segs] = np.arange(global_seg_id, global_seg_id + len(unique_segs))
                            segments = mapping[segments]
                            segments[~valid_mask_crop] = 0
                            global_seg_id += len(unique_segs)
                        else:
                            segments[~valid_mask_crop] = 0

                        out_band.WriteArray(segments.astype(np.int32), x, y)

            out_ds.FlushCache()
            out_ds = None
            if ds_foot:
                ds_foot = None
            print(f"    Segmentation Raster saved to {self.seg_tif}\n")
        except Exception as e:
            print(f"ERROR in Python segmentation: {e}")
            raise

    # --- Stage 2: Prepare Point Split ---
    def stage_2_split_samples(self, force_recompute=False, **kwargs):
        self._ensure_directories()
        params = self.stage2_params.copy()
        params.update(kwargs)
        stage = 2

        if not self.sample_shp.exists():
            print("ERROR: Input sample file not found.")
            return

        gdf = gpd.read_file(str(self.sample_shp), engine="pyogrio")

        if not self.seg_tif.exists():
            print("ERROR: Segmentation raster not found. Run Stage 1 first to do spatial-leakage-free split.")
            return

        print(f"    Aligning training samples with segmentation raster {self.seg_tif.name} to prevent spatial data leakage...")
        ds_seg = gdal.Open(str(self.seg_tif))
        gt = ds_seg.GetGeoTransform()
        inv_gt = gdal.InvGeoTransform(gt)
        raster_proj = ds_seg.GetProjection()
        cols = ds_seg.RasterXSize
        rows = ds_seg.RasterYSize
        seg_band = ds_seg.GetRasterBand(1)

        foot_band = None
        ds_foot = None
        if self.use_footprint_mask and self.footprint_mask.exists():
            print(f"    Filtering points by footprint mask {self.footprint_mask.name} to discard NoData areas...")
            ds_foot = gdal.Open(str(self.footprint_mask))
            foot_band = ds_foot.GetRasterBand(1)

        from pyproj import CRS
        if raster_proj and gdf.crs:
            target_crs = CRS.from_wkt(raster_proj)
            if gdf.crs != target_crs:
                print("    Reprojecting points to match segmentation CRS...")
                gdf = gdf.to_crs(target_crs)

        xs = gdf.geometry.x.values
        ys = gdf.geometry.y.values
        pxs = (inv_gt[0] + inv_gt[1] * xs + inv_gt[2] * ys).astype(int)
        pys = (inv_gt[3] + inv_gt[4] * xs + inv_gt[5] * ys).astype(int)

        seg_ids = []
        for px, py in zip(pxs, pys):
            if 0 <= px < cols and 0 <= py < rows:
                try:
                    if foot_band is not None:
                        is_active = foot_band.ReadAsArray(int(px), int(py), 1, 1)[0, 0] > 0
                        if not is_active:
                            seg_ids.append(0)
                            continue
                    
                    val = seg_band.ReadAsArray(int(px), int(py), 1, 1)[0, 0]
                    seg_ids.append(val)
                except:
                    seg_ids.append(0)
            else:
                seg_ids.append(0)

        gdf['seg_id'] = seg_ids
        ds_seg = None
        ds_foot = None

        gdf_valid = gdf[gdf['seg_id'] > 0].copy()
        dropped = len(gdf) - len(gdf_valid)
        if dropped > 0:
            print(f"    Warning: Dropped {dropped} points that fell outside valid segments.")

        if len(gdf_valid) == 0:
            print("ERROR: No points fell within any valid segments.")
            return

        unique_segs = gdf_valid['seg_id'].unique()
        print(f"    Found {len(unique_segs)} unique segments for {len(gdf_valid)} valid points.")

        np.random.seed(params['random_state'])
        np.random.shuffle(unique_segs)
        split_idx = int(len(unique_segs) * params['learn_frac'])
        train_segs = set(unique_segs[:split_idx])

        learn = gdf_valid[gdf_valid['seg_id'].isin(train_segs)].copy()
        control = gdf_valid[~gdf_valid['seg_id'].isin(train_segs)].copy()

        learn = learn.drop(columns=['seg_id'])
        control = control.drop(columns=['seg_id'])

        learn.to_file(str(self.learn_shp), engine="pyogrio")
        control.to_file(str(self.control_shp), engine="pyogrio")
        print(f"Completed stage {stage}. Total valid: {len(gdf_valid)}, Learn: {len(learn)}, Control: {len(control)}\n")

    # --- Stage 3: Feature Extraction (Hybrid Presto + Raw SAR) ---
    def stage_3_selection(self):
        stage = 3
        if self.sel_csv.exists():
            print(f"[Stage {stage}] Features already extracted, skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Extracting HYBRID (Raw SAR + Presto) features for training segments...")
        
        device = "cuda" if (HAS_TORCH and torch.cuda.is_available()) else "cpu"
        weights_path = presto_dir / "default_model.pt"
        
        if not weights_path.exists():
            raise FileNotFoundError(f"Presto model weights not found at {weights_path}")
            
        ds_tmp = gdal.Open(str(self.ras))
        if ds_tmp is None:
            raise FileNotFoundError(f"Raster not found: {self.ras}")
        nbands = ds_tmp.RasterCount
        num_dates = nbands // 2
        ds_tmp = None

        import single_file_presto
        model = single_file_presto.Presto.construct(max_sequence_length=max(24, num_dates))
        state_dict = torch.load(weights_path, map_location=device)
        state_dict.pop('encoder.pos_embed', None)
        state_dict.pop('decoder.pos_embed', None)
        model.load_state_dict(state_dict, strict=False)
        model.to(device)
        model.eval()
        
        if not self.learn_shp.exists():
            print("ERROR: Learn samples not found.")
            return

        gdf = gpd.read_file(str(self.learn_shp), engine="pyogrio")
        
        ds = gdal.Open(str(self.ras))
        gt = ds.GetGeoTransform()
        inv_gt = gdal.InvGeoTransform(gt)
        raster_proj = ds.GetProjection()
        cols = ds.RasterXSize
        rows = ds.RasterYSize
        nbands = ds.RasterCount
        num_dates = nbands // 2

        # Extract months from band descriptions
        months_list = []
        for b in range(1, num_dates + 1):
            desc = ds.GetRasterBand(b).GetDescription()
            months_list.append(parse_month_from_description(desc))
        month_tensor = torch.tensor(months_list).long().to(device)

        if raster_proj and gdf.crs:
            from pyproj import CRS
            target_crs = CRS.from_wkt(raster_proj)
            if gdf.crs != target_crs:
                print("    Reprojecting samples to Match Raster CRS...")
                gdf = gdf.to_crs(target_crs)

        seg_ds = gdal.Open(str(self.seg_tif))
        seg_band = seg_ds.GetRasterBand(1)
        seg_array = seg_band.ReadAsArray()
        
        gdf_wgs84 = gdf.to_crs("EPSG:4326")
        centroids_wgs84 = gdf_wgs84.geometry.centroid
        
        xs = gdf.geometry.x.values
        ys = gdf.geometry.y.values
        pxs = (inv_gt[0] + inv_gt[1] * xs + inv_gt[2] * ys).astype(int)
        pys = (inv_gt[3] + inv_gt[4] * xs + inv_gt[5] * ys).astype(int)
        crop_ids = gdf['crop_id'].values

        target_segments = {}
        segment_coords = {}
        for idx, (px, py, crop_id) in enumerate(zip(pxs, pys, crop_ids)):
            if 0 <= px < cols and 0 <= py < rows:
                seg_id = seg_array[py, px]
                if seg_id > 0:
                    target_segments[seg_id] = crop_id
                    lat = centroids_wgs84.iloc[idx].y
                    lon = centroids_wgs84.iloc[idx].x
                    segment_coords[seg_id] = (lat, lon)

        if not target_segments:
            print("ERROR: No valid samples found overlapping the raster.")
            return

        print(f"    Found {len(target_segments)} unique segments for training.")
        
        print("    Computing segment bounding boxes...")
        from scipy.ndimage import find_objects
        slices = find_objects(seg_array)
        
        features_list = []
        raw_means_list = []
        seg_ids_list = []
        crop_ids_list = []
        
        count = 0
        total = len(target_segments)
        
        batch_size = 64
        batch_x = []
        batch_latlons = []
        batch_seg_ids = []
        batch_raw_means = []
        
        for seg_id, crop_id in target_segments.items():
            if seg_id - 1 >= len(slices) or slices[seg_id - 1] is None:
                continue
                
            sl = slices[seg_id - 1]
            y_min, y_max = sl[0].start, sl[0].stop
            x_min, x_max = sl[1].start, sl[1].stop
            w = x_max - x_min
            h = y_max - y_min
            
            try:
                bands_data = []
                for b in range(1, nbands + 1):
                    band = ds.GetRasterBand(b)
                    arr = band.ReadAsArray(int(x_min), int(y_min), int(w), int(h))
                    arr = np.nan_to_num(arr)
                    bands_data.append(arr)
                bands_arr = np.stack(bands_data, axis=0)
                
                segment_mask_cropped = seg_array[int(y_min):int(y_max), int(x_min):int(x_max)]
                mask = (segment_mask_cropped == seg_id)
                
                if not np.any(mask):
                    continue

                # 1. Compute raw SAR band means
                raw_means = [np.mean(bands_arr[b][mask]) for b in range(nbands)]

                # 2. Compute Presto time-series profile
                temp_profile = np.zeros((num_dates, 17), dtype=np.float32)
                for d in range(num_dates):
                    vh_vals = bands_arr[d][mask]
                    vv_vals = bands_arr[num_dates + d][mask]
                    
                    mean_vh = np.mean(vh_vals) if len(vh_vals) > 0 else -15.0
                    mean_vv = np.mean(vv_vals) if len(vv_vals) > 0 else -10.0
                    
                    temp_profile[d, 0] = (mean_vv + 25.0) / 25.0
                    temp_profile[d, 1] = (mean_vh + 25.0) / 25.0
                    
                lat, lon = segment_coords.get(seg_id, (0.0, 0.0))
                
                batch_x.append(torch.from_numpy(temp_profile))
                batch_latlons.append(torch.tensor([lat, lon], dtype=torch.float32))
                batch_seg_ids.append(seg_id)
                batch_raw_means.append(raw_means)
            except Exception as e:
                continue
                
            if len(batch_x) == batch_size or (count + len(batch_x) == total):
                input_x = torch.stack(batch_x, dim=0).to(device)
                input_latlons = torch.stack(batch_latlons, dim=0).to(device)
                input_dw = torch.ones(input_x.shape[0], num_dates).long().to(device) * 9
                input_mask = torch.zeros_like(input_x, device=device).float()
                input_mask[:, :, 2:] = 1.0
                
                with torch.no_grad():
                    features = model.encoder(
                        x=input_x,
                        dynamic_world=input_dw,
                        latlons=input_latlons,
                        mask=input_mask,
                        month=month_tensor.unsqueeze(0).expand(input_x.shape[0], -1),
                        eval_task=True
                    )
                    pooled = features.cpu().numpy()
                    
                for idx, pooled_feat in enumerate(pooled):
                    curr_seg_id = batch_seg_ids[idx]
                    features_list.append(pooled_feat)
                    seg_ids_list.append(curr_seg_id)
                    crop_ids_list.append(target_segments[curr_seg_id])
                    raw_means_list.append(batch_raw_means[idx])
                    
                count += len(batch_x)
                sys.stdout.write(f"\r      Extracted features for {count}/{total} segments...  ")
                sys.stdout.flush()
                
                batch_x = []
                batch_latlons = []
                batch_seg_ids = []
                batch_raw_means = []
                
        print("\n    Aggregation complete. Saving features...")
        
        feature_data = {'crop_id': crop_ids_list, 'seg_id': seg_ids_list}
        
        # 1. Add raw SAR features
        raw_means_arr = np.array(raw_means_list)
        for b in range(nbands):
            feature_data[f'meanB{b}'] = raw_means_arr[:, b]

        # 2. Add Presto features
        features_arr = np.array(features_list)
        for f_idx in range(128):
            feature_data[f'feat{f_idx}'] = features_arr[:, f_idx]
            
        df_final = pd.DataFrame(feature_data)
        df_final.to_csv(self.sel_csv, index=False)
        print(f"    Presto Hybrid Features saved to {self.sel_csv}\n")
        
        ds = None
        seg_ds = None

    # --- Stage 4: Train Classifier ---
    def stage_4_train_classifier(self, **kwargs):
        self._ensure_directories()
        params = self.stage4_params.copy()
        params.update(kwargs)
        stage = 4

        if not self.sel_csv.exists():
            print("ERROR: Feature CSV not found.")
            return

        model_fn = self.model_pkl
        print(f"[Stage {stage}/{self.total_stages}] Training Classifier (ANN) on Hybrid Features...")

        df = pd.read_csv(self.sel_csv)
        local_samples = len(df)
        local_classes = df['crop_id'].nunique()

        feat_cols = [c for c in df.columns if c not in ['crop_id', 'seg_id']]
        self.feat_cols = feat_cols

        print("    Balancing classes (Capped Oversampling)...")
        threshold = params.get('balance_threshold', 1000)
        df_balanced = pd.DataFrame()
        for crop_id in df['crop_id'].unique():
            df_class = df[df['crop_id'] == crop_id]
            count = len(df_class)

            if count < threshold:
                df_resampled = resample(df_class, replace=True, n_samples=threshold, random_state=42)
                df_balanced = pd.concat([df_balanced, df_resampled])
            else:
                df_balanced = pd.concat([df_balanced, df_class])

        print(f"    Original samples: {len(df)}. Balanced samples: {len(df_balanced)}")

        X = df_balanced[feat_cols].values
        y = df_balanced['crop_id'].values

        if self.country == 'NL':
            print("    Applying crop aggregation for Netherlands training labels...")
            crop_aggregation = get_crop_aggregation(self.country, self.learn_shp)
            y = np.array([crop_aggregation.get(val, val) for val in y])

        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)

        all_classes = np.unique(y)
        class_weights = _calculate_class_weights(y, all_classes)

        hidden_sizes = tuple(map(int, str(params['sk_hidden_sizes']).split(',')))
        max_iter = params.get('sk_max_iter', 120)

        print(f"    Training TorchMLPClassifier on {X.shape[0]} samples with input dim {X.shape[1]}...")
        clf = TorchMLPClassifier(
            hidden_layer_sizes=hidden_sizes,
            max_iter=max_iter,
            batch_size=256,
            class_weights=class_weights,
            all_classes=all_classes
        )

        clf.fit(X_scaled, y)

        joblib.dump({'model': clf, 'scaler': scaler, 'feats': feat_cols}, model_fn)
        print(f"Model saved to {model_fn}")

        y_pred = clf.predict(X_scaled)
        labels = sorted(list(set(y)))
        cm = confusion_matrix(y, y_pred, labels=labels)
        print("\n--- Training Confusion Matrix ---")
        print(pd.DataFrame(cm, index=labels, columns=labels).to_string())
        print("\n")

    # --- Stage 5: Tiled Inference (Hybrid + ANN) ---
    def stage_5_classify_vector(self, force_recompute=False):
        self._ensure_directories()
        stage = 5

        model_file = self.model_pkl
        if not model_file.exists():
            print("ERROR: Model not found.")
            return

        if self.class_tif.exists() and not force_recompute:
            print(f"[Stage {stage}] Classification Raster exists, skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Running Object-Based Inference using Hybrid Classifier...")

        data = joblib.load(model_file)
        clf = data['model']
        scaler = data['scaler']
        
        device = "cuda" if (HAS_TORCH and torch.cuda.is_available()) else "cpu"
        weights_path = presto_dir / "default_model.pt"
        
        ds_tmp = gdal.Open(str(self.ras))
        if ds_tmp is None:
            raise FileNotFoundError(f"Raster not found: {self.ras}")
        nbands = ds_tmp.RasterCount
        num_dates = nbands // 2
        ds_tmp = None

        import single_file_presto
        model = single_file_presto.Presto.construct(max_sequence_length=max(24, num_dates))
        state_dict = torch.load(weights_path, map_location=device)
        state_dict.pop('encoder.pos_embed', None)
        state_dict.pop('decoder.pos_embed', None)
        model.load_state_dict(state_dict, strict=False)
        model.to(device)
        model.eval()
        
        # --- Calculate Bayesian Priors ---
        print("    Calculating Bayesian priors for Hybrid inference...")
        df_learn = pd.read_csv(self.sel_csv)
        y_train = df_learn['crop_id'].values
        if self.country == 'NL':
            print("    Applying crop aggregation for Netherlands priors...")
            crop_aggregation = get_crop_aggregation(self.country, self.learn_shp)
            y_train = np.array([crop_aggregation.get(val, val) for val in y_train])

        class_counts = pd.Series(y_train).value_counts().to_dict()
        classes = clf.classes_
        total_samples = len(df_learn)
        
        p_true = _get_priors_for_country(
            self.country, self.learn_shp, classes, class_counts, total_samples, 
            priors_file_override=self.base_dir / "priors.json"
        )
        
        threshold = 1000
        balanced_counts = {}
        for c in classes:
            count = len(df_learn[df_learn['crop_id'] == c])
            balanced_counts[c] = max(count, threshold)
        total_balanced = sum(balanced_counts.values())
        p_train = np.array([balanced_counts[c] / total_balanced for c in classes])
        
        correction = p_true / (p_train + 1e-9)
        correction = np.power(correction, 0.7)
        correction = np.clip(correction, 0.01, 10.0)
        priors_arr = correction / np.sum(correction)
        # ---------------------------------

        ds_stack_info = gdal.Open(str(self.ras))
        cols = ds_stack_info.RasterXSize
        rows = ds_stack_info.RasterYSize
        gt = ds_stack_info.GetGeoTransform()
        proj = ds_stack_info.GetProjection()
        nbands = ds_stack_info.RasterCount
        num_dates = nbands // 2
        ds_stack_info = None

        from pyproj import Transformer
        srs_ras = osr.SpatialReference()
        srs_ras.ImportFromWkt(proj)
        ras_epsg = srs_ras.GetAttrValue("AUTHORITY", 1) or "3857"
        transformer_to_wgs84 = Transformer.from_crs(f"EPSG:{ras_epsg}", "EPSG:4326", always_xy=True)

        driver = gdal.GetDriverByName('GTiff')
        ds_cls = driver.Create(str(self.class_tif), cols, rows, 1, gdal.GDT_Int32,
                               options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        ds_cls.SetGeoTransform(gt)
        ds_cls.SetProjection(proj)
        ds_cls.GetRasterBand(1).SetNoDataValue(0)

        ds_conf = driver.Create(str(self.conf_tif), cols, rows, 1, gdal.GDT_Float32,
                                options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        ds_conf.SetGeoTransform(gt)
        ds_conf.SetProjection(proj)
        ds_conf.GetRasterBand(1).SetNoDataValue(0)

        tile_size = 2048
        write_lock = threading.Lock()
        gpu_lock = threading.Lock()

        # Load entire segmentation database
        seg_ds = gdal.Open(str(self.seg_tif))
        seg_band = seg_ds.GetRasterBand(1)
        seg_array = seg_band.ReadAsArray()
        seg_ds = None
        
        print("    Computing full segmentation bounding boxes...")
        from scipy.ndimage import find_objects
        slices = find_objects(seg_array)

        # Extract months from band descriptions
        ds_desc = gdal.Open(str(self.ras))
        months_list = []
        for b in range(1, num_dates + 1):
            desc = ds_desc.GetRasterBand(b).GetDescription()
            months_list.append(parse_month_from_description(desc))
        month_tensor = torch.tensor(months_list).long().to(device)
        ds_desc = None

        def process_tile(x, y):
            xsize = min(tile_size, cols - x)
            ysize = min(tile_size, rows - y)

            ds_stack = gdal.Open(str(self.ras))
            ds_foot = gdal.Open(str(self.footprint_mask))

            try:
                sub_seg = seg_array[y:y+ysize, x:x+xsize]
                foot_arr = ds_foot.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                
                unique_ids = np.unique(sub_seg)
                unique_ids = unique_ids[unique_ids > 0]
                
                if len(unique_ids) == 0:
                    return

                features = []
                valid_raw_means = []
                valid_ids = []
                batch_x = []
                batch_latlons = []
                batch_ids = []
                batch_raw_means = []
                batch_size = 128
                
                for sid in unique_ids:
                    if sid - 1 >= len(slices) or slices[sid - 1] is None:
                        continue
                    sl = slices[sid - 1]
                    y_min, y_max = sl[0].start, sl[0].stop
                    x_min, x_max = sl[1].start, sl[1].stop
                    w = x_max - x_min
                    h = y_max - y_min
                    
                    try:
                        sub_seg_mask = seg_array[y_min:y_max, x_min:x_max] == sid
                        
                        if not np.any(sub_seg_mask):
                            continue

                        bands_data = []
                        for b in range(1, nbands + 1):
                            band = ds_stack.GetRasterBand(b)
                            arr = band.ReadAsArray(int(x_min), int(y_min), int(w), int(h))
                            arr = np.nan_to_num(arr)
                            bands_data.append(arr)
                        bands_arr = np.stack(bands_data, axis=0)
                        
                        # 1. Raw SAR Means
                        raw_means = [np.mean(bands_arr[b][sub_seg_mask]) for b in range(nbands)]

                        # 2. Presto profile
                        temp_profile = np.zeros((num_dates, 17), dtype=np.float32)
                        for d in range(num_dates):
                            vh_vals = bands_arr[d][sub_seg_mask]
                            vv_vals = bands_arr[num_dates + d][sub_seg_mask]
                            
                            mean_vh = np.mean(vh_vals) if len(vh_vals) > 0 else -15.0
                            mean_vv = np.mean(vv_vals) if len(vv_vals) > 0 else -10.0
                            
                            temp_profile[d, 0] = (mean_vv + 25.0) / 25.0
                            temp_profile[d, 1] = (mean_vh + 25.0) / 25.0
                            
                        cy = y_min + h / 2.0
                        cx = x_min + w / 2.0
                        mx = gt[0] + cx * gt[1] + cy * gt[2]
                        my = gt[3] + cx * gt[4] + cy * gt[5]
                        lon, lat = transformer_to_wgs84.transform(mx, my)
                        
                        batch_x.append(torch.from_numpy(temp_profile))
                        batch_latlons.append(torch.tensor([lat, lon], dtype=torch.float32))
                        batch_ids.append(sid)
                        batch_raw_means.append(raw_means)
                    except:
                        continue
                        
                    if len(batch_x) == batch_size:
                        input_x = torch.stack(batch_x, dim=0).to(device)
                        input_latlons = torch.stack(batch_latlons, dim=0).to(device)
                        input_dw = torch.ones(input_x.shape[0], num_dates).long().to(device) * 9
                        input_mask = torch.zeros_like(input_x, device=device).float()
                        input_mask[:, :, 2:] = 1.0
                        
                        with gpu_lock:
                            with torch.no_grad():
                                out = model.encoder(
                                    x=input_x,
                                    dynamic_world=input_dw,
                                    latlons=input_latlons,
                                    mask=input_mask,
                                    month=month_tensor.unsqueeze(0).expand(input_x.shape[0], -1),
                                    eval_task=True
                                )
                                pooled = out.cpu().numpy()
                        for idx, p_f in enumerate(pooled):
                            features.append(p_f)
                            valid_ids.append(batch_ids[idx])
                            valid_raw_means.append(batch_raw_means[idx])
                        batch_x = []
                        batch_latlons = []
                        batch_ids = []
                        batch_raw_means = []
                        
                if batch_x:
                    input_x = torch.stack(batch_x, dim=0).to(device)
                    input_latlons = torch.stack(batch_latlons, dim=0).to(device)
                    input_dw = torch.ones(input_x.shape[0], num_dates).long().to(device) * 9
                    input_mask = torch.zeros_like(input_x, device=device).float()
                    input_mask[:, :, 2:] = 1.0
                    
                    with gpu_lock:
                        with torch.no_grad():
                            out = model.encoder(
                                x=input_x,
                                dynamic_world=input_dw,
                                latlons=input_latlons,
                                mask=input_mask,
                                month=month_tensor.unsqueeze(0).expand(input_x.shape[0], -1),
                                eval_task=True
                            )
                            pooled = out.cpu().numpy()
                    for idx, p_f in enumerate(pooled):
                        features.append(p_f)
                        valid_ids.append(batch_ids[idx])
                        valid_raw_means.append(batch_raw_means[idx])
                        
                if not valid_ids:
                    return

                features_arr = np.stack(features, axis=0) # [n_valid, 128]
                raw_means_arr = np.stack(valid_raw_means, axis=0) # [n_valid, nbands]
                combined_features = np.concatenate([raw_means_arr, features_arr], axis=1) # [n_valid, nbands + 128]

                X_scaled = scaler.transform(combined_features)
                raw_probs = clf.predict_proba(X_scaled)
                
                corrected_probs = raw_probs * priors_arr
                corrected_probs = corrected_probs / np.sum(corrected_probs, axis=1, keepdims=True)
                
                preds = clf.classes_[np.argmax(corrected_probs, axis=1)]
                max_probs = np.max(corrected_probs, axis=1)

                id_to_pred = {sid: int(pred) for sid, pred in zip(valid_ids, preds)}
                id_to_prob = {sid: float(prob) for sid, prob in zip(valid_ids, max_probs)}

                pred_arr = np.zeros_like(sub_seg, dtype=np.int32)
                prob_arr = np.zeros_like(sub_seg, dtype=np.float32)

                for sid in unique_ids:
                    if sid in id_to_pred:
                        mask = (sub_seg == sid)
                        pred_arr[mask] = id_to_pred[sid]
                        prob_arr[mask] = id_to_prob[sid]

                pred_arr[foot_arr == 0] = 0
                prob_arr[foot_arr == 0] = 0

                with write_lock:
                    ds_cls.GetRasterBand(1).WriteArray(pred_arr, x, y)
                    ds_conf.GetRasterBand(1).WriteArray(prob_arr, x, y)

            except Exception as e:
                pass
            finally:
                ds_stack = None
                ds_foot = None

        tiles = []
        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                tiles.append((x, y))

        print(f"    Processing {len(tiles)} tiles sequentially (using all CPU cores via PyTorch internal threads)...")
        with ThreadPoolExecutor(max_workers=1) as executor:
            executor.map(lambda t: process_tile(*t), tiles)

        ds_cls.GetRasterBand(1).FlushCache()
        ds_conf.GetRasterBand(1).FlushCache()
        ds_cls = None
        ds_conf = None
        print("    Classification and Confidence map generated successfully!\n")

    # --- Stage 6: Apply Masks ---
    def stage_6_mask_classification(self, force_recompute=False):
        stage = 6
        if self.masked_class.exists() and self.masked_conf.exists() and not force_recompute:
            print("[Stage 6] Masked outputs already exist, skipping.")
            return

        print(f"[Stage 6/{self.total_stages}] Applying Arable & Data Footprint Mask...")
        ds_stack = gdal.Open(str(self.ras))
        cols = ds_stack.RasterXSize
        rows = ds_stack.RasterYSize
        gt = ds_stack.GetGeoTransform()
        proj = ds_stack.GetProjection()

        mask_tif = self.agri_mask
        ds_mask = None
        temp_mask_vrt = None
        if mask_tif and mask_tif.exists():
            print(f"    Warping country agricultural mask to match classification raster bounds...")
            minx = gt[0]
            maxy = gt[3]
            maxx = minx + gt[1] * cols
            miny = maxy + gt[5] * rows

            temp_mask_vrt = str(self.masked_class).replace('.tif', '_mask_temp.vrt')
            mask_opts = gdal.WarpOptions(
                format='VRT',
                outputBounds=(minx, miny, maxx, maxy),
                width=cols,
                height=rows,
                dstSRS=proj,
                resampleAlg=gdal.GRA_NearestNeighbour
            )
            ds_mask = gdal.Warp(temp_mask_vrt, str(mask_tif), options=mask_opts)
            if not ds_mask:
                raise RuntimeError("Failed to warp the arable mask.")
            print(f"    Applying country agricultural mask: {mask_tif.name}")
        else:
            print("    WARNING: Arable mask not found. Only applying data footprint mask.")

        if not self.class_tif.exists() or not self.conf_tif.exists():
            print("ERROR: Classification outputs not found. Run Stage 5 first.")
            if ds_mask:
                ds_mask = None
            if temp_mask_vrt and os.path.exists(temp_mask_vrt):
                os.remove(temp_mask_vrt)
            return

        ds_foot = gdal.Open(str(self.footprint_mask))
        ds_cls = gdal.Open(str(self.class_tif))
        ds_conf = gdal.Open(str(self.conf_tif))

        driver = gdal.GetDriverByName('GTiff')
        out_cls = driver.Create(str(self.masked_class), cols, rows, 1, gdal.GDT_Int32,
                                options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        out_cls.SetGeoTransform(gt)
        out_cls.SetProjection(proj)
        out_cls.GetRasterBand(1).SetNoDataValue(0)

        out_conf = driver.Create(str(self.masked_conf), cols, rows, 1, gdal.GDT_Float32,
                                 options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        out_conf.SetGeoTransform(gt)
        out_conf.SetProjection(proj)
        out_conf.GetRasterBand(1).SetNoDataValue(0)

        tile_size = 2048
        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                cls_arr = ds_cls.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                conf_arr = ds_conf.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                foot_arr = ds_foot.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)

                mask_arr = np.ones((ysize, xsize), dtype=np.uint8)
                if ds_mask:
                    mask_arr = ds_mask.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)

                combined_mask = (foot_arr > 0) & (mask_arr > 0)
                cls_arr[~combined_mask] = 0
                conf_arr[~combined_mask] = 0.0

                out_cls.GetRasterBand(1).WriteArray(cls_arr, x, y)
                out_conf.GetRasterBand(1).WriteArray(conf_arr, x, y)

        out_cls.GetRasterBand(1).FlushCache()
        out_conf.GetRasterBand(1).FlushCache()
        out_cls = None
        out_conf = None
        ds_mask = None
        ds_foot = None
        ds_cls = None
        ds_conf = None
        ds_stack = None
        
        if temp_mask_vrt and os.path.exists(temp_mask_vrt):
            try:
                os.remove(temp_mask_vrt)
            except Exception as e:
                print(f"Warning: Could not remove temp VRT: {e}")
        print("    Masking stage complete.\n")

    # --- Stage 7: Validation Metrics ---
    def stage_7_calculate_metrics(self):
        stage = 7
        print(f"[Stage {stage}/{self.total_stages}] Calculating Validation Metrics (OOB Control points)...")

        if not self.control_shp.exists():
            print("ERROR: Control samples not found.")
            return

        gdf = gpd.read_file(str(self.control_shp), engine="pyogrio")
        ds = gdal.Open(str(self.masked_class))
        gt = ds.GetGeoTransform()
        inv_gt = gdal.InvGeoTransform(gt)
        raster_proj = ds.GetProjection()

        if raster_proj and gdf.crs:
            from pyproj import CRS
            target_crs = CRS.from_wkt(raster_proj)
            if gdf.crs != target_crs:
                gdf = gdf.to_crs(target_crs)

        xs = gdf.geometry.x.values
        ys = gdf.geometry.y.values
        pxs = (inv_gt[0] + inv_gt[1] * xs + inv_gt[2] * ys).astype(int)
        pys = (inv_gt[3] + inv_gt[4] * xs + inv_gt[5] * ys).astype(int)
        
        true_labels = gdf['crop_id'].values
        if self.country == 'NL':
            print("    Applying crop aggregation for Netherlands validation labels...")
            crop_aggregation = get_crop_aggregation(self.country, self.control_shp)
            true_labels = np.array([crop_aggregation.get(val, val) for val in true_labels])

        pred_labels = []
        valid_indices = []

        band = ds.GetRasterBand(1)
        for i, (px, py) in enumerate(zip(pxs, pys)):
            if 0 <= px < ds.RasterXSize and 0 <= py < ds.RasterYSize:
                val = int(band.ReadAsArray(int(px), int(py), 1, 1)[0, 0])
                if val > 0:
                    if self.country == 'NL':
                        crop_aggregation = get_crop_aggregation(self.country, self.control_shp)
                        val = crop_aggregation.get(val, val)
                    pred_labels.append(val)
                    valid_indices.append(i)

        ds = None
        true_labels = true_labels[valid_indices]

        if len(pred_labels) == 0:
            print("ERROR: No valid control points overlap the classified raster mask.")
            return

        print(f"    Extracted {len(pred_labels)} validation points overlapping the mask.")
        
        labels = sorted(list(set(true_labels) | set(pred_labels)))
        cm = confusion_matrix(true_labels, pred_labels, labels=labels)
        precisions, recalls, f1s, _ = precision_recall_fscore_support(
            true_labels, pred_labels, labels=labels, zero_division=0
        )

        total = np.sum(cm)
        oa = np.trace(cm) / total
        sum_po = oa
        sum_pe = np.sum(np.sum(cm, axis=0) * np.sum(cm, axis=1)) / (total ** 2)
        kappa = (sum_po - sum_pe) / (1 - sum_pe) if (1 - sum_pe) != 0 else np.nan

        resx, resy = abs(gt[1]), abs(gt[5])
        area_ha = resx * resy / 10000

        # Read the classified raster for area calculations
        ds_metrics = gdal.Open(str(self.masked_class))
        band_metrics = ds_metrics.GetRasterBand(1)
        arr = band_metrics.ReadAsArray()
        ds_metrics = None

        if arr is not None:
            unique_classes, counts = np.unique(arr[arr > 0], return_counts=True)
            class_areas = dict(zip(unique_classes, counts))
            areas = [{'Class': c, 'Area_ha': round(class_areas.get(c, 0) * area_ha, 2)} for c in labels]
        else:
            areas = [{'Class': c, 'Area_ha': 0} for c in labels]

        wb = openpyxl.Workbook()
        sh = wb.active
        sh.title = 'Results'

        sh.cell(row=1, column=1, value='Confusion Matrix').font = Font(bold=True)
        sh.cell(row=2, column=1, value='True \\ Pred').font = Font(bold=True)
        for j, lbl in enumerate(labels, start=2):
            sh.cell(row=2, column=j, value=lbl).font = Font(bold=True)
        for i, lbl in enumerate(labels, start=3):
            sh.cell(row=i, column=1, value=lbl).font = Font(bold=True)
            for j, _ in enumerate(labels):
                sh.cell(row=i, column=j + 2, value=int(cm[i - 3, j]))

        # Compute weighted overall accuracy
        weighted_oa = None
        try:
            df_train = pd.read_csv(self.sel_csv)
            y_train = df_train['crop_id'].values
            if self.country == 'NL':
                crop_aggregation = get_crop_aggregation(self.country, self.learn_shp)
                y_train = np.array([crop_aggregation.get(val, val) for val in y_train])
            
            train_classes = sorted(list(set(y_train)))
            train_class_counts = pd.Series(y_train).value_counts()
            train_total_samples = len(y_train)
            
            p_true_all = _get_priors_for_country(
                country=self.country,
                learn_shp_path=self.learn_shp,
                classes=train_classes,
                class_counts=train_class_counts,
                total_samples=train_total_samples
            )
            p_true_dict = dict(zip(train_classes, p_true_all))
            
            active_priors = np.array([p_true_dict.get(lbl, 1e-5) for lbl in labels])
            active_priors = active_priors / np.sum(active_priors)
            
            weighted_oa = np.sum(active_priors * recalls)
            print(f"    Weighted Overall Accuracy ({self.country} Area-Adjusted): {weighted_oa:.4f}")
        except Exception as e:
            print(f"    [WARNING] Could not compute Weighted Overall Accuracy: {e}")

        base = 4 + len(labels)
        sh.cell(row=base, column=1, value='Overall Accuracy').font = Font(bold=True)
        sh.cell(row=base, column=2, value=round(oa, 4))
        sh.cell(row=base + 1, column=1, value='Kappa').font = Font(bold=True)
        sh.cell(row=base + 1, column=2, value=round(kappa, 4))

        if weighted_oa is not None:
            sh.cell(row=base + 2, column=1, value='Weighted Overall Accuracy').font = Font(bold=True)
            sh.cell(row=base + 2, column=2, value=round(weighted_oa, 4))
            start = base + 4
        else:
            start = base + 3
        headers = ['Class', 'Producer Acc (Recall)', 'User Acc (Precision)', 'F1-score']
        for j, h in enumerate(headers, start=1):
            sh.cell(row=start, column=j, value=h).font = Font(bold=True)
        for idx, c in enumerate(labels):
            row_idx = start + 1 + idx
            sh.cell(row=row_idx, column=1, value=c)
            sh.cell(row=row_idx, column=2, value=round(recalls[idx], 4))
            sh.cell(row=row_idx, column=3, value=round(precisions[idx], 4))
            sh.cell(row=row_idx, column=4, value=round(f1s[idx], 4))

        ar0 = start + 1 + len(labels) + 1
        sh.cell(row=ar0, column=1, value='Areas (ha)').font = Font(bold=True)
        sh.cell(row=ar0 + 1, column=1, value='Class').font = Font(bold=True)
        sh.cell(row=ar0 + 1, column=2, value='Area_ha').font = Font(bold=True)
        for idx, a in enumerate(areas, start=ar0 + 2):
            sh.cell(row=idx, column=1, value=a['Class'])
            sh.cell(row=idx, column=2, value=a['Area_ha'])

        wb.save(str(self.metrics_fp))
        print(f"    Metrics successfully saved to {self.metrics_fp}")
        print(f"    Overall Accuracy: {oa:.4f}\n")


# --- Interactive Menu Helpers ---

SAM_MODELS = {
    '1': {'name': 'vit_b  (Small, ~375 MB, FAST,   ~2 GB VRAM - recommended for testing)',
           'model_type': 'vit_b',  'checkpoint': 'sam_vit_b_01ec64.pth',
           'url': 'https://dl.fbaipublicfiles.com/segment_anything/sam_vit_b_01ec64.pth'},
    '2': {'name': 'vit_l  (Medium, ~1.2 GB, MEDIUM, ~6 GB VRAM)',
           'model_type': 'vit_l',  'checkpoint': 'sam_vit_l_0b3195.pth',
           'url': 'https://dl.fbaipublicfiles.com/segment_anything/sam_vit_l_0b3195.pth'},
    '3': {'name': 'vit_h  (Huge,  ~2.4 GB, SLOW,   ~10 GB VRAM - highest accuracy)',
           'model_type': 'vit_h',  'checkpoint': 'sam_vit_h_4b8939.pth',
           'url': 'https://dl.fbaipublicfiles.com/segment_anything/sam_vit_h_4b8939.pth'},
}


def get_stage1_params_sam(param_dict):
    new_params = param_dict.copy()
    current_method = new_params.get('method', 'python_sam')
    
    print("\n=== SELECT SEGMENTATION METHOD ===")
    print(f"  Current method: {current_method.upper()}")
    print()
    print("  [1] Meta SAM (Deep learning, default) [python_sam]")
    print("  [2] SLIC algorithm on full raster [python_slic]")
    print("  [3] LPIS boundary rasterization (Cadastral vector data) [lpis]")
    print("  [Enter] Keep current method")
    
    choice = input("Choose option (1-3): ").strip()
    
    method_mapping = {
        '1': 'python_sam',
        '2': 'python_slic',
        '3': 'lpis'
    }
    
    if choice in method_mapping:
        new_params['method'] = method_mapping[choice]
        method = new_params['method']
        print(f"  Selected method: {method.upper()}")
        
        if method == 'python_sam':
            new_params.setdefault('tile_size', 1024)
            new_params.setdefault('buffer', 64)
            new_params.setdefault('sam_checkpoint', str(aux_dir / 'SAM_models' / 'sam_vit_b_01ec64.pth'))
            new_params.setdefault('sam_model_type', 'vit_b')
            new_params.setdefault('sam_device', 'cuda' if (HAS_TORCH and torch.cuda.is_available()) else 'cpu')
            new_params.setdefault('points_per_side', 16)
            new_params.setdefault('crop_n_layers', 0)
        elif method == 'python_slic':
            new_params.setdefault('tile_size', 2048)
            new_params.setdefault('buffer', 64)
            new_params.setdefault('n_segments', 32000)  # Original: 15000, Tuned: 32000
            new_params.setdefault('compactness', 0.05)  # Original: 0.1, Tuned: 0.05
            new_params.setdefault('slic_sigma', 1.5)    # Original: 1.5, Tuned: 1.5
            
    method = new_params.get('method', 'python_sam')
    
    if method == 'python_sam':
        current_type = new_params.get('sam_model_type', 'vit_b')
        current_ckpt = new_params.get('sam_checkpoint', 'sam_vit_b_01ec64.pth')
        print("\n  Available SAM models:")
        for k, v in SAM_MODELS.items():
            marker = " <-- current" if v['model_type'] == current_type else ""
            print(f"    [{k}] {v['name']}{marker}")
        print()
        
        sam_choice = input("Choose SAM model (1/2/3) or Enter to keep current: ").strip()
        if sam_choice in SAM_MODELS:
            selected = SAM_MODELS[sam_choice]
            new_params['sam_model_type'] = selected['model_type']
            
            sam_models_dir = aux_dir / 'SAM_models'
            ckpt_fn = selected['checkpoint']
            ckpt_path = sam_models_dir / ckpt_fn
            new_params['sam_checkpoint'] = str(ckpt_path)

            if not ckpt_path.exists():
                print(f"\n  [WARNING] Checkpoint weight file '{ckpt_fn}' does not exist in {sam_models_dir}!")
                print(f"  Download it from: {selected['url']}")
                proceed = input("  Continue anyway? (y/n) [n]: ").strip().lower()
                if proceed != 'y':
                    return None
            else:
                print(f"  [OK] Checkpoint weight file '{ckpt_fn}' found.")
                
    show_keys = []
    if method == 'python_sam':
        show_keys = ['tile_size', 'buffer', 'sam_device']
    elif method == 'python_slic':
        show_keys = ['tile_size', 'buffer', 'n_segments', 'compactness', 'slic_sigma']
    elif method == 'lpis':
        show_keys = []

    if method == 'lpis':
        print("\n  No configurable parameters for LPIS rasterization.")
        print("  It will use the cadastral vector files in the country folder.")
    else:
        print(f"\n--- Edit parameters for: {method.upper()} ---")
        for key in show_keys:
            val = new_params.get(key)
            new_val_str = input(f"  {key} [{val}]: ").strip()
            if new_val_str:
                try:
                    if isinstance(val, bool):
                        new_params[key] = new_val_str.lower() in ('true', '1', 'y', 'yes')
                    else:
                        new_params[key] = type(val)(new_val_str) if val is not None else new_val_str
                except ValueError:
                    print("    Invalid value. Keeping default.")

    print("\n--- APPROVED SEGMENTATION PARAMETERS ---")
    print(f"  Method: {method.upper()}")
    for key in show_keys:
        print(f"  {key}: {new_params.get(key)}")
    print("==========================================\n")
    return new_params


def get_params(param_dict):
    new_params = param_dict.copy()
    print("--- Current Parameters ---")
    for key, val in new_params.items():
        print(f"  {key}: {val}")

    if input("Change parameters? (y/n) [n]: ").lower() != 'y':
        return new_params

    for key, val in new_params.items():
        new_val_str = input(f"Enter new value for '{key}' [{val}]: ")
        if not new_val_str:
            continue
        try:
            original_type = type(val)
            new_params[key] = original_type(new_val_str)
        except ValueError:
            print(f"Invalid value. Keeping default {val}.")
    return new_params


def get_classifier_params(param_dict):
    new_params = param_dict.copy()
    print("--- Current Parameters ---")
    for key, val in new_params.items():
        print(f"  {key}: {val}")

    if input("Change parameters? (y/n) [n]: ").lower() != 'y':
        return new_params

    print(f"\n--- Setting parameters for Sklearn MLPClassifier ---")
    for key in ['sk_hidden_sizes', 'sk_max_iter', 'balance_threshold']:
        val = new_params[key]
        new_val_str = input(f"Enter new value for '{key}' [{val}]: ")
        if new_val_str:
            try:
                new_params[key] = type(val)(new_val_str)
            except ValueError:
                print(f"Invalid value.")
    return new_params


def main_menu(pipeline):
    while True:
        seg_method = pipeline.stage1_params.get('method', 'python_sam')
        seg_desc = "LPIS Cadastral Vector Rasterization" if pipeline.seg_mode == 'lpis' else f"SAR Segmentation ({seg_method})"
        
        menu = f"""
    --- Raster-Based OBIA Pipeline (Presto-SAR Hybrid) ---
    Track: {pipeline.track} ({pipeline.country})
    Segmentation Mode: {pipeline.seg_mode.upper()} ({seg_method})

    [0] Stage 0: Generate Data Footprint
    [1] Stage 1: {seg_desc}
    [2] Stage 2: Prepare Point Split
    [3] Stage 3: Extract Hybrid Features
    [4] Stage 4: Train Hybrid ANN Classifier
    [5] Stage 5: Run Inference (Object-based)
    [6] Stage 6: Apply Agricultural Mask
    [7] Stage 7: Calculate Validation Metrics

    [A] Run All Stages (Forces overwrite of Stages 5-7)
    [Q] Quit

    Enter your choice: """
        try:
            choice = input(menu).strip().upper()
        except (EOFError, KeyboardInterrupt):
            print("\nExiting interactive menu due to standard input disconnection or interruption.")
            break
        try:
            if choice == '0':
                pipeline.stage_0_generate_footprint(force_recompute=True)
            elif choice == '1':
                new_params = get_stage1_params_sam(pipeline.stage1_params)
                if new_params is None:
                    print("  Segmentation parameter setup cancelled.")
                    continue
                pipeline.stage1_params.update(new_params)
                pipeline.update_paths(pipeline.stage1_params['method'])
                pipeline.stage_1_segmentation(force_recompute=True)
            elif choice == '2':
                new_params = get_params(pipeline.stage2_params)
                pipeline.stage2_params.update(new_params)
                pipeline.stage_2_split_samples(force_recompute=True, **pipeline.stage2_params)
            elif choice == '3':
                pipeline.stage_3_selection()
            elif choice == '4':
                new_params = get_classifier_params(pipeline.stage4_params)
                pipeline.stage4_params.update(new_params)
                pipeline.stage_4_train_classifier(**pipeline.stage4_params)
            elif choice == '5':
                pipeline.stage_5_classify_vector(force_recompute=True)
            elif choice == '6':
                pipeline.stage_6_mask_classification(force_recompute=True)
            elif choice == '7':
                pipeline.stage_7_calculate_metrics()
            elif choice == 'A':
                print(
                    "\nNOTE: Running 'A' will automatically force recomputation of Stages 5-7 to clear any corrupted old files.")
                pipeline.stage_0_generate_footprint(force_recompute=False)
                pipeline.stage_1_segmentation(force_recompute=False)
                pipeline.stage_2_split_samples(force_recompute=False, **pipeline.stage2_params)
                pipeline.stage_3_selection()
                pipeline.stage_4_train_classifier(**pipeline.stage4_params)
                pipeline.stage_5_classify_vector(force_recompute=True)
                pipeline.stage_6_mask_classification(force_recompute=True)
                pipeline.stage_7_calculate_metrics()
            elif choice == 'Q':
                break
        except Exception as e:
            print(f"\n--- ERROR ---: {e}")
            import traceback
            traceback.print_exc()


def main():
    parser = argparse.ArgumentParser(description="Object-Based Hybrid Crop Classification using NASA Harvest Presto")
    parser.add_argument('--track', required=True, help="Track name, e.g. NL/orbit_88 or PT/orbit_161")
    parser.add_argument('--stage', default=None, help="Stage to run: 'A' (all), '0', '1', '2', '3', '4', '5', '6', '7'")
    parser.add_argument('--seg_mode', default='sam', choices=['sam', 'lpis', 'slic'], help="Segmentation mode: sam, lpis, slic (default: sam)")
    args = parser.parse_args()

    pipeline = ProcessingPipeline(args.track, seg_mode=args.seg_mode)
    
    choice = args.stage
    if choice is None:
        main_menu(pipeline)
    else:
        choice = choice.strip().upper()
        print(f"Running in automated mode. Selected Stage/Choice: {choice} (Mode: {pipeline.seg_mode.upper()})")
        if choice == 'A':
            pipeline.stage_0_generate_footprint(force_recompute=False)
            pipeline.stage_1_segmentation(force_recompute=False)
            pipeline.stage_2_split_samples(force_recompute=False, **pipeline.stage2_params)
            pipeline.stage_3_selection()
            pipeline.stage_4_train_classifier(**pipeline.stage4_params)
            pipeline.stage_5_classify_vector(force_recompute=True)
            pipeline.stage_6_mask_classification(force_recompute=True)
            pipeline.stage_7_calculate_metrics()
        elif choice == '0':
            pipeline.stage_0_generate_footprint(force_recompute=True)
        elif choice == '1':
            pipeline.stage_1_segmentation(force_recompute=True)
        elif choice == '2':
            pipeline.stage_2_split_samples(force_recompute=True, **pipeline.stage2_params)
        elif choice == '3':
            pipeline.stage_3_selection()
        elif choice == '4':
            pipeline.stage_4_train_classifier(**pipeline.stage4_params)
        elif choice == '5':
            pipeline.stage_5_classify_vector(force_recompute=True)
        elif choice == '6':
            pipeline.stage_6_mask_classification(force_recompute=True)
        elif choice == '7':
            pipeline.stage_7_calculate_metrics()
        elif choice == 'Q':
            sys.exit(0)
        else:
            print("Invalid choice. Exiting.")


if __name__ == '__main__':
    main()
