import argparse
import os
from pathlib import Path
import numpy as np
from osgeo import gdal
import geopandas as gpd
from sklearn.metrics import confusion_matrix, precision_recall_fscore_support
import openpyxl
from openpyxl.styles import Font

# Minimal mapping of track → country; extend as needed
TRACK_REGIONS = {
    'P1':  'AT', 'P1a': 'AT',
    'P2':  'IE', 'P2a': 'IE',
    'P3':  'NL',
    'P4':  'PT', 'P4a': 'PT',
    'P5':  'XX', 'P5a': 'XX', 'P5b': 'XX',
    # add more as you create new tracks…
}

def find_masked_files(base_dir: Path, tr: str, country: str):
    """
    Look for [{country}_{tr}_classified_masked.tif,
              {country}_{tr}_confidence_masked.tif]
    in either:
      - classification_results/classification/
      - classification_results/
    Returns (cls_fp, conf_fp) or (None, None).
    """
    candidates = [
        base_dir / tr / 'classification_results' / 'classification',
        base_dir / tr / 'classification_results'
    ]
    cls_name  = f"{country}_{tr}_classified_masked.tif"
    conf_name = f"{country}_{tr}_confidence_masked.tif"
    for folder in candidates:
        cls_fp  = folder / cls_name
        conf_fp = folder / conf_name
        if cls_fp.exists() and conf_fp.exists():
            return cls_fp, conf_fp
    return None, None

def discover_tracks(base_dir: Path, prefix: str):
    """
    Automatically discover subfolders under base_dir whose
    names start with `prefix` and exist in TRACK_REGIONS.
    Returns list of (tr, country, cls_fp, conf_fp).
    """
    tracks = []
    for sub in base_dir.iterdir():
        tr = sub.name
        if not tr.startswith(prefix):
            continue
        country = TRACK_REGIONS.get(tr)
        if country is None:
            continue
        cls_fp, conf_fp = find_masked_files(base_dir, tr, country)
        if cls_fp:
            tracks.append((tr, country, cls_fp, conf_fp))
    return tracks

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--track', required=True,
                        help='Base track prefix (e.g. P1)')
    args = parser.parse_args()
    prefix = args.track

    base_dir = Path(os.environ.get("AIML_WORKING_DIR", r"D:\AIML_CropMapper_Cloud\workingDir"))
    tracks   = discover_tracks(base_dir, prefix)
    if not tracks:
        raise FileNotFoundError(
            f"No valid classification/confidence files for tracks starting with {prefix}"
        )

    print(f"Discovered tracks: {[t for t,_,_,_ in tracks]}")

    # --- compute union extent & grid ---------------------------------------
    ds0   = gdal.Open(str(tracks[0][2]))
    proj  = ds0.GetProjection()
    gt0   = ds0.GetGeoTransform()
    resX, resY = gt0[1], abs(gt0[5])

    extents = []
    for _, _, cls_fp, _ in tracks:
        ds = gdal.Open(str(cls_fp))
        gt = ds.GetGeoTransform()
        c, r = ds.RasterXSize, ds.RasterYSize
        minX = gt[0]
        maxY = gt[3]
        maxX = gt[0] + c * gt[1]
        minY = gt[3] + r * gt[5]
        extents.append((minX, maxY, maxX, minY))

    minX = min(e[0] for e in extents)
    maxY = max(e[1] for e in extents)
    maxX = max(e[2] for e in extents)
    minY = min(e[3] for e in extents)

    cols = int(np.ceil((maxX - minX) / resX))
    rows = int(np.ceil((maxY - minY) / resY))
    gt_global = (minX, resX, 0, maxY, 0, -resY)

    print(f"Global mosaic: {cols} cols × {rows} rows")

    # --- warp and stack using VRTs to avoid OOM ------------------------------
    ds_cls_list = []
    ds_conf_list = []
    nodata_vals = []
    
    for tr, country, cls_fp, conf_fp in tracks:
        # classification VRT
        vrt_cls = gdal.Warp(
            '', str(cls_fp), format='VRT',
            width=cols, height=rows,
            outputBounds=(minX, minY, maxX, maxY),
            dstSRS=proj,
            resampleAlg=gdal.GRA_NearestNeighbour
        )
        ds_cls_list.append(vrt_cls)
        nodata_vals.append(vrt_cls.GetRasterBand(1).GetNoDataValue())

        # confidence VRT
        vrt_conf = gdal.Warp(
            '', str(conf_fp), format='VRT',
            width=cols, height=rows,
            outputBounds=(minX, minY, maxX, maxY),
            dstSRS=proj,
            resampleAlg=gdal.GRA_NearestNeighbour
        )
        ds_conf_list.append(vrt_conf)

    # --- Prepare output file ------------------------------------------------
    base_tr, base_country, _, _ = tracks[0]
    out_dir = base_dir / base_tr / 'classification_results'
    out_tif = out_dir / f"{base_country}_final_classification.tif"
    
    drv = gdal.GetDriverByName('GTiff')
    ds_out = drv.Create(str(out_tif), cols, rows, 1, gdal.GDT_Int32, 
                        options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
    ds_out.SetGeoTransform(gt_global)
    ds_out.SetProjection(proj)
    band_out = ds_out.GetRasterBand(1)
    band_out.SetNoDataValue(0)
    
    # --- Block by block processing ------------------------------------------
    gdal.SetCacheMax(4 * 1024 * 1024 * 1024)
    tile_size = 4096
    print("Merging tracks block-by-block to conserve memory...")
    
    for y in range(0, rows, tile_size):
        for x in range(0, cols, tile_size):
            xsize = min(tile_size, cols - x)
            ysize = min(tile_size, rows - y)
            
            c_stack = []
            cf_stack = []
            
            for i in range(len(tracks)):
                c_arr = ds_cls_list[i].GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                cf_arr = ds_conf_list[i].GetRasterBand(1).ReadAsArray(x, y, xsize, ysize).astype(np.float32)
                nod = nodata_vals[i]
                
                if nod is not None:
                    cf_arr[c_arr == nod] = np.nan
                    c_arr = np.where(c_arr == nod, 0, c_arr).astype(np.int32)
                    
                c_stack.append(c_arr)
                cf_stack.append(cf_arr)
                
            c_stack = np.stack(c_stack, axis=0)
            cf_stack = np.stack(cf_stack, axis=0)
            
            cf_stack[np.isnan(cf_stack)] = -np.inf
            idx = np.argmax(cf_stack, axis=0)
            final_block = np.take_along_axis(c_stack, idx[None,:,:], axis=0)[0]
            final_block[np.all(np.isneginf(cf_stack), axis=0)] = 0
            
            band_out.WriteArray(final_block, x, y)

    ds_out.FlushCache()
    
    # --- apply morphological sieve filter to restore objects -------------
    print("Applying Sieve Filter to remove slivers and isolated pixels...")
    # Using threshold of 50 pixels (approx field size depending on resolution)
    # 8-connectedness is generally preferred for diagonals
    gdal.SieveFilter(band_out, None, band_out, 50, 8, callback=None)
    ds_out.FlushCache()
    print(f"Merged classification saved: {out_tif}")

    # --- compute metrics & areas --------------------------------------------
    print("Calculating metrics and areas...")
    ctrl_shp = out_dir / 'samples' / 'control.shp'
    ctrl     = gpd.read_file(str(ctrl_shp))
    inv      = gdal.InvGeoTransform(gt_global)

    x_coords = ctrl.geometry.x.values
    y_coords = ctrl.geometry.y.values
    crop_ids = ctrl['crop_id'].values

    px_vals = (inv[0] + inv[1]*x_coords + inv[2]*y_coords).astype(int)
    py_vals = (inv[3] + inv[4]*x_coords + inv[5]*y_coords).astype(int)

    valid_mask = (px_vals >= 0) & (px_vals < cols) & (py_vals >= 0) & (py_vals < rows)

    valid_px = px_vals[valid_mask]
    valid_py = py_vals[valid_mask]
    valid_crop_ids = crop_ids[valid_mask].astype(int)
    
    # Read predictions locally per point to avoid full raster memory
    valid_preds = np.zeros(len(valid_px), dtype=int)
    for i in range(len(valid_px)):
        val = band_out.ReadAsArray(int(valid_px[i]), int(valid_py[i]), 1, 1)
        if val is not None:
            valid_preds[i] = val[0, 0]
            
    final_mask = (valid_crop_ids > 0) & (valid_preds > 0)

    true_vals = valid_crop_ids[final_mask].tolist()
    pred_vals = valid_preds[final_mask].tolist()
    # Guarantee all classes from control set are represented in the matrix, even if completely missed
    all_control_classes = set(ctrl['crop_id'].unique().astype(int))
    all_control_classes.discard(0)
    labels = sorted(list(all_control_classes.union(set(pred_vals))))

    cm     = confusion_matrix(true_vals, pred_vals, labels=labels)
    prec, rec, f1, _ = precision_recall_fscore_support(
        true_vals, pred_vals,
        labels=labels,
        average=None,
        zero_division=0
    )
    total = cm.sum()
    exp   = (cm.sum(axis=0) * cm.sum(axis=1)).sum() / (total**2)
    oa    = np.trace(cm) / total
    kappa = (oa - exp) / (1 - exp) if (1 - exp) else np.nan

    # Calculate areas block-by-block
    areas_counts = {}
    for y in range(0, rows, tile_size):
        for x in range(0, cols, tile_size):
            xsize = min(tile_size, cols - x)
            ysize = min(tile_size, rows - y)
            block = band_out.ReadAsArray(x, y, xsize, ysize)
            unique, counts = np.unique(block, return_counts=True)
            for u, c in zip(unique, counts):
                if u != 0:
                    areas_counts[u] = areas_counts.get(u, 0) + c

    resx, resy = abs(gt_global[1]), abs(gt_global[5])
    area_ha    = resx * resy / 10000
    areas      = [{
        'Class':   c,
        'Area_ha': round(areas_counts.get(c, 0) * area_ha, 2)
    } for c in labels]

    # --- write Excel report -------------------------------------------------
    xlsx = out_dir / f"{base_country}_final_metrics.xlsx"
    wb   = openpyxl.Workbook()
    sh   = wb.active
    sh.title = 'Results'

    # Confusion matrix table
    sh.cell(1,1,'Confusion Matrix').font = Font(bold=True)
    for j, lbl in enumerate(labels, start=2):
        sh.cell(2,j,lbl).font = Font(bold=True)
    for i, lbl in enumerate(labels, start=3):
        sh.cell(i,1,lbl).font = Font(bold=True)
        for j in range(len(labels)):
            sh.cell(i,j+2,int(cm[i-3,j]))

    # Overall accuracy & kappa
    r0 = 3 + len(labels)
    sh.cell(r0,1,'Overall Accuracy').font = Font(bold=True)
    sh.cell(r0,2,round(oa,2))
    sh.cell(r0+1,1,'Kappa').font          = Font(bold=True)
    sh.cell(r0+1,2,round(kappa,2))

    # Per‐class recall/precision/F1
    r1 = r0 + 3
    hdrs = ['Class','Producer Acc','User Acc','F1-score']
    for j, h in enumerate(hdrs, start=1):
        sh.cell(r1,j,h).font = Font(bold=True)
    for idx, c in enumerate(labels, start=r1+1):
        sh.cell(idx,1,c)
        sh.cell(idx,2,round(rec[idx-r1-1],2))
        sh.cell(idx,3,round(prec[idx-r1-1],2))
        sh.cell(idx,4,round(f1[idx-r1-1],2))

    # Area per class
    ra = r1 + 1 + len(labels) + 1
    sh.cell(ra,1,'Areas (ha)').font = Font(bold=True)
    sh.cell(ra+1,1,'Class').font   = Font(bold=True)
    sh.cell(ra+1,2,'Area_ha').font = Font(bold=True)
    for i, a in enumerate(areas, start=ra+2):
        sh.cell(i,1,a['Class'])
        sh.cell(i,2,a['Area_ha'])

    wb.save(str(xlsx))
    print(f"Final metrics saved: {xlsx}")

if __name__ == '__main__':
    main()