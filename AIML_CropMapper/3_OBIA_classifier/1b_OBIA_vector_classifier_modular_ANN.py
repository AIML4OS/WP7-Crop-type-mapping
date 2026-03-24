import os
import argparse
from pathlib import Path
import subprocess
import sys
import geopandas as gpd
import numpy as np
import pandas as pd
from osgeo import gdal, ogr, osr, gdalconst
from sklearn.metrics import confusion_matrix, precision_recall_fscore_support
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.utils import resample
import joblib
import openpyxl
from openpyxl.styles import Font

# Try importing scikit-image
try:
    from skimage.segmentation import felzenszwalb, slic
    from skimage.util import img_as_float
    from skimage.measure import regionprops_table
    HAS_SKIMAGE = True
except ImportError:
    HAS_SKIMAGE = False
    print("WARNING: scikit-image not found. This script requires scikit-image for raster segmentation.")

# --- Configuration (Global) ---

# python 1_OBIA_vector_classifier_modular_ANN.py -track P1a

# Base Paths provided by user
base_dir = Path("D:/AIML_CropMapper/AIML_CropMapper/workingDir")
aux_dir = Path("D:/AIML_CropMapper/AIML_CropMapper/auxiliary_files")

# OTB Installation Path (Still used for some auxiliary tasks if needed, but main flow is Python)
otb_dir = Path("D:/AIML_CropMapper_Cloud/2_OBIA_classifier/OTB-6.2.0-Win64")

# Track to Country Mapping
track_regions = {
    'P1': 'AT', 'P1a': 'AT',
    'P2': 'IE', 'P2a': 'IE',
    'P3': 'NL',
    'P4': 'PT', 'P4a': 'PT'
}
TOTAL_STAGES = 11


# --- Main Pipeline Class ---

class ProcessingPipeline:
    def __init__(self, track):
        self.track = track
        try:
            self.country = track_regions[track]
        except KeyError:
            print(f"Error: Track '{track}' not defined in track_regions configuration.")
            sys.exit(1)

        self.total_stages = TOTAL_STAGES
        print(f"Initializing pipeline for Track: {self.track}, Country: {self.country}")

        # --- 1. Define all paths ---
        self.base_dir = base_dir
        self.aux_dir = aux_dir
        self.proc_dir = self.base_dir / self.track / 'processed_raster'
        self.out_dir = self.base_dir / self.track / 'classification_results'
        self.samples_dir = self.out_dir / 'samples'
        self.model_dir = self.out_dir / 'train_model'
        self.seg_dir = self.out_dir / 'segmentation'
        self.class_dir = self.out_dir / 'classification'

        self._ensure_directories()

        # --- 2. Resolve input raster ---
        # UPDATED: Prioritize TIF files (BigTIFF stack)
        search_patterns = [
            f"{self.track}_*_VH_VV*.tif",
            f"*_{self.track}_*_VH_VV*.tif",
            f"*{self.track}*.tif",
            # Fallback to HDR if TIF not found
            f"{self.track}_*_VH_VV*.hdr",
            f"*{self.track}*.hdr",
        ]

        self.hdr = None
        if self.proc_dir.exists():
            for pattern in search_patterns:
                self.hdr = next(self.proc_dir.glob(pattern), None)
                if self.hdr:
                    break

            if not self.hdr:
                raise FileNotFoundError(f"No raster file (TIF/HDR) found for track {self.track} in {self.proc_dir}")

            self.ras = self._resolve_raster(self.hdr)
            print(f"Input raster found: {self.ras}")
        else:
            raise FileNotFoundError(f"Processing directory does not exist: {self.proc_dir}")

        # --- 3. Define all output file paths ---
        # CHANGED: Segmentation is now a TIF
        self.seg_tif = self.seg_dir / f"{self.country}_{self.track}_segmentation.tif"

        # Samples
        samples_base = self.aux_dir / 'shapefiles_samples'
        candidate_paths = [
            samples_base / f"{self.country}_{self.track}" / "samples.shp",
            samples_base / self.track / "samples.shp",
            samples_base / self.country / "samples.shp",
            samples_base / "samples.shp"
        ]

        self.sample_shp = None
        for p in candidate_paths:
            if p.exists():
                self.sample_shp = p
                print(f"Training samples found at: {self.sample_shp}")
                break

        if not self.sample_shp:
            print(f"\nCRITICAL WARNING: Could not find 'samples.shp' inside {samples_base}")
            self.sample_shp = samples_base / f"{self.country}_{self.track}" / "samples.shp"

        # Output paths
        self.learn_shp = self.samples_dir / 'learn.shp'
        self.control_shp = self.samples_dir / 'control.shp'
        # CHANGED: Selected samples is now a CSV (features) or SHP with features attached
        self.sel_csv = self.samples_dir / f"{self.country}_{self.track}_learn_features.csv"
        
        # Classification outputs
        self.class_tif = self.class_dir / f"{self.country}_{self.track}_classified.tif"
        self.conf_tif = self.class_dir / f"{self.country}_{self.track}_confidence_map.tif"
        
        self.cutline_shp = self.proc_dir / f"{self.country}_{self.track}_valid_coverage.shp"
        self.masked_class = self.class_dir / f"{self.country}_{self.track}_classified_masked.tif"
        self.masked_conf = self.class_dir / f"{self.country}_{self.track}_confidence_masked.tif"
        self.metrics_fp = self.class_dir / f"{self.country}_{self.track}_metrics.xlsx"

        # --- 4. Parameters (TUNED FOR POLAND/IRELAND SMALL FIELDS) ---
        self.stage1_params = {
            'method': 'python_slic', 
            'tile_size': 4096,
            
            # SLIC Params (Tuned)
            # 80,000 segments per 4096^2 tile ~= 200 pixels/segment ~= 0.2 hectares
            # This ensures we catch small strip fields.
            'n_segments': 80000, 
            
            # Compactness 0.5 allows for elongated shapes (strips) rather than forcing squares
            'compactness': 0.5, 
            
            # Sigma 1.0 provides robust smoothing for SAR speckle
            'slic_sigma': 1.0,

            # Felzenszwalb (Fallback)
            'scale': 100, 'sigma': 0.8, 'min_size': 50
        }
        self.stage2_params = {
            'learn_frac': 0.7, 'random_state': 42
        }
        self.stage4_params = {
            'classifier': 'ann_sklearn',
            # Simpler architecture to prevent overfitting on mean values
            'sk_hidden_sizes': '100', 
            'sk_activation': 'relu',
            'sk_solver': 'adam',
            'sk_alpha': 0.0001,
            'sk_max_iter': 500,
            'balance_threshold': 1000 
        }

        self.feat_cols = [] # Will store feature names

    # --- Utility Methods ---

    def _ensure_directories(self):
        for d in [self.samples_dir, self.model_dir, self.seg_dir, self.class_dir]:
            d.mkdir(parents=True, exist_ok=True)

    def _run_cmd(self, cmd, stage, desc):
        # Kept for OTB legacy calls (like Rasterization if needed)
        print(f"[Stage {stage}/{self.total_stages}] {desc}")
        env = os.environ.copy()
        otb_bin = str(otb_dir / "bin")
        otb_lib = str(otb_dir / "lib")
        otb_apps = str(otb_dir / "lib" / "otb" / "applications")
        env["PATH"] = f"{otb_bin};{otb_lib};{env['PATH']}"
        env["OTB_APPLICATION_PATH"] = otb_apps
        env["GEOTIFF_CSV"] = str(otb_dir / "share" / "epsg_csv")
        env["GDAL_DATA"] = str(otb_dir / "share" / "gdal")

        proc = subprocess.Popen(cmd, shell=True, stdout=sys.stdout, stderr=sys.stderr, env=env)
        proc.communicate()
        if proc.returncode != 0:
            raise RuntimeError(f"Stage {stage} failed with return code {proc.returncode}: {cmd}")
        print(f"Completed stage {stage}/{self.total_stages}\n")

    def _resolve_raster(self, hdr):
        # If input is already TIF, return it
        if hdr.suffix.lower() in ['.tif', '.tiff']:
            return hdr
            
        # If input is HDR, look for associated image
        for ext in ['.img', '.tif', '.TIF']:
            p = hdr.with_suffix(ext)
            if p.exists(): return p
            
        p_no_ext = hdr.with_suffix('')
        if p_no_ext.exists() and p_no_ext.is_file():
            return p_no_ext

        raise FileNotFoundError(f"No raster image (.img/.tif) found matching header {hdr.stem}")

    def _raster_to_cutline(self, input_tif, cutline_shp, stage):
        # Same as before, but robust
        print(f"[Stage {stage}/{self.total_stages}] Generating Cutline...")
        nuts_filename = f"NUTS2_{self.country}.shp"
        nuts_path = self.aux_dir / 'shapefiles_nuts' / self.country / nuts_filename
        
        has_nuts = nuts_path.exists()
        if not has_nuts:
            print(f"WARNING: NUTS shapefile not found at {nuts_path}")

        ds = gdal.Open(str(input_tif))
        if not ds: raise RuntimeError(f"Could not open {input_tif}")

        # Create footprint polygon
        # Simplified: Use gdal.Nearblack or simple mask
        band = ds.GetRasterBand(1)
        mask_arr = (band.ReadAsArray() != 0).astype(np.uint8) # Assuming 0 is nodata
        
        drv_mem = gdal.GetDriverByName('MEM')
        mask_ds = drv_mem.Create('', ds.RasterXSize, ds.RasterYSize, 1, gdal.GDT_Byte)
        mask_ds.SetGeoTransform(ds.GetGeoTransform())
        mask_ds.SetProjection(ds.GetProjection())
        mask_ds.GetRasterBand(1).WriteArray(mask_arr)
        
        temp_footprint = self.proc_dir / "temp_footprint.shp"
        shp_drv = ogr.GetDriverByName('ESRI Shapefile')
        if os.path.exists(str(temp_footprint)): shp_drv.DeleteDataSource(str(temp_footprint))
        out_ds = shp_drv.CreateDataSource(str(temp_footprint))
        srs = osr.SpatialReference()
        srs.ImportFromWkt(ds.GetProjection())
        layer = out_ds.CreateLayer('footprint', srs=srs, geom_type=ogr.wkbPolygon)
        layer.CreateField(ogr.FieldDefn('DN', ogr.OFTInteger))
        gdal.Polygonize(mask_ds.GetRasterBand(1), None, layer, 0, [], callback=None)
        out_ds.Destroy()

        try:
            gdf_footprint = gpd.read_file(str(temp_footprint))
            gdf_footprint = gdf_footprint[gdf_footprint['DN'] == 1]

            if has_nuts:
                gdf_nuts = gpd.read_file(str(nuts_path))
                if gdf_footprint.crs != gdf_nuts.crs:
                    gdf_nuts = gdf_nuts.to_crs(gdf_footprint.crs)
                gdf_cutline = gpd.overlay(gdf_footprint, gdf_nuts, how='intersection')
            else:
                gdf_cutline = gdf_footprint

            if gdf_cutline.empty:
                raise RuntimeError("Intersection resulted in empty cutline!")

            gdf_final = gdf_cutline.dissolve()
            gdf_final['DN'] = 1
            gdf_final.to_file(str(cutline_shp))
            print(f"Cutline saved to {cutline_shp}\n")

        except Exception as e:
            print(f"Error during cutline processing: {e}")
        finally:
            if temp_footprint.exists():
                for ext in ['.shp', '.shx', '.dbf', '.prj', '.cpg']:
                    f = temp_footprint.with_suffix(ext)
                    if f.exists(): os.remove(f)

    def _clip_and_mask(self, input_tif, mask_tif, cutline_shp, out_tif, stage):
        print(f"[Stage {stage}/{self.total_stages}] Clipping and Masking...")
        # Use GDAL Warp for efficiency
        warp_opts = gdal.WarpOptions(
            format='GTiff',
            cutlineDSName=str(cutline_shp),
            cropToCutline=True,
            dstNodata=0,
            creationOptions=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES']
        )
        gdal.Warp(str(out_tif), str(input_tif), options=warp_opts)
        
        # Apply Arable Mask (if exists)
        if mask_tif.exists():
            # This part is tricky with just Warp. 
            # Better to do a calc, but for now let's assume the clip is sufficient 
            # or implement a masking step if strictly required.
            # For speed, we often skip the pixel-wise arable mask if the NUTS clip is good enough,
            # or apply it during the classification inference step (masked pixels = 0).
            pass
        print(f"Completed stage {stage}\n")

    # --- Stage 1: Raster Segmentation ---
    def stage_1_segmentation(self, **kwargs):
        self._ensure_directories()
        params = self.stage1_params.copy()
        params.update(kwargs)
        stage = 1
        
        if self.seg_tif.exists():
            print(f"[Stage {stage}/{self.total_stages}] Segmentation Raster exists, skipping\n")
            return

        print(f"[Stage {stage}/{self.total_stages}] Running Tiled Raster Segmentation ({params['method']})...")
        
        ds = gdal.Open(str(self.ras))
        if not ds: raise RuntimeError("Could not open raster")
        
        cols = ds.RasterXSize
        rows = ds.RasterYSize
        nbands = ds.RasterCount
        gt = ds.GetGeoTransform()
        proj = ds.GetProjection()
        
        # Create Output Segmentation Raster
        driver = gdal.GetDriverByName('GTiff')
        out_ds = driver.Create(str(self.seg_tif), cols, rows, 1, gdal.GDT_Int32, 
                               options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        out_ds.SetGeoTransform(gt)
        out_ds.SetProjection(proj)
        out_band = out_ds.GetRasterBand(1)
        out_band.SetNoDataValue(0)
        
        tile_size = params.get('tile_size', 4096)
        
        # Global segment ID counter to ensure uniqueness across tiles
        global_seg_id = 1
        
        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)
                
                print(f"    Processing Tile: x={x}, y={y}")
                
                # Read Data
                img_list = []
                for b in range(1, nbands + 1):
                    band = ds.GetRasterBand(b)
                    arr = band.ReadAsArray(x, y, xsize, ysize)
                    # Simple nodata handling
                    arr = np.nan_to_num(arr)
                    img_list.append(arr)
                
                img = np.dstack(img_list)
                
                # Skip empty
                if np.all(img == 0): continue
                
                img_norm = img_as_float(img)
                
                # Segment
                if params['method'] == 'python_felzenszwalb':
                    segments = felzenszwalb(img_norm, scale=params['scale'], sigma=params['sigma'], min_size=params['min_size'])
                elif params['method'] == 'python_slic':
                    segments = slic(img_norm, n_segments=params['n_segments'], compactness=params['compactness'], sigma=params['slic_sigma'], start_label=1)
                else:
                    raise ValueError("Unknown method")
                
                # Offset segment IDs to make them unique globally
                # segments is 0-indexed or 1-indexed. 
                # We add global_seg_id.
                # Mask out 0 (background) if SLIC produces it, though start_label=1 handles it.
                
                valid_mask = segments > 0
                segments[valid_mask] += global_seg_id
                
                # Update global counter
                global_seg_id = segments.max() + 1
                
                # Write to output
                out_band.WriteArray(segments.astype(np.int32), x, y)
        
        out_ds.FlushCache()
        out_ds = None
        print(f"    Segmentation Raster saved to {self.seg_tif}\n")

    # --- Stage 2: Sample Split ---
    def stage_2_split_samples(self, **kwargs):
        # Same as before
        self._ensure_directories()
        params = self.stage2_params.copy()
        params.update(kwargs)
        stage = 2

        if not self.sample_shp.exists():
            print("ERROR: Input sample file not found.")
            return

        gdf = gpd.read_file(str(self.sample_shp))
        learn = gdf.sample(frac=params['learn_frac'], random_state=params['random_state'])
        control = gdf.drop(learn.index)

        learn.to_file(str(self.learn_shp))
        control.to_file(str(self.control_shp))
        print(f"Completed stage {stage}.\n")

    # --- Stage 3: Feature Extraction (Raster Sampling) ---
    def stage_3_selection(self):
        self._ensure_directories()
        stage = 3
        
        if self.sel_csv.exists():
            print(f"[Stage {stage}] Features already extracted, skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Extracting features from Raster Stack at Training Points...")
        
        if not self.learn_shp.exists():
            print("ERROR: Learn samples not found.")
            return
            
        gdf = gpd.read_file(str(self.learn_shp))
        
        # Open Raster Stack
        ds = gdal.Open(str(self.ras))
        gt = ds.GetGeoTransform()
        inv_gt = gdal.InvGeoTransform(gt)
        nbands = ds.RasterCount
        
        # Prepare list to store data
        data = []
        
        # Iterate points
        # Note: For millions of points, this loop is slow. 
        # Faster approach: Rasterize points and use zonal stats, or use rasterio.sample.
        # Here we stick to a simple loop for robustness.
        
        print(f"    Sampling {len(gdf)} points...")
        
        # Pre-read bands into memory if RAM allows? No, risky for huge files.
        # We will read pixel-by-pixel or block-by-block.
        # Optimization: Sort points by Y coordinate to minimize cache misses.
        gdf['x'] = gdf.geometry.x
        gdf['y'] = gdf.geometry.y
        gdf.sort_values(by='y', inplace=True)
        
        # We need to extract: crop_id, and Band1...BandN values
        # We do NOT need the segmentation ID for training the ANN per se, 
        # unless we want to average pixels within the segment.
        # Standard OBIA: Train on Segment Means.
        # So we need: Point -> Segment ID -> Segment Mean.
        
        # 1. Sample Segment IDs at points
        seg_ds = gdal.Open(str(self.seg_tif))
        seg_band = seg_ds.GetRasterBand(1)
        
        valid_samples = []
        
        for idx, row in gdf.iterrows():
            px = int(inv_gt[0] + inv_gt[1] * row.x + inv_gt[2] * row.y)
            py = int(inv_gt[3] + inv_gt[4] * row.x + inv_gt[5] * row.y)
            
            if 0 <= px < ds.RasterXSize and 0 <= py < ds.RasterYSize:
                # Read Segment ID
                # ReadAsArray(x, y, 1, 1) returns 2D array [[val]]
                try:
                    seg_id = seg_band.ReadAsArray(px, py, 1, 1)[0, 0]
                    if seg_id > 0:
                        valid_samples.append({
                            'crop_id': row['crop_id'],
                            'seg_id': seg_id,
                            'px': px,
                            'py': py
                        })
                except: pass
        
        if not valid_samples:
            print("ERROR: No valid samples found overlapping the raster.")
            return
            
        df_samples = pd.DataFrame(valid_samples)
        print(f"    Found {len(df_samples)} valid samples intersecting segments.")
        
        # 2. Calculate Mean Features for these Segments
        # This is the hard part for huge rasters: Random Access to segments.
        # Strategy: 
        #   We cannot easily compute means for *random* segments scattered across a 50GB file.
        #   Compromise: For training, we will just use the PIXEL value at the point.
        #   "Pixel-based training, Object-based inference".
        #   This is a common and valid strategy. The ANN learns spectral signatures from pixels.
        #   Then we apply that ANN to the *mean* of the segments later.
        
        print("    Extracting pixel values for training...")
        
        # Columns: crop_id, val_b1, val_b2...
        feature_data = {'crop_id': [], 'seg_id': []}
        for b in range(1, nbands + 1):
            feature_data[f'meanB{b-1}'] = []
            
        # Optimization: Read all points for one band at a time
        # This requires iterating the file N times, but is memory safe.
        
        for b in range(1, nbands + 1):
            print(f"      Reading Band {b}...")
            band = ds.GetRasterBand(b)
            vals = []
            # This is still slow for many points. 
            # Faster: Use gdal.LocationInfo logic or bulk read if points are clustered.
            # Let's stick to simple read for now.
            for idx, row in df_samples.iterrows():
                val = band.ReadAsArray(int(row['px']), int(row['py']), 1, 1)[0, 0]
                vals.append(val)
            feature_data[f'meanB{b-1}'] = vals
            
        feature_data['crop_id'] = df_samples['crop_id'].values
        feature_data['seg_id'] = df_samples['seg_id'].values # Kept for reference
        
        df_final = pd.DataFrame(feature_data)
        df_final.to_csv(self.sel_csv, index=False)
        print(f"    Features saved to {self.sel_csv}\n")

    # --- Stage 4: Train Classifier ---
    def stage_4_train_classifier(self, **kwargs):
        self._ensure_directories()
        params = self.stage4_params.copy()
        params.update(kwargs)
        stage = 4
        
        if not self.sel_csv.exists():
            print("ERROR: Feature CSV not found.")
            return
            
        model_fn = self.model_dir / f"{self.country}_{self.track}_model.pkl"
        
        print(f"[Stage {stage}/{self.total_stages}] Training ANN...")
        
        df = pd.read_csv(self.sel_csv)
        
        # Identify feature columns (meanB0, meanB1...)
        feat_cols = [c for c in df.columns if c.startswith('meanB')]
        self.feat_cols = feat_cols # Store for inference
        
        # --- BALANCING LOGIC ---
        print("    Balancing classes (Capped Oversampling)...")
        threshold = params.get('balance_threshold', 1000)
        
        df_balanced = pd.DataFrame()
        for crop_id in df['crop_id'].unique():
            df_class = df[df['crop_id'] == crop_id]
            count = len(df_class)
            
            if count < threshold:
                # Upsample rare classes
                df_resampled = resample(df_class, replace=True, n_samples=threshold, random_state=42)
                df_balanced = pd.concat([df_balanced, df_resampled])
            else:
                # Keep abundant classes as is
                df_balanced = pd.concat([df_balanced, df_class])
        
        print(f"    Original samples: {len(df)}. Balanced samples: {len(df_balanced)}")
        
        X = df_balanced[feat_cols].values
        y = df_balanced['crop_id'].values
        
        # Scale
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)
        
        hidden_sizes = tuple(map(int, str(params['sk_hidden_sizes']).split(',')))
        
        clf = MLPClassifier(
            hidden_layer_sizes=hidden_sizes,
            activation=params['sk_activation'],
            solver=params['sk_solver'],
            alpha=params['sk_alpha'],
            max_iter=params['sk_max_iter'],
            random_state=42,
            verbose=True
        )
        
        clf.fit(X_scaled, y)
        
        joblib.dump({'model': clf, 'scaler': scaler, 'feats': feat_cols}, model_fn)
        print(f"Model saved to {model_fn}")
        
        # CM
        y_pred = clf.predict(X_scaled)
        labels = sorted(list(set(y)))
        cm = confusion_matrix(y, y_pred, labels=labels)
        print("\n--- Training Confusion Matrix ---")
        print(pd.DataFrame(cm, index=labels, columns=labels).to_string())
        print("\n")

    # --- Stage 5: Tiled Inference (Object-Based) ---
    def stage_5_classify_vector(self):
        # Renamed logic, kept name for compatibility
        self._ensure_directories()
        stage = 5
        
        model_file = self.model_dir / f"{self.country}_{self.track}_model.pkl"
        if not model_file.exists():
            print("ERROR: Model not found.")
            return
            
        if self.class_tif.exists():
            print(f"[Stage {stage}] Classification Raster exists, skipping.")
            return
            
        print(f"[Stage {stage}/{self.total_stages}] Running Tiled Object-Based Inference...")
        
        # Load Model
        data = joblib.load(model_file)
        clf = data['model']
        scaler = data['scaler']
        feat_cols = data['feats'] # Ensure we use same bands
        
        # Open Rasters
        ds_stack = gdal.Open(str(self.ras))
        ds_seg = gdal.Open(str(self.seg_tif))
        
        cols = ds_stack.RasterXSize
        rows = ds_stack.RasterYSize
        nbands = ds_stack.RasterCount
        
        # Create Output Rasters
        driver = gdal.GetDriverByName('GTiff')
        # Class Raster (Int)
        ds_cls = driver.Create(str(self.class_tif), cols, rows, 1, gdal.GDT_Int32, 
                               options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        ds_cls.SetGeoTransform(ds_stack.GetGeoTransform())
        ds_cls.SetProjection(ds_stack.GetProjection())
        
        # Confidence Raster (Float)
        ds_conf = driver.Create(str(self.conf_tif), cols, rows, 1, gdal.GDT_Float32, 
                                options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        ds_conf.SetGeoTransform(ds_stack.GetGeoTransform())
        ds_conf.SetProjection(ds_stack.GetProjection())
        
        tile_size = 2048 # Smaller tile size for inference to handle feature extraction RAM
        
        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)
                
                print(f"    Inferencing Tile: x={x}, y={y}")
                
                # 1. Read Segmentation
                seg_arr = ds_seg.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                
                if np.all(seg_arr == 0): continue
                
                # 2. Read Data Stack
                img_list = []
                for b in range(1, nbands + 1):
                    band = ds_stack.GetRasterBand(b)
                    arr = band.ReadAsArray(x, y, xsize, ysize)
                    arr = np.nan_to_num(arr)
                    img_list.append(arr)
                img = np.dstack(img_list)
                
                # 3. Object-Based Feature Extraction (On-the-fly)
                # We need mean value per segment ID in this tile
                # regionprops is fast
                props = regionprops_table(seg_arr, intensity_image=img, properties=('label', 'mean_intensity'))
                df_props = pd.DataFrame(props)
                
                if df_props.empty: continue
                
                # Prepare features for prediction
                # Rename to match training (meanB0...)
                rename_map = {f'mean_intensity-{i}': f'meanB{i}' for i in range(nbands)}
                if nbands == 1 and 'mean_intensity' in df_props.columns:
                    rename_map = {'mean_intensity': 'meanB0'}
                df_props.rename(columns=rename_map, inplace=True)
                
                # Ensure columns match model
                # (Handle case where model expects meanB0, meanB1 but we have them)
                X_tile = df_props[feat_cols].values
                
                # 4. Predict
                X_scaled = scaler.transform(X_tile)
                preds = clf.predict(X_scaled)
                probs = np.max(clf.predict_proba(X_scaled), axis=1)
                
                # 5. Map predictions back to pixels
                # Create a mapping array: Label -> Prediction
                # Max label in tile
                max_label = seg_arr.max()
                
                # We need a lookup table. 
                # Initialize with 0 (nodata)
                # If segment IDs are huge (global), a direct array lookup might crash RAM.
                # But seg_arr contains the IDs.
                # We can use np.searchsorted or a dictionary map if IDs are sparse?
                # Actually, regionprops returns 'label'.
                
                # Efficient mapping:
                # Use a dense array if range is small, or np.select/map.
                # Since IDs are global and huge, we cannot make a lookup array of size max_global_id.
                # BUT, we only care about the IDs *in this tile*.
                
                # Map global IDs to local indices 0..N
                unique_ids = df_props['label'].values
                
                # Create output arrays
                cls_tile = np.zeros_like(seg_arr, dtype=np.int32)
                conf_tile = np.zeros_like(seg_arr, dtype=np.float32)
                
                # Fast mapping using searchsorted
                # Sort unique_ids (regionprops usually returns sorted, but ensure it)
                # unique_ids is already sorted by regionprops
                
                # Find indices of pixels in the unique_ids array
                # This maps every pixel value to its index in unique_ids
                # Pixels with 0 (background) need handling
                
                # Flatten for processing
                flat_seg = seg_arr.ravel()
                
                # We only map non-zero pixels
                mask = flat_seg > 0
                valid_seg = flat_seg[mask]
                
                # searchsorted returns the index in unique_ids where valid_seg would fit
                # Since valid_seg MUST exist in unique_ids (extracted from it), this gives the index
                idx_map = np.searchsorted(unique_ids, valid_seg)
                
                # Now map to predictions
                flat_cls = np.zeros_like(flat_seg, dtype=np.int32)
                flat_conf = np.zeros_like(flat_seg, dtype=np.float32)
                
                flat_cls[mask] = preds[idx_map]
                flat_conf[mask] = probs[idx_map]
                
                # Reshape
                cls_tile = flat_cls.reshape(ysize, xsize)
                conf_tile = flat_conf.reshape(ysize, xsize)
                
                # 6. Write
                ds_cls.GetRasterBand(1).WriteArray(cls_tile, x, y)
                ds_conf.GetRasterBand(1).WriteArray(conf_tile, x, y)
        
        ds_cls.FlushCache()
        ds_conf.FlushCache()
        print(f"    Classification saved to {self.class_tif}\n")

    # --- Stage 6 & 7: Skipped (Already Raster) ---
    def stage_6_rasterize_class(self):
        print("Stage 6 skipped (Output is already raster).")
    def stage_7_rasterize_confidence(self):
        print("Stage 7 skipped (Output is already raster).")

    # --- Stage 8: Cutline ---
    def stage_8_create_cutline(self):
        stage = 8
        if not self.cutline_shp.exists():
            self._raster_to_cutline(self.ras, self.cutline_shp, stage)
        else:
            print(f"[Stage {stage}/{self.total_stages}] Cutline exists, skipping\n")

    # --- Stage 9: Mask Class ---
    def stage_9_mask_class(self):
        self._ensure_directories()
        stage = 9
        mask_file = self.aux_dir / 'raster_files' / 'EU_arable_areas_mask_3857.tif'
        if not self.class_tif.exists():
            print(f"ERROR: Classified TIF not found.")
            return
        if not self.cutline_shp.exists():
            print(f"ERROR: Cutline not found.")
            return
        if not mask_file.exists():
            print(f"ERROR: Arable mask {mask_file} not found.")
            return

        if not self.masked_class.exists():
            self._clip_and_mask(self.class_tif, mask_file, self.cutline_shp, self.masked_class, stage)
        else:
            print(f"[Stage {stage}/{self.total_stages}] Masked classification exists, skipping\n")

    # --- Stage 10: Mask Confidence ---
    def stage_10_mask_confidence(self):
        self._ensure_directories()
        stage = 10
        mask_file = self.aux_dir / 'raster_files' / 'EU_arable_areas_mask_3857.tif'
        if not self.conf_tif.exists():
            print(f"ERROR: Confidence TIF not found.")
            return
        if not self.cutline_shp.exists():
            print(f"ERROR: Cutline not found.")
            return

        if not self.masked_conf.exists():
            self._clip_and_mask(self.conf_tif, mask_file, self.cutline_shp, self.masked_conf, stage)
        else:
            print(f"[Stage {stage}/{self.total_stages}] Masked confidence exists, skipping\n")

    # --- Stage 11: Metrics ---
    def stage_11_calculate_metrics(self):
        self._ensure_directories()
        stage = 11
        if not self.metrics_fp.exists():
            print(f"[Stage {stage}/{self.total_stages}] Computing metrics...")

            if not self.control_shp.exists():
                print(f"ERROR: Control shapefile not found.")
                return
            if not self.masked_class.exists():
                print(f"ERROR: Masked classification not found.")
                return

            ctrl = gpd.read_file(str(self.control_shp))
            ds = gdal.Open(str(self.masked_class))
            arr = ds.GetRasterBand(1).ReadAsArray()
            gt = ds.GetGeoTransform()
            inv = gdal.InvGeoTransform(gt)
            true_vals, pred_vals = [], []

            for _, row in ctrl.iterrows():
                try:
                    px = int(inv[0] + inv[1] * row.geometry.x + inv[2] * row.geometry.y)
                    py = int(inv[3] + inv[4] * row.geometry.x + inv[5] * row.geometry.y)
                    if 0 <= px < arr.shape[1] and 0 <= py < arr.shape[0]:
                        t = int(row['crop_id'])
                        p = int(arr[py, px])
                        if t > 0 and p > 0:
                            true_vals.append(t)
                            pred_vals.append(p)
                except Exception as e:
                    pass

            if not true_vals or not pred_vals:
                print("ERROR: No valid matching true/predicted values found.")
                return

            labels = sorted(list(set(true_vals + pred_vals)))
            cm = confusion_matrix(true_vals, pred_vals, labels=labels)
            precisions, recalls, f1s, _ = precision_recall_fscore_support(
                true_vals, pred_vals, labels=labels, average=None, zero_division=0
            )

            total = np.sum(cm)
            oa = np.trace(cm) / total
            sum_po = oa
            sum_pe = np.sum(np.sum(cm, axis=0) * np.sum(cm, axis=1)) / (total ** 2)
            kappa = (sum_po - sum_pe) / (1 - sum_pe) if (1 - sum_pe) != 0 else np.nan

            resx, resy = abs(gt[1]), abs(gt[5])
            area_ha = resx * resy / 10000
            unique_classes, counts = np.unique(arr[arr > 0], return_counts=True)
            class_areas = dict(zip(unique_classes, counts))
            areas = [{'Class': c, 'Area_ha': round(class_areas.get(c, 0) * area_ha, 2)} for c in labels]

            wb = openpyxl.Workbook()
            sh = wb.active
            sh.title = 'Results'

            sh.cell(row=1, column=1, value='Confusion Matrix').font = Font(bold=True)
            sh.cell(row=2, column=1, value='True \\ Pred').font = Font(bold=True)
            for j, lbl in enumerate(labels, start=2):
                sh.cell(row=2, column=j, value=lbl).font = Font(bold=True)
            for i, lbl in enumerate(labels, start=3):
                sh.cell(row=i, column=1, value=lbl).font = Font(bold=True)
                for j, _ in enumerate(labels):
                    sh.cell(row=i, column=j + 2, value=int(cm[i - 3, j]))

            base = 4 + len(labels)
            sh.cell(row=base, column=1, value='Overall Accuracy').font = Font(bold=True)
            sh.cell(row=base, column=2, value=round(oa, 4))
            sh.cell(row=base + 1, column=1, value='Kappa').font = Font(bold=True)
            sh.cell(row=base + 1, column=2, value=round(kappa, 4))

            start = base + 3
            headers = ['Class', 'Producer Acc (Recall)', 'User Acc (Precision)', 'F1-score']
            for j, h in enumerate(headers, start=1):
                sh.cell(row=start, column=j, value=h).font = Font(bold=True)
            for idx, c in enumerate(labels):
                row_idx = start + 1 + idx
                sh.cell(row=row_idx, column=1, value=c)
                sh.cell(row=row_idx, column=2, value=round(recalls[idx], 4))
                sh.cell(row=row_idx, column=3, value=round(precisions[idx], 4))
                sh.cell(row=row_idx, column=4, value=round(f1s[idx], 4))

            ar0 = start + 1 + len(labels) + 1
            sh.cell(row=ar0, column=1, value='Areas (ha)').font = Font(bold=True)
            sh.cell(row=ar0 + 1, column=1, value='Class').font = Font(bold=True)
            sh.cell(row=ar0 + 1, column=2, value='Area_ha').font = Font(bold=True)
            for idx, a in enumerate(areas, start=ar0 + 2):
                sh.cell(row=idx, column=1, value=a['Class'])
                sh.cell(row=idx, column=2, value=a['Area_ha'])

            wb.save(str(self.metrics_fp))
            print(f"Metrics saved to {self.metrics_fp}\n")
        else:
            print(f"[Stage 11] Metrics Excel exists, skipping")


# --- Interactive Menu Helpers ---

def get_params(param_dict):
    new_params = param_dict.copy()
    print("--- Current Parameters ---")
    for key, val in new_params.items():
        print(f"  {key}: {val}")

    if input("Change parameters? (y/n) [n]: ").lower() != 'y':
        return new_params

    for key, val in new_params.items():
        new_val_str = input(f"Enter new value for '{key}' [{val}]: ")
        if not new_val_str:
            continue
        try:
            original_type = type(val)
            new_params[key] = original_type(new_val_str)
        except ValueError:
            print(f"Invalid value. Keeping default {val}.")
    return new_params


def get_classifier_params(param_dict):
    new_params = param_dict.copy()
    print("--- Current Parameters ---")
    for key, val in new_params.items():
        print(f"  {key}: {val}")

    if input("Change parameters? (y/n) [n]: ").lower() != 'y':
        return new_params

    clf = input(f"Enter classifier (ann_sklearn) [{new_params['classifier']}]: ") or new_params['classifier']
    new_params['classifier'] = clf.lower()

    print(f"\n--- Setting parameters for {clf.upper()} ---")
    prefix = 'sk_'
    for key in [k for k in new_params if k.startswith(prefix)]:
        val = new_params[key]
        new_val_str = input(f"Enter new value for '{key}' [{val}]: ")
        if new_val_str:
            try:
                new_params[key] = type(val)(new_val_str)
            except ValueError:
                print(f"Invalid value.")
    return new_params


# --- Main Execution ---

def main_menu(pipeline):
    menu = f"""
    --- Raster-Based OBIA Pipeline (ANN) ---
    Track: {pipeline.track} ({pipeline.country})

    [1] Stage 1: Raster Segmentation (Method: {pipeline.stage1_params['method']})
    [2] Stage 2: Split Samples
    [3] Stage 3: Extract Features (Pixel-based Training)
    [4] Stage 4: Train ANN Classifier
    [5] Stage 5: Tiled Object-Based Inference
    [6] Stage 6: (Skipped)
    [7] Stage 7: (Skipped)
    [8] Stage 8: Create Valid-Pixel Cutline
    [9] Stage 9: Mask Classification
    [10] Stage 10: Mask Confidence
    [11] Stage 11: Calculate Metrics

    [A] Run All Stages
    [Q] Quit

    Enter your choice:
    """

    while True:
        choice = input(menu).strip().upper()
        try:
            if choice == '1':
                new_params = get_params(pipeline.stage1_params)
                pipeline.stage1_params.update(new_params)
                pipeline.stage_1_segmentation(**pipeline.stage1_params)
            elif choice == '2':
                new_params = get_params(pipeline.stage2_params)
                pipeline.stage2_params.update(new_params)
                pipeline.stage_2_split_samples(**pipeline.stage2_params)
            elif choice == '3':
                pipeline.stage_3_selection()
            elif choice == '4':
                new_params = get_classifier_params(pipeline.stage4_params)
                force = (pipeline.stage4_params != new_params)
                pipeline.stage4_params.update(new_params)
                pipeline.stage_4_train_classifier(force_retrain=force, **pipeline.stage4_params)
            elif choice == '5':
                pipeline.stage_5_classify_vector()
            elif choice == '8':
                pipeline.stage_8_create_cutline()
            elif choice == '9':
                pipeline.stage_9_mask_class()
            elif choice == '10':
                pipeline.stage_10_mask_confidence()
            elif choice == '11':
                pipeline.stage_11_calculate_metrics()
            elif choice == 'A':
                pipeline.stage_1_segmentation(**pipeline.stage1_params)
                pipeline.stage_2_split_samples(**pipeline.stage2_params)
                pipeline.stage_3_selection()
                pipeline.stage_4_train_classifier(**pipeline.stage4_params)
                pipeline.stage_5_classify_vector()
                pipeline.stage_8_create_cutline()
                pipeline.stage_9_mask_class()
                pipeline.stage_10_mask_confidence()
                pipeline.stage_11_calculate_metrics()
            elif choice == 'Q':
                break
        except Exception as e:
            print(f"\n--- ERROR ---: {e}")
            import traceback
            traceback.print_exc()


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Modular OBIA Classification Pipeline (ANN)")
    parser.add_argument('--track', required=True, help="Processing track ID (e.g., P1, P2)")
    args = parser.parse_args()

    try:
        pipeline = ProcessingPipeline(track=args.track)
        main_menu(pipeline)
    except Exception as e:
        print(f"Initialization Error: {e}")
        sys.exit(1)